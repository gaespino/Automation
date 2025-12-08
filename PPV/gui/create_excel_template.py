"""
Create a sample Excel template for Experiment Builder imports

This script generates a sample .xlsx file demonstrating the expected
format for importing experiments into the Experiment Builder.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

def create_sample_template():
    """Create a sample Excel template with example experiments"""
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Sample Experiment 1: Basic Loop Test
    ws1 = wb.create_sheet("Loop_Test_Example")
    
    fields_exp1 = [
        ("Field Name", "Value"),
        ("Experiment", "Enabled"),
        ("Test Name", "Loop_Test_Example"),
        ("Test Mode", "Mesh"),
        ("Test Type", "Loops"),
        ("Visual ID", "75857N7H00175"),
        ("Bucket", "PPV"),
        ("COM Port", "8"),
        ("IP Address", "192.168.0.2"),
        ("TTL Folder", "C:\\TTL\\Tests"),
        ("Scripts File", "test_script.ttl"),
        ("Pass String", "Test Complete"),
        ("Fail String", "Test Failed"),
        ("Content", "Linux"),
        ("Test Number", "1"),
        ("Test Time", "30"),
        ("Reset", "TRUE"),
        ("Reset on PASS", "FALSE"),
        ("FastBoot", "TRUE"),
        ("Configuration (Mask)", "FirstPass"),
        ("Voltage Type", "vbump"),
        ("Voltage IA", "1.0"),
        ("Voltage CFC", "0.9"),
        ("Frequency IA", "2000"),
        ("Frequency CFC", "1800"),
        ("Loops", "10"),
        ("Linux Path", "C:\\Linux\\Content"),
        ("Linux Pre Command", "echo Starting test"),
        ("Linux Post Command", "echo Test completed"),
        ("Linux Pass String", "PASS"),
        ("Linux Fail String", "FAIL"),
        ("Linux Content Wait Time", "15")
    ]
    
    for row_idx, (field, value) in enumerate(fields_exp1, 1):
        ws1.cell(row_idx, 1, field)
        ws1.cell(row_idx, 2, value)
        
        # Style header row
        if row_idx == 1:
            ws1.cell(row_idx, 1).font = Font(bold=True)
            ws1.cell(row_idx, 2).font = Font(bold=True)
            ws1.cell(row_idx, 1).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws1.cell(row_idx, 2).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws1.cell(row_idx, 1).font = Font(color="FFFFFF", bold=True)
            ws1.cell(row_idx, 2).font = Font(color="FFFFFF", bold=True)
    
    # Adjust column widths
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 40
    
    # Sample Experiment 2: Voltage Sweep
    ws2 = wb.create_sheet("Voltage_Sweep_Example")
    
    fields_exp2 = [
        ("Field Name", "Value"),
        ("Experiment", "Enabled"),
        ("Test Name", "Voltage_Sweep_Example"),
        ("Test Mode", "Slice"),
        ("Test Type", "Sweep"),
        ("Visual ID", "75857N7H00176"),
        ("Bucket", "PPV_Characterization"),
        ("COM Port", "9"),
        ("IP Address", "192.168.0.3"),
        ("TTL Folder", "C:\\TTL\\Sweep"),
        ("Scripts File", "sweep_test.ttl"),
        ("Pass String", "Sweep Complete"),
        ("Fail String", "Sweep Failed"),
        ("Content", "Dragon"),
        ("Test Number", "1"),
        ("Test Time", "60"),
        ("Reset", "TRUE"),
        ("Reset on PASS", "TRUE"),
        ("FastBoot", "FALSE"),
        ("Voltage Type", "fixed"),
        ("Type", "Voltage"),
        ("Domain", "IA"),
        ("Start", "0.8"),
        ("End", "1.2"),
        ("Steps", "0.05"),
        ("Dragon Content Path", "C:\\Dragon\\Content"),
        ("Dragon Pre Command", "load_config"),
        ("Dragon Post Command", "save_results"),
        ("ULX Path", "C:\\ULX"),
        ("ULX CPU", "GNR")
    ]
    
    for row_idx, (field, value) in enumerate(fields_exp2, 1):
        ws2.cell(row_idx, 1, field)
        ws2.cell(row_idx, 2, value)
        
        if row_idx == 1:
            ws2.cell(row_idx, 1).font = Font(bold=True)
            ws2.cell(row_idx, 2).font = Font(bold=True)
            ws2.cell(row_idx, 1).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws2.cell(row_idx, 2).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws2.cell(row_idx, 1).font = Font(color="FFFFFF", bold=True)
            ws2.cell(row_idx, 2).font = Font(color="FFFFFF", bold=True)
    
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 40
    
    # Sample Experiment 3: Shmoo Test
    ws3 = wb.create_sheet("Shmoo_Test_Example")
    
    fields_exp3 = [
        ("Field Name", "Value"),
        ("Experiment", "Disabled"),
        ("Test Name", "Shmoo_Test_Example"),
        ("Test Mode", "Mesh"),
        ("Test Type", "Shmoo"),
        ("Visual ID", "75857N7H00177"),
        ("Bucket", "PPV_Shmoo"),
        ("COM Port", "10"),
        ("IP Address", "192.168.0.4"),
        ("TTL Folder", "C:\\TTL\\Shmoo"),
        ("Scripts File", ""),
        ("Pass String", "SHMOO PASS"),
        ("Fail String", "SHMOO FAIL"),
        ("Content", "PYSVConsole"),
        ("Test Number", "1"),
        ("Test Time", "120"),
        ("Reset", "FALSE"),
        ("Reset on PASS", "FALSE"),
        ("FastBoot", "TRUE"),
        ("Voltage Type", "ppvc"),
        ("ShmooFile", "C:\\Shmoo\\config.json"),
        ("ShmooLabel", "VF_Shmoo_IA_CFC"),
        ("Check Core", "0"),
        ("Stop on Fail", "TRUE")
    ]
    
    for row_idx, (field, value) in enumerate(fields_exp3, 1):
        ws3.cell(row_idx, 1, field)
        ws3.cell(row_idx, 2, value)
        
        if row_idx == 1:
            ws3.cell(row_idx, 1).font = Font(bold=True)
            ws3.cell(row_idx, 2).font = Font(bold=True)
            ws3.cell(row_idx, 1).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws3.cell(row_idx, 2).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws3.cell(row_idx, 1).font = Font(color="FFFFFF", bold=True)
            ws3.cell(row_idx, 2).font = Font(color="FFFFFF", bold=True)
    
    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 40
    
    # Instructions sheet
    ws_inst = wb.create_sheet("Instructions", 0)
    
    instructions = [
        ("PPV Experiment Builder - Excel Template", ""),
        ("", ""),
        ("How to Use This Template:", ""),
        ("1. Each sheet represents ONE experiment", ""),
        ("2. Sheet name becomes the experiment name in the JSON output", ""),
        ("3. Column A contains field names (must match Control Panel fields)", ""),
        ("4. Column B contains the field values", ""),
        ("5. The header row (Field Name | Value) is optional and will be skipped", ""),
        ("", ""),
        ("To Create Your Own Experiment:", ""),
        ("1. Duplicate an example sheet (right-click > Move or Copy)", ""),
        ("2. Rename the sheet to your desired experiment name", ""),
        ("3. Update the values in Column B", ""),
        ("4. Delete any fields you don't need", ""),
        ("5. Save the Excel file", ""),
        ("6. Import into Experiment Builder", ""),
        ("", ""),
        ("Supported Test Types:", ""),
        ("- Loops: Repeat test multiple times", ""),
        ("- Sweep: Single parameter sweep (voltage or frequency)", ""),
        ("- Shmoo: 2D characterization test", ""),
        ("", ""),
        ("Test Modes:", ""),
        ("- Mesh: Standard mesh testing", ""),
        ("- Slice: Slice-based testing", ""),
        ("", ""),
        ("Content Types:", ""),
        ("- Linux: Linux-based content", ""),
        ("- Dragon: Dragon content execution", ""),
        ("- PYSVConsole: Python SV console", ""),
        ("", ""),
        ("Boolean Fields:", ""),
        ("Use TRUE or FALSE (case insensitive)", ""),
        ("Examples: Reset, Reset on PASS, FastBoot, Stop on Fail", ""),
        ("", ""),
        ("Important Notes:", ""),
        ("- COM Port must be a number (0-256)", ""),
        ("- IP Address must be valid IPv4 format (e.g., 192.168.0.2)", ""),
        ("- Voltage and Frequency values should be numeric", ""),
        ("- File paths should use backslashes (\\) on Windows", ""),
        ("- Empty fields will use default values", ""),
        ("", ""),
        ("Example Experiments Included:", ""),
        ("1. Loop_Test_Example - Basic loop test with Linux content", ""),
        ("2. Voltage_Sweep_Example - Voltage sweep with Dragon content", ""),
        ("3. Shmoo_Test_Example - Shmoo test (disabled by default)", ""),
    ]
    
    for row_idx, (text, _) in enumerate(instructions, 1):
        cell = ws_inst.cell(row_idx, 1, text)
        if row_idx == 1:
            cell.font = Font(size=16, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        elif text and not text.startswith(" "):
            if any(text.startswith(prefix) for prefix in ["How to", "To Create", "Supported", "Test Modes", "Content Types", "Boolean", "Important", "Example"]):
                cell.font = Font(size=12, bold=True, color="203764")
    
    ws_inst.column_dimensions['A'].width = 80
    
    # Save workbook
    output_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(output_dir, "Experiment_Template_Sample.xlsx")
    
    wb.save(output_path)
    
    print(f"Sample Excel template created: {output_path}")
    return output_path

if __name__ == "__main__":
    create_sample_template()
