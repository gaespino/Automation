"""
PPV Experiment Builder - Create JSON configurations for Debug Framework Control Panel

This tool provides a user-friendly interface to create and edit experiment configurations
that can be consumed by the Debug Framework Control Panel.
"""

import sys
import os
import json
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext
from typing import Dict, Any, List
import openpyxl
from datetime import datetime

# Add parent directory for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


class ExperimentBuilderGUI:
    """Main GUI for building and managing experiment configurations"""
    
    def __init__(self, parent=None):
        """Initialize the Experiment Builder GUI"""
        if parent is None:
            self.root = tk.Tk()
            self.is_standalone = True
        else:
            self.root = tk.Toplevel(parent)
            self.is_standalone = False
            
        self.root.title("PPV Experiment Builder - Control Panel Config Generator")
        self.root.geometry("1400x900")
        
        # Data structures
        self.experiments = {}
        self.current_experiment = None
        self.config_template = self.load_config_template()
        self.field_widgets = {}
        
        # Setup UI
        self.setup_styles()
        self.create_main_layout()
        self.populate_default_template()
        
        if self.is_standalone:
            self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
            
    def setup_styles(self):
        """Configure ttk styles for the application"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Configure colors
        self.colors = {
            'primary': '#2c3e50',
            'secondary': '#34495e',
            'accent': '#3498db',
            'success': '#27ae60',
            'warning': '#f39c12',
            'danger': '#e74c3c',
            'light': '#ecf0f1',
            'dark': '#2c3e50'
        }
        
        # Button styles
        style.configure('Primary.TButton', 
                       background=self.colors['accent'],
                       foreground='white',
                       borderwidth=0,
                       focuscolor='none',
                       padding=10)
        
        style.configure('Success.TButton',
                       background=self.colors['success'],
                       foreground='white',
                       borderwidth=0,
                       padding=10)
        
        style.configure('Danger.TButton',
                       background=self.colors['danger'],
                       foreground='white',
                       borderwidth=0,
                       padding=10)
        
        # Frame styles
        style.configure('Card.TFrame', background='white', relief='raised')
        style.configure('Header.TFrame', background=self.colors['primary'])
        
    def load_config_template(self):
        """Load the configuration template from GNRControlPanelConfig.json"""
        # Try to find the config file
        possible_paths = [
            os.path.join(os.path.dirname(__file__), '..', '..', 'S2T', 'BASELINE', 'DebugFramework', 'UI', 'GNRControlPanelConfig.json'),
            os.path.join(os.path.dirname(__file__), '..', '..', 'S2T', 'BASELINE_DMR', 'DebugFramework', 'UI', 'GNRControlPanelConfig.json'),
        ]
        
        for path in possible_paths:
            if os.path.exists(path):
                try:
                    with open(path, 'r') as f:
                        return json.load(f)
                except Exception as e:
                    print(f"Error loading config from {path}: {e}")
                    
        # Fallback default template
        return self.get_default_template()
    
    def get_default_template(self):
        """Return a default configuration template if file not found"""
        return {
            "data_types": {
                "Experiment": ["str"],
                "Test Name": ["str"],
                "Test Mode": ["str"],
                "Test Type": ["str"],
                "Visual ID": ["str"],
                "Bucket": ["str"],
                "COM Port": ["int"],
                "IP Address": ["str"],
                "TTL Folder": ["str"],
                "Scripts File": ["str"],
                "Pass String": ["str"],
                "Fail String": ["str"],
                "Content": ["str"],
                "Test Number": ["int"],
                "Test Time": ["int"],
                "Reset": ["bool"],
                "Reset on PASS": ["bool"],
                "FastBoot": ["bool"],
                "Core License": ["str"],
                "600W Unit": ["bool"],
                "Pseudo Config": ["bool"],
                "Post Process": ["str"],
                "Configuration (Mask)": ["str"],
                "Boot Breakpoint": ["str"],
                "Disable 2 Cores": ["str"],
                "Check Core": ["int"],
                "Voltage Type": ["str"],
                "Voltage IA": ["float"],
                "Voltage CFC": ["float"],
                "Frequency IA": ["int"],
                "Frequency CFC": ["int"],
                "Loops": ["int"],
                "Type": ["str"],
                "Domain": ["str"],
                "Start": ["int", "float"],
                "End": ["int", "float"],
                "Steps": ["int", "float"],
                "ShmooFile": ["str"],
                "ShmooLabel": ["str"],
                "Linux Path": ["str"],
                "Linux Pre Command": ["str"],
                "Linux Post Command": ["str"],
                "Linux Pass String": ["str"],
                "Linux Fail String": ["str"],
                "Linux Content Wait Time": ["int"],
                "Startup Linux": ["str"],
                "Linux Content Line 0": ["str"],
                "Linux Content Line 1": ["str"],
                "Linux Content Line 2": ["str"],
                "Linux Content Line 3": ["str"],
                "Linux Content Line 4": ["str"],
                "Linux Content Line 5": ["str"],
                "Linux Content Line 6": ["str"],
                "Linux Content Line 7": ["str"],
                "Linux Content Line 8": ["str"],
                "Linux Content Line 9": ["str"],
                "Dragon Pre Command": ["str"],
                "Dragon Post Command": ["str"],
                "Startup Dragon": ["str"],
                "ULX Path": ["str"],
                "ULX CPU": ["str"],
                "Product Chop": ["str"],
                "VVAR0": ["str"],
                "VVAR1": ["str"],
                "VVAR2": ["str"],
                "VVAR3": ["str"],
                "VVAR_EXTRA": ["str"],
                "Dragon Content Path": ["str"],
                "Dragon Content Line": ["str"],
                "Stop on Fail": ["bool"],
                "Merlin Name": ["str"],
                "Merlin Drive": ["str"],
                "Merlin Path": ["str"]
            },
            "TEST_MODES": ["Mesh", "Slice"],
            "TEST_TYPES": ["Loops", "Sweep", "Shmoo"],
            "CONTENT_OPTIONS": ["Linux", "Dragon", "PYSVConsole"],
            "VOLTAGE_TYPES": ["vbump", "fixed", "ppvc"],
            "TYPES": ["Voltage", "Frequency"],
            "DOMAINS": ["CFC", "IA"]
        }
    
    def create_main_layout(self):
        """Create the main application layout"""
        # Header
        header_frame = tk.Frame(self.root, bg=self.colors['primary'], height=80)
        header_frame.pack(fill='x', side='top')
        header_frame.pack_propagate(False)
        
        title_label = tk.Label(header_frame, 
                              text="PPV Experiment Builder",
                              font=('Segoe UI', 20, 'bold'),
                              bg=self.colors['primary'],
                              fg='white')
        title_label.pack(pady=20)
        
        # Main container with paned window
        main_paned = tk.PanedWindow(self.root, orient=tk.HORIZONTAL, sashwidth=5)
        main_paned.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Left panel - Experiment list
        self.create_left_panel(main_paned)
        
        # Right panel - Experiment editor
        self.create_right_panel(main_paned)
        
        # Bottom panel - Action buttons
        self.create_bottom_panel()
        
    def create_left_panel(self, parent):
        """Create the left panel with experiment list"""
        left_frame = tk.Frame(parent, bg='white', width=300)
        parent.add(left_frame)
        
        # Toolbar
        toolbar = tk.Frame(left_frame, bg=self.colors['light'], height=50)
        toolbar.pack(fill='x', side='top')
        
        tk.Label(toolbar, text="Experiments", font=('Segoe UI', 12, 'bold'),
                bg=self.colors['light']).pack(side='left', padx=10, pady=10)
        
        # Buttons frame
        btn_frame = tk.Frame(toolbar, bg=self.colors['light'])
        btn_frame.pack(side='right', padx=5)
        
        tk.Button(btn_frame, text="+", font=('Segoe UI', 12, 'bold'),
                 bg=self.colors['success'], fg='white',
                 command=self.add_new_experiment,
                 width=3, relief='flat').pack(side='left', padx=2)
        
        tk.Button(btn_frame, text="-", font=('Segoe UI', 12, 'bold'),
                 bg=self.colors['danger'], fg='white',
                 command=self.delete_experiment,
                 width=3, relief='flat').pack(side='left', padx=2)
        
        tk.Button(btn_frame, text="📋", font=('Segoe UI', 10),
                 bg=self.colors['accent'], fg='white',
                 command=self.duplicate_experiment,
                 width=3, relief='flat').pack(side='left', padx=2)
        
        # Search box
        search_frame = tk.Frame(left_frame, bg='white')
        search_frame.pack(fill='x', padx=5, pady=5)
        
        tk.Label(search_frame, text="Search:", bg='white').pack(side='left')
        self.search_var = tk.StringVar()
        self.search_var.trace('w', self.filter_experiments)
        search_entry = tk.Entry(search_frame, textvariable=self.search_var)
        search_entry.pack(side='left', fill='x', expand=True, padx=5)
        
        # Experiment listbox with scrollbar
        list_frame = tk.Frame(left_frame, bg='white')
        list_frame.pack(fill='both', expand=True, padx=5, pady=5)
        
        scrollbar = tk.Scrollbar(list_frame)
        scrollbar.pack(side='right', fill='y')
        
        self.exp_listbox = tk.Listbox(list_frame, 
                                      yscrollcommand=scrollbar.set,
                                      font=('Segoe UI', 10),
                                      selectmode='single',
                                      activestyle='dotbox')
        self.exp_listbox.pack(side='left', fill='both', expand=True)
        scrollbar.config(command=self.exp_listbox.yview)
        
        self.exp_listbox.bind('<<ListboxSelect>>', self.on_experiment_select)
        
    def create_right_panel(self, parent):
        """Create the right panel with experiment editor"""
        right_frame = tk.Frame(parent, bg='white')
        parent.add(right_frame)
        
        # Notebook for tabbed interface
        self.notebook = ttk.Notebook(right_frame)
        self.notebook.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Create tabs
        self.create_basic_tab()
        self.create_test_config_tab()
        self.create_voltage_freq_tab()
        self.create_content_tab()
        self.create_advanced_tab()
        self.create_json_preview_tab()
        
    def create_basic_tab(self):
        """Create the Basic Information tab"""
        tab = tk.Frame(self.notebook, bg='white')
        self.notebook.add(tab, text='Basic Info')
        
        # Scrollable frame
        canvas = tk.Canvas(tab, bg='white')
        scrollbar = tk.Scrollbar(tab, orient='vertical', command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='white')
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # Fields
        self.create_field(scrollable_frame, "Experiment", "str", 0, 
                         options=["Enabled", "Disabled"],
                         description="Enable or disable this experiment")
        self.create_field(scrollable_frame, "Test Name", "str", 1,
                         description="Unique name for this test")
        self.create_field(scrollable_frame, "Test Mode", "str", 2,
                         options=self.config_template.get("TEST_MODES", ["Mesh", "Slice"]),
                         description="Test execution mode")
        self.create_field(scrollable_frame, "Test Type", "str", 3,
                         options=self.config_template.get("TEST_TYPES", ["Loops", "Sweep", "Shmoo"]),
                         description="Type of test to execute")
        self.create_field(scrollable_frame, "Visual ID", "str", 4,
                         description="Unit visual identifier")
        self.create_field(scrollable_frame, "Bucket", "str", 5,
                         description="Test bucket classification")
        
    def create_test_config_tab(self):
        """Create the Test Configuration tab"""
        tab = tk.Frame(self.notebook, bg='white')
        self.notebook.add(tab, text='Test Config')
        
        canvas = tk.Canvas(tab, bg='white')
        scrollbar = tk.Scrollbar(tab, orient='vertical', command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='white')
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # Fields
        self.create_field(scrollable_frame, "COM Port", "int", 0,
                         description="Serial COM port number")
        self.create_field(scrollable_frame, "IP Address", "str", 1,
                         description="Network IP address")
        self.create_field(scrollable_frame, "TTL Folder", "str", 2, browse=True,
                         description="Path to TTL files folder")
        self.create_field(scrollable_frame, "Scripts File", "str", 3, browse=True,
                         description="Path to test scripts file")
        self.create_field(scrollable_frame, "Pass String", "str", 4,
                         description="String indicating test pass")
        self.create_field(scrollable_frame, "Fail String", "str", 5,
                         description="String indicating test failure")
        self.create_field(scrollable_frame, "Test Number", "int", 6,
                         description="Number of test iterations")
        self.create_field(scrollable_frame, "Test Time", "int", 7,
                         description="Test duration in seconds")
        self.create_field(scrollable_frame, "Loops", "int", 8,
                         description="Number of test loops")
        self.create_field(scrollable_frame, "Reset", "bool", 9,
                         description="Reset system between tests")
        self.create_field(scrollable_frame, "Reset on PASS", "bool", 10,
                         description="Reset only on pass")
        self.create_field(scrollable_frame, "FastBoot", "bool", 11,
                         description="Enable fast boot mode")
        self.create_field(scrollable_frame, "Core License", "str", 12,
                         description="Core license (leave blank if not used)")
        self.create_field(scrollable_frame, "600W Unit", "bool", 13,
                         description="Set to True for 600W units")
        self.create_field(scrollable_frame, "Pseudo Config", "bool", 14,
                         description="Set if using Pseudo Mesh Content to disable HT")
        self.create_field(scrollable_frame, "Post Process", "str", 15, browse=True,
                         description="Script file to be loaded after test execution")
        
    def create_voltage_freq_tab(self):
        """Create the Voltage & Frequency tab"""
        tab = tk.Frame(self.notebook, bg='white')
        self.notebook.add(tab, text='Voltage/Freq')
        
        canvas = tk.Canvas(tab, bg='white')
        scrollbar = tk.Scrollbar(tab, orient='vertical', command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='white')
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # Fields
        self.create_field(scrollable_frame, "Voltage Type", "str", 0,
                         options=self.config_template.get("VOLTAGE_TYPES", ["vbump", "fixed", "ppvc"]),
                         description="Voltage control type")
        self.create_field(scrollable_frame, "Voltage IA", "float", 1,
                         description="IA domain voltage")
        self.create_field(scrollable_frame, "Voltage CFC", "float", 2,
                         description="CFC domain voltage")
        self.create_field(scrollable_frame, "Frequency IA", "int", 3,
                         description="IA domain frequency (MHz)")
        self.create_field(scrollable_frame, "Frequency CFC", "int", 4,
                         description="CFC domain frequency (MHz)")
        
        # Sweep/Shmoo specific fields
        tk.Label(scrollable_frame, text="Sweep/Shmoo Configuration",
                font=('Segoe UI', 11, 'bold'), bg='white').grid(
                    row=5, column=0, columnspan=3, sticky='w', padx=10, pady=(20, 10))
        
        self.create_field(scrollable_frame, "Type", "str", 6,
                         options=self.config_template.get("TYPES", ["Voltage", "Frequency"]),
                         description="Parameter to sweep")
        self.create_field(scrollable_frame, "Domain", "str", 7,
                         options=self.config_template.get("DOMAINS", ["CFC", "IA"]),
                         description="Domain to target")
        self.create_field(scrollable_frame, "Start", "float", 8,
                         description="Starting value")
        self.create_field(scrollable_frame, "End", "float", 9,
                         description="Ending value")
        self.create_field(scrollable_frame, "Steps", "float", 10,
                         description="Step size or count")
        self.create_field(scrollable_frame, "ShmooFile", "str", 11, browse=True,
                         description="Path to shmoo configuration file")
        self.create_field(scrollable_frame, "ShmooLabel", "str", 12,
                         description="Shmoo label identifier")
        
    def create_content_tab(self):
        """Create the Content Configuration tab"""
        tab = tk.Frame(self.notebook, bg='white')
        self.notebook.add(tab, text='Content')
        
        canvas = tk.Canvas(tab, bg='white')
        scrollbar = tk.Scrollbar(tab, orient='vertical', command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='white')
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # Fields
        self.create_field(scrollable_frame, "Content", "str", 0,
                         options=self.config_template.get("CONTENT_OPTIONS", ["Linux", "Dragon", "PYSVConsole"]),
                         description="Content type to execute")
        
        # Linux specific
        tk.Label(scrollable_frame, text="Linux Configuration",
                font=('Segoe UI', 11, 'bold'), bg='white').grid(
                    row=1, column=0, columnspan=3, sticky='w', padx=10, pady=(20, 10))
        
        self.create_field(scrollable_frame, "Linux Path", "str", 2, browse=True,
                         description="Path to Linux content")
        self.create_field(scrollable_frame, "Linux Pre Command", "str", 3,
                         description="Command to run before test")
        self.create_field(scrollable_frame, "Linux Post Command", "str", 4,
                         description="Command to run after test")
        self.create_field(scrollable_frame, "Linux Pass String", "str", 5,
                         description="String indicating Linux pass")
        self.create_field(scrollable_frame, "Linux Fail String", "str", 6,
                         description="String indicating Linux failure")
        self.create_field(scrollable_frame, "Linux Content Wait Time", "int", 7,
                         description="Wait time for Linux content")
        self.create_field(scrollable_frame, "Startup Linux", "str", 8, browse=True,
                         description="Script to start Linux from EFI (e.g., bootlinux.nsh)")
        self.create_field(scrollable_frame, "Linux Content Line 0", "str", 9,
                         description="Content command line (up to 10 can be set)")
        self.create_field(scrollable_frame, "Linux Content Line 1", "str", 10,
                         description="Optional content command line")
        self.create_field(scrollable_frame, "Linux Content Line 2", "str", 11,
                         description="Optional content command line")
        self.create_field(scrollable_frame, "Linux Content Line 3", "str", 12,
                         description="Optional content command line")
        self.create_field(scrollable_frame, "Linux Content Line 4", "str", 13,
                         description="Optional content command line")
        self.create_field(scrollable_frame, "Linux Content Line 5", "str", 14,
                         description="Optional content command line")
        self.create_field(scrollable_frame, "Linux Content Line 6", "str", 15,
                         description="Optional content command line")
        self.create_field(scrollable_frame, "Linux Content Line 7", "str", 16,
                         description="Optional content command line")
        self.create_field(scrollable_frame, "Linux Content Line 8", "str", 17,
                         description="Optional content command line")
        self.create_field(scrollable_frame, "Linux Content Line 9", "str", 18,
                         description="Optional content command line")
        
        # Dragon specific
        tk.Label(scrollable_frame, text="Dragon Configuration",
                font=('Segoe UI', 11, 'bold'), bg='white').grid(
                    row=19, column=0, columnspan=3, sticky='w', padx=10, pady=(20, 10))
        
        self.create_field(scrollable_frame, "Dragon Pre Command", "str", 20,
                         description="EFI Shell command or script before running content")
        self.create_field(scrollable_frame, "Dragon Post Command", "str", 21,
                         description="EFI Shell command or script after running content")
        self.create_field(scrollable_frame, "Startup Dragon", "str", 22, browse=True,
                         description="Script with preset conditions (e.g., startup_efi.nsh)")
        self.create_field(scrollable_frame, "ULX Path", "str", 23, browse=True,
                         description="Path to ULX script")
        self.create_field(scrollable_frame, "ULX CPU", "str", 24,
                         description="CPU to be used (GNR_B0, CWF)")
        self.create_field(scrollable_frame, "Product Chop", "str", 25,
                         description="Specify product (e.g., GNR)")
        self.create_field(scrollable_frame, "VVAR0", "str", 26,
                         description="Test runtime (e.g., 3 sec = 0x2DC6C0)")
        self.create_field(scrollable_frame, "VVAR1", "str", 27,
                         description="32-bit fixed point value (use 80064000 for System)")
        self.create_field(scrollable_frame, "VVAR2", "str", 28,
                         description="Thread count (Mesh: 0x1000000, Slice GNR: 0x1000002)")
        self.create_field(scrollable_frame, "VVAR3", "str", 29,
                         description="Debug flags (see Dragon VVDEBUG Decoder wiki)")
        self.create_field(scrollable_frame, "VVAR_EXTRA", "str", 30,
                         description="Optional additional VVAR parameters")
        self.create_field(scrollable_frame, "Dragon Content Path", "str", 31, browse=True,
                         description="Path to content")
        self.create_field(scrollable_frame, "Dragon Content Line", "str", 32,
                         description="Filters to content (comma-separated)")
        
    def create_advanced_tab(self):
        """Create the Advanced Settings tab"""
        tab = tk.Frame(self.notebook, bg='white')
        self.notebook.add(tab, text='Advanced')
        
        canvas = tk.Canvas(tab, bg='white')
        scrollbar = tk.Scrollbar(tab, orient='vertical', command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='white')
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # Fields
        self.create_field(scrollable_frame, "Configuration (Mask)", "str", 0,
                         description="Debug mask configuration")
        self.create_field(scrollable_frame, "Boot Breakpoint", "str", 1,
                         description="Boot breakpoint address (not used)")
        self.create_field(scrollable_frame, "Disable 2 Cores", "str", 2,
                         description="Disable 2 cores (not used)")
        self.create_field(scrollable_frame, "Check Core", "int", 3,
                         description="Core to check data (Mesh) or tested core (Slice)")
        self.create_field(scrollable_frame, "Stop on Fail", "bool", 4,
                         description="Stop execution after first failure")
        
        # Merlin Configuration
        tk.Label(scrollable_frame, text="Merlin Configuration",
                font=('Segoe UI', 11, 'bold'), bg='white').grid(
                    row=5, column=0, columnspan=3, sticky='w', padx=10, pady=(20, 10))
        
        self.create_field(scrollable_frame, "Merlin Name", "str", 6,
                         description="Name of the Merlin file (e.g., MerlinX.efi)")
        self.create_field(scrollable_frame, "Merlin Drive", "str", 7,
                         description="Drive where Merlin is located (e.g., FS1:)")
        self.create_field(scrollable_frame, "Merlin Path", "str", 8, browse=True,
                         description="Path to Merlin files")
        
    def create_json_preview_tab(self):
        """Create the JSON Preview tab"""
        tab = tk.Frame(self.notebook, bg='white')
        self.notebook.add(tab, text='JSON Preview')
        
        # Toolbar
        toolbar = tk.Frame(tab, bg=self.colors['light'], height=40)
        toolbar.pack(fill='x', side='top')
        
        tk.Button(toolbar, text="Refresh Preview", 
                 bg=self.colors['accent'], fg='white',
                 command=self.update_json_preview,
                 relief='flat', padx=15, pady=5).pack(side='left', padx=10, pady=5)
        
        tk.Button(toolbar, text="Copy to Clipboard",
                 bg=self.colors['success'], fg='white',
                 command=self.copy_json_to_clipboard,
                 relief='flat', padx=15, pady=5).pack(side='left', padx=5, pady=5)
        
        # JSON preview text area
        self.json_preview = scrolledtext.ScrolledText(tab, 
                                                      font=('Consolas', 10),
                                                      wrap=tk.WORD,
                                                      bg='#f8f9fa')
        self.json_preview.pack(fill='both', expand=True, padx=5, pady=5)
        
    def create_field(self, parent, label, field_type, row, options=None, 
                    browse=False, description=""):
        """Create a field widget in the form"""
        # Label
        lbl = tk.Label(parent, text=f"{label}:", 
                      font=('Segoe UI', 9), bg='white', anchor='w')
        lbl.grid(row=row, column=0, sticky='w', padx=10, pady=5)
        
        # Widget based on type
        if field_type == "bool":
            var = tk.BooleanVar()
            widget = tk.Checkbutton(parent, variable=var, bg='white')
            widget.grid(row=row, column=1, sticky='w', padx=5, pady=5)
        elif options:
            var = tk.StringVar()
            widget = ttk.Combobox(parent, textvariable=var, 
                                 values=options, state='readonly')
            widget.grid(row=row, column=1, sticky='ew', padx=5, pady=5)
        else:
            var = tk.StringVar()
            widget = tk.Entry(parent, textvariable=var, width=40)
            widget.grid(row=row, column=1, sticky='ew', padx=5, pady=5)
        
        # Browse button for file/folder paths
        if browse:
            btn = tk.Button(parent, text="📁", 
                           command=lambda: self.browse_path(label, var),
                           relief='flat', bg=self.colors['accent'], fg='white')
            btn.grid(row=row, column=2, padx=5, pady=5)
        
        # Store reference
        self.field_widgets[label] = {
            'var': var,
            'widget': widget,
            'type': field_type
        }
        
        # Tooltip
        if description:
            self.create_tooltip(widget, description)
        
        # Configure column weights
        parent.grid_columnconfigure(1, weight=1)
    
    def create_tooltip(self, widget, text):
        """Create a tooltip for a widget"""
        def on_enter(event):
            tooltip = tk.Toplevel()
            tooltip.wm_overrideredirect(True)
            tooltip.wm_geometry(f"+{event.x_root+10}+{event.y_root+10}")
            label = tk.Label(tooltip, text=text, 
                           background='#ffffe0', relief='solid',
                           borderwidth=1, font=('Segoe UI', 8))
            label.pack()
            widget.tooltip = tooltip
            
        def on_leave(event):
            if hasattr(widget, 'tooltip'):
                widget.tooltip.destroy()
                
        widget.bind('<Enter>', on_enter)
        widget.bind('<Leave>', on_leave)
    
    def create_bottom_panel(self):
        """Create the bottom action button panel"""
        bottom_frame = tk.Frame(self.root, bg=self.colors['light'], height=70)
        bottom_frame.pack(fill='x', side='bottom')
        bottom_frame.pack_propagate(False)
        
        # Left buttons
        left_buttons = tk.Frame(bottom_frame, bg=self.colors['light'])
        left_buttons.pack(side='left', padx=10)
        
        tk.Button(left_buttons, text="Import from Excel",
                 bg=self.colors['accent'], fg='white',
                 command=self.import_from_excel,
                 font=('Segoe UI', 10),
                 relief='flat', padx=20, pady=10).pack(side='left', padx=5)
        
        tk.Button(left_buttons, text="Import from JSON",
                 bg=self.colors['accent'], fg='white',
                 command=self.import_from_json,
                 font=('Segoe UI', 10),
                 relief='flat', padx=20, pady=10).pack(side='left', padx=5)
        
        # Right buttons
        right_buttons = tk.Frame(bottom_frame, bg=self.colors['light'])
        right_buttons.pack(side='right', padx=10)
        
        tk.Button(right_buttons, text="Export to JSON",
                 bg=self.colors['success'], fg='white',
                 command=self.export_to_json,
                 font=('Segoe UI', 10, 'bold'),
                 relief='flat', padx=30, pady=10).pack(side='left', padx=5)
        
        tk.Button(right_buttons, text="Validate All",
                 bg=self.colors['warning'], fg='white',
                 command=self.validate_all_experiments,
                 font=('Segoe UI', 10),
                 relief='flat', padx=20, pady=10).pack(side='left', padx=5)
    
    def browse_path(self, field_name, var):
        """Browse for file or folder path"""
        if "Folder" in field_name or "Path" in field_name:
            path = filedialog.askdirectory(title=f"Select {field_name}")
        else:
            path = filedialog.askopenfilename(title=f"Select {field_name}")
        
        if path:
            var.set(path)
    
    def populate_default_template(self):
        """Populate with a default experiment template"""
        default_exp = self.create_default_experiment_data()
        self.experiments["Default_Experiment"] = default_exp
        self.refresh_experiment_list()
    
    def create_default_experiment_data(self):
        """Create default experiment data structure"""
        data = {}
        data_types = self.config_template.get('data_types', {})
        
        for field, field_types in data_types.items():
            if field == 'Experiment':
                data[field] = 'Enabled'
            elif field == 'Test Name':
                data[field] = 'New_Experiment'
            elif 'bool' in field_types:
                data[field] = False
            elif 'int' in field_types:
                if field == 'Test Number':
                    data[field] = 1
                elif field == 'Test Time':
                    data[field] = 30
                elif field == 'COM Port':
                    data[field] = 1
                else:
                    data[field] = 0
            elif 'float' in field_types:
                data[field] = 0.0
            else:
                if field == 'Test Mode':
                    data[field] = 'Mesh'
                elif field == 'Test Type':
                    data[field] = 'Loops'
                elif field == 'Content':
                    data[field] = 'Linux'
                elif field == 'Voltage Type':
                    data[field] = 'vbump'
                elif field == 'Pass String':
                    data[field] = 'Test Complete'
                elif field == 'Fail String':
                    data[field] = 'Test Failed'
                else:
                    data[field] = ''
        
        return data
    
    def add_new_experiment(self):
        """Add a new experiment"""
        # Get unique name
        base_name = "New_Experiment"
        counter = 1
        exp_name = base_name
        
        while exp_name in self.experiments:
            exp_name = f"{base_name}_{counter}"
            counter += 1
        
        # Create new experiment
        self.experiments[exp_name] = self.create_default_experiment_data()
        self.experiments[exp_name]["Test Name"] = exp_name
        
        # Refresh and select
        self.refresh_experiment_list()
        
        # Select the new experiment
        items = self.exp_listbox.get(0, tk.END)
        for i, item in enumerate(items):
            if item == exp_name:
                self.exp_listbox.selection_clear(0, tk.END)
                self.exp_listbox.selection_set(i)
                self.exp_listbox.see(i)
                self.load_experiment_data(exp_name)
                break
    
    def delete_experiment(self):
        """Delete selected experiment"""
        selection = self.exp_listbox.curselection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select an experiment to delete.")
            return
        
        exp_name = self.exp_listbox.get(selection[0])
        
        if len(self.experiments) <= 1:
            messagebox.showwarning("Cannot Delete", 
                                  "Cannot delete the last experiment. At least one must remain.")
            return
        
        if messagebox.askyesno("Confirm Delete", 
                              f"Are you sure you want to delete '{exp_name}'?"):
            del self.experiments[exp_name]
            self.refresh_experiment_list()
            
            if self.experiments:
                self.exp_listbox.selection_set(0)
                self.load_experiment_data(list(self.experiments.keys())[0])
    
    def duplicate_experiment(self):
        """Duplicate selected experiment"""
        selection = self.exp_listbox.curselection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select an experiment to duplicate.")
            return
        
        exp_name = self.exp_listbox.get(selection[0])
        
        # Create unique name
        base_name = f"{exp_name}_copy"
        counter = 1
        new_name = base_name
        
        while new_name in self.experiments:
            new_name = f"{base_name}_{counter}"
            counter += 1
        
        # Copy data
        self.experiments[new_name] = self.experiments[exp_name].copy()
        self.experiments[new_name]["Test Name"] = new_name
        
        # Refresh and select
        self.refresh_experiment_list()
        
        items = self.exp_listbox.get(0, tk.END)
        for i, item in enumerate(items):
            if item == new_name:
                self.exp_listbox.selection_clear(0, tk.END)
                self.exp_listbox.selection_set(i)
                self.exp_listbox.see(i)
                self.load_experiment_data(new_name)
                break
    
    def filter_experiments(self, *args):
        """Filter experiment list based on search"""
        search_text = self.search_var.get().lower()
        
        self.exp_listbox.delete(0, tk.END)
        
        for exp_name in self.experiments.keys():
            if search_text in exp_name.lower():
                self.exp_listbox.insert(tk.END, exp_name)
    
    def refresh_experiment_list(self):
        """Refresh the experiment listbox"""
        self.exp_listbox.delete(0, tk.END)
        
        for exp_name in sorted(self.experiments.keys()):
            self.exp_listbox.insert(tk.END, exp_name)
    
    def on_experiment_select(self, event):
        """Handle experiment selection"""
        selection = self.exp_listbox.curselection()
        if selection:
            exp_name = self.exp_listbox.get(selection[0])
            
            # Save current experiment first
            if self.current_experiment and self.current_experiment in self.experiments:
                self.save_current_experiment()
            
            # Load new experiment
            self.load_experiment_data(exp_name)
    
    def load_experiment_data(self, exp_name):
        """Load experiment data into the form"""
        self.current_experiment = exp_name
        data = self.experiments[exp_name]
        
        # Update all field widgets
        for field_name, field_info in self.field_widgets.items():
            var = field_info['var']
            field_type = field_info['type']
            value = data.get(field_name, "")
            
            if field_type == "bool":
                var.set(bool(value))
            else:
                var.set(str(value) if value is not None else "")
        
        # Update JSON preview
        self.update_json_preview()
    
    def save_current_experiment(self):
        """Save current form data to experiment"""
        if not self.current_experiment:
            return
        
        data = {}
        
        for field_name, field_info in self.field_widgets.items():
            var = field_info['var']
            field_type = field_info['type']
            
            try:
                if field_type == "bool":
                    data[field_name] = var.get()
                elif field_type == "int":
                    value = var.get().strip()
                    data[field_name] = int(value) if value else 0
                elif field_type == "float":
                    value = var.get().strip()
                    data[field_name] = float(value) if value else 0.0
                else:
                    data[field_name] = var.get()
            except ValueError:
                # Keep the string value if conversion fails
                data[field_name] = var.get()
        
        # Check if Test Name changed
        old_name = self.current_experiment
        new_name = data.get("Test Name", old_name)
        
        if new_name != old_name and new_name in self.experiments:
            messagebox.showwarning("Duplicate Name", 
                                  f"An experiment named '{new_name}' already exists.")
            data["Test Name"] = old_name
            self.field_widgets["Test Name"]['var'].set(old_name)
            new_name = old_name
        
        # Update experiments dict
        if new_name != old_name:
            del self.experiments[old_name]
            self.experiments[new_name] = data
            self.current_experiment = new_name
            self.refresh_experiment_list()
            
            # Re-select
            items = self.exp_listbox.get(0, tk.END)
            for i, item in enumerate(items):
                if item == new_name:
                    self.exp_listbox.selection_clear(0, tk.END)
                    self.exp_listbox.selection_set(i)
                    break
        else:
            self.experiments[old_name] = data
    
    def update_json_preview(self):
        """Update the JSON preview tab"""
        if not self.current_experiment:
            return
        
        # Save current data first
        self.save_current_experiment()
        
        # Generate JSON preview
        preview_data = {self.current_experiment: self.experiments[self.current_experiment]}
        json_str = json.dumps(preview_data, indent=4)
        
        # Update text widget
        self.json_preview.delete(1.0, tk.END)
        self.json_preview.insert(1.0, json_str)
    
    def copy_json_to_clipboard(self):
        """Copy JSON preview to clipboard"""
        json_text = self.json_preview.get(1.0, tk.END)
        self.root.clipboard_clear()
        self.root.clipboard_append(json_text)
        messagebox.showinfo("Copied", "JSON copied to clipboard!")
    
    def import_from_excel(self):
        """Import experiments from Excel file"""
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if not file_path:
            return
        
        try:
            data = self.process_excel_file(file_path)
            
            if not data:
                messagebox.showwarning("No Data", "No experiment data found in Excel file.")
                return
            
            # Ask if user wants to merge or replace
            if self.experiments:
                response = messagebox.askyesnocancel(
                    "Import Mode",
                    "Merge with existing experiments?\n\n" +
                    "Yes: Merge (keep existing)\n" +
                    "No: Replace all experiments\n" +
                    "Cancel: Abort import"
                )
                
                if response is None:
                    return
                elif response:  # Yes - merge
                    self.experiments.update(data)
                else:  # No - replace
                    self.experiments = data
            else:
                self.experiments = data
            
            self.refresh_experiment_list()
            
            if self.experiments:
                self.exp_listbox.selection_set(0)
                self.load_experiment_data(list(self.experiments.keys())[0])
            
            messagebox.showinfo("Success", 
                              f"Successfully imported {len(data)} experiment(s)!")
            
        except Exception as e:
            messagebox.showerror("Import Error", f"Failed to import Excel file:\n{str(e)}")
    
    def process_excel_file(self, file_path):
        """Process Excel file and extract experiment data"""
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        all_data = {}
        
        # Iterate over each sheet (each sheet is an experiment)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            data = {}
            
            # Try to find a table or use column A and B
            # Assuming format: Field Name | Value
            for row in sheet.iter_rows(min_row=1, values_only=True):
                if row[0] and len(row) > 1:
                    field_name = str(row[0]).strip()
                    field_value = row[1]
                    
                    # Skip headers
                    if field_name.lower() in ['field', 'fields', 'parameter']:
                        continue
                    
                    # Store the value
                    data[field_name] = field_value if field_value is not None else ''
            
            if data:
                all_data[sheet_name] = data
        
        return all_data
    
    def import_from_json(self):
        """Import experiments from JSON file"""
        file_path = filedialog.askopenfilename(
            title="Select JSON File",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )
        
        if not file_path:
            return
        
        try:
            with open(file_path, 'r') as f:
                data = json.load(f)
            
            if not isinstance(data, dict):
                messagebox.showerror("Invalid Format", 
                                   "JSON file must contain a dictionary of experiments.")
                return
            
            # Ask if user wants to merge or replace
            if self.experiments:
                response = messagebox.askyesnocancel(
                    "Import Mode",
                    "Merge with existing experiments?\n\n" +
                    "Yes: Merge (keep existing)\n" +
                    "No: Replace all experiments\n" +
                    "Cancel: Abort import"
                )
                
                if response is None:
                    return
                elif response:  # Yes - merge
                    self.experiments.update(data)
                else:  # No - replace
                    self.experiments = data
            else:
                self.experiments = data
            
            self.refresh_experiment_list()
            
            if self.experiments:
                self.exp_listbox.selection_set(0)
                self.load_experiment_data(list(self.experiments.keys())[0])
            
            messagebox.showinfo("Success", 
                              f"Successfully imported {len(data)} experiment(s)!")
            
        except Exception as e:
            messagebox.showerror("Import Error", f"Failed to import JSON file:\n{str(e)}")
    
    def export_to_json(self):
        """Export all experiments to JSON file"""
        # Save current experiment
        if self.current_experiment:
            self.save_current_experiment()
        
        if not self.experiments:
            messagebox.showwarning("No Data", "No experiments to export.")
            return
        
        # Ask for file location
        file_path = filedialog.asksaveasfilename(
            title="Save JSON File",
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            initialfile=f"experiments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        )
        
        if not file_path:
            return
        
        try:
            with open(file_path, 'w') as f:
                json.dump(self.experiments, f, indent=4)
            
            messagebox.showinfo("Success", 
                              f"Successfully exported {len(self.experiments)} experiment(s)!\n\n" +
                              f"File: {os.path.basename(file_path)}")
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export JSON file:\n{str(e)}")
    
    def validate_all_experiments(self):
        """Validate all experiments and show report"""
        # Save current experiment
        if self.current_experiment:
            self.save_current_experiment()
        
        issues = []
        
        for exp_name, exp_data in self.experiments.items():
            exp_issues = self.validate_experiment(exp_name, exp_data)
            if exp_issues:
                issues.extend(exp_issues)
        
        if issues:
            report = "Validation Issues Found:\n\n"
            for issue in issues:
                report += f"• {issue}\n"
            
            messagebox.showwarning("Validation Results", report)
        else:
            messagebox.showinfo("Validation Results", 
                              "All experiments validated successfully! ✓")
    
    def validate_experiment(self, exp_name, exp_data):
        """Validate a single experiment"""
        issues = []
        
        # Check required fields
        if not exp_data.get("Test Name"):
            issues.append(f"{exp_name}: Missing Test Name")
        
        if not exp_data.get("Test Mode"):
            issues.append(f"{exp_name}: Missing Test Mode")
        
        if not exp_data.get("Test Type"):
            issues.append(f"{exp_name}: Missing Test Type")
        
        # Validate numeric fields
        if exp_data.get("COM Port") is not None:
            try:
                port = int(exp_data["COM Port"])
                if port < 0 or port > 256:
                    issues.append(f"{exp_name}: COM Port out of range (0-256)")
            except (ValueError, TypeError):
                issues.append(f"{exp_name}: COM Port must be numeric")
        
        # Validate IP address format
        ip = exp_data.get("IP Address", "")
        if ip:
            parts = ip.split('.')
            if len(parts) != 4:
                issues.append(f"{exp_name}: Invalid IP Address format")
            else:
                try:
                    for part in parts:
                        num = int(part)
                        if num < 0 or num > 255:
                            issues.append(f"{exp_name}: Invalid IP Address range")
                            break
                except ValueError:
                    issues.append(f"{exp_name}: Invalid IP Address format")
        
        return issues
    
    def on_closing(self):
        """Handle window closing"""
        if messagebox.askokcancel("Quit", "Do you want to quit? Unsaved changes will be lost."):
            self.root.quit()
            self.root.destroy()
    
    def run(self):
        """Run the application (for standalone mode)"""
        if self.is_standalone:
            self.root.mainloop()


def main():
    """Main entry point"""
    app = ExperimentBuilderGUI()
    app.run()


if __name__ == "__main__":
    main()
