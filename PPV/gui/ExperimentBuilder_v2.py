"""
PPV Experiment Builder v2.0 - Excel-like Interface for Control Panel Config Generator

Features:
- Each tab represents one experiment (Excel-like workflow)
- Left panel with Templates and Unit Data (shared across experiments)
- Grouped sections with visual separation
- Conditional field enabling/disabling based on Test Type and Content
- Unit Data overrides experiment-specific values
"""

import sys
import os
import json
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext
from typing import Dict, Any, List, Optional
import openpyxl
from datetime import datetime

# Add parent directory for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


class ExperimentBuilderGUI:
    """Main GUI for building and managing experiment configurations - Excel-like interface"""
    
    def __init__(self, parent=None):
        """Initialize the Experiment Builder GUI"""
        if parent is None:
            self.root = tk.Tk()
            self.is_standalone = True
        else:
            self.root = tk.Toplevel(parent)
            self.is_standalone = False
            
        self.root.title("PPV Experiment Builder v2.0 - Excel-like Interface")
        self.root.geometry("1600x900")
        
        # Data structures
        self.experiments = {}
        self.templates = {}
        self.unit_data = {}  # Shared unit configuration
        self.current_product = "GNR"  # Current product selection
        self.current_experiment = None
        self.config_template = self.load_config_template(self.current_product)
        self.field_widgets = {}  # Current experiment's widgets
        
        # Setup UI
        self.setup_styles()
        self.create_main_layout()
        self.populate_default_data()
        
        if self.is_standalone:
            self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
            
    def setup_styles(self):
        """Configure ttk styles for the application"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Configure colors
        self.colors = {
            'primary': '#2c3e50',
            'secondary': '#34495e',
            'accent': '#3498db',
            'success': '#27ae60',
            'warning': '#f39c12',
            'danger': '#e74c3c',
            'light': '#ecf0f1',
            'dark': '#2c3e50',
            'section': '#e8f4f8',
            'section_border': '#b0d4e3'
        }
        
        # Section frame style
        style.configure('Section.TLabelframe', 
                       background='white',
                       relief='solid',
                       borderwidth=2)
        style.configure('Section.TLabelframe.Label',
                       font=('Segoe UI', 10, 'bold'),
                       foreground=self.colors['primary'],
                       background=self.colors['section'])
        
    def load_config_template(self, product="GNR"):
        """Load the configuration template from PPV/configs folder"""
        config_dir = os.path.join(os.path.dirname(__file__), '..', 'configs')
        config_file = os.path.join(config_dir, f'{product}ControlPanelConfig.json')
        
        if os.path.exists(config_file):
            with open(config_file, 'r') as f:
                return json.load(f)
                    
        return self.get_default_template()
    
    def get_default_template(self):
        """Return a default configuration template"""
        return {
            "data_types": {
                "Experiment": ["str"],
                "Test Name": ["str"],
                "Test Mode": ["str"],
                "Test Type": ["str"],
                "Visual ID": ["str"],
                "Bucket": ["str"],
                "COM Port": ["int"],
                "IP Address": ["str"],
                "TTL Folder": ["str"],
                "Scripts File": ["str"],
                "Pass String": ["str"],
                "Fail String": ["str"],
                "Content": ["str"],
                "Test Number": ["int"],
                "Test Time": ["int"],
                "Reset": ["bool"],
                "Reset on PASS": ["bool"],
                "FastBoot": ["bool"],
                "Core License": ["str"],
                "600W Unit": ["bool"],
                "Pseudo Config": ["bool"],
                "Post Process": ["str"],
                "Configuration (Mask)": ["str"],
                "Boot Breakpoint": ["str"],
                "Disable 2 Cores": ["str"],
                "Check Core": ["int"],
                "Voltage Type": ["str"],
                "Voltage IA": ["float"],
                "Voltage CFC": ["float"],
                "Frequency IA": ["int"],
                "Frequency CFC": ["int"],
                "Loops": ["int"],
                "Type": ["str"],
                "Domain": ["str"],
                "Start": ["int", "float"],
                "End": ["int", "float"],
                "Steps": ["int", "float"],
                "ShmooFile": ["str"],
                "ShmooLabel": ["str"],
                "Linux Path": ["str"],
                "Linux Pre Command": ["str"],
                "Linux Post Command": ["str"],
                "Linux Pass String": ["str"],
                "Linux Fail String": ["str"],
                "Linux Content Wait Time": ["int"],
                "Startup Linux": ["str"],
                "Linux Content Line 0": ["str"],
                "Linux Content Line 1": ["str"],
                "Linux Content Line 2": ["str"],
                "Linux Content Line 3": ["str"],
                "Linux Content Line 4": ["str"],
                "Linux Content Line 5": ["str"],
                "Linux Content Line 6": ["str"],
                "Linux Content Line 7": ["str"],
                "Linux Content Line 8": ["str"],
                "Linux Content Line 9": ["str"],
                "Dragon Pre Command": ["str"],
                "Dragon Post Command": ["str"],
                "Startup Dragon": ["str"],
                "ULX Path": ["str"],
                "ULX CPU": ["str"],
                "Product Chop": ["str"],
                "VVAR0": ["str"],
                "VVAR1": ["str"],
                "VVAR2": ["str"],
                "VVAR3": ["str"],
                "VVAR_EXTRA": ["str"],
                "Dragon Content Path": ["str"],
                "Dragon Content Line": ["str"],
                "Stop on Fail": ["bool"],
                "Merlin Name": ["str"],
                "Merlin Drive": ["str"],
                "Merlin Path": ["str"],
                "Fuse File": ["str"],
                "Bios File": ["str"]
            },
            "TEST_MODES": ["Mesh", "Slice"],
            "TEST_TYPES": ["Loops", "Sweep", "Shmoo"],
            "CONTENT_OPTIONS": ["Linux", "Dragon", "PYSVConsole"],
            "VOLTAGE_TYPES": ["vbump", "fixed", "ppvc"],
            "TYPES": ["Voltage", "Frequency"],
            "DOMAINS": ["CFC", "IA"]
        }
    
    def create_main_layout(self):
        """Create the main application layout"""
        # Header
        header_frame = tk.Frame(self.root, bg=self.colors['primary'], height=60)
        header_frame.pack(fill='x', side='top')
        header_frame.pack_propagate(False)
        
        title_label = tk.Label(header_frame, 
                              text="PPV Experiment Builder v2.0 - Excel-like Interface",
                              font=('Segoe UI', 16, 'bold'),
                              bg=self.colors['primary'],
                              fg='white')
        title_label.pack(side='left', padx=20, pady=15)
        
        # Main container with paned window
        main_paned = tk.PanedWindow(self.root, orient=tk.HORIZONTAL, sashwidth=5)
        main_paned.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Left panel - Templates and Unit Data
        self.create_left_panel(main_paned)
        
        # Right panel - Experiment tabs (Excel-like)
        self.create_right_panel(main_paned)
        
        # Bottom panel - Action buttons
        self.create_bottom_panel()
        
    def create_left_panel(self, parent):
        """Create the left panel with Templates and Unit Data"""
        left_frame = tk.Frame(parent, bg='white', width=350)
        parent.add(left_frame)
        
        # === UNIT DATA SECTION ===
        unit_data_frame = ttk.LabelFrame(left_frame, text="📋 Unit Data (Shared)", 
                                         style='Section.TLabelframe', padding=10)
        unit_data_frame.pack(fill='both', padx=5, pady=5, expand=False)
        
        # Create scrollable frame for unit data
        unit_canvas = tk.Canvas(unit_data_frame, bg='white', height=250)
        unit_scrollbar = tk.Scrollbar(unit_data_frame, orient='vertical', command=unit_canvas.yview)
        self.unit_data_fields_frame = tk.Frame(unit_canvas, bg='white')
        
        self.unit_data_fields_frame.bind(
            "<Configure>",
            lambda e: unit_canvas.configure(scrollregion=unit_canvas.bbox("all"))
        )
        
        unit_canvas.create_window((0, 0), window=self.unit_data_fields_frame, anchor='nw')
        unit_canvas.configure(yscrollcommand=unit_scrollbar.set)
        
        unit_canvas.pack(side='left', fill='both', expand=True)
        unit_scrollbar.pack(side='right', fill='y')
        
        # Unit data fields
        self.unit_data_widgets = {}
        self.create_unit_data_fields()
        
        # Apply button
        apply_btn = tk.Button(unit_data_frame, text="Apply to Current Experiment",
                             bg=self.colors['success'], fg='white',
                             command=self.apply_unit_data,
                             relief='flat', padx=10, pady=5)
        apply_btn.pack(fill='x', pady=(10, 0))
        
        # === TEMPLATES SECTION ===
        templates_frame = ttk.LabelFrame(left_frame, text="📁 Templates", 
                                        style='Section.TLabelframe', padding=10)
        templates_frame.pack(fill='both', padx=5, pady=5, expand=True)
        
        # Template toolbar
        template_toolbar = tk.Frame(templates_frame, bg='white')
        template_toolbar.pack(fill='x', pady=(0, 5))
        
        tk.Button(template_toolbar, text="+ New", font=('Segoe UI', 9),
                 bg=self.colors['success'], fg='white',
                 command=self.save_as_template,
                 width=8, relief='flat', pady=3).pack(side='left', padx=2)
        
        tk.Button(template_toolbar, text="Load", font=('Segoe UI', 9),
                 bg=self.colors['accent'], fg='white',
                 command=self.load_template,
                 width=8, relief='flat', pady=3).pack(side='left', padx=2)
        
        tk.Button(template_toolbar, text="Delete", font=('Segoe UI', 9),
                 bg=self.colors['danger'], fg='white',
                 command=self.delete_template,
                 width=8, relief='flat', pady=3).pack(side='left', padx=2)
        
        # Templates listbox
        template_list_frame = tk.Frame(templates_frame, bg='white')
        template_list_frame.pack(fill='both', expand=True)
        
        template_scrollbar = tk.Scrollbar(template_list_frame)
        template_scrollbar.pack(side='right', fill='y')
        
        self.template_listbox = tk.Listbox(template_list_frame,
                                          yscrollcommand=template_scrollbar.set,
                                          font=('Segoe UI', 9),
                                          selectmode='single')
        self.template_listbox.pack(side='left', fill='both', expand=True)
        template_scrollbar.config(command=self.template_listbox.yview)
        
    def create_unit_data_fields(self):
        """Create unit data input fields"""
        # Product selector (first field)
        lbl = tk.Label(self.unit_data_fields_frame, text="Product:", 
                      font=('Segoe UI', 9, 'bold'), bg='white', anchor='w')
        lbl.grid(row=0, column=0, sticky='w', padx=5, pady=5)
        
        self.product_var = tk.StringVar(value="GNR")
        product_combo = ttk.Combobox(self.unit_data_fields_frame, textvariable=self.product_var,
                                    values=["GNR", "CWF", "DMR"], width=18, state='readonly')
        product_combo.grid(row=0, column=1, sticky='ew', padx=5, pady=5)
        product_combo.bind('<<ComboboxSelected>>', self.on_product_change)
        
        self.unit_data_widgets["Product"] = {
            'var': self.product_var,
            'widget': product_combo,
            'type': 'str'
        }
        
        # Other unit fields
        unit_fields = [
            ("Visual ID", "str", "Unit Visual ID"),
            ("Bucket", "str", "Unit Assigned Bucket"),
            ("COM Port", "int", "Serial Port to communicate to MB"),
            ("IP Address", "str", "IP Address to communicate to MB"),
        ]
        
        for idx, (field_name, field_type, description) in enumerate(unit_fields):
            self.create_simple_field(self.unit_data_fields_frame, field_name, 
                                    field_type, idx + 1, self.unit_data_widgets, description)
    
    def create_right_panel(self, parent):
        """Create the right panel with experiment tabs"""
        right_frame = tk.Frame(parent, bg='white')
        parent.add(right_frame)
        
        # Experiment tab toolbar
        toolbar = tk.Frame(right_frame, bg=self.colors['light'], height=40)
        toolbar.pack(fill='x', side='top')
        
        tk.Label(toolbar, text="Experiments:", font=('Segoe UI', 10, 'bold'),
                bg=self.colors['light']).pack(side='left', padx=10)
        
        tk.Button(toolbar, text="+ New Experiment", font=('Segoe UI', 9),
                 bg=self.colors['success'], fg='white',
                 command=self.add_new_experiment,
                 relief='flat', padx=15, pady=5).pack(side='left', padx=5)
        
        tk.Button(toolbar, text="✕ Delete", font=('Segoe UI', 9),
                 bg=self.colors['danger'], fg='white',
                 command=self.delete_current_experiment,
                 relief='flat', padx=15, pady=5).pack(side='left', padx=5)
        
        tk.Button(toolbar, text="📋 Duplicate", font=('Segoe UI', 9),
                 bg=self.colors['accent'], fg='white',
                 command=self.duplicate_current_experiment,
                 relief='flat', padx=15, pady=5).pack(side='left', padx=5)
        
        # Notebook for experiment tabs (each tab = one experiment)
        self.experiment_notebook = ttk.Notebook(right_frame)
        self.experiment_notebook.pack(fill='both', expand=True, padx=5, pady=5)
        self.experiment_notebook.bind('<<NotebookTabChanged>>', self.on_experiment_tab_change)
        
    def create_experiment_tab(self, exp_name):
        """Create a new experiment tab with all fields organized in sections"""
        # Create main frame for this experiment
        tab_frame = tk.Frame(self.experiment_notebook, bg='white')
        self.experiment_notebook.add(tab_frame, text=exp_name)
        
        # Create scrollable canvas
        canvas = tk.Canvas(tab_frame, bg='white')
        scrollbar = tk.Scrollbar(tab_frame, orient='vertical', command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='white')
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # Store reference to widgets for this experiment
        exp_widgets = {}
        self.experiments[exp_name]['widgets'] = exp_widgets
        self.experiments[exp_name]['tab_frame'] = tab_frame
        
        # Create sections
        row = 0
        
        # === BASIC INFORMATION SECTION ===
        row = self.create_basic_info_section(scrollable_frame, exp_widgets, row)
        
        # === TEST CONFIGURATION SECTION ===
        row = self.create_test_config_section(scrollable_frame, exp_widgets, row)
        
        # === VOLTAGE & FREQUENCY SECTION ===
        row = self.create_voltage_freq_section(scrollable_frame, exp_widgets, row)
        
        # === SWEEP/SHMOO CONFIGURATION SECTION ===
        row = self.create_sweep_shmoo_section(scrollable_frame, exp_widgets, row)
        
        # === CONTENT CONFIGURATION SECTION ===
        row = self.create_content_section(scrollable_frame, exp_widgets, row)
        
        # === LINUX CONFIGURATION SECTION ===
        row = self.create_linux_section(scrollable_frame, exp_widgets, row)
        
        # === DRAGON CONFIGURATION SECTION ===
        row = self.create_dragon_section(scrollable_frame, exp_widgets, row)
        
        # === ADVANCED CONFIGURATION SECTION ===
        row = self.create_advanced_section(scrollable_frame, exp_widgets, row)
        
        # === MERLIN CONFIGURATION SECTION ===
        row = self.create_merlin_section(scrollable_frame, exp_widgets, row)
        
        return tab_frame
    
    def create_basic_info_section(self, parent, widgets, start_row):
        """Create Basic Information section"""
        section = ttk.LabelFrame(parent, text="🔹 Basic Information", 
                                style='Section.TLabelframe', padding=15)
        section.grid(row=start_row, column=0, sticky='ew', padx=10, pady=10)
        parent.grid_columnconfigure(0, weight=1)
        
        fields = [
            ("Experiment", "str", ["Enabled", "Disabled"], "Enable or disable this experiment"),
            ("Test Name", "str", None, "Unique name for this test"),
            ("Test Mode", "str", self.config_template.get("TEST_MODES", ["Mesh", "Slice"]), 
             "Test execution mode"),
            ("Test Type", "str", self.config_template.get("TEST_TYPES", ["Loops", "Sweep", "Shmoo"]), 
             "Type of test to execute"),
        ]
        
        for idx, (field_name, field_type, options, desc) in enumerate(fields):
            self.create_field_in_section(section, field_name, field_type, idx, widgets, options, desc)
        
        return start_row + 1
    
    def create_test_config_section(self, parent, widgets, start_row):
        """Create Test Configuration section"""
        section = ttk.LabelFrame(parent, text="🔹 Test Configuration", 
                                style='Section.TLabelframe', padding=15)
        section.grid(row=start_row, column=0, sticky='ew', padx=10, pady=10)
        
        fields = [
            ("TTL Folder", "str", "Path to TTL files folder", True),
            ("Scripts File", "str", "Path to test scripts file", True),
            ("Post Process", "str", "Script file after test execution", True),
            ("Pass String", "str", "String indicating test pass", False),
            ("Fail String", "str", "String indicating test failure", False),
            ("Test Number", "int", "Number of test iterations", False),
            ("Test Time", "int", "Test duration in seconds", False),
            ("Loops", "int", "Number of test loops", False),
            ("Reset", "bool", "Reset system between tests", False),
            ("Reset on PASS", "bool", "Reset only on pass", False),
            ("FastBoot", "bool", "Enable fast boot mode", False),
            ("Core License", "str", "Core license (leave blank if not used)", False),
            ("600W Unit", "bool", "Set to True for 600W units", False),
            ("Pseudo Config", "bool", "Disable HT for Pseudo Mesh Content", False),
        ]
        
        for idx, field_data in enumerate(fields):
            if len(field_data) == 4:
                field_name, field_type, desc, browse = field_data
                self.create_field_in_section(section, field_name, field_type, idx, widgets, 
                                            None, desc, browse)
            else:
                field_name, field_type, desc, browse = field_data
                self.create_field_in_section(section, field_name, field_type, idx, widgets, 
                                            None, desc, browse)
        
        return start_row + 1
    
    def create_voltage_freq_section(self, parent, widgets, start_row):
        """Create Voltage & Frequency section"""
        section = ttk.LabelFrame(parent, text="🔹 Voltage & Frequency", 
                                style='Section.TLabelframe', padding=15)
        section.grid(row=start_row, column=0, sticky='ew', padx=10, pady=10)
        
        fields = [
            ("Voltage Type", "str", self.config_template.get("VOLTAGE_TYPES", ["vbump", "fixed", "ppvc"]), 
             "Voltage control type"),
            ("Voltage IA", "float", None, "IA domain voltage"),
            ("Voltage CFC", "float", None, "CFC domain voltage"),
            ("Frequency IA", "int", None, "IA domain frequency (MHz)"),
            ("Frequency CFC", "int", None, "CFC domain frequency (MHz)"),
        ]
        
        for idx, (field_name, field_type, options, desc) in enumerate(fields):
            self.create_field_in_section(section, field_name, field_type, idx, widgets, options, desc)
        
        return start_row + 1
    
    def create_sweep_shmoo_section(self, parent, widgets, start_row):
        """Create Sweep/Shmoo Configuration section"""
        section = ttk.LabelFrame(parent, text="🔹 Sweep/Shmoo Configuration (Conditional)", 
                                style='Section.TLabelframe', padding=15)
        section.grid(row=start_row, column=0, sticky='ew', padx=10, pady=10)
        
        fields = [
            ("Type", "str", self.config_template.get("TYPES", ["Voltage", "Frequency"]), 
             "Parameter to sweep"),
            ("Domain", "str", self.config_template.get("DOMAINS", ["CFC", "IA"]), 
             "Domain to target"),
            ("Start", "float", None, "Starting value"),
            ("End", "float", None, "Ending value"),
            ("Steps", "float", None, "Step size or count"),
            ("ShmooFile", "str", None, "Path to shmoo configuration file", True),
            ("ShmooLabel", "str", None, "Shmoo label identifier"),
        ]
        
        for idx, field_data in enumerate(fields):
            if len(field_data) == 5:
                field_name, field_type, options, desc, browse = field_data
                self.create_field_in_section(section, field_name, field_type, idx, widgets, 
                                            options, desc, browse)
            else:
                field_name, field_type, options, desc = field_data
                self.create_field_in_section(section, field_name, field_type, idx, widgets, options, desc)
        
        # Mark section for conditional enabling
        widgets['sweep_shmoo_section'] = section
        
        return start_row + 1
    
    def create_content_section(self, parent, widgets, start_row):
        """Create Content Configuration section"""
        section = ttk.LabelFrame(parent, text="🔹 Content Selection", 
                                style='Section.TLabelframe', padding=15)
        section.grid(row=start_row, column=0, sticky='ew', padx=10, pady=10)
        
        self.create_field_in_section(section, "Content", "str", 0, widgets,
                                     self.config_template.get("CONTENT_OPTIONS", ["Linux", "Dragon", "PYSVConsole"]),
                                     "Content type to execute")
        
        return start_row + 1
    
    def create_linux_section(self, parent, widgets, start_row):
        """Create Linux Configuration section"""
        section = ttk.LabelFrame(parent, text="🔹 Linux Configuration (Conditional)", 
                                style='Section.TLabelframe', padding=15)
        section.grid(row=start_row, column=0, sticky='ew', padx=10, pady=10)
        
        fields = [
            ("Linux Path", "str", "Path to Linux content", True),
            ("Startup Linux", "str", "Script to start Linux from EFI", True),
            ("Linux Pre Command", "str", "Command before test", False),
            ("Linux Post Command", "str", "Command after test", False),
            ("Linux Pass String", "str", "String indicating Linux pass", False),
            ("Linux Fail String", "str", "String indicating Linux failure", False),
            ("Linux Content Wait Time", "int", "Wait time for Linux content", False),
            ("Linux Content Line 0", "str", "Content command line 0", False),
            ("Linux Content Line 1", "str", "Content command line 1", False),
        ]
        
        for idx, field_data in enumerate(fields):
            if len(field_data) == 4:
                field_name, field_type, desc, browse = field_data
                self.create_field_in_section(section, field_name, field_type, idx, widgets, 
                                            None, desc, browse)
            else:
                field_name, field_type, desc, browse = field_data
                self.create_field_in_section(section, field_name, field_type, idx, widgets, 
                                            None, desc, browse)
        
        # Mark section for conditional enabling
        widgets['linux_section'] = section
        
        return start_row + 1
    
    def create_dragon_section(self, parent, widgets, start_row):
        """Create Dragon Configuration section"""
        section = ttk.LabelFrame(parent, text="🔹 Dragon Configuration (Conditional)", 
                                style='Section.TLabelframe', padding=15)
        section.grid(row=start_row, column=0, sticky='ew', padx=10, pady=10)
        
        fields = [
            ("Dragon Content Path", "str", "Path to content", True),
            ("Startup Dragon", "str", "Preset conditions script", True),
            ("Dragon Pre Command", "str", "Command before Dragon", False),
            ("Dragon Post Command", "str", "Command after Dragon", False),
            ("ULX Path", "str", "Path to ULX script", True),
            ("ULX CPU", "str", "CPU to be used (GNR_B0, CWF)", False),
            ("Product Chop", "str", "Product (e.g., GNR)", False),
            ("VVAR0", "str", "Test runtime (hex)", False),
            ("VVAR1", "str", "32-bit fixed point (80064000 for System)", False),
            ("VVAR2", "str", "Thread count (0x1000000 for Mesh)", False),
            ("VVAR3", "str", "Debug flags (see wiki)", False),
            ("VVAR_EXTRA", "str", "Additional VVAR parameters", False),
            ("Dragon Content Line", "str", "Content filters (comma-separated)", False),
        ]
        
        for idx, field_data in enumerate(fields):
            if len(field_data) == 4:
                field_name, field_type, desc, browse = field_data
                self.create_field_in_section(section, field_name, field_type, idx, widgets, 
                                            None, desc, browse)
            else:
                field_name, field_type, desc, browse = field_data
                self.create_field_in_section(section, field_name, field_type, idx, widgets, 
                                            None, desc, browse)
        
        # Mark section for conditional enabling
        widgets['dragon_section'] = section
        
        return start_row + 1
    
    def create_advanced_section(self, parent, widgets, start_row):
        """Create Advanced Configuration section"""
        section = ttk.LabelFrame(parent, text="🔹 Advanced Configuration", 
                                style='Section.TLabelframe', padding=15)
        section.grid(row=start_row, column=0, sticky='ew', padx=10, pady=10)
        
        fields = [
            ("Configuration (Mask)", "str", None, "Debug mask configuration", False),
            ("Boot Breakpoint", "str", None, "Boot breakpoint address", False),
            ("Disable 2 Cores", "str", None, "Disable cores setting", False),
            ("Check Core", "int", None, "Core to check (Mesh/Slice)", False),
            ("Stop on Fail", "bool", None, "Stop after first failure", False),
            ("Fuse File", "str", None, "External fuse file to be loaded into the system", True),
            ("Bios File", "str", None, "BIOS file to be loaded at experiment start", True),
        ]
        
        for idx, field_data in enumerate(fields):
            if len(field_data) == 5:
                field_name, field_type, options, desc, browse = field_data
                self.create_field_in_section(section, field_name, field_type, idx, widgets, 
                                            options, desc, browse)
            else:
                field_name, field_type, options, desc = field_data
                self.create_field_in_section(section, field_name, field_type, idx, widgets, options, desc)
        
        return start_row + 1
    
    def create_merlin_section(self, parent, widgets, start_row):
        """Create Merlin Configuration section"""
        section = ttk.LabelFrame(parent, text="🔹 Merlin Configuration", 
                                style='Section.TLabelframe', padding=15)
        section.grid(row=start_row, column=0, sticky='ew', padx=10, pady=10)
        
        fields = [
            ("Merlin Name", "str", None, "Merlin file name (MerlinX.efi)"),
            ("Merlin Drive", "str", None, "Drive location (FS1:)"),
            ("Merlin Path", "str", None, "Path to Merlin files", True),
        ]
        
        for idx, field_data in enumerate(fields):
            if len(field_data) == 5:
                field_name, field_type, options, desc, browse = field_data
                self.create_field_in_section(section, field_name, field_type, idx, widgets, 
                                            options, desc, browse)
            else:
                field_name, field_type, options, desc = field_data
                self.create_field_in_section(section, field_name, field_type, idx, widgets, options, desc)
        
        return start_row + 1
    
    def create_field_in_section(self, section, field_name, field_type, row, widgets, 
                                options=None, description="", browse=False):
        """Create a field widget within a section"""
        # Label
        lbl = tk.Label(section, text=f"{field_name}:", 
                      font=('Segoe UI', 9), bg='white', anchor='w', width=25)
        lbl.grid(row=row, column=0, sticky='w', padx=5, pady=3)
        
        # Widget based on type
        if field_type == "bool":
            var = tk.BooleanVar(value=False)
            widget = tk.Checkbutton(section, variable=var, bg='white')
            widget.grid(row=row, column=1, sticky='w', padx=5, pady=3)
        elif options:
            var = tk.StringVar()
            widget = ttk.Combobox(section, textvariable=var, values=options, width=30)
            if options:
                var.set(options[0])
            widget.grid(row=row, column=1, sticky='ew', padx=5, pady=3)
            # Bind change event for conditional fields
            widget.bind('<<ComboboxSelected>>', lambda e: self.update_conditional_sections())
        else:
            var = tk.StringVar()
            widget = tk.Entry(section, textvariable=var, width=35, font=('Segoe UI', 9))
            widget.grid(row=row, column=1, sticky='ew', padx=5, pady=3)
        
        # Browse button for file/folder paths
        if browse:
            browse_btn = tk.Button(section, text="📁", font=('Segoe UI', 8),
                                   command=lambda: self.browse_path(field_name, var),
                                   width=3, relief='flat', bg=self.colors['light'])
            browse_btn.grid(row=row, column=2, padx=5)
        
        # Store reference
        widgets[field_name] = {
            'var': var,
            'widget': widget,
            'type': field_type,
            'label': lbl
        }
        
        # Tooltip
        if description:
            self.create_tooltip(widget, description)
        
        # Configure column weights
        section.grid_columnconfigure(1, weight=1)
    
    def create_simple_field(self, parent, field_name, field_type, row, widget_dict, description=""):
        """Create a simple field (for unit data panel)"""
        lbl = tk.Label(parent, text=f"{field_name}:", 
                      font=('Segoe UI', 9), bg='white', anchor='w')
        lbl.grid(row=row, column=0, sticky='w', padx=5, pady=5)
        
        if field_type == "int":
            var = tk.StringVar()
            widget = tk.Entry(parent, textvariable=var, width=20, font=('Segoe UI', 9))
        else:
            var = tk.StringVar()
            widget = tk.Entry(parent, textvariable=var, width=20, font=('Segoe UI', 9))
        
        widget.grid(row=row, column=1, sticky='ew', padx=5, pady=5)
        
        widget_dict[field_name] = {
            'var': var,
            'widget': widget,
            'type': field_type
        }
        
        if description:
            self.create_tooltip(widget, description)
        
        parent.grid_columnconfigure(1, weight=1)
    
    def on_product_change(self, event=None):
        """Handle product selection change"""
        new_product = self.product_var.get()
        if new_product != self.current_product:
            self.current_product = new_product
            self.config_template = self.load_config_template(new_product)
            messagebox.showinfo("Product Changed", 
                              f"Configuration updated for {new_product}.\n"
                              f"This affects dropdown options for new experiments.")
    
    def create_tooltip(self, widget, text):
        """Create a tooltip for a widget"""
        def on_enter(event):
            tooltip = tk.Toplevel()
            tooltip.wm_overrideredirect(True)
            tooltip.wm_geometry(f"+{event.x_root+10}+{event.y_root+10}")
            label = tk.Label(tooltip, text=text, background="#ffffe0", 
                           relief='solid', borderwidth=1, font=('Segoe UI', 8))
            label.pack()
            widget.tooltip = tooltip
            
        def on_leave(event):
            if hasattr(widget, 'tooltip'):
                widget.tooltip.destroy()
                
        widget.bind('<Enter>', on_enter)
        widget.bind('<Leave>', on_leave)
    
    def create_bottom_panel(self):
        """Create the bottom action button panel"""
        bottom_frame = tk.Frame(self.root, bg=self.colors['light'], height=60)
        bottom_frame.pack(fill='x', side='bottom')
        bottom_frame.pack_propagate(False)
        
        # Left buttons
        left_buttons = tk.Frame(bottom_frame, bg=self.colors['light'])
        left_buttons.pack(side='left', padx=10)
        
        tk.Button(left_buttons, text="📂 Import from Excel",
                 bg=self.colors['accent'], fg='white',
                 command=self.import_from_excel,
                 font=('Segoe UI', 10),
                 relief='flat', padx=20, pady=8).pack(side='left', padx=5)
        
        tk.Button(left_buttons, text="📂 Import from JSON",
                 bg=self.colors['accent'], fg='white',
                 command=self.import_from_json,
                 font=('Segoe UI', 10),
                 relief='flat', padx=20, pady=8).pack(side='left', padx=5)
        
        # Right buttons
        right_buttons = tk.Frame(bottom_frame, bg=self.colors['light'])
        right_buttons.pack(side='right', padx=10)
        
        tk.Button(right_buttons, text="💾 Export to JSON",
                 bg=self.colors['success'], fg='white',
                 command=self.export_to_json,
                 font=('Segoe UI', 10, 'bold'),
                 relief='flat', padx=30, pady=8).pack(side='left', padx=5)
        
        tk.Button(right_buttons, text="✓ Validate All",
                 bg=self.colors['warning'], fg='white',
                 command=self.validate_all_experiments,
                 font=('Segoe UI', 10),
                 relief='flat', padx=20, pady=8).pack(side='left', padx=5)
    
    def populate_default_data(self):
        """Populate with default experiment and unit data"""
        # Default unit data
        self.unit_data = {
            "Product": "GNR",
            "Visual ID": "TestUnitData",
            "Bucket": "IDIPAR",
            "COM Port": "11",
            "IP Address": "10.250.0.2"
        }
        
        # Update unit data widgets
        for field_name, value in self.unit_data.items():
            if field_name in self.unit_data_widgets:
                self.unit_data_widgets[field_name]['var'].set(value)
        
        # Create default experiment
        self.add_new_experiment("Baseline")
    
    def add_new_experiment(self, name=None):
        """Add a new experiment tab"""
        if name is None:
            base_name = "New_Experiment"
            counter = 1
            name = base_name
            while name in self.experiments:
                name = f"{base_name}_{counter}"
                counter += 1
        
        # Create experiment data structure
        exp_data = self.create_default_experiment_data()
        exp_data["Test Name"] = name
        
        self.experiments[name] = {
            'data': exp_data,
            'widgets': {},
            'tab_frame': None
        }
        
        # Create the tab
        self.create_experiment_tab(name)
        
        # Load data into widgets
        self.load_experiment_data(name)
        
        # Switch to new tab
        self.experiment_notebook.select(len(self.experiment_notebook.tabs()) - 1)
        
        # Update conditional sections
        self.update_conditional_sections()
    
    def delete_current_experiment(self):
        """Delete the currently selected experiment"""
        if len(self.experiments) <= 1:
            messagebox.showwarning("Cannot Delete", 
                                  "Cannot delete the last experiment. At least one must remain.")
            return
        
        current_tab = self.experiment_notebook.index('current')
        tab_id = self.experiment_notebook.tabs()[current_tab]
        exp_name = self.experiment_notebook.tab(tab_id, 'text')
        
        if messagebox.askyesno("Confirm Delete", 
                              f"Are you sure you want to delete experiment '{exp_name}'?"):
            del self.experiments[exp_name]
            self.experiment_notebook.forget(current_tab)
    
    def duplicate_current_experiment(self):
        """Duplicate the currently selected experiment"""
        current_tab = self.experiment_notebook.index('current')
        tab_id = self.experiment_notebook.tabs()[current_tab]
        exp_name = self.experiment_notebook.tab(tab_id, 'text')
        
        # Save current data
        self.save_current_experiment()
        
        # Create new name
        base_name = f"{exp_name}_copy"
        counter = 1
        new_name = base_name
        while new_name in self.experiments:
            new_name = f"{base_name}_{counter}"
            counter += 1
        
        # Copy experiment data
        exp_data = self.experiments[exp_name]['data'].copy()
        exp_data["Test Name"] = new_name
        
        self.experiments[new_name] = {
            'data': exp_data,
            'widgets': {},
            'tab_frame': None
        }
        
        # Create tab
        self.create_experiment_tab(new_name)
        self.load_experiment_data(new_name)
        
        # Switch to new tab
        self.experiment_notebook.select(len(self.experiment_notebook.tabs()) - 1)
    
    def on_experiment_tab_change(self, event):
        """Handle experiment tab change"""
        current_tab = self.experiment_notebook.index('current')
        if current_tab >= 0:
            tab_id = self.experiment_notebook.tabs()[current_tab]
            exp_name = self.experiment_notebook.tab(tab_id, 'text')
            self.current_experiment = exp_name
            self.update_conditional_sections()
    
    def update_conditional_sections(self):
        """Enable/disable sections based on Test Type and Content"""
        if not self.current_experiment or self.current_experiment not in self.experiments:
            return
        
        widgets = self.experiments[self.current_experiment]['widgets']
        
        # Get current values
        test_type = widgets.get('Test Type', {}).get('var', tk.StringVar()).get()
        content = widgets.get('Content', {}).get('var', tk.StringVar()).get()
        
        # Sweep/Shmoo section - only enabled for Sweep or Shmoo
        if 'sweep_shmoo_section' in widgets:
            section = widgets['sweep_shmoo_section']
            if test_type in ['Sweep', 'Shmoo']:
                self.enable_section(section, widgets, ['Type', 'Domain', 'Start', 'End', 'Steps', 'ShmooFile', 'ShmooLabel'])
            else:
                self.disable_section(section, widgets, ['Type', 'Domain', 'Start', 'End', 'Steps', 'ShmooFile', 'ShmooLabel'])
        
        # Linux section - only enabled when Content is Linux
        if 'linux_section' in widgets:
            section = widgets['linux_section']
            linux_fields = [f'Linux {x}' for x in ['Path', 'Pre Command', 'Post Command', 'Pass String', 'Fail String', 'Content Wait Time']]
            linux_fields += ['Startup Linux'] + [f'Linux Content Line {i}' for i in range(10)]
            
            if content == 'Linux':
                self.enable_section(section, widgets, linux_fields)
            else:
                self.disable_section(section, widgets, linux_fields)
        
        # Dragon section - only enabled when Content is Dragon
        if 'dragon_section' in widgets:
            section = widgets['dragon_section']
            dragon_fields = ['Dragon Content Path', 'Startup Dragon', 'Dragon Pre Command', 'Dragon Post Command',
                           'ULX Path', 'ULX CPU', 'Product Chop', 'VVAR0', 'VVAR1', 'VVAR2', 'VVAR3', 
                           'VVAR_EXTRA', 'Dragon Content Line']
            
            if content == 'Dragon':
                self.enable_section(section, widgets, dragon_fields)
            else:
                self.disable_section(section, widgets, dragon_fields)
    
    def enable_section(self, section, widgets, field_names):
        """Enable a section and its fields"""
        section.configure(text=section.cget('text').replace('(Conditional)', '').strip())
        for field in field_names:
            if field in widgets:
                widgets[field]['widget'].configure(state='normal')
                widgets[field]['label'].configure(fg='black')
    
    def disable_section(self, section, widgets, field_names):
        """Disable a section and its fields"""
        if '(Conditional)' not in section.cget('text'):
            section.configure(text=section.cget('text') + ' (Conditional)')
        for field in field_names:
            if field in widgets:
                try:
                    widgets[field]['widget'].configure(state='disabled')
                    widgets[field]['label'].configure(fg='gray')
                except:
                    pass  # Some widgets don't support state
    
    def apply_unit_data(self):
        """Apply unit data to current experiment"""
        if not self.current_experiment:
            messagebox.showwarning("No Experiment", "Please select an experiment first.")
            return
        
        # Get unit data from widgets
        unit_data = {}
        for field_name, field_info in self.unit_data_widgets.items():
            unit_data[field_name] = field_info['var'].get()
        
        # Apply to current experiment widgets
        widgets = self.experiments[self.current_experiment]['widgets']
        for field_name, value in unit_data.items():
            if field_name in widgets:
                widgets[field_name]['var'].set(value)
        
        messagebox.showinfo("Applied", f"Unit data applied to '{self.current_experiment}'")
    
    def save_as_template(self):
        """Save current experiment as a template"""
        if not self.current_experiment:
            messagebox.showwarning("No Experiment", "Please select an experiment first.")
            return
        
        self.save_current_experiment()
        
        # Ask for template name
        template_name = tk.simpledialog.askstring("Save Template", 
                                                  "Enter template name:",
                                                  initialvalue=self.current_experiment)
        if template_name:
            self.templates[template_name] = self.experiments[self.current_experiment]['data'].copy()
            self.refresh_template_list()
            messagebox.showinfo("Saved", f"Template '{template_name}' saved successfully.")
    
    def load_template(self):
        """Load selected template into current experiment"""
        selection = self.template_listbox.curselection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a template to load.")
            return
        
        if not self.current_experiment:
            messagebox.showwarning("No Experiment", "Please select an experiment first.")
            return
        
        template_name = self.template_listbox.get(selection[0])
        template_data = self.templates[template_name].copy()
        
        # Keep current experiment name
        template_data["Test Name"] = self.current_experiment
        
        # Update experiment data
        self.experiments[self.current_experiment]['data'] = template_data
        
        # Reload widgets
        self.load_experiment_data(self.current_experiment)
        
        messagebox.showinfo("Loaded", f"Template '{template_name}' loaded successfully.")
    
    def delete_template(self):
        """Delete selected template"""
        selection = self.template_listbox.curselection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a template to delete.")
            return
        
        template_name = self.template_listbox.get(selection[0])
        
        if messagebox.askyesno("Confirm Delete", f"Delete template '{template_name}'?"):
            del self.templates[template_name]
            self.refresh_template_list()
    
    def refresh_template_list(self):
        """Refresh the template listbox"""
        self.template_listbox.delete(0, tk.END)
        for template_name in sorted(self.templates.keys()):
            self.template_listbox.insert(tk.END, template_name)
    
    def create_default_experiment_data(self):
        """Create default experiment data structure"""
        data = {}
        data_types = self.config_template.get('data_types', {})
        
        for field, field_types in data_types.items():
            field_type = field_types[0]
            
            if field_type == "str":
                if field == "Experiment":
                    data[field] = "Enabled"
                elif field == "Test Mode":
                    data[field] = "Mesh"
                elif field == "Test Type":
                    data[field] = "Loops"
                elif field == "Content":
                    data[field] = "Dragon"
                elif field == "Voltage Type":
                    data[field] = "vbump"
                else:
                    data[field] = ""
            elif field_type == "int":
                data[field] = 0
            elif field_type == "float":
                data[field] = 0.0
            elif field_type == "bool":
                data[field] = False
            elif field_type == "dict":
                data[field] = {}
        
        return data
    
    def load_experiment_data(self, exp_name):
        """Load experiment data into the form"""
        if exp_name not in self.experiments:
            return
        
        data = self.experiments[exp_name]['data']
        widgets = self.experiments[exp_name]['widgets']
        
        for field_name, field_info in widgets.items():
            if field_name in ['sweep_shmoo_section', 'linux_section', 'dragon_section']:
                continue  # Skip section markers
            
            value = data.get(field_name, "")
            var = field_info['var']
            
            try:
                if field_info['type'] == 'bool':
                    var.set(bool(value))
                else:
                    var.set(str(value))
            except:
                pass
    
    def save_current_experiment(self):
        """Save current experiment data from widgets"""
        if not self.current_experiment or self.current_experiment not in self.experiments:
            return
        
        widgets = self.experiments[self.current_experiment]['widgets']
        data = self.experiments[self.current_experiment]['data']
        
        for field_name, field_info in widgets.items():
            if field_name in ['sweep_shmoo_section', 'linux_section', 'dragon_section']:
                continue
            
            try:
                value = field_info['var'].get()
                field_type = field_info['type']
                
                if field_type == 'int':
                    data[field_name] = int(value) if value else 0
                elif field_type == 'float':
                    data[field_name] = float(value) if value else 0.0
                elif field_type == 'bool':
                    data[field_name] = bool(value)
                else:
                    data[field_name] = str(value)
            except:
                pass
    
    def browse_path(self, field_name, var):
        """Browse for file or folder path"""
        if "Folder" in field_name or "Path" in field_name:
            path = filedialog.askdirectory()
        else:
            path = filedialog.askopenfilename()
        
        if path:
            var.set(path)
    
    def import_from_excel(self):
        """Import experiments from Excel file"""
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if file_path:
            try:
                self.process_excel_file(file_path)
                messagebox.showinfo("Success", "Excel file imported successfully!")
            except Exception as e:
                messagebox.showerror("Import Error", f"Failed to import Excel file:\n{str(e)}")
    
    def process_excel_file(self, file_path):
        """Process Excel file and create experiments"""
        wb = openpyxl.load_workbook(file_path)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            exp_data = {}
            
            # Read field-value pairs from columns A and B
            for row in range(1, sheet.max_row + 1):
                field = sheet.cell(row, 1).value
                value = sheet.cell(row, 2).value
                
                if field and field in self.config_template.get('data_types', {}):
                    exp_data[field] = value if value is not None else ""
            
            if exp_data:
                # Create experiment
                exp_name = exp_data.get("Test Name", sheet_name)
                self.experiments[exp_name] = {
                    'data': exp_data,
                    'widgets': {},
                    'tab_frame': None
                }
                
                self.create_experiment_tab(exp_name)
                self.load_experiment_data(exp_name)
    
    def import_from_json(self):
        """Import experiments from JSON file"""
        file_path = filedialog.askopenfilename(
            title="Select JSON File",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )
        
        if file_path:
            try:
                with open(file_path, 'r') as f:
                    data = json.load(f)
                
                # Handle both single experiment and experiment list
                if isinstance(data, dict):
                    if "experiments" in data:
                        experiments = data["experiments"]
                    else:
                        experiments = [data]
                else:
                    experiments = data
                
                for exp_data in experiments:
                    exp_name = exp_data.get("Test Name", "Imported_Experiment")
                    self.experiments[exp_name] = {
                        'data': exp_data,
                        'widgets': {},
                        'tab_frame': None
                    }
                    
                    self.create_experiment_tab(exp_name)
                    self.load_experiment_data(exp_name)
                
                messagebox.showinfo("Success", f"Imported {len(experiments)} experiment(s)!")
            except Exception as e:
                messagebox.showerror("Import Error", f"Failed to import JSON:\n{str(e)}")
    
    def export_to_json(self):
        """Export all experiments to JSON file"""
        if not self.experiments:
            messagebox.showwarning("No Data", "No experiments to export.")
            return
        
        # Save current experiment
        self.save_current_experiment()
        
        file_path = filedialog.asksaveasfilename(
            title="Save JSON File",
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )
        
        if file_path:
            try:
                # Collect all experiment data
                experiments_data = []
                for exp_name, exp_info in self.experiments.items():
                    experiments_data.append(exp_info['data'])
                
                # Create output structure
                output = {
                    "experiments": experiments_data,
                    "generated_date": datetime.now().isoformat(),
                    "tool": "PPV Experiment Builder v2.0"
                }
                
                with open(file_path, 'w') as f:
                    json.dump(output, f, indent=4)
                
                messagebox.showinfo("Success", f"Exported {len(experiments_data)} experiments to:\n{file_path}")
            except Exception as e:
                messagebox.showerror("Export Error", f"Failed to export:\n{str(e)}")
    
    def validate_all_experiments(self):
        """Validate all experiments"""
        self.save_current_experiment()
        
        errors = []
        warnings = []
        
        for exp_name, exp_info in self.experiments.items():
            exp_errors, exp_warnings = self.validate_experiment(exp_name, exp_info['data'])
            if exp_errors:
                errors.extend([f"{exp_name}: {e}" for e in exp_errors])
            if exp_warnings:
                warnings.extend([f"{exp_name}: {w}" for w in exp_warnings])
        
        message = f"Validation Results:\n\n"
        message += f"Experiments: {len(self.experiments)}\n"
        message += f"Errors: {len(errors)}\n"
        message += f"Warnings: {len(warnings)}\n\n"
        
        if errors:
            message += "ERRORS:\n" + "\n".join(errors[:10])
            if len(errors) > 10:
                message += f"\n... and {len(errors) - 10} more"
        
        if warnings:
            message += "\n\nWARNINGS:\n" + "\n".join(warnings[:10])
            if len(warnings) > 10:
                message += f"\n... and {len(warnings) - 10} more"
        
        if not errors and not warnings:
            message += "✓ All experiments validated successfully!"
            messagebox.showinfo("Validation Complete", message)
        else:
            messagebox.showwarning("Validation Issues", message)
    
    def validate_experiment(self, exp_name, exp_data):
        """Validate a single experiment"""
        errors = []
        warnings = []
        
        # Required fields
        if not exp_data.get("Test Name"):
            errors.append("Test Name is required")
        
        # IP Address format
        ip = exp_data.get("IP Address", "")
        if ip and not self.validate_ip(ip):
            errors.append(f"Invalid IP Address format: {ip}")
        
        # Test Type specific validation
        test_type = exp_data.get("Test Type", "")
        if test_type in ["Sweep", "Shmoo"]:
            if not exp_data.get("Start"):
                warnings.append("Start value should be set for Sweep/Shmoo")
            if not exp_data.get("End"):
                warnings.append("End value should be set for Sweep/Shmoo")
        
        return errors, warnings
    
    def validate_ip(self, ip):
        """Validate IP address format"""
        parts = ip.split('.')
        if len(parts) != 4:
            return False
        try:
            return all(0 <= int(part) <= 255 for part in parts)
        except:
            return False
    
    def on_closing(self):
        """Handle window close"""
        self.save_current_experiment()
        self.root.destroy()
    
    def run(self):
        """Run the application"""
        self.root.mainloop()


def main():
    """Main entry point"""
    app = ExperimentBuilderGUI()
    app.run()


if __name__ == "__main__":
    main()
