from openpyxl import load_workbook
import json, pathlib
import csv

#def load_workbook(path, data_only=False):
#	return load_workbook(path=path, data_only=data_only)


wb = load_workbook(
	r"C:\Users\gaespino\OneDrive - Intel Corporation\TLO DTD - GNR and SRF Debug - DT\FACR\ErrorChecks\MCA Analysis.xlsx",
	data_only=False  # <-- keeps formulas, not cached values
)

def collect_formula_data(workbook):
	formula_data = {}
	for sheet_name in workbook.sheetnames:
		ws = workbook[sheet_name]
		formula_data[sheet_name] = {}
		for row in ws.iter_rows():
			for cell in row:
				if cell.value is not None and cell.data_type == 'f':  # 'f' indicates a formula
					formula_data[sheet_name][cell.coordinate] = cell.value
	return formula_data

def write_formula_data_to_csv(formula_data, filename):
	with open(filename, 'w', newline='') as csvfile:
		writer = csv.writer(csvfile)
		for sheet_name, cells in formula_data.items():
			writer.writerow([f"Sheet: {sheet_name}"])
			for coord, formula in cells.items():
				writer.writerow([coord, formula])

#data = {}  # to store cell coordinates and values
#for sheet_name in wb.sheetnames:
#    ws = wb[sheet_name]
#    data[sheet_name] = {}
#    print(f"\n=== Sheet: {sheet_name} ===")
#    for row in ws.iter_rows():
#        for cell in row:
#            if cell.value is not None:
#                data[sheet_name][cell.coordinate] = cell.value
#                print(f"  {cell.coordinate}: {cell.value}")

# save data into a .csv file

#with open('mca_analysis_data.csv', 'w', newline='') as csvfile:#
#	writer = csv.writer(csvfile)
#	for sheet_name, cells in data.items():
#		writer.writerow([f"Sheet: {sheet_name}"])
#		for coord, value in cells.items():
#			writer.writerow([coord, value])
#

def collect_computes_data(workbook, sheet_name='Computes'):
    ws = workbook[sheet_name]

    def safe_convert(value):
        """Convert Excel values to JSON-serializable types"""
        if value is None:
            return ""
        # Handle ArrayFormula and other openpyxl objects
        if hasattr(value, '__str__'):
            return str(value)
        return value

    rows = list(ws.iter_rows(values_only=True))
    # Row 0 = headers; data starts row 1
    # Cols: B=DIE(0), C=CORE(1), D=Row(2), E=Col(3), S=IP_string(18), T=IP_name(19), AB=Device(27), AC=Device_name(28), AD=offset(29)
    data = {}
    for row in rows[1:]:
        core_id = row[2]  # col C
        if core_id is None:
            continue

        # Convert core_id safely
        try:
            core_key = str(int(float(safe_convert(core_id))))
        except (ValueError, TypeError):
            core_key = safe_convert(core_id)

        data[core_key] = {
            "compute": safe_convert(row[1]),   # DIE / Compute name
            "row":     safe_convert(row[3]),   # D
            "col":     safe_convert(row[4]),   # E
            #"ip":      safe_convert(row[18]),  # S
            #"ip_name": safe_convert(row[19]),  # T
            #"device":  safe_convert(row[27]),  # AB
            #"dev_name":safe_convert(row[28]),  # AC
            #"offset":  safe_convert(row[29]) if row[29] is not None else 0  # AD
        }

    pathlib.Path(__file__).resolve().parents[1].joinpath("analysis", "GNR_layout.json").write_text(json.dumps(data, indent=2))
    print(f"Exported {len(data)} compute entries")


def collect_map_data(workbook, sheet_name='Map'):
	ws = workbook[sheet_name]
	data = {}
	print(f"Dimensions: {ws.dimensions}")
	for row in ws.iter_rows(min_row=1, max_row=30, values_only=True):
		data[row[0]] = row

	pathlib.Path(__file__).resolve().parents[1].joinpath("analysis", "GNR_ip_map.json").write_text(json.dumps(data, indent=2))
	print(f"Exported {len(data)} compute entries")


#formula_data = collect_formula_data(wb)
#write_formula_data_to_csv(formula_data, 'mca_analysis_formulas.csv')

#collect_computes_data(wb)
collect_map_data(wb)
