import zipfile
import os
import pandas as pd
import colorama
from tabulate import tabulate
from colorama import Fore, Style, Back
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.table import Table, TableStyleInfo


class LogFileParser:

	def __init__(self, zip_file_path, content_type, exclusion_string='pysv', casesens=False):
		self.zip_file_path = zip_file_path
		self.content_type = content_type.lower()  # Ensure content type is case-insensitive
		self.exclusion_string = exclusion_string
		self.casesens = casesens
		self.pass_strings = ["Test Complete"]
		self.fail_strings = ["exit: fail", "not ok"]
		self.hang_strings = ["running MerlinX.efi"]
		self.check_strings = [r"CHANGING BACK TO DIR: FS1:\EFI"]
		self.skip_strings = []

	def set_zip_file_path(self, zip_path):
		self.zip_file_path = zip_path

	def set_pass_strings(self, pass_strings):
		self.pass_strings = [p.lower() for p in pass_strings] if not self.casesens else pass_strings

	def set_fail_strings(self, fail_strings):
		self.fail_strings = [f.lower() for f in fail_strings] if not self.casesens else fail_strings

	def set_hang_strings(self, hang_strings):
		self.hang_strings = [h.lower() for h in hang_strings] if not self.casesens else hang_strings

	def set_check_strings(self, check_strings):
		self.check_strings = [c.lower() for c in check_strings] if not self.casesens else check_strings

	def set_skip_strings(self, skip_strings):
		self.skip_strings = [s.lower() for s in skip_strings] if not self.casesens else skip_strings

	def parse_log_files_in_zip(self):
		zipped_log_data = {}
		with zipfile.ZipFile(self.zip_file_path, 'r') as zip_ref:
			for log_file in zip_ref.namelist():
				if log_file.endswith('.log') and self.exclusion_string not in log_file:
					with zip_ref.open(log_file) as log_file_ref:
						log_content = log_file_ref.read().decode('utf-8')
						# Use the content type to determine which parsing method to call
						if self.content_type == 'efi':
							result = self.parse_efi_log(log_content)
						elif self.content_type == 'linux':
							result = self.parse_linux_log(log_content)
						elif self.content_type == 'sandstone':
							result = self.parse_sandstone_log(log_content)
						elif self.content_type == 'imunch':
							result = self.parse_linux_log(log_content)
						elif self.content_type == 'tsl':
							result = self.parse_linux_log(log_content)
						elif self.content_type == 'python':
							result = self.parse_python_log(log_content)
						elif self.content_type == 'other':
							result = self.parse_efi_log(log_content)
						else:
							result = []
						print(f"Result for {log_file}: {result}")
						zipped_log_data[log_file] = result #','.join(result)
		
		return zipped_log_data

	def parse_efi_log(self, log_content):
		lines = log_content.splitlines()
		fail_results = set()

		for line in lines:
			if any((skip_string in line) if self.casesens else (skip_string.lower() in line.lower()) for skip_string in self.skip_strings):
				continue

			for search_string in self.fail_strings:
				if (search_string in line) if self.casesens else (search_string.lower() in line.lower()):
					obj_file = self.extract_obj_from_line(line)
					fail_results.add(obj_file.replace('.obj', '') + '_FAIL' if obj_file else "FAIL")

		if not fail_results:
			for line in reversed(lines):
				if any((skip_string in line) if self.casesens else (skip_string.lower() in line.lower()) for skip_string in self.skip_strings):
					continue

				for search_string in self.pass_strings:
					if (search_string in line) if self.casesens else (search_string.lower() in line.lower()):
						fail_results.add("PASS")
						break
				else:
					for search_string in self.check_strings:
						if (search_string in line) if self.casesens else (search_string.lower() in line.lower()):
							fail_results.add("_CHECK_MCA")
							break
					else:
						for search_string in self.hang_strings:
							if (search_string in line) if self.casesens else (search_string.lower() in line.lower()):
								obj_file = self.extract_obj_from_line(line)
								fail_results.add(obj_file.replace('.obj', '') + '_HANG' if obj_file else "HANG")
								break

				if "PASS" in fail_results or "_CHECK_MCA" in fail_results or any(obj.endswith('_HANG') for obj in fail_results):
					break

		return list(fail_results)

	def extract_obj_from_line(self, line):
		start = line.rfind('\\') + 1
		end = line.rfind('.obj')
		if start != -1 and end != -1:
			return line[start:end + 4]
		return None

	def parse_sandstone_log(self, log_content):
		lines = log_content.splitlines()
		fail_results = set()

		# Check for fail conditions
		for line in lines:
			if any((skip_string in line) if self.casesens else (skip_string.lower() in line.lower()) for skip_string in self.skip_strings):
				continue

			for search_string in self.fail_strings:
				if (search_string in line) if self.casesens else (search_string.lower() in line.lower()):
					# Extract the failing test information
					parts = line.strip(search_string).split(" ")
					if len(parts) >= 2:
						fail_results.add(f"{parts[1]}")
					else:
						fail_results.add("FAIL")

		# If no fail conditions are found, check for pass conditions
		if not fail_results:
			for line in lines:
				if any((skip_string in line) if self.casesens else (skip_string.lower() in line.lower()) for skip_string in self.skip_strings):
					continue

				for search_string in self.pass_strings:
					if (search_string in line) if self.casesens else (search_string.lower() in line.lower()):
						fail_results.add("PASS")
						break

		# If neither pass nor fail conditions are found, default to "HANG"
		if not fail_results:
			fail_results.add("HANG")

		return list(fail_results)

	def parse_linux_log(self, log_content):
		lines = log_content.splitlines()
		fail_results = set()

		# Check for fail conditions
		for line in lines:
			if any((skip_string in line) if self.casesens else (skip_string.lower() in line.lower()) for skip_string in self.skip_strings):
				continue

			for search_string in self.fail_strings:
				if (search_string in line) if self.casesens else (search_string.lower() in line.lower()):
					# Extract the failing test information
					parts = line.strip(search_string)
					if len(parts) > 2:
						fail_results.add(f"{parts}")
					else:
						fail_results.add("FAIL")

		# If no fail conditions are found, check for pass conditions
		if not fail_results:
			for line in lines:
				if any((skip_string in line) if self.casesens else (skip_string.lower() in line.lower()) for skip_string in self.skip_strings):
					continue

				for search_string in self.pass_strings:
					if (search_string in line) if self.casesens else (search_string.lower() in line.lower()):
						fail_results.add("PASS")
						break

		# If neither pass nor fail conditions are found, default to "HANG"
		if not fail_results:
			fail_results.add("HANG")

		return list(fail_results)

	def parse_python_log(self, log_content):
		lines = log_content.splitlines()
		fail_results = set()

		for line in lines:
			if any((skip_string in line) if self.casesens else (skip_string.lower() in line.lower()) for skip_string in self.skip_strings):
				continue

			for search_string in self.fail_strings:
				if (search_string in line) if self.casesens else (search_string.lower() in line.lower()):
					fail_results.add("FAIL")

		if not fail_results:
			for line in reversed(lines):
				if any((skip_string in line) if self.casesens else (skip_string.lower() in line.lower()) for skip_string in self.skip_strings):
					continue

				for search_string in self.pass_strings:
					if (search_string in line) if self.casesens else (search_string.lower() in line.lower()):
						fail_results.add("PASS")
						break
				else:
					for search_string in self.hang_strings:
						if (search_string in line) if self.casesens else (search_string.lower() in line.lower()):
							fail_results.add("HANG")
							break

		return list(fail_results)

class LogSummaryParser:
	
	def __init__(self, excel_path_dict, test_df, product = 'GNR'):
		self.excel_path_dict = excel_path_dict
		self.test_df = test_df
		self.product = product
		self.search_word = 'CORE' if product == 'GNR' else 'MODULE'

	def parse_mca_tabs_from_files(self):
		# Initialize a dictionary to store results
		report_data = []
		excel_path_dict = self.excel_path_dict
		# Initialize a list to store MCA data

		# Define the tabs to parse
		tabs_to_parse = ['CHA_MCAS', 'LLC_MCAS', 'CORE_MCAS', 'UBOX']

		# Iterate over each Excel file in the path dictionary
		for experiment_name, info in excel_path_dict.items():
			excel_file_path = info['path']
			# Load the Excel file
			print(excel_file_path)
			excel_data = pd.read_excel(excel_file_path, sheet_name=None, engine='openpyxl')

			# Iterate over each tab
			for tab in tabs_to_parse:
				if tab in excel_data:
					df = excel_data[tab]
					for _, row in df.iterrows():
						# Extract the run value from the 'Run' column
						run_value = str(row['Run']).split('-')[0]

						# Extract MCA data based on the tab
						if tab == 'CHA_MCAS':
							report_data.append({'Failed MCA': f"CDIE{row['Compute']}::{row['CHA']}::{row['MC DECODE']}", 'Content': 'CHA_MCAs', 'Experiment': experiment_name, 'Run Value': run_value})
						elif tab == 'LLC_MCAS':
							report_data.append({'Failed MCA': f"CDIE{row['Compute']}::{row['LLC']}::{row['MC DECODE']}", 'Content': 'LLC_MCAs', 'Experiment': experiment_name, 'Run Value': run_value})
						elif tab == 'CORE_MCAS':
							report_data.append({'Failed MCA': f"CDIE{row['Compute']}::{row[self.search_word]}::{row['ErrorType']}::{row['MCACOD (ErrDecode)']}", 'Content': 'CORE_MCAs', 'Experiment': experiment_name, 'Run Value': run_value})
						elif tab == 'UBOX':
							location = row['FirstError - Location']
							if pd.notna(location) and location != '0x0':
								report_data.append({'Failed MCA': f"{row['NCEVENT'].split('__')[-1]}::{location}", 'Content': 'UBOX', 'Experiment': experiment_name, 'Run Value': run_value})

		# Convert the MCA data to a DataFrame
		mca_df = pd.DataFrame(report_data)

		# Check if mca_df is empty and initialize with default columns if necessary
		if mca_df.empty:
			mca_df = pd.DataFrame(columns=['Failed MCA', 'Content', 'Experiment', 'Run Value'])
			unique_mcas_df = pd.DataFrame(columns=['Failed MCA', 'Content', 'Fail_Count'])
		
		else:
			# Generate unique MCA values and their counts
			
			unique_mcas_df = mca_df.groupby('Failed MCA').agg(
				Content=('Content', lambda x: ', '.join(set(x))),
				Fail_Count=('Failed MCA', 'size')
			).reset_index()

		return unique_mcas_df, mca_df

	def get_experiment_name(self, experiment_number):
		# Filter the DataFrame directly, ensuring 'Test File' column does not contain NaN values
		matching_row = self.test_df[self.test_df['Test File'].notna() & self.test_df['Test File'].str.startswith(f"{experiment_number}_")]
		
		if not matching_row.empty:
			return matching_row.iloc[0]['Experiment']
		return f"Experiment {experiment_number}"

######################################################################################################################
###########################     Framework Report Scripts Below
######################################################################################################################

def framework_merge(file_dict, output_file, prefix = '', skip=[]):
	# List to hold dataframes
	all_data = {}

	# Iterate over all Excel files in the input folder
	for k, v in file_dict.items():
		file = file_dict[k]['path']
		testype = file_dict[k]['test_type']
		if testype in skip:
			formatted_print(f' >>> {testype.upper()} <<< {file}', Fore.MAGENTA, Back.LIGHTYELLOW_EX)
			continue
		
		if prefix != None:
			keyword = prefix
			excelfile = (file.endswith(f'{keyword}.xlsx') or file.endswith(f'{keyword}.xls')) or (prefix in file and '.xls' in file)
		else:
			keyword = ''
			excelfile = (file.endswith(f'{keyword}.xlsx') or file.endswith(f'{keyword}.xls'))
	   
				
		if excelfile:
			file_path = file
			excel_data = pd.read_excel(file_path, sheet_name=None)
			
			for sheet_name, df in excel_data.items():
				if sheet_name not in all_data:
					all_data[sheet_name] = []
				all_data[sheet_name].append(df)
		
		formatted_print(f' >>> DONE <<< {file}', Fore.BLACK, Back.LIGHTGREEN_EX)

	# Create a writer object to write the merged data to a new Excel file
	with pd.ExcelWriter(output_file) as writer:
		for sheet_name, df_list in all_data.items():
			merged_df = pd.concat(df_list, ignore_index=True)
			merged_df.to_excel(writer, sheet_name=sheet_name, index=False)
	
	print(f"Framework Data Merged Excel file saved as {output_file}")

def parse_log_files(log_dict):
	# List to store test data
	test_data = []

	# Keywords to look for in the log files
	keywords = [
		"Unit QDF:",
		"Configuration:",
		"Voltage set to:",
		"HT Disabled (BigCore):",
		"Dis 2 Cores (Atomcore):",
		"Core License:",
		"Core Freq:",
		"Core Voltage:",
		"Mesh Freq:",
		"Mesh Voltage:",
		"Unit 600w Fuses Applied",
		"Running Content:",
		"Pass String:",
		"Fail String:"
	]

	# Process each log file
	for filename, info in log_dict.items():
		log_path = info['path']
		test_type = info['test_type']
		content_info = info['content']
		comments = info['comments']

		with open(log_path, 'r') as file:
			content = file.read()

		# Split the content based on the test start string
		tests = content.split("-- Test Start ---")

		# Extract path information
		#logfile_path = log_path.split(os.sep)
		path_parts = os.path.normpath(log_path).split(os.sep)
		if len(path_parts) >= 6:
			vid = path_parts[-5]
			platform = path_parts[-4]
			date = path_parts[-3]
			folder = path_parts[-2]
			
		# Process each test section
		for test in tests[1:]:  # Skip the first split part as it is before the first test start
			test_info = {
				'VID': vid,
				'Platform': platform,
				'Date': date,
				'Folder': folder,
				'Type': test_type
			}

			# Extract tdata information
			tdata_start = test.find("tdata_")
			if tdata_start != -1:
				tdata_end = test.find('\n', tdata_start)
				tdata_line = test[tdata_start:tdata_end].strip()
				tdata_parts = tdata_line.split("::")
				if len(tdata_parts) >= 5:
					test_info['Test Number'] = tdata_parts[0]
					test_info['Experiment'] = tdata_parts[1]
					test_info['PostCode'] = tdata_parts[3]
					test_info['Result'] = tdata_parts[2]
					test_info['Content Status'] = tdata_parts[4]
					test_info['Test File'] = f'{tdata_parts[0].strip("tdata_")}_{tdata_parts[1]}.log'


			# Extract keyword values
			for keyword in keywords:
				start = test.find(keyword)
				if start != -1:
					end = test.find('\n', start)
					value = test[start + len(keyword):end].strip()
					test_info[keyword.strip(':')] = value
				else:
					test_info[keyword.strip(':')] = None
		
			# Check for 600W Fuses Applied
			test_info['Unit 600w Fuses Applied'] = 'True' if "Unit 600w Fuses Applied" in test else 'False'
			test_info['Content Detail'] = content_info
			test_info['Comments'] = comments
			# Append the test information to the list
			test_data.append(test_info)

	# Create a DataFrame from the test data list
	test_df = pd.DataFrame(test_data)

	return test_df

def check_zip_data(zip_path_dict, skip_array, test_df):
	zip_results = []
	pass_array = []
	fail_array = []

	# Group by 'Folder' to aggregate pass and fail strings
	grouped_df = test_df.groupby('Folder').agg({
		'Pass String': lambda x: ','.join(x.dropna()),
		'Fail String': lambda x: ','.join(x.dropna())
	}).reset_index()

	for experiment, zip_files in zip_path_dict.items():
		zip_path = zip_files['path']
		test_type = zip_files['test_type']
		content_info = zip_files['content']
		comments = zip_files['comments']
		exp_df = test_df[(test_df['Folder'] == experiment)]
		
		if not exp_df.empty:
			pass_string = exp_df.iloc[0]['Pass String']
			fail_string = exp_df.iloc[0]['Fail String']
			
			# Split the concatenated strings and convert to sets to ensure uniqueness, then back to lists
			try:
				pass_array = list(set(pass_string.split(",")))
			except:
				pass_array = []
			
			try:
				fail_array = list(set(fail_string.split(",")))
			except:
				fail_array = []
		#fail_info_df[(fail_info_df['Log File'] == test_file) & (fail_info_df['Experiment'] == folder)]
		#for zip_file in zip_files:
				
		results = parse_zip_files(zip_file_path=zip_path, content=content_info, pass_array = pass_array, fail_array = fail_array, skip_array = skip_array, exclusion_string='pysv', casesens=False)
			
		for filename, data in results.items():
			zip_results.append({'Experiment': experiment, 'Content': content_info, 'Log File': filename, 'Log Fails': ', '.join(data)})

	return pd.DataFrame(zip_results)

def parse_zip_files(zip_file_path, content, pass_array = [], fail_array = [], skip_array = [], exclusion_string='pysv', casesens=False):

	linux_rules = {	'pass': ['exit: pass', "test_result: passed" , "Passed", "Result=SUCCESS"], 
		  		'Fail': ['exit: fail', "test_result: failed", "failed", "Result=FAILED"], 'hang': [], 'check': []}
	
	tsl_rules = {	'pass': ['exit: pass'], 
		  		'fail': ['exit: fail', "not ok"], 'hang': [], 'check': []}
	
	sandstone_rules = {	'pass': ['exit: pass'], 
		  		'fail': ['exit: fail', "not ok"], 'hang': [], 'check': []}
	
	imunch_rules = {	'pass': ['exit: pass'], 
		  		'fail': ['exit: fail', "not ok"], 'hang': [], 'check': []}

	efi_rules = {	'pass': ['Test Complete'], 
		  		'fail': ['Test Failed'], 'hang': [], 'check': []}

	dragon_rules = {	'pass': ['Test Complete'], 
		  		'fail': ['Test Failed'], 'hang': ["running MerlinX.efi", "MerlinX.efi"], 'check':[r"CHANGING BACK TO DIR: FS1:\EFI"]}

	python_rules = {	'pass': ['Test Complete'], 
		  		'fail': ['Test Failed'], 'hang': [], 'check': []}

	other_rules = {	'pass': ['Test Complete'], 
		  		'fail': ['Test Failed'], 'hang': [], 'check': []}
	
	## Content division for Failing Checks
	efi_content = ["EFI"]
	dragon_content = ["DBM", "Pseudo Slice", "Pseudo Mesh"]
	linux_content = [ "Linux"]
	tsl_content = [ "TSL"]
	sandstone_content = ["Sandstone"]
	imunch_content = ["Imunch"]
	python_content = ["Python"]
	other_content = ["Other"]

	if content in efi_content:
		content_selection = "EFI"
		content_type = "efi"
	elif content in dragon_content:
		content_selection = "Dragon"
		content_type = "efi"
	elif content in linux_content:
		content_selection = "Linux"
		content_type = "linux"
	elif content in sandstone_content:
		content_selection = "Sandstone"
		content_type = "sandstone"
	elif content in imunch_content:
		content_selection = "Imunch"
		content_type = "imunch"
	elif content in tsl_content:
		content_selection = "TSL"
		content_type = "tsl"
	elif content in python_content:
		content_selection = "Python"
		content_type = "python"
	elif content in other_content:
		content_selection = "Other"
		content_type = "efi"
	else:
		print(' -- No valid content selected..')
		return {}

	content_strings = {		"EFI": efi_rules,
							"Dragon": dragon_rules,
							"Linux": linux_rules,
							"Sandstone": sandstone_rules,
							"Imunch": imunch_rules,
							"TSL": tsl_rules,
							"Python": python_rules,
							"Other": other_rules} # will use same as EFI rules for now Placeholder

	zip_parser = LogFileParser(zip_file_path, content_type=content_type, exclusion_string=exclusion_string, casesens=casesens)

	# Set Strings into parser
	zip_parser.set_pass_strings(array_merge(content_strings[content_selection]['pass'], pass_array))
	zip_parser.set_fail_strings(array_merge(content_strings[content_selection]['fail'], fail_array))
	zip_parser.set_hang_strings(content_strings[content_selection]['hang'])
	zip_parser.set_check_strings(content_strings[content_selection]['check'])
	zip_parser.set_skip_strings(skip_array)

	# Execute Parser
	zipped_log_data = zip_parser.parse_log_files_in_zip()

	return zipped_log_data

def find_files(base_folder, excel_keyword='Summary', zip_keyword='ExperimentData', log_keyword='DebugFrameworkLogger.log'):
	# Lists to store results
	data = []

	# Walk through the directory tree
	for root, dirs, files in os.walk(base_folder):
		# Initialize variables to store file paths
		excel_path = None
		zip_path = None
		log_path = None

		# Check for Excel files
		for file in files:
			if file.endswith('.xlsx'):
				if excel_keyword is None or excel_keyword in file:
					excel_path = os.path.join(root, file)

					# Extract information from the path
					path_parts = os.path.normpath(excel_path).split(os.sep)
					if len(path_parts) >= 6:
						product = path_parts[-6]
						vid = path_parts[-5]
						platform = path_parts[-4]
						date = path_parts[-3]
						experiment = path_parts[-2]
						summary_file = path_parts[-1]

						# Append to data list
						data.append({
							'Product': product,
							'VID': vid,
							'Platform': platform,
							'Date': date,
							'Experiment': experiment,
							'Summary File': summary_file,
							'Excel': excel_path,
							'ZIP': None,  # Placeholder for ZIP path
							'Log': None   # Placeholder for log path
						})

		# If an Excel file was found, check for ZIP and log files
		if excel_path:
			for file in files:
				# Check for ZIP files
				if file.endswith('.zip'):
					if zip_keyword is None or zip_keyword in file:
						zip_path = os.path.join(root, file)

				# Check for log files
				if file == log_keyword:
					log_path = os.path.join(root, file)

			# Update the last entry in the data list with ZIP and log paths
			if data:
				data[-1]['ZIP'] = zip_path
				data[-1]['Log'] = log_path

	# Create a DataFrame from the data list
	df = pd.DataFrame(data)

	return df

def create_file_dict(df, file_type, test_types=None, content_types=None, comments_types=None):
	file_dict = {}
	for _, row in df.iterrows():
		file_path = row[f'{file_type}']
		if pd.notna(file_path) and not os.path.basename(file_path).startswith('~$'):
			folder_name = os.path.basename(os.path.dirname(file_path))
			file_dict[folder_name] = {
				'path': file_path,
				'test_type': test_types.get(row['Experiment'], 'Base') if test_types else None,
				'content': content_types.get(row['Experiment'], '') if content_types else None,
				'comments': comments_types.get(row['Experiment'], '') if comments_types else None
			}
	
	return file_dict

def create_summary_df(test_df):
	summary_data = []
	experiment_index_map = {}  # Map to store index values for each experiment

	for _, row in test_df.iterrows():
		# Extract values from the current row
		# If line experiment is Invalid skip it
		exptype = row['Type']
		if exptype == 'Invalid':
			continue
		running_content = row['Running Content']
		content_detail = row['Content Detail']
		comments = row['Comments']
		folder =  row['Folder']
		experiment = row['Experiment']
		exptype = row['Type']
		voltage_set_to = row['Voltage set to']
		postcode = row['PostCode']
		ht_disabled = row['HT Disabled (BigCore)']
		dis_cores = row['Dis 2 Cores (Atomcore)']
		core_freq = row['Core Freq']
		core_voltage = row['Core Voltage']
		mesh_freq = row['Mesh Freq']
		mesh_voltage = row['Mesh Voltage']
		fuses_applied = row['Unit 600w Fuses Applied']
		mask = row['Configuration']
		status = row['Result']
		log_file = row['Test File']
		content_results = row['Content Status']
		mca_status = row.get('MCA Status', '')  # Get MCA Status from test_df
	
		# Assign an index to the experiment if not already assigned
		if folder not in experiment_index_map:
			experiment_index_map[folder] = len(experiment_index_map)

		# Construct Defeature String
		Exclude_array = ['None', 'False', None, False]

		run_content_framework = running_content if running_content not in Exclude_array else ''
		run_content_detail = content_detail if content_detail not in Exclude_array else ''
		
		used_content = f'{run_content_framework}::{run_content_detail}' if run_content_detail != run_content_framework else run_content_framework

		CFC_Values_V = format_voltage(mesh_voltage, voltage_set_to.upper(), Exclude_array) 
		CFC_Values_F = format_frequency(mesh_freq, Exclude_array) 
		IA_Values_V = format_voltage(core_voltage, voltage_set_to.upper(), Exclude_array)
		IA_Values_F = format_frequency(core_freq, Exclude_array)
		
		HTDIS = f'HTDIS::{ht_disabled.upper()} - ' if ht_disabled not in Exclude_array else ''
		DisableModules = f'DISMOD::{dis_cores.upper()} - ' if dis_cores not in Exclude_array else ''
		Fuses600w = f'600W::{fuses_applied.upper()} - ' if fuses_applied not in Exclude_array else ''

		CFC_String = f'CFC{CFC_Values_V}{CFC_Values_F} - ' if any([CFC_Values_V, CFC_Values_F]) else ''
		IA_String = f'IA{IA_Values_V}{IA_Values_F} - ' if any([IA_Values_V, IA_Values_F]) else ''

		DefeatureString = CFC_String + IA_String  + HTDIS + DisableModules +  Fuses600w

		# Append the row data to the summary list
		summary_data.append({
			'#': experiment_index_map[folder],  # Use the mapped index
			'Experiment': experiment,
			'Type': exptype,
			'PostCode': postcode,
			'Defeature': DefeatureString,
			'Mask': mask,
			'Status': status,
			'Used Content': used_content,
			'Content Results': content_results,
			'MCAs': mca_status,  # Add MCA presence
			'Comments': comments
		})

	summary_df = pd.DataFrame(summary_data)
	return summary_df

def format_voltage(value, volt_type, exclude_array):
	
	if value not in exclude_array:
		float_value = float(value)
		formatted_value = f'({value})' if float_value < 0 else value
		return_value = f'::{volt_type} {formatted_value}'
	else:
		return_value = ''
	
	return return_value

def format_frequency(value, exclude_array):
	
	if value not in exclude_array:
		return_value = f'::F{value}'
	else:
		return_value = ''
	
	return return_value

def save_to_excel(initial_df, test_df, summary_df, fail_info_df=None, unique_fails_df=None, unique_mcas_df=None, filename='output.xlsx'):
	
	def color_cell(df_data, cell, column_name = "Result", c1 = 'FAIL', fill_1 = 'FFC7CE', font_1 = '9C0006', c2 = 'PASS', fill_2 = 'C6EFCE', font_2 = '006100'):
			
			if df_data.columns[c_idx-1] == column_name:
				if value == c1:
					cell.fill = PatternFill(start_color=fill_1, end_color=fill_1, fill_type='solid')
					cell.font = Font(color=font_1)
				elif value == c2:
					cell.fill = PatternFill(start_color=fill_2, end_color=fill_2, fill_type='solid')
					cell.font = Font(color=font_2)

	# Create a workbook and add worksheets
	wb = Workbook()
	ws_data = wb.active
	ws_data.title = 'FrameworkData'  # Renamed to 'Data'
	ws_experiment_report = wb.create_sheet(title='ExperimentReport')


	# Build Table Styles
	style_data = TableStyleInfo(name="TableStyleLight8", showFirstColumn=False,
						   showLastColumn=False, showRowStripes=False, showColumnStripes=False)

	style_report = TableStyleInfo(name="TableStyleDark8", showFirstColumn=False,
						   showLastColumn=False, showRowStripes=False, showColumnStripes=False)
	
	style_file = TableStyleInfo(name="TableStyleLight1", showFirstColumn=False,
						   showLastColumn=False, showRowStripes=False, showColumnStripes=False)

	style_fails = TableStyleInfo(name="TableStyleLight9", showFirstColumn=False,
								 showLastColumn=False, showRowStripes=False, showColumnStripes=False)

	style_mcas = TableStyleInfo(name="TableStyleLight10", showFirstColumn=False,
								showLastColumn=False, showRowStripes=False, showColumnStripes=False)

	# Write test_df to the Data sheet
	for r_idx, row in enumerate(test_df.itertuples(), start=1):
		for c_idx, value in enumerate(row[1:], start=1):
			cell = ws_data.cell(row=r_idx+1, column=c_idx, value=value)  # Start from row 2 for headers
			if test_df.columns[c_idx-1] == 'Result':
				if value == 'FAIL':
					cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
					cell.font = Font(color='9C0006')
				elif value == 'PASS':
					cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
					cell.font = Font(color='006100')

	# Add headers to the Data sheet
	for c_idx, header in enumerate(test_df.columns, start=1):
		ws_data.cell(row=1, column=c_idx, value=header)

	# Adjust column widths in the Data sheet
	for column in ws_data.columns:
		max_length = max(len(str(cell.value)) for cell in column)
		adjusted_width = max_length + 2  # Add some padding
		ws_data.column_dimensions[column[0].column_letter].width = adjusted_width

	# Create a table in the Data sheet
	data_table = Table(displayName='FrameworkData', ref=f"A1:{ws_data.cell(row=len(test_df)+1, column=len(test_df.columns)).coordinate}")
	data_table.tableStyleInfo = style_data
	ws_data.add_table(data_table)

	# Write summary_df to the ExperimentReport sheet
	for r_idx, row in enumerate(summary_df.itertuples(), start=1):
		for c_idx, value in enumerate(row[1:], start=1):
			cell = ws_experiment_report.cell(row=r_idx+1, column=c_idx, value=value)  # Start from row 2 for headers
			
			color_cell(df_data=summary_df, cell=cell, column_name='Status')
			color_cell(df_data=summary_df, cell=cell, column_name='Content Results')
			color_cell(df_data=summary_df, cell=cell, column_name='MCAs', c1='YES', c2='NO')
			#if summary_df.columns[c_idx-1] == 'Status':
			#	if value == 'FAIL':
			#		cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
			#		cell.font = Font(color='9C0006')
			#	elif value == 'PASS':
			#		cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
			#		cell.font = Font(color='006100')

		#for c_idx, value in enumerate(row[1:], start=1):
		#	ws_experiment_report.cell(row=r_idx+1, column=c_idx, value=value)  # Start from row 2 for headers

	# Add headers to the ExperimentReport sheet
	for c_idx, header in enumerate(summary_df.columns, start=1):
		ws_experiment_report.cell(row=1, column=c_idx, value=header)

	# Adjust column widths in the ExperimentReport sheet
	for column in ws_experiment_report.columns:
		max_length = max(len(str(cell.value)) for cell in column)
		adjusted_width = max_length + 2
		ws_experiment_report.column_dimensions[column[0].column_letter].width = adjusted_width

	# Create a table in the ExperimentReport sheet
	experiment_report_table = Table(displayName='FrameworkSummary', ref=f"A1:{ws_experiment_report.cell(row=len(summary_df)+1, column=len(summary_df.columns)).coordinate}")
	experiment_report_table.tableStyleInfo = style_report
	ws_experiment_report.add_table(experiment_report_table)

	# Write fail_info_df to the FrameworkFails sheet if not empty or None
	if fail_info_df is not None and not fail_info_df.empty:
		
		ws_framework_fails = wb.create_sheet(title='FrameworkFails')
		
		for r_idx, row in enumerate(fail_info_df.itertuples(), start=1):
			for c_idx, value in enumerate(row[1:], start=1):
				ws_framework_fails.cell(row=r_idx+1, column=c_idx, value=value)  # Start from row 2 for headers

		# Add headers to the FrameworkFails sheet
		for c_idx, header in enumerate(fail_info_df.columns, start=1):
			ws_framework_fails.cell(row=1, column=c_idx, value=header)

		# Adjust column widths in the FrameworkFails sheet
		for column in ws_framework_fails.columns:
			max_length = max(len(str(cell.value)) for cell in column)
			adjusted_width = max_length + 2
			ws_framework_fails.column_dimensions[column[0].column_letter].width = adjusted_width

		# Create a table in the FrameworkFails sheet
		framework_fails_table = Table(displayName='FrameworkFails', ref=f"A1:{ws_framework_fails.cell(row=len(fail_info_df)+1, column=len(fail_info_df.columns)).coordinate}")
		framework_fails_table.tableStyleInfo = style_fails
		ws_framework_fails.add_table(framework_fails_table)

		# Write unique_fails_df to the UniqueFails sheet if not empty or None
		if unique_fails_df is not None and not unique_fails_df.empty:
			unique_fails_df = unique_fails_df.sort_values(by='Fail_Count', ascending=False)

			ws_unique_fails = wb.create_sheet(title='UniqueFails')
			for r_idx, row in enumerate(unique_fails_df.itertuples(), start=1):
				for c_idx, value in enumerate(row[1:], start=1):
					ws_unique_fails.cell(row=r_idx+1, column=c_idx, value=value)  # Start from row 2 for headers

			# Add headers to the UniqueFails sheet
			for c_idx, header in enumerate(unique_fails_df.columns, start=1):
				ws_unique_fails.cell(row=1, column=c_idx, value=header)

			# Adjust column widths in the UniqueFails sheet
			for column in ws_unique_fails.columns:
				max_length = max(len(str(cell.value)) for cell in column)
				adjusted_width = max_length + 2
				ws_unique_fails.column_dimensions[column[0].column_letter].width = adjusted_width

			# Create a table in the UniqueFails sheet
			unique_fails_table = Table(displayName='UniqueFails', ref=f"A1:{ws_unique_fails.cell(row=len(unique_fails_df)+1, column=len(unique_fails_df.columns)).coordinate}")
			unique_fails_table.tableStyleInfo = style_fails
			ws_unique_fails.add_table(unique_fails_table)

			# Write unique_mcas_df next to the UniqueFails table
			if unique_mcas_df is not None and not unique_mcas_df.empty:
				unique_mcas_df = unique_mcas_df.sort_values(by='Fail_Count', ascending=False)

				for r_idx, row in enumerate(unique_mcas_df.itertuples(), start=1):
					for c_idx, value in enumerate(row[1:], start=1):
						ws_unique_fails.cell(row=r_idx+1, column=c_idx+len(unique_fails_df.columns)+2, value=value)

				# Add headers for the unique_mcas_df
				for c_idx, header in enumerate(unique_mcas_df.columns, start=1):
					ws_unique_fails.cell(row=1, column=c_idx+len(unique_fails_df.columns)+2, value=header)

				# Adjust column widths for the unique_mcas_df
				for column in ws_unique_fails.iter_cols(min_col=len(unique_fails_df.columns)+3, max_col=len(unique_fails_df.columns)+2+len(unique_mcas_df.columns)):
					max_length = max(len(str(cell.value)) for cell in column)
					adjusted_width = max_length + 2
					ws_unique_fails.column_dimensions[column[0].column_letter].width = adjusted_width

				# Create a table for the unique_mcas_df
				unique_mcas_table = Table(displayName='UniqueMCAs', ref=f"{ws_unique_fails.cell(row=1, column=len(unique_fails_df.columns)+3).coordinate}:{ws_unique_fails.cell(row=len(unique_mcas_df)+1, column=len(unique_fails_df.columns)+2+len(unique_mcas_df.columns)).coordinate}")
				unique_mcas_table.tableStyleInfo = style_mcas
				ws_unique_fails.add_table(unique_mcas_table)

	# Create Framework Files Tab
	ws_framework_files = wb.create_sheet(title='FrameworkFiles')		
	# Write initial_df to the Framework Files sheet
	for r_idx, row in enumerate(initial_df.itertuples(), start=1):
		for c_idx, value in enumerate(row[1:], start=1):
			ws_framework_files.cell(row=r_idx+1, column=c_idx, value=value)  # Start from row 2 for headers

	# Add headers to the Framework Files sheet
	for c_idx, header in enumerate(initial_df.columns, start=1):
		ws_framework_files.cell(row=1, column=c_idx, value=header)

	# Adjust column widths in the Framework Files sheet
	for column in ws_framework_files.columns:
		max_length = max(len(str(cell.value)) for cell in column)
		adjusted_width = max_length + 2
		ws_framework_files.column_dimensions[column[0].column_letter].width = adjusted_width

	# Create a table in the Framework Files sheet
	framework_files_table = Table(displayName='FrameworkFiles', ref=f"A1:{ws_framework_files.cell(row=len(initial_df)+1, column=len(initial_df.columns)).coordinate}")
	framework_files_table.tableStyleInfo = style_file
	ws_framework_files.add_table(framework_files_table)

	# Save the workbook
	wb.save(filename)

def add_framework_fails_tab(save_path, zip_results):
	with pd.ExcelWriter(save_path, mode='a', engine='openpyxl') as writer:
		df = pd.DataFrame([
			{'#': idx + 1, 'Experiment': result['Experiment'], 'Fail Data': ', '.join(result['Data'])}
			for idx, result in enumerate(zip_results.values())
		])
		df.to_excel(writer, sheet_name='FrameworkFails', index=False)

def formatted_print(text, fore = Fore.RESET, back = Back.RESET):
	
	pretty_text = back + fore + text + Fore.RESET + Back.RESET
	
	print(pretty_text)

def generate_unique_fails(fail_info_df):

	# Split the fail data by commas and explode the DataFrame
	fail_info_df['Log Fails'] = fail_info_df['Log Fails'].apply(lambda x: x.split(','))
	exploded_df = fail_info_df.explode('Log Fails')

	# Filter out "PASS" entries
	#exploded_df = exploded_df[exploded_df['Fails'] != "PASS"]

	#unique_fails = fail_info_df.explode('Log Fails')
	unique_fails = exploded_df.groupby('Log Fails').agg(
		Content=('Content', lambda x: ', '.join(set(x))),
		Fail_Count=('Log Fails', 'size')
	).reset_index().rename(columns={'Log Fails': 'Failed Content'})
	
	# Return it to the previous format
	fail_info_df['Log Fails'] = fail_info_df['Log Fails'].apply(lambda x: ','.join(x))
	
	print(unique_fails)
	return unique_fails

def update_content_results(base_df, fail_info_df):
	for idx, row in base_df.iterrows():
		test_file = row['Test File']
		folder = row['Folder']
		fail_data = fail_info_df[(fail_info_df['Log File'] == test_file) & (fail_info_df['Experiment'] == folder)]
		if not fail_data.empty:
			fails = fail_data['Log Fails'].tolist()
			fails_list = fails[0].split(",")
			base_df.at[idx, 'Content Status'] = (', '.join(fails_list[:2]) + '...') if len(fails_list) > 2 else ','.join(fails_list)
	return base_df

def update_mca_results(test_df, fail_info_df, mca_df):
	# Call check_mca_presence to get MCA presence for each log file
	mca_presence = check_mca_presence(fail_info_df, mca_df)

	# Iterate over each row in test_df and update MCA Status
	for idx, row in test_df.iterrows():
		log_file = row['Test File']
		test_df.at[idx, 'MCA Status'] = mca_presence.get(log_file, 'NO')

	return test_df

def check_mca_presence(fail_info_df, mca_df):
	# Initialize a dictionary to store MCA presence for each log file
	mca_presence = {}

	# Iterate over each row in fail_info_df
	for _, row in fail_info_df.iterrows():
		experiment_name = row['Experiment']
		log_file = row['Log File']
		run_value = log_file.split('_')[0]

		# Check if there is any MCA data for the experiment and run_value
		mca_exists = not mca_df[(mca_df['Experiment'] == experiment_name) & (mca_df['Run Value'] == run_value)].empty

		# Set MCA presence to 'YES' or 'NO'
		mca_presence[log_file] = 'YES' if mca_exists else 'NO'

	return mca_presence

def array_merge(array1, array2, unique = True):
	new_array = list(set(array1 + array2)) if unique else array1 + array2
	return new_array

def test():
	
	#zip_file_path = r'R:/DebugFramework/GNR/754F39Y300403/cr03tppv0007en/20250707/20250707_192129_T0_BaselineChecks-VoltageSweeps_Sweep/ExperimentData.zip'
	#zip_file_path = r'R:/DebugFramework/GNR/754F39Y300403/cr03tppv0162en/20250728/20250728_112725_T8_Pseudo_HTDIS_Loops\ExperimentData.zip'
	zip_file_path = r'R:/DebugFramework/GNR/74698UA600040\cr03tppv0165en\20250725\20250724_154318_T1_Shmoo_Core_Volt_vs_Freq_Shmoo\ExperimentData.zip'

	content_type = 'Sandstone'  # Set the content type here
	pass_array = []#["exit: pass"]
	fail_array = []#["exit: fail", "not ok"]
	skip_array = ["0Sanity"]

	data = parse_zip_files(zip_file_path=zip_file_path, content=content_type, pass_array = pass_array, fail_array = fail_array, skip_array = skip_array , exclusion_string='pysv', casesens=False)

	print(data)
	
def test_log_file_parser():
	path = r'R:\DebugFramework\GNR\75VP061900080'
	Excel_kw = 'Summary'
	zip_kw = 'ExperimentData'
	output_file = r'C:\ParsingFiles\DebugFramework\Framework_Report.xlsx'

	data = find_files(base_folder=path, excel_keyword=Excel_kw, zip_keyword=zip_kw)
	
	log_path_dict = create_file_dict(data, 'Log')
	excel_path_dict = create_file_dict(data, 'Excel')
	zip_path_dict = create_file_dict(data, 'ZIP')
	
	print('Excel Files', excel_path_dict , '\n')
	print('Zip Files', zip_path_dict , '\n')  
	print('Log Files', log_path_dict , '\n')  
	print('Data Files', data , '\n')

	# Parse log files and create test data DataFrame
	test_df = parse_log_files(log_path_dict)

	# Example usage:
		
	#excel_files = [info['path'] for info in excel_path_dict.values()]
	print('file:', excel_path_dict)
	log_summary = LogSummaryParser(excel_path_dict=excel_path_dict, test_df=test_df)
	report_df, mca_df = log_summary.parse_mca_tabs_from_files()
	print(report_df)
	print("")
	print(mca_df)
	
	# Output results using tabulate
	#print("\nTest Data DataFrame:")
	#print(tabulate(test_df, headers='keys', tablefmt='grid'))
	
	# Save DataFrames to Excel
	# save_to_excel(data, test_df, filename=output_file)

if __name__ == "__main__":
	test_log_file_parser()