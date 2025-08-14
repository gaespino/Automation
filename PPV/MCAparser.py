#import argparse
#from traceback import extract_tb
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import shutil
#import win32com.client as win32
import xlwings as xw
import sys
import os
#import json
#import re
import Decoder.decoder as mcparse


def init_select_data(product):
		
	atomlist = ['SRF','CWF']
	coretype = 'atom' if product in atomlist else 'bigcore'
	reduced_data_cha = {}
	reduced_data_core = {}
	reduced_data_others = {}

	if coretype == 'bigcore':## Declaring reduced data Information
		reduced_data_cha = {
								'UTIL__MC_STATUS': '0X20000000000000',
								'UTIL__MC_ADDR':None, 
								'__MCI_STATUS':'0X20000000000000',
								'__MCI_MISC':'0X80',
								'__MCI_ADDR':None,
								'UTIL__MC_MISC':'0X80',
	#							'__MC_MISC3':None,
	#							'BIOS':None,
								'UBOX':None,
								'FW_ERR_CAUSE': None,
								'S3M_ERR_STS' : None,
								'PTPCFSMS__MC_STATUS':None}
		reduced_data_core = {
								'ML2_CR_MC3': '0X7F', # ML2 MCAs 
								'ML3_CR_PIC_EXTENDED_LOCAL_APIC_ID':None,
								'IFU_CR_MC0':'0X1FFF', # IFU MCAs 
								'DCU_CR_MC1':'0X1F', # DCU MCAs 
								'DTLB_CR_MC2':'0X3F', # DTLB MCAs 
								'ROB1_CR_MC':None, # ROB1 MCAs -- to be included in excel dashboard
								'C6SRAM_MCA_STATUS':None, # C6SRAM MCAs -- to be included in excel dashboard
								'PMSB':None}
			## Will use this for some misc fails, moving PM and MEM errors here
		reduced_data_others = {
								'MEMSS__B2CMI': '0X7F', # ML2 MCAs 
								'ML3_CR_PIC_EXTENDED_LOCAL_APIC_ID':None,
								'IFU_CR_MC0':'0X1FFF', # IFU MCAs 
								'DCU_CR_MC1':'0X1F', # DCU MCAs 
								'DTLB_CR_MC2':'0X3F', # DTLB MCAs 
								'ROB1_CR_MC':None, # DTLB MCAs
								'C6SRAM_MCA_STATUS':None, # DTLB MCAs
								'PMSB':None}
		
	if coretype == 'atom':## Declaring reduced data Information
		reduced_data_cha = {
								'UTIL__MC_STATUS': '0X20000000000000',
								'UTIL__MC_ADDR':None, 
								'__MCI_STATUS':'0X20000000000000',
								'__MCI_MISC':'0X80',
								'__MCI_ADDR':None,
								'UTIL__MC_MISC':'0X80',
	#							'__MC_MISC3':None,
	#							'BIOS':None,
								'UBOX':None,
								'FW_ERR_CAUSE': None,
								'S3M_ERR_STS' : None,
								'PTPCFSMS__MC_STATUS':None}
		reduced_data_core = {
								'_CR_MCI_CTRL': '0X500000000', # CORE Control register
								'_CR_MCI_STATUS':'0X20000000000000', # CORE MCAS
								'_RESULT__':'PASS',
								'_TEST__':None, # IFU MCAs 
								'__ID___CORE__':None, # DCU MCAs 
								'__ID___LOGICAL__':None, # DTLB MCAs 
								'__ID___PACKAGE__':None, # ROB1 MCAs -- to be included in excel dashboard
								'__ID___THREAD__':None, # C6SRAM MCAs -- to be included in excel dashboard
								'__MESSAGES___0___TEXT__':None,
								'__MESSAGES___0___LEVEL__':None,
								'__AVG_FREQ_MHZ__':None,
								'__FAIL___TIME_TO_FAIL__':None,
								'__FAIL___SEED__':None,
								'__FAIL___CPU_MASK__':None,
								}
			## Will use this for some misc fails, moving PM and MEM errors here
		reduced_data_others = {
								'MEMSS__B2CMI': '0X7F', # ML2 MCAs 
								'ML3_CR_PIC_EXTENDED_LOCAL_APIC_ID':None,
								'IFU_CR_MC0':'0X1FFF', # IFU MCAs 
								'DCU_CR_MC1':'0X1F', # DCU MCAs 
								'DTLB_CR_MC2':'0X3F', # DTLB MCAs 
								'ROB1_CR_MC':None, # DTLB MCAs
								'C6SRAM_MCA_STATUS':None, # DTLB MCAs
								'PMSB':None}

	return reduced_data_cha, reduced_data_core, reduced_data_others

class ppv_report():
	def __init__(self, name, week, label, source_file, report, data_core = None, data_cha = None, reduced = False, mcdetail = True, overview = False, decode = False, mode='Bucketer'):

		self.source_file = rf'{source_file}'
		self.source_sheet = 'raw_data'
		#self.data_file = data_file
		#self.mca_file = mca_file
		self.sheet_CHA = 'CHA'
		self.sheet_CORE = 'CORE'
		self.sheet_PPV = 'PPV'
		#self.raw_data = 'raw_data'
		self.table_cha = 'cha_mc'
		self.table_core = 'core_mc'
		self.table_ppv = 'ppv'
		#self.data_core = data_core
		#self.data_cha = data_cha
		self.reduced = reduced
		self.ovw = overview
		self.decode = decode
		self.mcfile = mcdetail
		## File Initialization

		self.name = name
		self.product = name.upper() ## Setting it the same as the name, can be changed later
		self.week = week
		self.label = label
		self.output = report

		## Templates management
		self.base_dir = os.path.dirname(os.path.abspath(__file__))
		self.templates_dir = os.path.join(self.base_dir, 'MCChecker')
		self.templates_dir = os.path.join(self.templates_dir, self.product)
		self.data_template = '##Name##_##w##_##LABEL##_PPV_Data.xlsx'
		self.mca_template = '##Name##_##w##_##LABEL##_PPV_MC_Checker.xlsm'
		self.ovw_template = '##Name##_##w##_##LABEL##_PPV_Unit_Overview.xlsx'
		self.template_file_MCchk = os.path.join(self.templates_dir, self.mca_template) ## Not used in CWF and newer products
		self.template_file = os.path.join(self.templates_dir, self.data_template)

		## New Data and MCA file from templates
		self.data_file, self.mca_file, self.ovw_file = self.file_rename()
		self.data_file = os.path.join(self.output, self.data_file)
		self.mca_file = os.path.join(self.output, self.mca_file)

		if mode == 'Bucketer': filecopy(self.template_file, self.data_file)
		if self.mcfile: filecopy(self.template_file_MCchk, self.mca_file)

		## OVerview file option selected -- Only works with Reduced selected
		if self.ovw: 
			if self.reduced:
				
				self.template_file_ovw = os.path.join(self.templates_dir, self.ovw_template)
				self.ovw_file = os.path.join(self.output, self.ovw_file)
				filecopy(self.template_file_ovw, self.ovw_file)
			else:
				print(' -- Overview file selected but not in reduced data mode, select reduced mode to create the file. Skipping... ')

		## Declaring non reduced data
		data_CHA = ['CHA', 'LLC', 'BIOS']
		data_CORE = ['CPU', 'PMSB']

		## Collects strings for reduced data
		reduced_data_cha, reduced_data_core, reduced_data_others = init_select_data(self.product)
		## Declaring reduced data Information
		#reduced_data_cha = {
		#					'UTIL__MC_STATUS': '0X20000000000000',
		#					'UTIL__MC_ADDR':None, 
		#					'__MCI_STATUS':'0X20000000000000',
		#					'__MCI_MISC':'0X80',
		#					'__MCI_ADDR':None,
		#					'UTIL__MC_MISC':'0X80',
#							'__MC_MISC3':None,
#							'BIOS':None,
		#					'UBOX':None,
		#					'FW_ERR_CAUSE': None,
		#					'S3M_ERR_STS' : None,
		#					'PTPCFSMS__MC_STATUS':None}
		#reduced_data_core = {
		#					'ML2_CR_MC3': '0X7F', # ML2 MCAs 
		#					'ML3_CR_PIC_EXTENDED_LOCAL_APIC_ID':None,
		#					'IFU_CR_MC0':'0X1FFF', # IFU MCAs 
		#					'DCU_CR_MC1':'0X1F', # DCU MCAs 
		#					'DTLB_CR_MC2':'0X3F', # DTLB MCAs 
		#					'ROB1_CR_MC':None, # ROB1 MCAs -- to be included in excel dashboard
		#					'C6SRAM_MCA_STATUS':None, # C6SRAM MCAs -- to be included in excel dashboard
		#					'PMSB':None}
		## Will use this for some misc fails, moving PM and MEM errors here
		#reduced_data_others = {
		#					'MEMSS__B2CMI': '0X7F', # ML2 MCAs 
		#					'ML3_CR_PIC_EXTENDED_LOCAL_APIC_ID':None,
		#					'IFU_CR_MC0':'0X1FFF', # IFU MCAs 
		#					'DCU_CR_MC1':'0X1F', # DCU MCAs 
		#					'DTLB_CR_MC2':'0X3F', # DTLB MCAs 
		#					'ROB1_CR_MC':None, # DTLB MCAs
		#					'C6SRAM_MCA_STATUS':None, # DTLB MCAs
		#					'PMSB':None}

		if data_cha == None:
			if self.reduced:
				self.data_cha = reduced_data_cha
			else:
				self.data_cha = data_CHA

		if data_core == None:
			if self.reduced:
				self.data_core = reduced_data_core
			else:
				self.data_core = data_CORE

		#self.target_sheet = target_sheet


	def run(self, options = ['MESH', 'CORE']):

		# Update Data table

		ovw_on = self.ovw & self.reduced
		mcfile_on = self.mcfile
		decode = self.decode

		for option in options:
			self.parse_data(option=option)

		# Adds the PPV Final Bucket info to the file
		self.bucket_info()
		options.append('PPV')
		
		# Add MCA data if selected --- CHA only for now
		if 'MESH' in options and decode:
			print(' -- Parsing MCA Data for CHA in tab CHA_MCAs...')
			self.parse_mcas(self.data_file, self.sheet_CHA)
		if 'CORE' in options and decode:
			self.parse_CORE_mcas(self.data_file, self.sheet_CORE)
		
		self.gen_auxfiles(data_file = self.data_file, mca_file=self.mca_file, ovw_file=self.ovw_file, mcfile_on=mcfile_on, ovw_on= ovw_on, options = options)
		
		print(f' !!! New file report created succesfully !!!')

	def gen_auxfiles(self, data_file, mca_file, ovw_file, mcfile_on=False, ovw_on= True, options = ['MESH', 'CORE']):
    
		# Open the Data file and MCA file to start moving the new table to the report
		if mcfile_on or ovw_on: 
			source_wb = file_open(file=data_file)
		if mcfile_on: 
			target_wb = file_open(file=mca_file)

		if ovw_on: 
			ovw_wb = file_open(file=ovw_file)
		# Update MCA Report
			
		for option in options:
			if mcfile_on: self.copy_table_data(source_wb, target_wb, option = option)
			if ovw_on: self.copy_table_data(source_wb, ovw_wb, option = option)
		
		# Close the files
		if mcfile_on or ovw_on: file_close(source_wb, save=False)
		if mcfile_on: file_close(target_wb)
		if ovw_on: file_close(ovw_wb)
		

	def parse_data(self, option = 'MESH'):

		# Variables Init
		source_file = self.source_file
		source_sheet = self.source_sheet
		target_file = self.data_file
		reduced = self.reduced
		
		print(f' -- Filtering data from dpmb source file to create data for {option} MCA...')
		
		## Selects which table to update
		if option == 'MESH': 
			data_to_port = self.data_cha
			target_sheet_name = self.sheet_CHA
			table_name = self.table_cha

		elif option == 'CORE': 
			data_to_port = self.data_core
			target_sheet_name = self.sheet_CORE
			table_name = self.table_core

		else: 
			print('No valid option selected, use: MESH or CORE')
			sys.exit()

		# Read data from the source Excel file and sheet
		source_data = pd.read_excel(source_file, sheet_name=source_sheet)
		#extra_data = pd.DataFrame(columns=['PPVRun_MC'])
		source_data['PPVRun_MC'] = source_data['LotsSeqKey'].astype(str) + "-" + source_data['UnitTestingSeqKey'].astype(str)
		
		#source_data = pd.concat([source_data, extra_data], ignore_index=True)	
		# Filter the data based on the conditions
		if reduced:
			filtered_data = pd.DataFrame()
			for key in data_to_port:
				keyisthere = source_data['TestName'].str.contains(key)
				if len(keyisthere) > 0:
					if data_to_port[key] is not None:
						_filtered_data = source_data[(source_data['TestName'].str.contains(key)) & (source_data['TestValue'] != data_to_port[key])]
						filtered_data = pd.concat([filtered_data, _filtered_data])
					else:
						_filtered_data = source_data[(source_data['TestName'].str.contains(key))]
						filtered_data = pd.concat([filtered_data, _filtered_data])
		else:
			filtered_data = source_data[(source_data['TestName'].str.contains('|'.join(data_to_port)))]

		# Load the target Excel file
		target_workbook = load_workbook(target_file)
		
		# Get the target sheet
		target_sheet = target_workbook[target_sheet_name]
		
		# Find the named table in the target sheet
		table = None
		for tbl in target_sheet.tables.values():
			if tbl.name == table_name:
				table = tbl
				break
		
		#if table is None:
			raise ValueError(f"Table {table_name} does not exist in the target file.")
		
		# Get the headers from the table
		table_range = target_sheet[table.ref]
		headers = [cell.value for cell in table_range[0]]
		
		# Prepare the data to be updated in the table
		data_to_update = []
		for _, row in filtered_data.iterrows():
			data_row = []
			for header in headers[:10]:
				data_row.append(row[header])
			data_to_update.append(data_row)
		
		# Update the table data
		start_row = table_range[1][0].row
		start_col = table_range[1][0].column
		
		for i, data_row in enumerate(data_to_update):
			for j, value in enumerate(data_row):
				target_sheet.cell(row=start_row + i, column=start_col + j, value=value)
		
		# Update the table range to include the new data
		end_row = start_row + len(data_to_update) - 1
		table.ref = f"{table_range[0][0].coordinate}:{target_sheet.cell(row=end_row, column=start_col + len(headers) - 1).coordinate}"
		
		# Replicate the formula in the last column to the newly added rows
		last_col_letter = target_sheet.cell(row=start_row, column=start_col + len(headers) - 1).column_letter
		formula_cell = target_sheet[f"{last_col_letter}{start_row}"]
		if formula_cell.data_type == 'f':  # Check if the cell contains a formula
			formula = formula_cell.value
			for row in range(start_row + 1, end_row + 1):
				target_sheet[f"{last_col_letter}{row}"].value = formula

		# Save the updated target workbook
		target_workbook.save(target_file)

	def bucket_info(self):

		# Variables Init
		source_file = self.source_file
		#source_sheet = self.source_sheet
		target_file = self.data_file
		#reduced = self.reduced
		
		print(f' -- Checking for final bucket data...')

		# Read data from the source Excel file and sheet
		source_data = pd.read_excel(source_file, sheet_name='final_bucket')
		content_data = pd.read_excel(source_file, sheet_name='results')
		
		# Add the BinData to the source_data dataframe

		ppvBin = content_data.groupby('VisualId')['BinDesc'].apply(lambda x: ', '.join(x.dropna().astype(str))).reset_index()

		source_data = source_data.merge(ppvBin, on='VisualId', how='left')

		# Load the target Excel file
		target_workbook = load_workbook(target_file)
		
		# Get the target sheet
		target_sheet = target_workbook['PPV']
		
		# Find the named table in the target sheet
		table = None
		for tbl in target_sheet.tables.values():
			if tbl.name == 'ppv':
				table = tbl
				break
		
		if table is None:
			raise ValueError(f"Table {'ppv'} does not exist in the target file.")
		
		# Get the headers from the table
		table_range = target_sheet[table.ref]
		headers = [cell.value for cell in table_range[0]]
		
		# Prepare the data to be updated in the table
		data_to_update = []
		for _, row in source_data.iterrows():
			data_row = []
			for header in headers:
				data_row.append(row[header])
			data_to_update.append(data_row)
		
		# Update the table data
		start_row = table_range[1][0].row
		start_col = table_range[1][0].column
		
		for i, data_row in enumerate(data_to_update):
			for j, value in enumerate(data_row):
				target_sheet.cell(row=start_row + i, column=start_col + j, value=value)
		
		# Update the table range to include the new data
		end_row = start_row + len(data_to_update) - 1
		table.ref = f"{table_range[0][0].coordinate}:{target_sheet.cell(row=end_row, column=start_col + len(headers) - 1).coordinate}"
		
		# Save the updated target workbook
		target_workbook.save(target_file)

	def copy_table_data(self, source_wb, target_wb, option = 'MESH'):

		# Variables Init
		#source_wb = self.data_file
		#target_wb = self.mca_file
		
		## Selects which table to update
		if option == 'MESH': 
			sheet_name = self.sheet_CHA
			table_name = self.table_cha

		elif option == 'CORE': 
			sheet_name = self.sheet_CORE
			table_name = self.table_core

		elif option == 'PPV': 
			sheet_name = self.sheet_PPV
			table_name = self.table_ppv

		else: 
			print('No valid option selected, use: MESH or CORE')
			sys.exit()

		# Get the source and target tables
		source_table = source_wb.sheets[sheet_name].tables[table_name]
		target_table = target_wb.sheets[sheet_name].tables[table_name]

		# Get the range of the source table
		source_range = source_table.data_body_range

		# Clear the target table data
		target_table.data_body_range.clear_contents()

		# Copy the data from the source table to the target table
		target_table.data_body_range.value = source_range.value

		print(f"Data copied successfully from source to target workbook {sheet_name} table {table_name}.")

	def file_rename(self):
		name = self.name
		week = self.week
		label = self.label
		# Generate the target file name based on user input
		target_file = self.data_template.replace("##Name##", name).replace("##w##", week).replace("##LABEL##", label)
		target_file_MC = self.mca_template.replace("##Name##", name).replace("##w##", week).replace("##LABEL##", label)
		target_file_ovw = self.ovw_template.replace("##Name##", name).replace("##w##", week).replace("##LABEL##", label)

		return target_file, target_file_MC, target_file_ovw

	def parse_mcas(self, source_file, source_sheet):
		
		mcas = pd.read_excel(source_file, sheet_name=source_sheet)

		
		# Call decoder
		mc = mcparse.decoder(data= mcas, product=self.product)
		
		cha_df = mc.cha()
		llc_df = mc.llc()
		ubox_df = mc.portids()

		# Save the new dataframe to a new worksheet named "CHA_MCAS"
		with pd.ExcelWriter(source_file, engine='openpyxl', mode='a') as writer:
			cha_df.to_excel(writer, sheet_name='CHA_MCAS', index=False)
			llc_df.to_excel(writer, sheet_name='LLC_MCAS', index=False)
			ubox_df.to_excel(writer, sheet_name='UBOX', index=False)
		#with pd.ExcelWriter(source_file, engine='openpyxl', mode='a') as writer:
			
		
		addtable(df=cha_df, excel_file=source_file, sheet='CHA_MCAS', table_name='chadecode')
		addtable(df=llc_df, excel_file=source_file, sheet='LLC_MCAS', table_name='llcdecode')
		addtable(df=ubox_df, excel_file=source_file, sheet='UBOX', table_name='uboxdecode')

	def parse_CORE_mcas(self, source_file, source_sheet):
    		
		mcas = pd.read_excel(source_file, sheet_name=source_sheet)
	
		# Call decoder
		mc = mcparse.decoder(data= mcas, product=self.product)
		core_df = mc.core()

		with pd.ExcelWriter(source_file, engine='openpyxl', mode='a') as writer:
			core_df.to_excel(writer, sheet_name='CORE_MCAS', index=False)

		addtable(df=core_df, excel_file=source_file, sheet='CORE_MCAS', table_name='coredecode')

# File manipulation scripts
def file_open(file):
	# Variables Init
	#source_file = self.data_file
	#target_file = self.mca_file

	# Open the source and target workbooks
	wb = xw.Book(file)
	#target_wb = xw.Book(target_file)

	return wb

def file_close(file, save = True): #source_wb, target_wb):
		
	# Variables Init
	#source_wb = self.source_wb
	#target_wb = self.target_wb
		
	# Save and close the workbooks
	if save: file.save()
	file.close()
	#target_wb.close()
		
def filecopy(src, dst):
	print(f' -- Duplicating file from template, new file located in: {dst}.')
	shutil.copy(src, dst)

# Used for decoder to add tables to templates
def addtable(df, excel_file, sheet, table_name ):
	# Load the workbook and the sheet
	wb = load_workbook(excel_file)
	ws = wb[sheet]

	# Define the table range, in case there is no data add blanks to the first column, this is just to be consistent with data 
	table_range = f'A1:{chr(64+len(df.columns))}{len(df)+1 if not df.empty else 2}'

	# Create a table
	table = Table(displayName=table_name, ref=table_range)

	# Add a default style with striped rows and banded columns
	style = TableStyleInfo(
		name="TableStyleMedium4",
		showFirstColumn=False,
		showLastColumn=False,
		showRowStripes=False,
		showColumnStripes=False
	)
	table.tableStyleInfo = style

	# Add the table to the sheet
	ws.add_table(table)
		# Save the workbook
	wb.save(excel_file)
	wb.close()

# Not used for now
def load_dataframe_to_excel(df, excel_file, sheet_name, table_name):
	"""
	Load a DataFrame into an existing Excel table.

	:param df: The DataFrame to load.
	:param excel_file: The path to the Excel file.
	:param sheet_name: The name of the sheet containing the table.
	:param table_name: The name of the table in the sheet.
	"""
	# Load the existing workbook
	workbook = load_workbook(excel_file)
		
	# Select the sheet
	sheet = workbook[sheet_name]
		
	# Find the table in the sheet
	table = None
	if table_name in sheet.tables.values():
		table = table_name
		
	else:
		table_range = f"A1:{chr(64+len(df.columns))}{len(df)+1}"
		table = table_name
		table = Table(displayName="chadecode", ref=table_range)
		# Add a default style with striped rows and banded columns
		style = TableStyleInfo(
			name="TableStyleMedium9",
			showFirstColumn=False,
			showLastColumn=False,
			showRowStripes=True,
			showColumnStripes=True
		)
		table.tableStyleInfo = style	
		
		# Add the table to the sheet
		sheet.add_table(table)

	if table is None:
		raise ValueError(f"Table {table_name} not found in sheet {sheet_name}")
	
	# Get the headers from the table
	table_range = sheet[table.ref]
	# Update the table data
	start_row = table_range[0][0].row
	start_col = table_range[0][0].column
	
	# Write the DataFrame to the sheet starting from the table's starting cell
	for r_idx, row in enumerate(df.itertuples(index=False), 1):
		for c_idx, value in enumerate(row, 1):
			sheet.cell(row=start_row + r_idx, column=start_col + c_idx, value=value)
		
	# Save the workbook
	workbook.save(excel_file)

#if __name__ == "__main__":
def test(): ## Comment and run with above line, not setting args for this one, use the UI
	# Example usage
	source_file = r'C:\ParsingFiles\GNR'
	filename = r'\GNR3_output_result_combined.xlsx'
	source_file = source_file + filename
	path = r'C:\ParsingFiles\GNR'
	#source_sheet = 'raw_data'
	#template_file_MCchk = r'C:\Git\Automation\##Name##_##w##_##LABEL##_PPV_MC_Checker.xlsm'
	#template_file = r'C:\Git\Automation\##Name##_##w##_##LABEL##_PPV_Data.xlsx'


	#target_file_template = '##Name##_##w##_##LABEL##_PPV_Data.xlsx'
	#MCcheck_file = '##Name##_##w##_##LABEL##_PPV_MC_Checker.xlsm'
	#table_name = 'cha_mc'

	# User inputs
	name = 'GNR'
	week = 'WW24'
	label = 'LLC_DATA_PAR'
	
	reduced_data_cha = {'UTIL__MC_STATUS': '0X20000000000000', 'LLC':'0X20000000000000', 'BIOS':None}
	reduced_data_core = {'ML2_CR_MC3_STATUS': '0X20000000000000', 'PMSB':None}
	table_dict = {'CHA':'cha_mc', 'CORE':'core_mc'}


	data_CHA = ['CHA', 'LLC', 'BIOS']
	target_sheet = 'CHA'
	# Generate the target file name based on user input
	#target_file = template_file.replace("##Name##", name).replace("##w##", week).replace("##LABEL##", label)
	#target_file_MC = template_file_MCchk.replace("##Name##", name).replace("##w##", week).replace("##LABEL##", label)

	#shutil.copy(template_file, target_file)
	#shutil.copy(template_file_MCchk, target_file_MC)

	# Copy the template file to the new target file
	PPVMCAs = ppv_report(name=name, week=week, label=label, source_file=source_file, report = path, reduced = True, mcdetail=False, overview = True, decode = True)
	PPVMCAs.run()


	# Update Data File
	#PPVMCAs.parse_data(option = 'MESH')
	#PPVMCAs.parse_data(option = 'CORE')


	# 
	## Update MCA file with the PPV MCA data 
	#source_wb, target_wb = file_open(data_file=target_file, mca_file=target_file_MC)
	#PPVMCAs.copy_table_data(source_wb=source_wb, target_wb=target_wb)
	#file_close(source_wb=source_wb, target_wb=target_wb)

	#target_sheet = 'CORE'
	#data_CORE = ['CPU', 'PMSB']
	#table_name = 'core_mc'
	# Update Core Data
	#parse_data(source_file, source_sheet, target_sheet, target_file, table_name, data_to_port=data_CORE, reduced=False)

	#update_table_data(source_file, source_sheet, target_sheet, target_file, table_name, data_to_port=data_CORE, reduced=False)

	#source_wb, target_wb = file_open(source_file=target_file, target_file=target_file_MC)
	#copy_table_data(source_wb, target_wb , table_name='cha_mc', sheet_name ='CHA' )
	#copy_table_data(source_wb, target_wb , table_name='core_mc', sheet_name ='CORE' )

	#file_close(source_wb, target_wb)


	# Refresh file formulas
	#refresh_formulas2(target_file)
	# Update MCAs in Sheet
	#run_vba_macro(target_file, macro_name='HighlightCells')

if __name__ == "__main__":

	test()