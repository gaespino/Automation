"""
Framework File Visual ID Fix
=============================
Development tool for correcting incorrect Visual IDs saved into framework output folders.

Supports two modes:
  - Single mode : manually specify a folder, old word, and new word.
  - Batch mode  : automatically scan a root folder tree for Summary_<TOKEN>_*.xlsx files,
                  build a per-folder row table, then execute replacements in bulk.

All original files are backed up to C:\\Temp\\FrameworkBackup\\<timestamp> before any change.
Processing logs are streamed to an in-window log panel in real time.
"""

import os
import re
import shutil
import zipfile
import threading
import datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
TEMP_DIR      = r'C:\Temp\FrameworkFileFix'
BACKUP_ROOT   = r'C:\Temp\FrameworkBackup'
ACCENT_COLOR  = '#1abc9c'
HEADER_BG     = '#2c3e50'
BTN_PRIMARY   = '#1abc9c'
BTN_SCAN      = '#3498db'
BTN_EXEC      = '#1abc9c'
BTN_DANGER    = '#e74c3c'
BTN_NEUTRAL   = '#95a5a6'
FONT_UI       = ('Segoe UI', 9)
FONT_UI_BOLD  = ('Segoe UI', 9, 'bold')
FONT_HEADER   = ('Segoe UI', 13, 'bold')
FONT_LOG      = ('Consolas', 9)

# Matches: Summary_TOKEN_*.xlsx  OR  Summary_[TOKEN]_*.xlsx
_SUMMARY_RE = re.compile(r'^Summary_\[?([^\]_]+)\]?_.*\.xlsx$', re.IGNORECASE)


# ---------------------------------------------------------------------------
# Backend utilities
# ---------------------------------------------------------------------------

def replace_word_in_file(file_path, old_word, new_word, log=None):
    """Replace all occurrences of old_word in a plain-text file (UTF-8)."""
    try:
        with open(file_path, 'r', encoding='utf-8') as fh:
            content = fh.read()
        updated = content.replace(old_word, new_word)
        with open(file_path, 'w', encoding='utf-8') as fh:
            fh.write(updated)
        if log:
            log(f'  Text replaced in: {os.path.basename(file_path)}', 'info')
    except Exception as exc:
        if log:
            log(f'  ERROR replacing text in {os.path.basename(file_path)}: {exc}', 'error')
        raise


def replace_word_in_excel(file_path, old_word, new_word, log=None):
    """Replace all string-cell occurrences of old_word in every sheet of an xlsx file."""
    try:
        wb = load_workbook(file_path)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and old_word in cell.value:
                        cell.value = cell.value.replace(old_word, new_word)
        wb.save(file_path)
        wb.close()
        if log:
            log(f'  Excel updated: {os.path.basename(file_path)}', 'info')
    except Exception as exc:
        if log:
            log(f'  ERROR updating Excel {os.path.basename(file_path)}: {exc}', 'error')
        raise


def process_zip_file(zip_path, old_word, new_word, temp_dir, log=None):
    """Extract zip → apply replacements → rename files inside → repack with new name."""
    zip_temp_dir = os.path.join(temp_dir, 'zip_temp')
    os.makedirs(zip_temp_dir, exist_ok=True)
    try:
        with zipfile.ZipFile(zip_path, 'r') as zip_in:
            zip_in.extractall(zip_temp_dir)

        for root, _dirs, files in os.walk(zip_temp_dir):
            for fname in files:
                fpath = os.path.join(root, fname)
                if fname.lower().endswith(('.txt', '.log')):
                    replace_word_in_file(fpath, old_word, new_word, log)
                elif fname.lower().endswith('.xlsx'):
                    replace_word_in_excel(fpath, old_word, new_word, log)
                new_fname = fname.replace(old_word, new_word)
                new_fpath = os.path.join(root, new_fname)
                if fpath != new_fpath:
                    os.rename(fpath, new_fpath)

        new_zip_name = os.path.basename(zip_path).replace(old_word, new_word)
        new_zip_path = os.path.join(temp_dir, new_zip_name)
        with zipfile.ZipFile(new_zip_path, 'w', compression=zipfile.ZIP_DEFLATED) as zip_out:
            for root, _dirs, files in os.walk(zip_temp_dir):
                for fname in files:
                    fpath = os.path.join(root, fname)
                    zip_out.write(fpath, os.path.relpath(fpath, zip_temp_dir))
        if log:
            log(f'  ZIP repacked: {new_zip_name}', 'info')
    finally:
        shutil.rmtree(zip_temp_dir, ignore_errors=True)


def replace_word_in_folder(folder_path, old_word, new_word, log=None):
    """
    Core replacement routine:
      1. Walk folder_path and process .log / .zip / .xlsx files via temp staging dir.
      2. Move processed files back; delete original .xlsx (replaced by updated copy).
    """
    os.makedirs(TEMP_DIR, exist_ok=True)

    if log:
        log(f'Processing folder: {folder_path}', 'info')

    for root, _dirs, files in os.walk(folder_path):
        for fname in files:
            src_path   = os.path.join(root, fname)
            new_fname  = fname.replace(old_word, new_word)
            dst_path   = os.path.join(TEMP_DIR, new_fname)

            if fname.lower().endswith('.log'):
                shutil.copy2(src_path, dst_path)
                replace_word_in_file(dst_path, old_word, new_word, log)
            elif fname.lower().endswith('.zip'):
                process_zip_file(src_path, old_word, new_word, TEMP_DIR, log)
            elif fname.lower().endswith('.xlsx'):
                shutil.copy2(src_path, dst_path)
                replace_word_in_excel(dst_path, old_word, new_word, log)

    # Move processed files back to source folder
    processed = os.listdir(TEMP_DIR)
    for fname in processed:
        src = os.path.join(TEMP_DIR, fname)
        dst = os.path.join(folder_path, fname)
        shutil.move(src, dst)
        if log:
            log(f'  Moved to folder: {fname}', 'success')

        # Remove original xlsx (it was replaced by the updated copy)
        if fname.lower().endswith('.xlsx'):
            original = os.path.join(folder_path, fname.replace(new_word, old_word))
            if os.path.exists(original) and original != dst:
                os.remove(original)
                if log:
                    log(f'  Removed original: {os.path.basename(original)}', 'info')

    shutil.rmtree(TEMP_DIR, ignore_errors=True)
    if log:
        log(f'Folder complete: {folder_path}', 'success')


def backup_folder(source_path, log=None):
    """
    Copy source_path into C:\\Temp\\FrameworkBackup\\<YYYYMMDD_HHMMSS>\\<folder_name>
    before any modification.  Returns the backup destination path.
    """
    timestamp   = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    folder_name = os.path.basename(os.path.normpath(source_path))
    dest        = os.path.join(BACKUP_ROOT, timestamp, folder_name)
    os.makedirs(dest, exist_ok=True)
    try:
        for item in os.listdir(source_path):
            src_item = os.path.join(source_path, item)
            dst_item = os.path.join(dest, item)
            if os.path.isfile(src_item):
                shutil.copy2(src_item, dst_item)
            elif os.path.isdir(src_item):
                shutil.copytree(src_item, dst_item)
        if log:
            log(f'Backup created: {dest}', 'success')
    except Exception as exc:
        if log:
            log(f'WARNING: backup incomplete for {source_path}: {exc}', 'warning')
    return dest


def scan_for_summary_files(root_path):
    """
    Walk root_path recursively.  For every file matching Summary_<TOKEN>_*.xlsx
    (with or without square brackets), record:
      - path     : the folder containing the file
      - old_word : extracted TOKEN (brackets stripped)
      - filename : matched filename for display

    Returns a list of dicts, deduplicated by (path, old_word).
    """
    seen    = set()
    results = []
    for dirpath, _dirs, files in os.walk(root_path):
        for fname in files:
            m = _SUMMARY_RE.match(fname)
            if m:
                token = m.group(1)
                key   = (dirpath, token)
                if key not in seen:
                    seen.add(key)
                    results.append({
                        'path'    : dirpath,
                        'old_word': token,
                        'filename': fname,
                    })
    return results


# ---------------------------------------------------------------------------
# Tooltip helper
# ---------------------------------------------------------------------------

class _ToolTip:
    """Simple hover tooltip for any widget."""
    def __init__(self, widget, text):
        self._widget = widget
        self._text   = text
        self._win    = None
        widget.bind('<Enter>', self._show)
        widget.bind('<Leave>', self._hide)

    def _show(self, _event=None):
        x = self._widget.winfo_rootx() + 20
        y = self._widget.winfo_rooty() + self._widget.winfo_height() + 4
        self._win = tk.Toplevel(self._widget)
        self._win.wm_overrideredirect(True)
        self._win.wm_geometry(f'+{x}+{y}')
        lbl = tk.Label(self._win, text=self._text, bg='#ffffe0', relief='solid',
                       borderwidth=1, font=('Segoe UI', 8), padx=6, pady=3,
                       wraplength=500, justify='left')
        lbl.pack()

    def _hide(self, _event=None):
        if self._win:
            self._win.destroy()
            self._win = None


# ---------------------------------------------------------------------------
# Scrollable frame helper (Canvas-based, allows embedded Entry widgets)
# ---------------------------------------------------------------------------

class _ScrollableFrame(tk.Frame):
    def __init__(self, parent, **kwargs):
        super().__init__(parent, **kwargs)
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)

        self._canvas = tk.Canvas(self, bg='white', highlightthickness=0)
        self._vsb    = ttk.Scrollbar(self, orient='vertical', command=self._canvas.yview)
        self._canvas.configure(yscrollcommand=self._vsb.set)

        self._canvas.grid(row=0, column=0, sticky='nsew')
        self._vsb.grid(row=0, column=1, sticky='ns')

        self.inner = tk.Frame(self._canvas, bg='white')
        self._window_id = self._canvas.create_window((0, 0), window=self.inner, anchor='nw')

        self.inner.bind('<Configure>', self._on_inner_configure)
        self._canvas.bind('<Configure>', self._on_canvas_configure)
        self._canvas.bind_all('<MouseWheel>', self._on_mousewheel)

    def _on_inner_configure(self, _event=None):
        self._canvas.configure(scrollregion=self._canvas.bbox('all'))

    def _on_canvas_configure(self, event):
        self._canvas.itemconfig(self._window_id, width=event.width)

    def _on_mousewheel(self, event):
        # Only scroll when pointer is over this canvas
        widget = event.widget
        while widget is not None:
            if widget is self._canvas:
                self._canvas.yview_scroll(int(-1 * (event.delta / 120)), 'units')
                break
            widget = getattr(widget, 'master', None)


# ---------------------------------------------------------------------------
# Status bar (inline – avoids relative import issues in standalone mode)
# ---------------------------------------------------------------------------

class _StatusBar:
    _COLORS = {
        'info'   : ('#3498db', '#ecf0f1', '#2c3e50'),
        'success': ('#27ae60', '#d5f4e6', '#145a32'),
        'warning': ('#f39c12', '#fef5e7', '#7d6608'),
        'error'  : ('#e74c3c', '#fadbd8', '#78281f'),
    }

    def __init__(self, parent, height=40):
        self._parent = parent
        self._timer  = None
        self.frame   = tk.Frame(parent, bg='#ecf0f1', relief='ridge', bd=2, height=height)
        self.frame.grid_propagate(False)
        self.frame.columnconfigure(0, weight=1)
        self._lbl = tk.Label(self.frame, text='Ready', font=('Segoe UI', 9),
                             bg='#ecf0f1', fg='#2c3e50', anchor='w',
                             padx=12, pady=6, justify='left')
        self._lbl.pack(fill='both', expand=True)

    def show(self, message, level='info', auto_clear=True):
        border, bg, fg = self._COLORS.get(level, self._COLORS['info'])
        bold = 'bold' if level in ('error', 'warning') else 'normal'
        self.frame.config(bg=border, bd=2)
        self._lbl.config(text=message, bg=bg, fg=fg,
                         font=('Segoe UI', 9, bold))
        if self._timer:
            try:
                self._parent.after_cancel(self._timer)
            except Exception:
                pass
            self._timer = None
        if auto_clear and level in ('info', 'success'):
            self._timer = self._parent.after(
                6000, lambda: self.show('Ready', 'info', False))

    def grid(self, **kw):
        self.frame.grid(**kw)


# ---------------------------------------------------------------------------
# Main application
# ---------------------------------------------------------------------------

class WordReplacementApp:
    """
    Framework File Visual ID Fix – enhanced UI with single and batch modes.
    """

    # Log text tag colours (dark-theme log panel)
    _LOG_TAGS = {
        'info'   : '#9cdcfe',
        'success': '#4ec9b0',
        'warning': '#dcdcaa',
        'error'  : '#f48771',
        'header' : '#c586c0',
    }

    def __init__(self, root):
        self.root = root
        self.root.title('Framework File Visual ID Fix')
        self.root.geometry('1000x780')
        self.root.minsize(860, 600)

        self._batch_mode  = tk.BooleanVar(value=False)
        self._batch_rows  = []   # list of dicts: {path, old_word, enabled_var, newword_var, row_frame}
        self._is_running  = False

        self._build_ui()
        self._center_window()

    # ------------------------------------------------------------------
    # Layout
    # ------------------------------------------------------------------

    def _build_ui(self):
        self.root.rowconfigure(0, weight=0)  # accent bar
        self.root.rowconfigure(1, weight=0)  # title bar
        self.root.rowconfigure(2, weight=0)  # mode toggle
        self.root.rowconfigure(3, weight=0)  # mode content
        self.root.rowconfigure(4, weight=1)  # log panel  ← expands
        self.root.rowconfigure(5, weight=0)  # status bar
        self.root.columnconfigure(0, weight=1)

        # -- Accent bar
        accent = tk.Frame(self.root, bg=ACCENT_COLOR, height=10)
        accent.grid(row=0, column=0, sticky='ew')
        accent.grid_propagate(False)

        # -- Title bar
        title_bar = tk.Frame(self.root, bg=HEADER_BG, pady=8)
        title_bar.grid(row=1, column=0, sticky='ew')
        title_bar.columnconfigure(0, weight=1)
        tk.Label(title_bar, text='Framework File Visual ID Fix',
                 font=FONT_HEADER, bg=HEADER_BG, fg='white').grid(
            row=0, column=0, padx=18, sticky='w')
        tk.Label(title_bar,
                 text='Replace Visual IDs in .log / .xlsx / .zip framework output files',
                 font=('Segoe UI', 8), bg=HEADER_BG, fg='#95a5a6').grid(
            row=1, column=0, padx=18, sticky='w')

        # -- Mode toggle
        toggle_bar = tk.Frame(self.root, bg='#f0f0f0', pady=6, padx=14)
        toggle_bar.grid(row=2, column=0, sticky='ew')
        ttk.Checkbutton(toggle_bar, text='  Batch Mode  —  scan folder tree for Summary files',
                        variable=self._batch_mode,
                        command=self._toggle_mode).pack(side='left')

        # -- Mode content container
        self._content_frame = tk.Frame(self.root, padx=12, pady=8)
        self._content_frame.grid(row=3, column=0, sticky='ew')
        self._content_frame.columnconfigure(0, weight=1)

        self._build_normal_frame()
        self._build_batch_frame()
        self._normal_frame.grid(row=0, column=0, sticky='ew')  # show normal by default

        # -- Log panel
        log_outer = tk.LabelFrame(self.root, text='  Activity Log',
                                  font=FONT_UI_BOLD, padx=8, pady=6)
        log_outer.grid(row=4, column=0, sticky='nsew', padx=12, pady=(0, 4))
        log_outer.rowconfigure(0, weight=1)
        log_outer.columnconfigure(0, weight=1)

        self._log_text = scrolledtext.ScrolledText(
            log_outer, height=10, font=FONT_LOG, wrap='word',
            bg='#1e1e1e', fg='#d4d4d4', relief='flat',
            borderwidth=0, highlightthickness=0,
            insertbackground='white', state='disabled')
        self._log_text.grid(row=0, column=0, sticky='nsew')

        for tag, color in self._LOG_TAGS.items():
            self._log_text.tag_configure(tag, foreground=color)

        # log toolbar
        log_toolbar = tk.Frame(log_outer, bg='#1e1e1e')
        log_toolbar.grid(row=1, column=0, sticky='ew', pady=(2, 0))
        tk.Button(log_toolbar, text='Clear Log', command=self._clear_log,
                  font=('Segoe UI', 8), bg='#3c3c3c', fg='#cccccc',
                  relief='flat', bd=0, padx=8, pady=2,
                  activebackground='#505050', cursor='hand2').pack(side='right', padx=4)

        # -- Status bar
        self._status = _StatusBar(self.root, height=36)
        self._status.grid(row=5, column=0, sticky='ew', padx=0, pady=0)

    # ------------------------------------------------------------------
    # Normal (single-folder) frame
    # ------------------------------------------------------------------

    def _build_normal_frame(self):
        self._normal_frame = tk.LabelFrame(
            self._content_frame, text='  Single Folder',
            font=FONT_UI_BOLD, padx=12, pady=10)
        self._normal_frame.columnconfigure(1, weight=1)

        # Folder row
        tk.Label(self._normal_frame, text='Folder:', font=FONT_UI).grid(
            row=0, column=0, sticky='w', pady=3)
        self._single_folder_var = tk.StringVar()
        folder_entry = tk.Entry(self._normal_frame, textvariable=self._single_folder_var,
                                font=FONT_UI)
        folder_entry.grid(row=0, column=1, sticky='ew', padx=(6, 4))
        tk.Button(self._normal_frame, text='Browse…', font=FONT_UI,
                  command=self._browse_single,
                  bg='#ecf0f1', relief='flat', padx=8, cursor='hand2').grid(
            row=0, column=2, padx=(0, 4))

        # Old / New word row
        words_frame = tk.Frame(self._normal_frame)
        words_frame.grid(row=1, column=0, columnspan=3, sticky='ew', pady=4)
        tk.Label(words_frame, text='Old Word:', font=FONT_UI).pack(side='left')
        self._single_old_var = tk.StringVar()
        tk.Entry(words_frame, textvariable=self._single_old_var,
                 font=FONT_UI, width=28).pack(side='left', padx=(4, 20))
        tk.Label(words_frame, text='New Word:', font=FONT_UI).pack(side='left')
        self._single_new_var = tk.StringVar()
        tk.Entry(words_frame, textvariable=self._single_new_var,
                 font=FONT_UI, width=28).pack(side='left', padx=4)

        # Buttons
        btn_frame = tk.Frame(self._normal_frame)
        btn_frame.grid(row=2, column=0, columnspan=3, pady=(8, 2), sticky='w')
        self._single_submit_btn = tk.Button(
            btn_frame, text='Execute', font=FONT_UI_BOLD,
            bg=BTN_PRIMARY, fg='white', relief='flat',
            padx=18, pady=5, cursor='hand2',
            command=self._run_single)
        self._single_submit_btn.pack(side='left', padx=(0, 8))
        tk.Button(btn_frame, text='Exit', font=FONT_UI,
                  bg=BTN_NEUTRAL, fg='white', relief='flat',
                  padx=14, pady=5, cursor='hand2',
                  command=self.root.quit).pack(side='left')

    # ------------------------------------------------------------------
    # Batch frame
    # ------------------------------------------------------------------

    def _build_batch_frame(self):
        self._batch_frame = tk.LabelFrame(
            self._content_frame, text='  Batch Processing',
            font=FONT_UI_BOLD, padx=12, pady=10)
        self._batch_frame.columnconfigure(1, weight=1)

        # Scan root row
        tk.Label(self._batch_frame, text='Root Folder:', font=FONT_UI).grid(
            row=0, column=0, sticky='w', pady=3)
        self._batch_root_var = tk.StringVar()
        tk.Entry(self._batch_frame, textvariable=self._batch_root_var,
                 font=FONT_UI).grid(row=0, column=1, sticky='ew', padx=(6, 4))
        tk.Button(self._batch_frame, text='Browse…', font=FONT_UI,
                  command=self._browse_batch_root,
                  bg='#ecf0f1', relief='flat', padx=8, cursor='hand2').grid(
            row=0, column=2, padx=(0, 4))
        self._scan_btn = tk.Button(
            self._batch_frame, text='Scan', font=FONT_UI_BOLD,
            bg=BTN_SCAN, fg='white', relief='flat',
            padx=14, pady=3, cursor='hand2',
            command=self._run_scan)
        self._scan_btn.grid(row=0, column=3, padx=(0, 4))

        # Global new-word row
        gw_frame = tk.Frame(self._batch_frame)
        gw_frame.grid(row=1, column=0, columnspan=4, sticky='w', pady=(4, 2))
        tk.Label(gw_frame, text='Global New Word:', font=FONT_UI).pack(side='left')
        self._global_new_var = tk.StringVar()
        tk.Entry(gw_frame, textvariable=self._global_new_var,
                 font=FONT_UI, width=26).pack(side='left', padx=(6, 10))
        tk.Button(gw_frame, text='Apply to All Rows', font=FONT_UI,
                  bg='#ecf0f1', relief='flat', padx=8, cursor='hand2',
                  command=self._apply_global_new_word).pack(side='left')

        # Table container (scrollable)
        table_container = tk.Frame(self._batch_frame, bg='#dee2e6',
                                   relief='solid', bd=1)
        table_container.grid(row=2, column=0, columnspan=4,
                              sticky='ew', pady=(4, 0))
        table_container.columnconfigure(0, weight=1)

        # Column headers
        hdr = tk.Frame(table_container, bg='#dee2e6')
        hdr.grid(row=0, column=0, sticky='ew')
        for col_idx, (text, w) in enumerate([
                ('', 3), ('Path', 40), ('Old Word', 18), ('New Word', 24)]):
            tk.Label(hdr, text=text, font=FONT_UI_BOLD,
                     bg='#dee2e6', fg='#2c3e50',
                     width=w, anchor='w', padx=4).grid(
                row=0, column=col_idx, sticky='w', padx=2, pady=3)

        # Scrollable inner region
        self._batch_scroll = _ScrollableFrame(table_container, bg='white')
        self._batch_scroll.grid(row=1, column=0, sticky='ew')
        self._batch_scroll.configure(height=180)
        self._batch_scroll.grid_propagate(False)
        self._batch_scroll.columnconfigure(0, weight=1)

        self._no_rows_label = tk.Label(
            self._batch_scroll.inner,
            text='No Summary files found.  Enter a root folder and click Scan.',
            font=('Segoe UI', 9, 'italic'), fg='#888888', bg='white', pady=10)
        self._no_rows_label.grid(row=0, column=0, columnspan=4, sticky='ew')

        # Footer action buttons
        footer = tk.Frame(self._batch_frame)
        footer.grid(row=3, column=0, columnspan=4, sticky='w', pady=(6, 0))
        self._exec_btn = tk.Button(
            footer, text='Execute Batch', font=FONT_UI_BOLD,
            bg=BTN_EXEC, fg='white', relief='flat',
            padx=18, pady=5, cursor='hand2',
            command=self._run_batch)
        self._exec_btn.pack(side='left', padx=(0, 8))
        tk.Button(footer, text='Select All', font=FONT_UI,
                  bg='#ecf0f1', relief='flat', padx=10, pady=5, cursor='hand2',
                  command=lambda: self._toggle_all(True)).pack(side='left', padx=(0, 4))
        tk.Button(footer, text='Deselect All', font=FONT_UI,
                  bg='#ecf0f1', relief='flat', padx=10, pady=5, cursor='hand2',
                  command=lambda: self._toggle_all(False)).pack(side='left', padx=(0, 4))
        tk.Button(footer, text='Clear Table', font=FONT_UI,
                  bg='#ecf0f1', relief='flat', padx=10, pady=5, cursor='hand2',
                  command=self._clear_table).pack(side='left', padx=(0, 8))
        tk.Button(footer, text='Exit', font=FONT_UI,
                  bg=BTN_NEUTRAL, fg='white', relief='flat',
                  padx=14, pady=5, cursor='hand2',
                  command=self.root.quit).pack(side='left')

    # ------------------------------------------------------------------
    # Mode toggle
    # ------------------------------------------------------------------

    def _toggle_mode(self):
        if self._batch_mode.get():
            self._normal_frame.grid_remove()
            self._batch_frame.grid(row=0, column=0, sticky='ew')
        else:
            self._batch_frame.grid_remove()
            self._normal_frame.grid(row=0, column=0, sticky='ew')
        self._clear_log()
        self._status.show('Ready', 'info', False)

    # ------------------------------------------------------------------
    # Scan
    # ------------------------------------------------------------------

    def _browse_single(self):
        path = filedialog.askdirectory()
        if path:
            self._single_folder_var.set(path)

    def _browse_batch_root(self):
        path = filedialog.askdirectory()
        if path:
            self._batch_root_var.set(path)

    def _run_scan(self):
        root_path = self._batch_root_var.get().strip()
        if not root_path or not os.path.isdir(root_path):
            self._status.show('Please enter a valid root folder before scanning.', 'warning')
            return
        self._clear_table()
        self._log(f'Scanning: {root_path}', 'header')
        results = scan_for_summary_files(root_path)
        if not results:
            self._log('No Summary_<TOKEN>_*.xlsx files found.', 'warning')
            self._status.show('Scan complete — no matching files found.', 'warning')
            return
        self._log(f'Found {len(results)} item(s).', 'success')
        global_new = self._global_new_var.get().strip()
        for item in results:
            self._add_batch_row(item['path'], item['old_word'],
                                item['filename'], global_new)
        self._status.show(
            f'Scan complete — {len(results)} row(s) added.  '
            'Set New Word and click Execute Batch.', 'success')

    def _add_batch_row(self, path, old_word, filename, preset_new=''):
        """Add one row to the batch table."""
        if self._no_rows_label.winfo_manager():
            self._no_rows_label.grid_remove()

        row_idx       = len(self._batch_rows)
        enabled_var   = tk.BooleanVar(value=True)
        newword_var   = tk.StringVar(value=preset_new)
        bg            = '#ffffff' if row_idx % 2 == 0 else '#f7f7f7'

        frame = tk.Frame(self._batch_scroll.inner, bg=bg)
        frame.grid(row=row_idx, column=0, sticky='ew', padx=2, pady=1)
        frame.columnconfigure(1, weight=1)

        ttk.Checkbutton(frame, variable=enabled_var).grid(
            row=0, column=0, padx=(4, 2))

        # Truncated path label with tooltip
        display = path if len(path) <= 55 else '…' + path[-52:]
        path_lbl = tk.Label(frame, text=display, font=FONT_UI, bg=bg,
                            width=42, anchor='w', padx=2)
        path_lbl.grid(row=0, column=1, sticky='w')
        _ToolTip(path_lbl, f'Full path: {path}\nTrigger file: {filename}')

        tk.Label(frame, text=old_word, font=FONT_UI, bg=bg,
                 width=18, anchor='w', padx=2).grid(
            row=0, column=2, sticky='w')

        tk.Entry(frame, textvariable=newword_var, font=FONT_UI,
                 width=24).grid(row=0, column=3, padx=(0, 6), sticky='w')

        self._batch_rows.append({
            'path'       : path,
            'old_word'   : old_word,
            'enabled_var': enabled_var,
            'newword_var': newword_var,
            'frame'      : frame,
        })

    def _apply_global_new_word(self):
        value = self._global_new_var.get().strip()
        for row in self._batch_rows:
            row['newword_var'].set(value)

    def _toggle_all(self, state: bool):
        for row in self._batch_rows:
            row['enabled_var'].set(state)

    def _clear_table(self):
        for row in self._batch_rows:
            row['frame'].destroy()
        self._batch_rows.clear()
        if not self._no_rows_label.winfo_manager():
            self._no_rows_label.grid(row=0, column=0, columnspan=4, sticky='ew')

    # ------------------------------------------------------------------
    # Execute single
    # ------------------------------------------------------------------

    def _run_single(self):
        folder   = self._single_folder_var.get().strip()
        old_word = self._single_old_var.get().strip()
        new_word = self._single_new_var.get().strip()
        if not folder or not old_word or not new_word:
            self._status.show('Please fill in all fields before executing.', 'warning')
            return
        if not os.path.isdir(folder):
            self._status.show('Folder does not exist.', 'error')
            return
        self._single_submit_btn.config(state='disabled')
        self._log('=== Single Mode — Starting ===', 'header')
        threading.Thread(target=self._worker_single,
                         args=(folder, old_word, new_word),
                         daemon=True).start()

    def _worker_single(self, folder, old_word, new_word):
        try:
            self._status_threadsafe('Backing up original files…', 'info')
            backup_folder(folder, self._log)
            self._status_threadsafe(f'Processing {folder}…', 'info')
            replace_word_in_folder(folder, old_word, new_word, self._log)
            self._log('=== Single Mode — Completed ===', 'header')
            self._status_threadsafe('Completed successfully.', 'success')
        except Exception as exc:
            self._log(f'ERROR: {exc}', 'error')
            self._status_threadsafe(f'Error: {exc}', 'error')
        finally:
            self.root.after(0, lambda: self._single_submit_btn.config(state='normal'))

    # ------------------------------------------------------------------
    # Execute batch
    # ------------------------------------------------------------------

    def _run_batch(self):
        selected = [r for r in self._batch_rows if r['enabled_var'].get()]
        if not selected:
            self._status.show('No rows selected.  Check at least one row.', 'warning')
            return
        missing_new = [r for r in selected if not r['newword_var'].get().strip()]
        if missing_new:
            self._status.show(
                f'{len(missing_new)} selected row(s) have no New Word set.', 'warning')
            return
        self._exec_btn.config(state='disabled')
        self._scan_btn.config(state='disabled')
        self._log(f'=== Batch Mode — {len(selected)} row(s) selected ===', 'header')
        threading.Thread(target=self._worker_batch,
                         args=(selected,), daemon=True).start()

    def _worker_batch(self, selected):
        errors  = 0
        success = 0
        try:
            # Backup phase
            unique_paths = list({r['path'] for r in selected})
            self._status_threadsafe(f'Backing up {len(unique_paths)} folder(s)…', 'info')
            for path in unique_paths:
                backup_folder(path, self._log)

            # Replace phase
            total = len(selected)
            for idx, row in enumerate(selected, 1):
                path     = row['path']
                old_word = row['old_word']
                new_word = row['newword_var'].get().strip()
                self._status_threadsafe(
                    f'Processing {idx}/{total}: {old_word} → {new_word} in {os.path.basename(path)}',
                    'info')
                self._log(f'--- Row {idx}/{total}: {old_word} → {new_word}', 'header')
                try:
                    replace_word_in_folder(path, old_word, new_word, self._log)
                    success += 1
                except Exception as exc:
                    self._log(f'  FAILED: {exc}', 'error')
                    errors += 1
        except Exception as exc:
            self._log(f'Batch aborted: {exc}', 'error')
            errors += 1
        finally:
            summary = f'Batch complete — {success} succeeded, {errors} failed.'
            self._log(f'=== {summary} ===', 'header')
            level = 'success' if errors == 0 else ('warning' if success > 0 else 'error')
            self._status_threadsafe(summary, level)
            self.root.after(0, lambda: self._exec_btn.config(state='normal'))
            self.root.after(0, lambda: self._scan_btn.config(state='normal'))

    # ------------------------------------------------------------------
    # Log helpers
    # ------------------------------------------------------------------

    def _log(self, message, level='info'):
        """Thread-safe log append with timestamp and colour tagging."""
        ts      = datetime.datetime.now().strftime('%H:%M:%S')
        line    = f'[{ts}] {message}\n'
        tag     = level if level in self._LOG_TAGS else 'info'

        def _append():
            self._log_text.config(state='normal')
            self._log_text.insert('end', line, tag)
            self._log_text.see('end')
            self._log_text.config(state='disabled')

        self.root.after(0, _append)

    def _clear_log(self):
        self._log_text.config(state='normal')
        self._log_text.delete('1.0', 'end')
        self._log_text.config(state='disabled')

    def _status_threadsafe(self, message, level='info'):
        self.root.after(0, lambda m=message, lv=level: self._status.show(m, lv))

    # ------------------------------------------------------------------
    # Utility
    # ------------------------------------------------------------------

    def _center_window(self):
        self.root.update_idletasks()
        w = self.root.winfo_width()
        h = self.root.winfo_height()
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        x = (sw - w) // 2
        y = (sh - h) // 2
        self.root.geometry(f'{w}x{h}+{x}+{y}')


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == '__main__':
    root = tk.Tk()
    app  = WordReplacementApp(root)
    root.mainloop()
