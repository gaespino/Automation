"""
PPV Experiment Builder v2.0 - Excel-like Interface for Control Panel Config Generator

Features:
- Each tab represents one experiment (Excel-like workflow)
- Left panel with Templates and Unit Data (shared across experiments)
- Grouped sections with visual separation
- Conditional field enabling/disabling based on Test Type and Content
- Unit Data overrides experiment-specific values
"""

import sys
import os
import json
import subprocess
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk, scrolledtext
    _TKINTER_AVAILABLE = True
except ImportError:
    _TKINTER_AVAILABLE = False
from typing import Dict, Any, List, Optional
import openpyxl
from datetime import datetime

# Add parent directory for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


def show_error(title, message):
    """Show error dialog and log to console"""
    print(f"ERROR - {title}: {message}")
    messagebox.showerror(title, message)


def show_warning(title, message):
    """Show warning dialog and log to console"""
    print(f"WARNING - {title}: {message}")
    messagebox.showwarning(title, message)


class ExperimentBuilderGUI:
    """Main GUI for building and managing experiment configurations - Excel-like interface"""

    def __init__(self, parent=None, default_product="GNR"):
        """Initialize the Experiment Builder GUI"""
        if parent is None:
            self.root = tk.Tk()
            self.is_standalone = True
        else:
            self.root = tk.Toplevel(parent)
            self.is_standalone = False

        self.root.title("Framework Experiment Builder")
        self.root.geometry("1600x900")

        # Data structures
        self.experiments = {}
        self.templates = {}
        self.current_product = default_product  # Use provided default product
        self.current_experiment = None
        self.config_template = self.load_config_template(self.current_product)
        self.field_widgets = {}  # Current experiment's widgets
        self.view_mode = "single"  # "single" or "side-by-side"
        self.experiment_notebook_secondary = None
        self.secondary_field_checkboxes = {}  # Track checkboxes for field selection
        self.secondary_notebook_frames = {}  # Track secondary notebook frames
        self.secondary_tab_buttons = {}  # Track secondary tab buttons
        self._updating_sections = False  # Flag to prevent recursion
        self._track_changes = True  # Flag to enable/disable change tracking

        # Setup UI
        self.setup_styles()
        self.create_main_layout()
        self.populate_default_data()

        # Initialize view button state
        self.update_view_button()

        if self.is_standalone:
            self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    def setup_styles(self):
        """Configure ttk styles for the application"""
        style = ttk.Style()
        style.theme_use('clam')

        # Configure colors
        self.colors = {
            'primary': '#2c3e50',
            'secondary': '#34495e',
            'accent': '#3498db',
            'success': '#10b981',  # Modern green
            'success_hover': '#059669',  # Darker green for hover
            'warning': '#f39c12',
            'danger': '#ef4444',  # Modern red
            'danger_hover': '#dc2626',  # Darker red for hover
            'light': '#ecf0f1',
            'dark': '#2c3e50',
            'section': '#e8f4f8',
            'section_border': '#b0d4e3',
            'disabled_tab': '#94a3b8',  # Slate gray for disabled tabs
            'disabled_tab_text': '#ffffff'
        }

        # Section frame style
        style.configure('Section.TLabelframe',
                       background='white',
                       relief='solid',
                       borderwidth=2)
        style.configure('Section.TLabelframe.Label',
                       font=('Segoe UI', 10, 'bold'),
                       foreground=self.colors['primary'],
                       background=self.colors['section'])

        # Disabled section style
        style.configure('Disabled.TLabelframe',
                       background='#f5f5f5',
                       relief='solid',
                       borderwidth=2)
        style.configure('Disabled.TLabelframe.Label',
                       font=('Segoe UI', 10, 'bold'),
                       foreground='#888888',
                       background='#e8e8e8')

        # Notebook tab style - use a custom style instead of globally hiding tabs
        # This prevents affecting other notebooks in the application
        style.configure('CustomHidden.TNotebook.Tab',
                       padding=[0, 0])
        style.layout('CustomHidden.TNotebook.Tab', [])  # Hide only CustomHidden notebook tabs

    def load_config_template(self, product="GNR"):
        """Load the configuration template from PPV/configs folder"""
        config_dir = os.path.join(os.path.dirname(__file__), '..', 'configs')
        config_file = os.path.join(config_dir, f'{product}ControlPanelConfig.json')

        if os.path.exists(config_file):
            with open(config_file, 'r') as f:
                config = json.load(f)
                # Migrate old format to new format if needed
                return self.migrate_config_format(config)

        return self.get_default_template()

    def migrate_config_format(self, config):
        """Migrate old data_types format to new field_configs format"""
        if 'field_configs' in config:
            # Already in new format
            return config

        if 'data_types' not in config:
            # No data types, return as-is
            return config

        # Convert old format to new format
        new_config = config.copy()
        new_config['field_configs'] = {}

        for field_name, type_list in config['data_types'].items():
            field_type = type_list[0] if type_list else 'str'
            new_config['field_configs'][field_name] = {
                'section': 'Basic Information',  # Default section
                'type': field_type,
                'default': self._get_default_for_type(field_type),
                'description': f'{field_name} field',
                'required': False
            }

        # Remove old data_types
        if 'data_types' in new_config:
            del new_config['data_types']

        return new_config

    def _get_default_for_type(self, field_type):
        """Get default value for a field type"""
        defaults = {
            'str': '',
            'int': 0,
            'float': 0.0,
            'bool': False
        }
        return defaults.get(field_type, '')

    def get_default_template(self):
        """Return a default configuration template with enhanced field metadata"""
        return {
            "field_configs": {
                "Experiment": {"section": "Status", "type": "bool", "default": True, "description": "Experiment enabled/disabled status", "required": True},
                "Test Name": {"section": "Basic Information", "type": "str", "default": "", "description": "Unique test identifier", "required": True},
                "Test Mode": {"section": "Basic Information", "type": "str", "default": "Mesh", "description": "Test execution mode", "required": True, "options": ["Mesh", "Slice"]},
                "Test Type": {"section": "Basic Information", "type": "str", "default": "Loops", "description": "Type of test to run", "required": True, "options": ["Loops", "Sweep", "Shmoo"]},
                "Product": {"section": "Unit Data", "type": "str", "default": "GNR", "description": "Product type", "required": True, "options": ["GNR", "CWF", "DMR"]},
                "Visual ID": {"section": "Unit Data", "type": "str", "default": "", "description": "Unit Visual ID", "required": False},
                "Bucket": {"section": "Unit Data", "type": "str", "default": "", "description": "Unit Assigned Bucket", "required": False},
                "COM Port": {"section": "Unit Data", "type": "int", "default": 1, "description": "Serial Port to communicate to MB", "required": False},
                "IP Address": {"section": "Unit Data", "type": "str", "default": "", "description": "IP Address to communicate to MB", "required": False},
                "TTL Folder": {"section": "Advanced Configuration", "type": "str", "default": "", "description": "TTL scripts folder path", "required": False},
                "Scripts File": {"section": "Advanced Configuration", "type": "str", "default": "", "description": "Test scripts file", "required": False},
                "Pass String": {"section": "Advanced Configuration", "type": "str", "default": "", "description": "String indicating test pass", "required": False},
                "Fail String": {"section": "Advanced Configuration", "type": "str", "default": "", "description": "String indicating test failure", "required": False},
                "Content": {"section": "Test Configuration", "type": "str", "default": "Linux", "description": "Content type to execute", "required": False, "options": ["Linux", "Dragon", "PYSVConsole"]},
                "Test Number": {"section": "Test Configuration", "type": "int", "default": 0, "description": "Test sequence number", "required": False},
                "Test Time": {"section": "Test Configuration", "type": "int", "default": 0, "description": "Test timeout in seconds", "required": False},
                "Reset": {"section": "Test Configuration", "type": "bool", "default": False, "description": "Reset before test", "required": False},
                "Reset on PASS": {"section": "Test Configuration", "type": "bool", "default": False, "description": "Reset after successful test", "required": False},
                "FastBoot": {"section": "Test Configuration", "type": "bool", "default": False, "description": "Enable fast boot mode", "required": False},
                "Core License": {"section": "Test Configuration", "type": "str", "default": "", "description": "Core license key", "required": False, "options": ["1: SSE/128", "2: AVX2/256 Light", "3: AVX2/256 Heavy", "4: AVX3/512 Light", "5: AVX3/512 Heavy", "6: TMUL Light", "7: TMUL Heavy"]},
                "600W Unit": {"section": "Test Configuration", "type": "bool", "default": False, "description": "Use 600W power unit", "required": False},
                "Pseudo Config": {"section": "Test Configuration", "type": "bool", "default": False, "description": "Enable pseudo configuration", "required": False},
                "Post Process": {"section": "Test Configuration", "type": "str", "default": "", "description": "Post-processing script", "required": False},
                "Configuration (Mask)": {
                    "section": "Test Configuration",
                    "type": "str",
                    "default": "",
                    "description": "Mesh: Mask options | Slice: Core number",
                    "required": False,
                    "conditional_options": {
                        "field": "Test Mode",
                        "Mesh": {"type": "str", "options": ["RowPass1", "RowPass2", "RowPass3", "FirstPass", "SecondPass", "ThirdPass"]},
                        "Slice": {"type": "int", "range": {"GNR": [0, 179], "CWF": [0, 179], "DMR": [0, 128]}}
                    }
                },
                "Boot Breakpoint": {"section": "Test Configuration", "type": "str", "default": "", "description": "Boot breakpoint location", "required": False},
                "Disable 2 Cores": {"section": "Test Configuration", "type": "str", "default": "", "description": "Cores to disable (2)", "required": False, "options": ["0x3", "0xc", "0x9", "0xa", "0x5"]},
                "Disable 1 Core": {"section": "Test Configuration", "type": "str", "default": "", "description": "Core to disable (1)", "required": False, "options": ["0x1", "0x2"]},
                "Check Core": {"section": "Test Configuration", "type": "int", "default": 0, "description": "Core to check", "required": False},
                "Voltage Type": {"section": "Voltage & Frequency", "type": "str", "default": "vbump", "description": "Voltage control type", "required": False, "options": ["vbump", "fixed", "ppvc"]},
                "Voltage IA": {"section": "Voltage & Frequency", "type": "float", "default": 0.0, "description": "IA voltage level", "required": False},
                "Voltage CFC": {"section": "Voltage & Frequency", "type": "float", "default": 0.0, "description": "CFC voltage level", "required": False},
                "Frequency IA": {"section": "Voltage & Frequency", "type": "int", "default": 0, "description": "IA frequency (MHz)", "required": False},
                "Frequency CFC": {"section": "Voltage & Frequency", "type": "int", "default": 0, "description": "CFC frequency (MHz)", "required": False},
                "Loops": {"section": "Loops", "type": "int", "default": 1, "description": "Number of test iterations", "required": False, "condition": {"field": "Test Type", "value": "Loops"}},
                "Type": {"section": "Sweep", "type": "str", "default": "Voltage", "description": "Sweep parameter type", "required": False, "options": ["Voltage", "Frequency"], "condition": {"field": "Test Type", "value": "Sweep"}},
                "Domain": {"section": "Sweep", "type": "str", "default": "IA", "description": "Domain to sweep", "required": False, "options": ["CFC", "IA"], "condition": {"field": "Test Type", "value": "Sweep"}},
                "Start": {"section": "Sweep", "type": "float", "default": 0.0, "description": "Sweep start value", "required": False, "condition": {"field": "Test Type", "value": "Sweep"}},
                "End": {"section": "Sweep", "type": "float", "default": 0.0, "description": "Sweep end value", "required": False, "condition": {"field": "Test Type", "value": "Sweep"}},
                "Steps": {"section": "Sweep", "type": "float", "default": 1.0, "description": "Sweep step size", "required": False, "condition": {"field": "Test Type", "value": "Sweep"}},
                "ShmooFile": {"section": "Shmoo", "type": "str", "default": "", "description": "Shmoo configuration file", "required": False, "condition": {"field": "Test Type", "value": "Shmoo"}},
                "ShmooLabel": {"section": "Shmoo", "type": "str", "default": "", "description": "Shmoo label identifier", "required": False, "condition": {"field": "Test Type", "value": "Shmoo"}},
                "Linux Path": {"section": "Linux", "type": "str", "default": "", "description": "Linux binary path", "required": False, "condition": {"field": "Content", "value": "Linux"}},
                "Linux Pre Command": {"section": "Linux", "type": "str", "default": "", "description": "Command before test", "required": False, "condition": {"field": "Content", "value": "Linux"}},
                "Linux Post Command": {"section": "Linux", "type": "str", "default": "", "description": "Command after test", "required": False, "condition": {"field": "Content", "value": "Linux"}},
                "Linux Pass String": {"section": "Linux", "type": "str", "default": "", "description": "Linux pass indicator", "required": False, "condition": {"field": "Content", "value": "Linux"}},
                "Linux Fail String": {"section": "Linux", "type": "str", "default": "", "description": "Linux fail indicator", "required": False, "condition": {"field": "Content", "value": "Linux"}},
                "Linux Content Wait Time": {"section": "Linux", "type": "int", "default": 0, "description": "Wait time in seconds", "required": False, "condition": {"field": "Content", "value": "Linux"}},
                "Startup Linux": {"section": "Linux", "type": "str", "default": "", "description": "Linux startup script", "required": False, "condition": {"field": "Content", "value": "Linux"}},
                "Linux Content Line 0": {"section": "Linux", "type": "str", "default": "", "description": "Content line 0", "required": False, "condition": {"field": "Content", "value": "Linux"}},
                "Linux Content Line 1": {"section": "Linux", "type": "str", "default": "", "description": "Content line 1", "required": False, "condition": {"field": "Content", "value": "Linux"}},
                "Linux Content Line 2": {"section": "Linux", "type": "str", "default": "", "description": "Content line 2", "required": False, "condition": {"field": "Content", "value": "Linux"}},
                "Linux Content Line 3": {"section": "Linux", "type": "str", "default": "", "description": "Content line 3", "required": False, "condition": {"field": "Content", "value": "Linux"}},
                "Linux Content Line 4": {"section": "Linux", "type": "str", "default": "", "description": "Content line 4", "required": False, "condition": {"field": "Content", "value": "Linux"}},
                "Linux Content Line 5": {"section": "Linux", "type": "str", "default": "", "description": "Content line 5", "required": False, "condition": {"field": "Content", "value": "Linux"}},
                "Linux Content Line 6": {"section": "Linux", "type": "str", "default": "", "description": "Content line 6", "required": False, "condition": {"field": "Content", "value": "Linux"}},
                "Linux Content Line 7": {"section": "Linux", "type": "str", "default": "", "description": "Content line 7", "required": False, "condition": {"field": "Content", "value": "Linux"}},
                "Linux Content Line 8": {"section": "Linux", "type": "str", "default": "", "description": "Content line 8", "required": False, "condition": {"field": "Content", "value": "Linux"}},
                "Linux Content Line 9": {"section": "Linux", "type": "str", "default": "", "description": "Content line 9", "required": False, "condition": {"field": "Content", "value": "Linux"}},
                "Dragon Pre Command": {"section": "Dragon", "type": "str", "default": "", "description": "Dragon pre-command", "required": False, "condition": {"field": "Content", "value": "Dragon"}},
                "Dragon Post Command": {"section": "Dragon", "type": "str", "default": "", "description": "Dragon post-command", "required": False, "condition": {"field": "Content", "value": "Dragon"}},
                "Startup Dragon": {"section": "Dragon", "type": "str", "default": "", "description": "Dragon startup script", "required": False, "condition": {"field": "Content", "value": "Dragon"}},
                "ULX Path": {"section": "Dragon", "type": "str", "default": "", "description": "ULX file path", "required": False, "condition": {"field": "Content", "value": "Dragon"}},
                "ULX CPU": {"section": "Dragon", "type": "str", "default": "", "description": "ULX CPU identifier", "required": False, "condition": {"field": "Content", "value": "Dragon"}},
                "Product Chop": {"section": "Dragon", "type": "str", "default": "", "description": "Product chop identifier", "required": False, "condition": {"field": "Content", "value": "Dragon"}},
                "VVAR0": {"section": "Dragon", "type": "str", "default": "", "description": "Variable 0", "required": False, "condition": {"field": "Content", "value": "Dragon"}},
                "VVAR1": {"section": "Dragon", "type": "str", "default": "", "description": "Variable 1", "required": False, "condition": {"field": "Content", "value": "Dragon"}},
                "VVAR2": {"section": "Dragon", "type": "str", "default": "", "description": "Variable 2", "required": False, "condition": {"field": "Content", "value": "Dragon"}},
                "VVAR3": {"section": "Dragon", "type": "str", "default": "", "description": "Variable 3", "required": False, "condition": {"field": "Content", "value": "Dragon"}},
                "VVAR_EXTRA": {"section": "Dragon", "type": "str", "default": "", "description": "Extra variables", "required": False, "condition": {"field": "Content", "value": "Dragon"}},
                "Dragon Content Path": {"section": "Dragon", "type": "str", "default": "", "description": "Dragon content path", "required": False, "condition": {"field": "Content", "value": "Dragon"}},
                "Dragon Content Line": {"section": "Dragon", "type": "str", "default": "", "description": "Dragon content line", "required": False, "condition": {"field": "Content", "value": "Dragon"}},
                "Stop on Fail": {"section": "Advanced Configuration", "type": "bool", "default": False, "description": "Stop test on failure", "required": False},
                "Merlin Name": {"section": "Merlin", "type": "str", "default": "", "description": "Merlin configuration name", "required": False},
                "Merlin Drive": {"section": "Merlin", "type": "str", "default": "", "description": "Merlin drive letter", "required": False},
                "Merlin Path": {"section": "Merlin", "type": "str", "default": "", "description": "Merlin executable path", "required": False},
                "Fuse File": {"section": "Merlin", "type": "str", "default": "", "description": "Fuse configuration file", "required": False},
                "Bios File": {"section": "Merlin", "type": "str", "default": "", "description": "BIOS file path", "required": False}
            },
            "TEST_MODES": ["Mesh", "Slice"],
            "TEST_TYPES": ["Loops", "Sweep", "Shmoo"],
            "CONTENT_OPTIONS": ["Linux", "Dragon", "PYSVConsole"],
            "VOLTAGE_TYPES": ["vbump", "fixed", "ppvc"],
            "TYPES": ["Voltage", "Frequency"],
            "DOMAINS": ["CFC", "IA"],
            "CORE_DISABLE_OPTIONS": ["None", "Disable 1 Core", "Disable 2 Cores"]  # Store core disable options separately
        }

    def create_main_layout(self):
        """Create the main application layout"""
        # Colored accent line at top (matching AutomationDesigner and Framework Parser)
        accent_line = tk.Frame(self.root, bg='#1abc9c', height=12)
        accent_line.pack(fill='x', side='top')
        accent_line.pack_propagate(False)

        # Header
        header_frame = tk.Frame(self.root, bg=self.colors['primary'], height=60)
        header_frame.pack(fill='x', side='top')
        header_frame.pack_propagate(False)

        title_label = tk.Label(header_frame,
                              text="Framework Experiment Builder v2.0",
                              font=('Segoe UI', 16, 'bold'),
                              bg=self.colors['primary'],
                              fg='white')
        title_label.pack(side='left', padx=20, pady=15)

        # Main container with paned window
        main_paned = tk.PanedWindow(self.root, orient=tk.HORIZONTAL, sashwidth=5)
        main_paned.pack(fill='both', expand=True, padx=5, pady=5)

        # Left panel - Templates and Unit Data
        self.create_left_panel(main_paned)

        # Right panel - Experiment tabs (Excel-like)
        self.create_right_panel(main_paned)

        # Bottom panel - Action buttons
        self.create_bottom_panel()

    def create_left_panel(self, parent):
        """Create the left panel with Templates and Unit Data"""
        left_frame = tk.Frame(parent, bg='white', width=350)
        parent.add(left_frame)

        # === FILE OPERATIONS SECTION ===
        file_ops_frame = tk.Frame(left_frame, bg='white')
        file_ops_frame.pack(fill='x', padx=5, pady=5)

        tk.Label(file_ops_frame, text="📁 File Operations",
                font=('Segoe UI', 10, 'bold'), bg='white',
                fg=self.colors['primary']).pack(side='left', padx=5)

        # Create dropdown menu button
        self.file_ops_menu_button = tk.Menubutton(file_ops_frame, text="≡ Menu",
                                                  font=('Segoe UI', 9),
                                                  bg=self.colors['accent'], fg='white',
                                                  relief='flat', padx=15, pady=5,
                                                  activebackground=self.colors['primary'],
                                                  activeforeground='white')
        self.file_ops_menu_button.pack(side='right', padx=5)

        # Create dropdown menu
        self.file_ops_menu = tk.Menu(self.file_ops_menu_button, tearoff=0, font=('Segoe UI', 9))
        self.file_ops_menu.add_command(label="📄 New Template", command=self.create_new_template)
        self.file_ops_menu.add_separator()
        self.file_ops_menu.add_command(label="💾 Save Configuration", command=self.save_configuration)
        self.file_ops_menu.add_command(label="📂 Load Configuration", command=self.load_configuration)
        self.file_ops_menu.add_separator()
        self.file_ops_menu.add_command(label="📥 Import from JSON", command=self.import_from_json)
        self.file_ops_menu.add_command(label="📥 Import from Excel", command=self.import_from_excel)
        self.file_ops_menu.add_separator()
        self.file_ops_menu.add_command(label="📤 Export to JSON", command=self.export_to_json)

        self.file_ops_menu_button['menu'] = self.file_ops_menu

        # Utils menu button
        self.utils_menu_button = tk.Menubutton(file_ops_frame, text="⚙ Utils",
                                              font=('Segoe UI', 9),
                                              bg=self.colors['secondary'], fg='white',
                                              relief='flat', padx=15, pady=5,
                                              activebackground=self.colors['primary'],
                                              activeforeground='white')
        self.utils_menu_button.pack(side='right', padx=5)

        # Create utils dropdown menu
        self.utils_menu = tk.Menu(self.utils_menu_button, tearoff=0, font=('Segoe UI', 9))
        self.utils_menu.add_command(label="🛠 Developer Environment", command=self.launch_developer_environment)
        self.utils_menu.add_separator()
        self.utils_menu.add_command(label="ℹ️ About", command=self.show_about)

        self.utils_menu_button['menu'] = self.utils_menu

        # Separator line
        separator = tk.Frame(left_frame, bg='#cccccc', height=1)
        separator.pack(fill='x', padx=5, pady=5)

        # === UNIT DATA SECTION ===
        unit_data_frame = ttk.LabelFrame(left_frame, text="📋 Unit Data (Shared)",
                                         style='Section.TLabelframe', padding=10)
        unit_data_frame.pack(fill='both', padx=5, pady=5, expand=False)

        # Info label
        info_label = tk.Label(unit_data_frame,
                             text="Unit Data applies to all experiments automatically",
                             font=('Segoe UI', 8, 'italic'),
                             bg='white', fg='#666666')
        info_label.pack(fill='x', padx=5, pady=(0, 5))

        # Create scrollable frame for unit data
        unit_canvas = tk.Canvas(unit_data_frame, bg='white', height=250)
        unit_scrollbar = tk.Scrollbar(unit_data_frame, orient='vertical', command=unit_canvas.yview)
        self.unit_data_fields_frame = tk.Frame(unit_canvas, bg='white')

        self.unit_data_fields_frame.bind(
            "<Configure>",
            lambda e: unit_canvas.configure(scrollregion=unit_canvas.bbox("all"))
        )

        unit_canvas.create_window((0, 0), window=self.unit_data_fields_frame, anchor='nw')
        unit_canvas.configure(yscrollcommand=unit_scrollbar.set)

        unit_canvas.pack(side='left', fill='both', expand=True)
        unit_scrollbar.pack(side='right', fill='y')

        # Unit data fields
        self.unit_data_widgets = {}
        self.create_unit_data_fields()

        # === TEMPLATES SECTION ===
        templates_frame = ttk.LabelFrame(left_frame, text="📁 Templates",
                                        style='Section.TLabelframe', padding=10)
        templates_frame.pack(fill='both', padx=5, pady=5, expand=True)

        # Template toolbar
        template_toolbar = tk.Frame(templates_frame, bg='white')
        template_toolbar.pack(fill='x', pady=(0, 5))

        tk.Button(template_toolbar, text="+ New", font=('Segoe UI', 9),
                 bg=self.colors['success'], fg='white',
                 command=self.save_as_template,
                 width=8, relief='flat', pady=3).pack(side='left', padx=2)

        tk.Button(template_toolbar, text="Load", font=('Segoe UI', 9),
                 bg=self.colors['accent'], fg='white',
                 command=self.load_template,
                 width=8, relief='flat', pady=3).pack(side='left', padx=2)

        tk.Button(template_toolbar, text="Delete", font=('Segoe UI', 9),
                 bg=self.colors['danger'], fg='white',
                 command=self.delete_template,
                 width=8, relief='flat', pady=3).pack(side='left', padx=2)

        tk.Button(template_toolbar, text="Import", font=('Segoe UI', 9),
                 bg=self.colors['accent'], fg='white',
                 command=self.import_templates,
                 width=8, relief='flat', pady=3).pack(side='left', padx=2)

        # Templates listbox
        template_list_frame = tk.Frame(templates_frame, bg='white')
        template_list_frame.pack(fill='both', expand=True)

        template_scrollbar = tk.Scrollbar(template_list_frame)
        template_scrollbar.pack(side='right', fill='y')

        self.template_listbox = tk.Listbox(template_list_frame,
                                          yscrollcommand=template_scrollbar.set,
                                          font=('Segoe UI', 9),
                                          selectmode='single')
        self.template_listbox.pack(side='left', fill='both', expand=True)
        template_scrollbar.config(command=self.template_listbox.yview)

    def create_unit_data_fields(self):
        """Create unit data input fields from configuration"""
        # Get Unit Data fields from config
        field_configs = self.config_template.get('field_configs', {})
        unit_data_fields = [
            (field_name, field_config)
            for field_name, field_config in field_configs.items()
            if field_config.get('section') == 'Unit Data'
        ]

        # Sort fields to ensure Product appears first, then Visual ID
        def sort_key(item):
            field_name = item[0]
            if field_name == "Product":
                return (0, field_name)  # Product first
            elif field_name == "Visual ID":
                return (1, field_name)  # Visual ID second
            else:
                return (2, field_name)  # Others alphabetically after

        unit_data_fields = sorted(unit_data_fields, key=sort_key)

        # Create fields in order
        for idx, (field_name, field_config) in enumerate(unit_data_fields):
            field_type = field_config.get('type', 'str')
            description = field_config.get('description', '')
            default_value = field_config.get('default', '')
            options = field_config.get('options', [])

            # Special handling for Product field with product change callback
            if field_name == "Product":
                lbl = tk.Label(self.unit_data_fields_frame, text=f"{field_name}:",
                              font=('Segoe UI', 9, 'bold'), bg='white', anchor='w')
                lbl.grid(row=idx, column=0, sticky='w', padx=5, pady=5)

                # Use current_product as the default value
                self.product_var = tk.StringVar(value=self.current_product)
                product_combo = ttk.Combobox(self.unit_data_fields_frame, textvariable=self.product_var,
                                            values=options, width=18, state='readonly')
                product_combo.grid(row=idx, column=1, sticky='ew', padx=5, pady=5)
                product_combo.bind('<<ComboboxSelected>>', self.on_product_change)

                self.unit_data_widgets[field_name] = {
                    'var': self.product_var,
                    'widget': product_combo,
                    'type': field_type
                }
            else:
                self.create_simple_field(self.unit_data_fields_frame, field_name,
                                        field_type, idx, self.unit_data_widgets, description)

        # Bind unit data changes to auto-apply to all experiments
        for field_name, field_info in self.unit_data_widgets.items():
            if field_name == "Product":
                continue  # Product handled separately
            widget = field_info['widget']
            widget.bind('<FocusOut>', lambda e: self.auto_apply_unit_data())
            if isinstance(widget, ttk.Combobox):
                widget.bind('<<ComboboxSelected>>', lambda e: self.auto_apply_unit_data())

    def create_right_panel(self, parent):
        """Create the right panel with experiment tabs"""
        right_frame = tk.Frame(parent, bg='white')
        parent.add(right_frame)

        # Experiment tab toolbar
        toolbar = tk.Frame(right_frame, bg=self.colors['light'], height=40)
        toolbar.pack(fill='x', side='top')

        tk.Label(toolbar, text="Experiments:", font=('Segoe UI', 10, 'bold'),
                bg=self.colors['light']).pack(side='left', padx=10)

        tk.Button(toolbar, text="+ New Experiment", font=('Segoe UI', 9),
                 bg=self.colors['success'], fg='white',
                 command=self.add_new_experiment,
                 relief='flat', padx=15, pady=5).pack(side='left', padx=5)

        tk.Button(toolbar, text="💾 Save", font=('Segoe UI', 9),
                 bg=self.colors['success'], fg='white',
                 command=self.save_and_update_tab,
                 relief='flat', padx=15, pady=5).pack(side='left', padx=5)

        tk.Button(toolbar, text="↺ Restore", font=('Segoe UI', 9),
                 bg=self.colors['warning'], fg='white',
                 command=self.restore_current_experiment,
                 relief='flat', padx=15, pady=5).pack(side='left', padx=5)

        tk.Button(toolbar, text="✕ Delete", font=('Segoe UI', 9),
                 bg=self.colors['danger'], fg='white',
                 command=self.delete_current_experiment,
                 relief='flat', padx=15, pady=5).pack(side='left', padx=5)

        tk.Button(toolbar, text="📋 Duplicate", font=('Segoe UI', 9),
                 bg=self.colors['accent'], fg='white',
                 command=self.duplicate_current_experiment,
                 relief='flat', padx=15, pady=5).pack(side='left', padx=5)

        tk.Button(toolbar, text="🗑 Clear", font=('Segoe UI', 9),
                 bg=self.colors['warning'], fg='white',
                 command=self.clear_current_experiment,
                 relief='flat', padx=15, pady=5).pack(side='left', padx=5)

        tk.Button(toolbar, text="📄 Apply Template", font=('Segoe UI', 9),
                 bg=self.colors['success'], fg='white',
                 command=self.apply_template_to_current,
                 relief='flat', padx=15, pady=5).pack(side='left', padx=5)

        tk.Button(toolbar, text="🔄 Refresh Config", font=('Segoe UI', 9),
                 bg=self.colors['secondary'], fg='white',
                 command=self.refresh_configuration,
                 relief='flat', padx=15, pady=5).pack(side='left', padx=5)

        # View mode toggle button
        self.view_mode_button = tk.Button(toolbar, text="⊞ Side-by-Side View", font=('Segoe UI', 9),
                                          bg=self.colors['accent'], fg='white',
                                          command=self.toggle_view_mode,
                                          relief='flat', padx=15, pady=5)
        self.view_mode_button.pack(side='right', padx=5)

        # Container for notebooks (to support side-by-side view)
        self.notebooks_container = tk.Frame(right_frame, bg='white')
        self.notebooks_container.pack(fill='both', expand=True, padx=5, pady=5)

        # Use PanedWindow for proper space distribution in side-by-side view
        self.paned_window = tk.PanedWindow(self.notebooks_container, orient=tk.HORIZONTAL,
                                          sashwidth=8, sashrelief=tk.RAISED, bg='#cccccc')
        self.paned_window.pack(fill='both', expand=True)

        # Primary notebook container with navigation arrows
        self.primary_notebook_container = tk.Frame(self.paned_window, bg='white')
        self.paned_window.add(self.primary_notebook_container, stretch='always', minsize=300)

        # Navigation frame for primary notebook
        self.primary_nav_frame = tk.Frame(self.primary_notebook_container, bg=self.colors['light'], height=35)
        self.primary_nav_frame.pack(fill='x', side='top')
        self.primary_nav_frame.pack_propagate(False)

        tk.Button(self.primary_nav_frame, text="◄", font=('Segoe UI', 10, 'bold'),
                 bg=self.colors['primary'], fg='white',
                 command=lambda: self.scroll_tabs('primary', -1),
                 relief='flat', width=3).pack(side='left', padx=2, pady=2)

        tk.Label(self.primary_nav_frame, text="Experiments",
                font=('Segoe UI', 9, 'bold'), bg=self.colors['light'],
                fg=self.colors['primary']).pack(side='left', padx=10)

        tk.Button(self.primary_nav_frame, text="►", font=('Segoe UI', 10, 'bold'),
                 bg=self.colors['primary'], fg='white',
                 command=lambda: self.scroll_tabs('primary', 1),
                 relief='flat', width=3).pack(side='left', padx=2, pady=2)

        tk.Button(self.primary_nav_frame, text="≡", font=('Segoe UI', 12, 'bold'),
                 bg=self.colors['accent'], fg='white',
                 command=lambda: self.show_tab_menu('primary'),
                 relief='flat', width=3).pack(side='right', padx=2, pady=2)

        # Create custom scrollable tabs container
        self.primary_tabs_container = tk.Frame(self.primary_notebook_container, bg='white')
        self.primary_tabs_container.pack(fill='x', side='top')

        # Scrollable tabs canvas
        self.primary_tabs_canvas = tk.Canvas(self.primary_tabs_container, bg=self.colors['light'], height=35, highlightthickness=0)
        self.primary_tabs_canvas.pack(fill='x', expand=True)

        # Frame inside canvas to hold tab buttons
        self.primary_tabs_frame = tk.Frame(self.primary_tabs_canvas, bg=self.colors['light'])
        self.primary_tabs_window = self.primary_tabs_canvas.create_window((0, 0), window=self.primary_tabs_frame, anchor='nw')

        # Bind canvas configuration
        self.primary_tabs_frame.bind('<Configure>', lambda e: self.primary_tabs_canvas.configure(scrollregion=self.primary_tabs_canvas.bbox('all')))

        # Notebook for experiment tabs (each tab = one experiment)
        # Use custom style to hide tabs (we use custom buttons instead)
        self.experiment_notebook = ttk.Notebook(self.primary_notebook_container,
                                                style='CustomHidden.TNotebook')
        self.experiment_notebook.pack(fill='both', expand=True)
        self.experiment_notebook.bind('<<NotebookTabChanged>>', self.on_experiment_tab_change)

        # Track tab buttons for primary notebook
        self.primary_tab_buttons = {}

        # Enable mouse wheel scrolling on tabs
        self.experiment_notebook.bind('<Button-4>', lambda e: self._scroll_notebook(e, -1))
        self.experiment_notebook.bind('<Button-5>', lambda e: self._scroll_notebook(e, 1))
        self.experiment_notebook.bind('<MouseWheel>', self._on_notebook_mousewheel)

        # Copy controls frame (will be added to PanedWindow in side-by-side mode)
        self.copy_controls_frame = None

    def create_experiment_tab(self, exp_name):
        """Create a new experiment tab with all fields organized in sections"""
        # Create main frame for this experiment
        tab_frame = tk.Frame(self.experiment_notebook, bg='white')
        self.experiment_notebook.add(tab_frame, text=exp_name)

        # Create custom tab button in scrollable container
        self.create_custom_tab_button(exp_name, 'primary')

        # Store reference to widgets for this experiment (create early for status toggle)
        exp_widgets = {}
        self.experiments[exp_name]['widgets'] = exp_widgets
        self.experiments[exp_name]['tab_frame'] = tab_frame
        self.experiments[exp_name]['modified'] = False  # Track if experiment has unsaved changes

        # Create Experiment Status Toggle at the top
        status_frame = tk.Frame(tab_frame, bg=self.colors['light'], height=50)
        status_frame.pack(fill='x', side='top', padx=5, pady=5)
        status_frame.pack_propagate(False)

        # Status indicator and toggle
        status_indicator_frame = tk.Frame(status_frame, bg=self.colors['light'])
        status_indicator_frame.pack(side='left', padx=20, pady=10)

        tk.Label(status_indicator_frame, text="Experiment Status:",
                font=('Segoe UI', 10, 'bold'), bg=self.colors['light'],
                fg=self.colors['primary']).pack(side='left', padx=5)

        # Create status variable and checkbox
        experiment_var = tk.BooleanVar(value=True)  # Default to Enabled (True)
        exp_widgets['Experiment'] = {
            'var': experiment_var,
            'type': 'bool'
        }

        # Status label that updates based on checkbox
        status_label = tk.Label(status_indicator_frame, text="✓ ENABLED",
                               font=('Segoe UI', 10, 'bold'),
                               bg=self.colors['success'], fg='white',
                               padx=20, pady=8, relief='flat', borderwidth=0)
        status_label.pack(side='left', padx=5)

        # Toggle checkbox
        def on_status_change():
            enabled = experiment_var.get()
            if enabled:
                status_label.configure(text="✓ ENABLED", bg=self.colors['success'])
                self.enable_experiment(exp_widgets)
                # Update tab color to normal
                self.update_tab_color(exp_name, enabled=True)
            else:
                status_label.configure(text="✕ DISABLED", bg=self.colors['danger'])
                self.gray_out_experiment(exp_widgets)
                # Update tab color to disabled
                self.update_tab_color(exp_name, enabled=False)
            # Update conditional sections after enabling/disabling
            self.update_conditional_sections()

        toggle_check = tk.Checkbutton(status_indicator_frame,
                                     text="Enable this experiment",
                                     variable=experiment_var,
                                     command=on_status_change,
                                     font=('Segoe UI', 9),
                                     bg=self.colors['light'],
                                     activebackground=self.colors['light'])
        toggle_check.pack(side='left', padx=10)

        exp_widgets['Experiment']['widget'] = toggle_check
        exp_widgets['Experiment']['label'] = status_label

        # Info text
        tk.Label(status_indicator_frame,
                text="(Disabled experiments will be skipped during test execution)",
                font=('Segoe UI', 8, 'italic'), bg=self.colors['light'],
                fg='#666666').pack(side='left', padx=10)

        # Create scrollable canvas
        canvas = tk.Canvas(tab_frame, bg='white')
        scrollbar = tk.Scrollbar(tab_frame, orient='vertical', command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='white')

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        # Create sections - use dynamic creation if field_configs exists
        if 'field_configs' in self.config_template:
            # New dynamic section creation based on field_configs
            self.create_all_sections_dynamically(scrollable_frame, exp_widgets)
        else:
            # Legacy section creation for backward compatibility
            row = 0
            row = self.create_basic_info_section(scrollable_frame, exp_widgets, row)
            row = self.create_advanced_section(scrollable_frame, exp_widgets, row)
            row = self.create_test_config_section(scrollable_frame, exp_widgets, row)
            row = self.create_voltage_freq_section(scrollable_frame, exp_widgets, row)
            row = self.create_loops_section(scrollable_frame, exp_widgets, row)
            row = self.create_sweep_section(scrollable_frame, exp_widgets, row)
            row = self.create_shmoo_section(scrollable_frame, exp_widgets, row)
            row = self.create_content_section(scrollable_frame, exp_widgets, row)
            row = self.create_linux_section(scrollable_frame, exp_widgets, row)
            row = self.create_dragon_section(scrollable_frame, exp_widgets, row)
            row = self.create_merlin_section(scrollable_frame, exp_widgets, row)

        # Initialize Configuration (Mask) options based on current Test Mode
        if 'Test Mode' in exp_widgets and 'Configuration (Mask)' in exp_widgets:
            self.update_configuration_mask_options(exp_widgets)

        # If side-by-side view is active, create tab in secondary notebook too
        if self.view_mode == "side-by-side" and self.experiment_notebook_secondary:
            self.create_experiment_tab_in_notebook(exp_name, self.experiment_notebook_secondary)

        return tab_frame

    def create_dynamic_section(self, parent, widgets, start_row, section_name, icon="🔹"):
        """Dynamically create a section based on field_configs"""
        if 'field_configs' not in self.config_template:
            return start_row

        # Get all fields for this section
        section_fields = []
        for field_name, field_config in self.config_template['field_configs'].items():
            if field_config.get('section') == section_name:
                # Skip Experiment field - it's now a toggle at the top
                if field_name == 'Experiment':
                    continue
                section_fields.append((field_name, field_config))

        if not section_fields:
            return start_row

        # Create section frame
        section = ttk.LabelFrame(parent, text=f"{icon} {section_name}",
                                style='Section.TLabelframe', padding=15)
        section.grid(row=start_row, column=0, sticky='ew', padx=10, pady=10)
        parent.grid_columnconfigure(0, weight=1)

        # Create fields in section
        for idx, (field_name, field_config) in enumerate(section_fields):
            field_type = field_config.get('type', 'str')
            options = field_config.get('options', None)
            description = field_config.get('description', '')

            # Check if field needs browse button
            browse = 'Path' in field_name or 'File' in field_name or 'Folder' in field_name

            self.create_field_in_section(section, field_name, field_type, idx, widgets,
                                        options, description, browse)

        # Special handling: If this is Basic Information section, bind Test Mode callback
        if section_name == "Basic Information" and 'Test Mode' in widgets:
            test_mode_widget = widgets['Test Mode'].get('widget')
            if test_mode_widget and isinstance(test_mode_widget, ttk.Combobox):
                # Bind to update both conditional sections and Configuration (Mask)
                def on_test_mode_change(e):
                    self.update_conditional_sections()
                    self.update_configuration_mask_options(widgets)
                test_mode_widget.bind('<<ComboboxSelected>>', on_test_mode_change)

        # Store section reference for conditional display (for grayout functionality)
        section_key = f"{section_name.lower().replace(' ', '_')}_section"
        widgets[section_key] = section

        # Store section metadata for conditional display
        section.field_name = section_name
        if section_name in ["Loops", "Sweep", "Shmoo"]:
            section.condition_field = "Test Type"
            section.condition_value = section_name
        elif section_name in ["Linux", "Dragon"]:
            section.condition_field = "Content"
            section.condition_value = section_name
        elif section_name == "Merlin":
            section.condition_field = "Content"
            section.condition_value = "Dragon"

        return start_row + 1

    def get_field_config(self, field_name):
        """Get field configuration by name"""
        if 'field_configs' in self.config_template:
            return self.config_template['field_configs'].get(field_name, {})
        return {}

    def get_conditional_options_for_field(self, field_name, default_mode="Mesh"):
        """Get options for a field with conditional_options based on current mode"""
        field_config = self.get_field_config(field_name)
        conditional_options = field_config.get('conditional_options', {})

        if not conditional_options:
            return field_config.get('options', [])

        # Return options by default for initial creation
        if default_mode in conditional_options:
            mode_config = conditional_options[default_mode]

            # Handle new format with type specifications
            if isinstance(mode_config, dict):
                if mode_config.get('type') == 'str':
                    return mode_config.get('options', [])
                elif mode_config.get('type') == 'int':
                    # Get range for current product
                    product = self.current_product
                    range_config = mode_config.get('range', {})
                    range_values = range_config.get(product, [0, 179])
                    if len(range_values) == 2:
                        return [str(i) for i in range(range_values[0], range_values[1] + 1)]
            # Fallback for old list format
            elif isinstance(mode_config, list):
                return mode_config

        return field_config.get('options', [])

    def get_all_sections(self):
        """Get list of all unique sections from field_configs"""
        if 'field_configs' not in self.config_template:
            return []

        sections = []
        for field_config in self.config_template['field_configs'].values():
            section = field_config.get('section', 'Basic Information')
            if section not in sections:
                sections.append(section)

        # Define section order
        section_order = [
            'Basic Information',
            'Unit Data',  # Rendered separately in left panel, not in experiment forms
            'Advanced Configuration',
            'Test Configuration',
            'Voltage & Frequency',
            'Loops',
            'Sweep',
            'Shmoo',
            'Content',
            'Linux',
            'Dragon',
            'Merlin'
        ]

        # Sort sections by defined order, putting undefined at end
        ordered_sections = []
        for sec in section_order:
            if sec in sections:
                ordered_sections.append(sec)

        # Add any remaining sections not in order
        for sec in sections:
            if sec not in ordered_sections:
                ordered_sections.append(sec)

        return ordered_sections

    def create_all_sections_dynamically(self, parent, widgets):
        """Create all sections dynamically from field_configs"""
        if 'field_configs' not in self.config_template:
            # Fall back to old method if no field_configs
            return self.create_sections_legacy(parent, widgets)

        row = 0
        sections = self.get_all_sections()

        # Exclude "Unit Data" section - it's rendered separately in left panel
        for section_name in sections:
            if section_name == "Unit Data":
                continue
            row = self.create_dynamic_section(parent, widgets, row, section_name)

        return row

    def create_basic_info_section(self, parent, widgets, start_row):
        """Create Basic Information section"""
        section = ttk.LabelFrame(parent, text="🔹 Basic Information",
                                style='Section.TLabelframe', padding=15)
        section.grid(row=start_row, column=0, sticky='ew', padx=10, pady=10)
        parent.grid_columnconfigure(0, weight=1)

        fields = [
            ("Test Name", "str", None, "Unique name for this test"),
            ("Test Mode", "str", self.config_template.get("TEST_MODES", ["Mesh", "Slice"]),
             "Test execution mode"),
            ("Test Type", "str", self.config_template.get("TEST_TYPES", ["Loops", "Sweep", "Shmoo"]),
             "Type of test to execute"),
        ]

        for idx, (field_name, field_type, options, desc) in enumerate(fields):
            self.create_field_in_section(section, field_name, field_type, idx, widgets, options, desc)

        # Add callback for Test Mode to update Configuration (Mask) options
        if 'Test Mode' in widgets and 'widget' in widgets['Test Mode']:
            test_mode_widget = widgets['Test Mode']['widget']
            if isinstance(test_mode_widget, ttk.Combobox):
                test_mode_widget.bind('<<ComboboxSelected>>',
                                     lambda e: self.update_configuration_mask_options(widgets))

        return start_row + 1

    def create_test_config_section(self, parent, widgets, start_row):
        """Create Test Configuration section"""
        section = ttk.LabelFrame(parent, text="🔹 Test Configuration",
                                style='Section.TLabelframe', padding=15)
        section.grid(row=start_row, column=0, sticky='ew', padx=10, pady=10)

        fields = [
            ("TTL Folder", "str", None, "Path to TTL files folder", True),
            ("Scripts File", "str", None, "Path to test scripts file", True),
            ("Post Process", "str", None, "Script file after test execution", True),
            ("Pass String", "str", None, "String indicating test pass", False),
            ("Fail String", "str", None, "String indicating test failure", False),
            ("Test Number", "int", None, "Number of test iterations", False),
            ("Test Time", "int", None, "Test duration in seconds", False),
            ("Reset", "bool", None, "Reset system between tests", False),
            ("Reset on PASS", "bool", None, "Reset only on pass", False),
            ("FastBoot", "bool", None, "Enable fast boot mode", False),
            ("Core License", "str", self.get_field_config("Core License").get('options'), "GNR/DMR: Core license setting (1-7)", False),
            ("600W Unit", "bool", None, "Set to True for 600W units", False),
        ]

        for idx, field_data in enumerate(fields):
            if len(field_data) == 5:
                field_name, field_type, options, desc, browse = field_data
                self.create_field_in_section(section, field_name, field_type, idx, widgets,
                                            options, desc, browse)
            else:
                field_name, field_type, options, desc, browse = field_data
                self.create_field_in_section(section, field_name, field_type, idx, widgets,
                                            options, desc, browse)

        return start_row + 1

    def create_voltage_freq_section(self, parent, widgets, start_row):
        """Create Voltage & Frequency section"""
        section = ttk.LabelFrame(parent, text="🔹 Voltage & Frequency",
                                style='Section.TLabelframe', padding=15)
        section.grid(row=start_row, column=0, sticky='ew', padx=10, pady=10)

        fields = [
            ("Voltage Type", "str", self.config_template.get("VOLTAGE_TYPES", ["vbump", "fixed", "ppvc"]),
             "Voltage control type"),
            ("Voltage IA", "float", None, "IA domain voltage"),
            ("Voltage CFC", "float", None, "CFC domain voltage"),
            ("Frequency IA", "int", None, "IA domain frequency (MHz)"),
            ("Frequency CFC", "int", None, "CFC domain frequency (MHz)"),
        ]

        for idx, (field_name, field_type, options, desc) in enumerate(fields):
            self.create_field_in_section(section, field_name, field_type, idx, widgets, options, desc)

        return start_row + 1

    def create_loops_section(self, parent, widgets, start_row):
        """Create Loops Configuration section (enabled when Test Type = Loops)"""
        section = ttk.LabelFrame(parent, text="🔹 Loops Configuration (Test Type = Loops)",
                                style='Section.TLabelframe', padding=15)
        section.grid(row=start_row, column=0, sticky='ew', padx=10, pady=10)

        self.create_field_in_section(section, "Loops", "int", 0, widgets,
                                    None, "Number of test loops")

        # Mark section for conditional enabling
        widgets['loops_section'] = section

        return start_row + 1

    def create_sweep_section(self, parent, widgets, start_row):
        """Create Sweep Configuration section (enabled when Test Type = Sweep)"""
        section = ttk.LabelFrame(parent, text="🔹 Sweep Configuration (Test Type = Sweep)",
                                style='Section.TLabelframe', padding=15)
        section.grid(row=start_row, column=0, sticky='ew', padx=10, pady=10)

        fields = [
            ("Type", "str", self.config_template.get("TYPES", ["Voltage", "Frequency"]),
             "Parameter to sweep"),
            ("Domain", "str", self.config_template.get("DOMAINS", ["CFC", "IA"]),
             "Domain to target"),
            ("Start", "float", None, "Starting value"),
            ("End", "float", None, "Ending value"),
            ("Steps", "float", None, "Step size or count"),
        ]

        for idx, (field_name, field_type, options, desc) in enumerate(fields):
            self.create_field_in_section(section, field_name, field_type, idx, widgets, options, desc)

        # Mark section for conditional enabling
        widgets['sweep_section'] = section

        return start_row + 1

    def create_shmoo_section(self, parent, widgets, start_row):
        """Create Shmoo Configuration section (enabled when Test Type = Shmoo)"""
        section = ttk.LabelFrame(parent, text="🔹 Shmoo Configuration (Test Type = Shmoo)",
                                style='Section.TLabelframe', padding=15)
        section.grid(row=start_row, column=0, sticky='ew', padx=10, pady=10)

        fields = [
            ("ShmooFile", "str", None, "Path to shmoo configuration file", True),
            ("ShmooLabel", "str", None, "Shmoo label identifier", False),
        ]

        for idx, field_data in enumerate(fields):
            if len(field_data) == 5:
                field_name, field_type, options, desc, browse = field_data
                self.create_field_in_section(section, field_name, field_type, idx, widgets,
                                            options, desc, browse)
            else:
                field_name, field_type, options, desc, browse = field_data
                self.create_field_in_section(section, field_name, field_type, idx, widgets,
                                            options, desc, browse)

        # Mark section for conditional enabling
        widgets['shmoo_section'] = section

        return start_row + 1

    def create_content_section(self, parent, widgets, start_row):
        """Create Content Configuration section"""
        section = ttk.LabelFrame(parent, text="🔹 Content Selection",
                                style='Section.TLabelframe', padding=15)
        section.grid(row=start_row, column=0, sticky='ew', padx=10, pady=10)

        self.create_field_in_section(section, "Content", "str", 0, widgets,
                                     self.config_template.get("CONTENT_OPTIONS", ["Linux", "Dragon", "PYSVConsole"]),
                                     "Content type to execute")

        return start_row + 1

    def create_linux_section(self, parent, widgets, start_row):
        """Create Linux Configuration section"""
        section = ttk.LabelFrame(parent, text="🔹 Linux Configuration (Conditional)",
                                style='Section.TLabelframe', padding=15)
        section.grid(row=start_row, column=0, sticky='ew', padx=10, pady=10)

        fields = [
            ("Linux Path", "str", "Path to Linux content", True),
            ("Startup Linux", "str", "Script to start Linux from EFI", True),
            ("Linux Pre Command", "str", "Command before test", False),
            ("Linux Post Command", "str", "Command after test", False),
            ("Linux Pass String", "str", "String indicating Linux pass", False),
            ("Linux Fail String", "str", "String indicating Linux failure", False),
            ("Linux Content Wait Time", "int", "Wait time for Linux content", False),
            ("Linux Content Line 0", "str", "Content command line 0", False),
            ("Linux Content Line 1", "str", "Content command line 1", False),
        ]

        for idx, field_data in enumerate(fields):
            if len(field_data) == 4:
                field_name, field_type, desc, browse = field_data
                self.create_field_in_section(section, field_name, field_type, idx, widgets,
                                            None, desc, browse)
            else:
                field_name, field_type, desc, browse = field_data
                self.create_field_in_section(section, field_name, field_type, idx, widgets,
                                            None, desc, browse)

        # Mark section for conditional enabling
        widgets['linux_section'] = section

        return start_row + 1

    def create_dragon_section(self, parent, widgets, start_row):
        """Create Dragon Configuration section"""
        section = ttk.LabelFrame(parent, text="🔹 Dragon Configuration (Conditional)",
                                style='Section.TLabelframe', padding=15)
        section.grid(row=start_row, column=0, sticky='ew', padx=10, pady=10)

        fields = [
            ("Dragon Content Path", "str", "Path to content", True),
            ("Startup Dragon", "str", "Preset conditions script", True),
            ("Dragon Pre Command", "str", "Command before Dragon", False),
            ("Dragon Post Command", "str", "Command after Dragon", False),
            ("ULX Path", "str", "Path to ULX script", True),
            ("ULX CPU", "str", "CPU to be used (GNR_B0, CWF)", False),
            ("Product Chop", "str", "Product (e.g., GNR)", False),
            ("VVAR0", "str", "Test runtime (hex)", False),
            ("VVAR1", "str", "32-bit fixed point (80064000 for System)", False),
            ("VVAR2", "str", "Thread count (0x1000000 for Mesh)", False),
            ("VVAR3", "str", "Debug flags (see wiki)", False),
            ("VVAR_EXTRA", "str", "Additional VVAR parameters", False),
            ("Dragon Content Line", "str", "Content filters (comma-separated)", False),
        ]

        for idx, field_data in enumerate(fields):
            if len(field_data) == 4:
                field_name, field_type, desc, browse = field_data
                self.create_field_in_section(section, field_name, field_type, idx, widgets,
                                            None, desc, browse)
            else:
                field_name, field_type, desc, browse = field_data
                self.create_field_in_section(section, field_name, field_type, idx, widgets,
                                            None, desc, browse)

        # Mark section for conditional enabling
        widgets['dragon_section'] = section

        return start_row + 1

    def create_advanced_section(self, parent, widgets, start_row):
        """Create Advanced Configuration section"""
        section = ttk.LabelFrame(parent, text="🔹 Advanced Configuration",
                                style='Section.TLabelframe', padding=15)
        section.grid(row=start_row, column=0, sticky='ew', padx=10, pady=10)

        fields = [
            ("Configuration (Mask)", "str", self.get_conditional_options_for_field("Configuration (Mask)", "Mesh"), "Mesh: Mask options | Slice: Core number (GNR:0-179, CWF:0-179, DMR:0-128)", False),
            ("Boot Breakpoint", "str", None, "Boot breakpoint address", False),
            ("Pseudo Config", "bool", None, "GNR only - Disable HT for Pseudo Mesh Content", False),
            ("Disable 2 Cores", "str", self.get_field_config("Disable 2 Cores").get('options'), "CWF only - Disable 2 cores (leave blank if not used)", False),
            ("Disable 1 Core", "str", self.get_field_config("Disable 1 Core").get('options'), "DMR only - Disable 1 core (leave blank if not used)", False),
            ("Check Core", "int", None, "Core to check (Mesh/Slice)", False),
            ("Stop on Fail", "bool", None, "Stop after first failure", False),
            ("Fuse File", "str", None, "External fuse file to be loaded into the system", True),
            ("Bios File", "str", None, "BIOS file to be loaded at experiment start", True),
        ]

        for idx, field_data in enumerate(fields):
            if len(field_data) == 5:
                field_name, field_type, options, desc, browse = field_data
                self.create_field_in_section(section, field_name, field_type, idx, widgets,
                                            options, desc, browse)
            else:
                field_name, field_type, options, desc = field_data
                self.create_field_in_section(section, field_name, field_type, idx, widgets, options, desc)

        return start_row + 1

    def create_merlin_section(self, parent, widgets, start_row):
        """Create Merlin Configuration section"""
        section = ttk.LabelFrame(parent, text="🔹 Merlin Configuration (Conditional)",
                                style='Section.TLabelframe', padding=15)
        section.grid(row=start_row, column=0, sticky='ew', padx=10, pady=10)

        fields = [
            ("Merlin Name", "str", None, "Merlin file name (MerlinX.efi)"),
            ("Merlin Drive", "str", None, "Drive location (FS1:)"),
            ("Merlin Path", "str", None, "Path to Merlin files", True),
        ]

        for idx, field_data in enumerate(fields):
            if len(field_data) == 5:
                field_name, field_type, options, desc, browse = field_data
                self.create_field_in_section(section, field_name, field_type, idx, widgets,
                                            options, desc, browse)
            else:
                field_name, field_type, options, desc = field_data
                self.create_field_in_section(section, field_name, field_type, idx, widgets, options, desc)

        return start_row + 1

    def create_field_in_section(self, section, field_name, field_type, row, widgets,
                                options=None, description="", browse=False):
        """Create a field widget within a section"""
        # Label
        lbl = tk.Label(section, text=f"{field_name}:",
                      font=('Segoe UI', 9), bg='white', anchor='w', width=25)
        lbl.grid(row=row, column=0, sticky='w', padx=5, pady=3)

        # Check if field has conditional options
        field_config = self.get_field_config(field_name)
        conditional_options = field_config.get('conditional_options', {})

        # If field has conditional options and no options provided, get default
        if conditional_options and options is None:
            options = self.get_conditional_options_for_field(field_name, "Mesh")

        # Widget based on type
        if field_type == "bool":
            var = tk.BooleanVar(value=False)
            widget = tk.Checkbutton(section, variable=var, bg='white')
            widget.grid(row=row, column=1, sticky='w', padx=5, pady=3)
        elif options:
            var = tk.StringVar()
            widget = ttk.Combobox(section, textvariable=var, values=options, width=30)
            if options:
                var.set(options[0])
            widget.grid(row=row, column=1, sticky='ew', padx=5, pady=3)
            # Bind change event for conditional fields
            widget.bind('<<ComboboxSelected>>', lambda e: self.update_conditional_sections())
        else:
            var = tk.StringVar()
            widget = tk.Entry(section, textvariable=var, width=60, font=('Segoe UI', 9))
            widget.grid(row=row, column=1, sticky='ew', padx=5, pady=3)

        # Browse button for file/folder paths
        if browse:
            browse_btn = tk.Button(section, text="📁", font=('Segoe UI', 8),
                                   command=lambda: self.browse_path(field_name, var),
                                   width=3, relief='flat', bg=self.colors['light'])
            browse_btn.grid(row=row, column=2, padx=5)

        # Store reference
        widgets[field_name] = {
            'var': var,
            'widget': widget,
            'type': field_type,
            'label': lbl
        }

        # Bind change tracking to the variable
        var.trace_add('write', lambda *args: self.mark_experiment_modified())

        # Tooltip
        if description:
            self.create_tooltip(widget, description)

        # Configure column weights
        section.grid_columnconfigure(1, weight=1)

    def create_simple_field(self, parent, field_name, field_type, row, widget_dict, description=""):
        """Create a simple field (for unit data panel)"""
        lbl = tk.Label(parent, text=f"{field_name}:",
                      font=('Segoe UI', 9, 'bold'), bg='white', anchor='w')
        lbl.grid(row=row, column=0, sticky='w', padx=5, pady=5)

        if field_type == "int":
            var = tk.StringVar()
            widget = tk.Entry(parent, textvariable=var, width=18, font=('Segoe UI', 9))
        else:
            var = tk.StringVar()
            widget = tk.Entry(parent, textvariable=var, width=18, font=('Segoe UI', 9))

        widget.grid(row=row, column=1, sticky='ew', padx=5, pady=5)

        widget_dict[field_name] = {
            'var': var,
            'widget': widget,
            'type': field_type,
            'label': lbl
        }

        if description:
            self.create_tooltip(widget, description)

        parent.grid_columnconfigure(1, weight=1)

    def get_config_mask_options(self):
        """DEPRECATED: Get Configuration Mask options from config instead"""
        # For backward compatibility, read from config
        return self.get_field_config("Configuration (Mask)").get('options', [])

    def get_core_license_options(self):
        """DEPRECATED: Get Core License options from config instead"""
        # For backward compatibility, read from config
        return self.get_field_config("Core License").get('options', [])

    def update_configuration_mask_options(self, exp_widgets):
        """Update Configuration (Mask) options based on Test Mode and Product"""
        if 'Test Mode' not in exp_widgets or 'Configuration (Mask)' not in exp_widgets:
            return

        test_mode = exp_widgets['Test Mode']['var'].get()
        product = self.product_var.get() if hasattr(self, 'product_var') else self.current_product

        field_config = self.get_field_config('Configuration (Mask)')
        conditional_options = field_config.get('conditional_options', {})

        if not conditional_options:
            return

        config_widget = exp_widgets['Configuration (Mask)']['widget']

        if test_mode == 'Mesh':
            # Use Mesh options (string type)
            mesh_config = conditional_options.get('Mesh', {})
            if isinstance(mesh_config, dict):
                mesh_options = mesh_config.get('options', [])
            else:
                mesh_options = mesh_config  # Fallback for old format

            if isinstance(config_widget, ttk.Combobox):
                config_widget['values'] = mesh_options
                config_widget.set('')  # Clear current value
            else:
                # Need to replace with combobox
                self.replace_widget_with_combobox(exp_widgets, 'Configuration (Mask)', mesh_options, 'str')

            # Update the widget type to string
            exp_widgets['Configuration (Mask)']['type'] = 'str'

        elif test_mode == 'Slice':
            # Use Slice range based on product (integer type)
            slice_config = conditional_options.get('Slice', {})
            if slice_config.get('type') == 'int':
                range_config = slice_config.get('range', {})
                range_values = range_config.get(product, [0, 179])
                if len(range_values) == 2:
                    start, end = range_values
                    options = [str(i) for i in range(start, end + 1)]
                    if isinstance(config_widget, ttk.Combobox):
                        config_widget['values'] = options
                        config_widget.set('')  # Clear current value
                    else:
                        # Need to replace with combobox
                        self.replace_widget_with_combobox(exp_widgets, 'Configuration (Mask)', options, 'int')

                    # Update the widget type to integer
                    exp_widgets['Configuration (Mask)']['type'] = 'int'

    def replace_widget_with_combobox(self, exp_widgets, field_name, options, field_type='str'):
        """Replace a widget with a combobox (for dynamic option switching)"""
        if field_name not in exp_widgets:
            return

        widget_info = exp_widgets[field_name]
        old_widget = widget_info['widget']
        var = widget_info['var']

        # Get grid info before destroying
        grid_info = old_widget.grid_info()
        parent = old_widget.master

        # Destroy old widget
        old_widget.destroy()

        # Create new combobox
        new_combo = ttk.Combobox(parent, textvariable=var, values=options, width=30)
        new_combo.grid(**grid_info)

        # Update reference
        widget_info['widget'] = new_combo
        widget_info['type'] = field_type  # Update the type

        # Restore tooltip if exists
        field_config = self.get_field_config(field_name)
        description = field_config.get('description', '')
        if description:
            self.create_tooltip(new_combo, description)

    def mark_experiment_modified(self):
        """Mark the current experiment as modified (has unsaved changes)"""
        if not self._track_changes or not self.current_experiment:
            return

        if self.current_experiment in self.experiments:
            # Set modified flag
            self.experiments[self.current_experiment]['modified'] = True
            # Update tab appearance
            self.update_tab_modified_indicator(self.current_experiment)

    def update_tab_modified_indicator(self, exp_name):
        """Update tab button to show modified state"""
        if exp_name not in self.primary_tab_buttons:
            return

        btn = self.primary_tab_buttons[exp_name]
        is_modified = self.experiments.get(exp_name, {}).get('modified', False)
        is_current = (exp_name == self.current_experiment)
        is_disabled = self.is_experiment_disabled(exp_name)

        # Update button text to include asterisk if modified
        base_text = exp_name
        display_text = f"* {base_text}" if is_modified else base_text
        btn.configure(text=display_text)

        # Update colors
        if is_current:
            if is_modified:
                # Current tab with unsaved changes - use warning color
                btn.configure(bg=self.colors['warning'], fg='white', relief='sunken')
            elif is_disabled:
                btn.configure(bg=self.colors['disabled_tab'], fg=self.colors['disabled_tab_text'], relief='sunken')
            else:
                btn.configure(bg=self.colors['primary'], fg='white', relief='sunken')
        else:
            if is_modified:
                # Non-current tab with unsaved changes - lighter warning color
                btn.configure(bg='#fbbf24', fg='black', relief='raised')
            elif is_disabled:
                btn.configure(bg='#cbd5e1', fg='#64748b', relief='raised')
            else:
                btn.configure(bg=self.colors['light'], fg='black', relief='raised')

    def refresh_configuration(self):
        """Refresh all configuration fields for all experiments based on current product and config"""
        # Reload config template to ensure we have the latest
        self.config_template = self.load_config_template(self.current_product)

        # Update all experiments
        for exp_name, exp_data in self.experiments.items():
            if 'widgets' not in exp_data:
                continue

            exp_widgets = exp_data['widgets']

            # Update Configuration (Mask) based on Test Mode
            if 'Test Mode' in exp_widgets and 'Configuration (Mask)' in exp_widgets:
                self.update_configuration_mask_options(exp_widgets)

            # Update all combobox fields that have options from config
            field_configs = self.config_template.get('field_configs', {})
            for field_name, widget_info in exp_widgets.items():
                if field_name in field_configs:
                    field_config = field_configs[field_name]
                    widget = widget_info.get('widget')

                    # Update combobox options
                    if isinstance(widget, ttk.Combobox):
                        # Check for regular options
                        options = field_config.get('options', [])
                        if options:
                            current_value = widget_info['var'].get()
                            widget['values'] = options
                            # Restore value if it's still valid
                            if current_value not in options and options:
                                widget_info['var'].set(options[0] if options else '')

                        # Check for conditional options (like Configuration Mask)
                        elif field_config.get('conditional_options'):
                            # Trigger update for conditional fields
                            if field_name == 'Configuration (Mask)':
                                self.update_configuration_mask_options(exp_widgets)

    def on_product_change(self, event=None):
        """Handle product selection change"""
        new_product = self.product_var.get()
        if new_product != self.current_product:
            self.current_product = new_product
            self.config_template = self.load_config_template(new_product)

            # Refresh all configuration fields
            self.refresh_configuration()

            messagebox.showinfo("Product Changed",
                              f"Configuration updated for {new_product}.\n"
                              "All fields have been refreshed for the new product.")

    def create_tooltip(self, widget, text):
        """Create a tooltip for a widget"""
        def on_enter(event):
            tooltip = tk.Toplevel()
            tooltip.wm_overrideredirect(True)
            tooltip.wm_geometry(f"+{event.x_root+10}+{event.y_root+10}")
            label = tk.Label(tooltip, text=text, background="#ffffe0",
                           relief='solid', borderwidth=1, font=('Segoe UI', 8))
            label.pack()
            widget.tooltip = tooltip

        def on_leave(event):
            if hasattr(widget, 'tooltip'):
                try:
                    widget.tooltip.destroy()
                    del widget.tooltip
                except:
                    pass

        widget.bind('<Enter>', on_enter)
        widget.bind('<Leave>', on_leave)

    def create_bottom_panel(self):
        """Create the bottom action button panel"""
        bottom_frame = tk.Frame(self.root, bg=self.colors['light'], height=60)
        bottom_frame.pack(fill='x', side='bottom')
        bottom_frame.pack_propagate(False)

        # Left buttons
        left_buttons = tk.Frame(bottom_frame, bg=self.colors['light'])
        left_buttons.pack(side='left', padx=10)

        tk.Button(left_buttons, text="📂 Import from Excel",
                 bg=self.colors['accent'], fg='white',
                 command=self.import_from_excel,
                 font=('Segoe UI', 10),
                 relief='flat', padx=20, pady=8).pack(side='left', padx=5)

        tk.Button(left_buttons, text="📂 Import from JSON",
                 bg=self.colors['accent'], fg='white',
                 command=self.import_from_json,
                 font=('Segoe UI', 10),
                 relief='flat', padx=20, pady=8).pack(side='left', padx=5)

        # Right buttons
        right_buttons = tk.Frame(bottom_frame, bg=self.colors['light'])
        right_buttons.pack(side='right', padx=10)

        tk.Button(right_buttons, text="💾 Export to JSON",
                 bg=self.colors['success'], fg='white',
                 command=self.export_to_json,
                 font=('Segoe UI', 10, 'bold'),
                 relief='flat', padx=30, pady=8).pack(side='left', padx=5)

        tk.Button(right_buttons, text="✓ Validate All",
                 bg=self.colors['warning'], fg='white',
                 command=self.validate_all_experiments,
                 font=('Segoe UI', 10),
                 relief='flat', padx=20, pady=8).pack(side='left', padx=5)

    def populate_default_data(self):
        """Populate with default experiment and unit data"""
        # Default unit data values
        default_values = {
            "Product": self.current_product,  # Use the selected default product
            "Visual ID": "TestUnitData",
            "Bucket": "IDIPAR",
            "COM Port": "11",
            "IP Address": "10.250.0.2"
        }

        # Update unit data widgets
        for field_name, value in default_values.items():
            if field_name in self.unit_data_widgets:
                self.unit_data_widgets[field_name]['var'].set(value)

        # Don't create default experiment - start with empty tabs
        # User will create first experiment manually

    def add_new_experiment(self, name=None):
        """Add a new experiment tab"""
        if name is None:
            base_name = "New_Experiment"
            counter = 1
            name = base_name
            while name in self.experiments:
                name = f"{base_name}_{counter}"
                counter += 1

        # Create experiment data structure
        exp_data = self.create_default_experiment_data()
        exp_data["Test Name"] = name

        self.experiments[name] = {
            'data': exp_data,
            'widgets': {},
            'tab_frame': None,
            'modified': False
        }

        # Create the tab
        self.create_experiment_tab(name)

        # Load data into widgets
        self.load_experiment_data(name)

        # Switch to new tab
        self.experiment_notebook.select(len(self.experiment_notebook.tabs()) - 1)

        # Update conditional sections
        self.update_conditional_sections()

    def delete_current_experiment(self):
        """Delete the currently selected experiment"""
        if len(self.experiments) <= 1:
            show_warning("Cannot Delete",
                                  "Cannot delete the last experiment. At least one must remain.")
            return

        current_tab = self.experiment_notebook.index('current')
        tab_id = self.experiment_notebook.tabs()[current_tab]
        exp_name = self.experiment_notebook.tab(tab_id, 'text')

        if messagebox.askyesno("Confirm Delete",
                              f"Are you sure you want to delete experiment '{exp_name}'?"):
            del self.experiments[exp_name]
            self.experiment_notebook.forget(current_tab)

            # Remove custom tab button from primary
            if exp_name in self.primary_tab_buttons:
                self.primary_tab_buttons[exp_name].destroy()
                del self.primary_tab_buttons[exp_name]

            # Remove from secondary notebook if in side-by-side view
            if self.view_mode == "side-by-side" and self.experiment_notebook_secondary:
                # Find and remove from secondary notebook
                for i, tab_id in enumerate(self.experiment_notebook_secondary.tabs()):
                    if self.experiment_notebook_secondary.tab(tab_id, 'text') == exp_name:
                        self.experiment_notebook_secondary.forget(i)
                        break

                # Remove custom tab button from secondary
                if hasattr(self, 'secondary_tab_buttons') and exp_name in self.secondary_tab_buttons:
                    self.secondary_tab_buttons[exp_name].destroy()
                    del self.secondary_tab_buttons[exp_name]

                # Clean up related data
                if exp_name in self.secondary_field_checkboxes:
                    del self.secondary_field_checkboxes[exp_name]
                if exp_name in self.secondary_notebook_frames:
                    del self.secondary_notebook_frames[exp_name]

    def duplicate_current_experiment(self):
        """Duplicate the currently selected experiment"""
        current_tab = self.experiment_notebook.index('current')
        tab_id = self.experiment_notebook.tabs()[current_tab]
        exp_name = self.experiment_notebook.tab(tab_id, 'text')

        # Save current data
        self.save_current_experiment()

        # Create new name
        base_name = f"{exp_name}_copy"
        counter = 1
        new_name = base_name
        while new_name in self.experiments:
            new_name = f"{base_name}_{counter}"
            counter += 1

        # Copy experiment data
        exp_data = self.experiments[exp_name]['data'].copy()
        exp_data["Test Name"] = new_name

        self.experiments[new_name] = {
            'data': exp_data,
            'widgets': {},
            'tab_frame': None,
            'modified': False
        }

        # Create tab
        self.create_experiment_tab(new_name)
        self.load_experiment_data(new_name)

        # If in side-by-side view, also add to secondary notebook
        if self.view_mode == "side-by-side" and self.experiment_notebook_secondary:
            self.create_experiment_tab_in_notebook(new_name, self.experiment_notebook_secondary)

        # Switch to new tab
        self.experiment_notebook.select(len(self.experiment_notebook.tabs()) - 1)

    def clear_current_experiment(self):
        """Clear all fields in current experiment to default values"""
        if not self.current_experiment:
            show_warning("No Experiment", "Please select an experiment first.")
            return

        if messagebox.askyesno("Confirm Clear",
                              f"Clear all fields in '{self.current_experiment}' to default values?"):
            # Reset to default data
            default_data = self.create_default_experiment_data()
            default_data["Test Name"] = self.current_experiment

            self.experiments[self.current_experiment]['data'] = default_data
            self.load_experiment_data(self.current_experiment)

            messagebox.showinfo("Cleared", f"Experiment '{self.current_experiment}' has been reset.")

    def apply_template_to_current(self):
        """Apply selected template to current experiment"""
        if not self.current_experiment:
            show_warning("No Experiment", "Please select an experiment first.")
            return

        selection = self.template_listbox.curselection()
        if not selection:
            show_warning("No Template", "Please select a template from the Templates panel.")
            return

        template_name = self.template_listbox.get(selection[0])
        template_data = self.templates[template_name].copy()

        # Keep current experiment name
        template_data["Test Name"] = self.current_experiment

        # Update experiment data
        self.experiments[self.current_experiment]['data'] = template_data

        # Reload widgets
        self.load_experiment_data(self.current_experiment)

        messagebox.showinfo("Applied", f"Template '{template_name}' applied to '{self.current_experiment}'.")

    def create_custom_tab_button(self, exp_name, notebook_type):
        """Create a custom tab button in the scrollable tabs container"""
        if notebook_type == 'primary':
            parent_frame = self.primary_tabs_frame
            tab_buttons = self.primary_tab_buttons
            notebook = self.experiment_notebook
        else:
            parent_frame = self.secondary_tabs_frame
            tab_buttons = self.secondary_tab_buttons
            notebook = self.experiment_notebook_secondary

        # Create tab button
        btn = tk.Button(parent_frame, text=exp_name,
                       font=('Segoe UI', 9),
                       bg=self.colors['light'], fg='black',
                       relief='raised', bd=1,
                       padx=10, pady=5,
                       command=lambda: self.select_custom_tab(exp_name, notebook_type))
        btn.pack(side='left', padx=1, pady=2)

        # Store button reference
        tab_buttons[exp_name] = btn

        # Update button appearance if this is the current tab
        current_tab = notebook.index('current')
        tab_id = notebook.tabs()[current_tab] if notebook.tabs() else None
        if tab_id and notebook.tab(tab_id, 'text') == exp_name:
            is_disabled = self.is_experiment_disabled(exp_name)
            if is_disabled:
                btn.configure(bg=self.colors['disabled_tab'], fg=self.colors['disabled_tab_text'], relief='sunken')
            else:
                btn.configure(bg=self.colors['primary'], fg='white', relief='sunken')

    def select_custom_tab(self, exp_name, notebook_type):
        """Select a tab when custom button is clicked"""
        if notebook_type == 'primary':
            notebook = self.experiment_notebook
            tab_buttons = self.primary_tab_buttons
        else:
            notebook = self.experiment_notebook_secondary
            tab_buttons = self.secondary_tab_buttons

        # Find and select the tab
        for i, tab_id in enumerate(notebook.tabs()):
            if notebook.tab(tab_id, 'text') == exp_name:
                notebook.select(i)
                break

        # Update button appearances
        for name, btn in tab_buttons.items():
            if name == exp_name:
                # Check if experiment is disabled
                is_disabled = self.is_experiment_disabled(name)
                if is_disabled:
                    btn.configure(bg=self.colors['disabled_tab'], fg=self.colors['disabled_tab_text'], relief='sunken')
                else:
                    btn.configure(bg=self.colors['primary'], fg='white', relief='sunken')
            else:
                # Check if experiment is disabled for non-selected tabs
                is_disabled = self.is_experiment_disabled(name)
                if is_disabled:
                    btn.configure(bg='#cbd5e1', fg='#64748b', relief='raised')
                else:
                    btn.configure(bg=self.colors['light'], fg='black', relief='raised')

    def show_tab_menu(self, notebook_type):
        """Show dropdown menu with all tabs"""
        if notebook_type == 'primary':
            notebook = self.experiment_notebook
        else:
            notebook = self.experiment_notebook_secondary
            if not notebook:
                return

        # Create popup menu
        menu = tk.Menu(self.root, tearoff=0)

        # Add all tabs to menu
        for i, tab_id in enumerate(notebook.tabs()):
            tab_name = notebook.tab(tab_id, 'text')
            is_current = (i == notebook.index('current'))
            label = f"• {tab_name}" if is_current else f"  {tab_name}"
            menu.add_command(label=label,
                           command=lambda idx=i, nt=notebook_type: self.select_tab_from_menu(idx, nt))

        # Show menu at mouse position
        try:
            if notebook_type == 'primary':
                x = self.primary_nav_frame.winfo_rootx() + self.primary_nav_frame.winfo_width() - 50
                y = self.primary_nav_frame.winfo_rooty() + self.primary_nav_frame.winfo_height()
            else:
                x = self.secondary_nav_frame.winfo_rootx() + self.secondary_nav_frame.winfo_width() - 50
                y = self.secondary_nav_frame.winfo_rooty() + self.secondary_nav_frame.winfo_height()
            menu.post(x, y)
        except:
            pass

    def select_tab_from_menu(self, index, notebook_type):
        """Select a tab from the dropdown menu"""
        if notebook_type == 'primary':
            notebook = self.experiment_notebook
        else:
            notebook = self.experiment_notebook_secondary

        notebook.select(index)

    def scroll_tabs(self, notebook_type, direction):
        """Scroll through tabs using navigation arrows"""
        try:
            if notebook_type == 'primary':
                notebook = self.experiment_notebook
            else:  # secondary
                notebook = self.experiment_notebook_secondary
                if not notebook:
                    return

            current = notebook.index('current')
            tabs_count = len(notebook.tabs())
            if tabs_count == 0:
                return

            new_index = (current + direction) % tabs_count
            notebook.select(new_index)
        except:
            pass

    def _scroll_notebook(self, event, direction):
        """Scroll through notebook tabs (Linux)"""
        try:
            current = self.experiment_notebook.index('current')
            tabs_count = len(self.experiment_notebook.tabs())
            new_index = (current + direction) % tabs_count
            self.experiment_notebook.select(new_index)
        except:
            pass

    def _on_notebook_mousewheel(self, event):
        """Scroll through notebook tabs (Windows)"""
        try:
            current = self.experiment_notebook.index('current')
            tabs_count = len(self.experiment_notebook.tabs())
            direction = -1 if event.delta > 0 else 1
            new_index = (current + direction) % tabs_count
            self.experiment_notebook.select(new_index)
        except:
            pass

    def on_experiment_tab_change(self, event):
        """Handle experiment tab change"""
        current_tab = self.experiment_notebook.index('current')
        if current_tab >= 0:
            tab_id = self.experiment_notebook.tabs()[current_tab]
            exp_name = self.experiment_notebook.tab(tab_id, 'text')
            self.current_experiment = exp_name

            # Update custom tab button appearances
            for name, btn in self.primary_tab_buttons.items():
                is_disabled = self.is_experiment_disabled(name)
                is_modified = self.experiments.get(name, {}).get('modified', False)

                # Update text with asterisk if modified
                display_text = f"* {name}" if is_modified else name
                btn.configure(text=display_text)

                if name == exp_name:
                    # Selected tab
                    if is_modified:
                        btn.configure(bg=self.colors['warning'], fg='white', relief='sunken')
                    elif is_disabled:
                        btn.configure(bg=self.colors['disabled_tab'], fg=self.colors['disabled_tab_text'], relief='sunken')
                    else:
                        btn.configure(bg=self.colors['primary'], fg='white', relief='sunken')
                else:
                    # Non-selected tab
                    if is_modified:
                        btn.configure(bg='#fbbf24', fg='black', relief='raised')
                    elif is_disabled:
                        btn.configure(bg='#cbd5e1', fg='#64748b', relief='raised')
                    else:
                        btn.configure(bg=self.colors['light'], fg='black', relief='raised')

            # Scroll to show the selected tab button
            if exp_name in self.primary_tab_buttons:
                btn = self.primary_tab_buttons[exp_name]
                self.primary_tabs_canvas.update_idletasks()
                x = btn.winfo_x()
                canvas_width = self.primary_tabs_canvas.winfo_width()
                if self.primary_tabs_frame.winfo_width() > 0:
                    self.primary_tabs_canvas.xview_moveto(max(0, (x - canvas_width/2) / self.primary_tabs_frame.winfo_width()))

            # Refresh status indicator to match current state
            self.refresh_status_indicator(exp_name)

            self.update_conditional_sections()

            # If in side-by-side view, also update secondary notebook's current tab
            if self.view_mode == 'side-by-side' and self.experiment_notebook_secondary:
                try:
                    secondary_tab = self.experiment_notebook_secondary.index('current')
                    if secondary_tab >= 0:
                        secondary_tab_id = self.experiment_notebook_secondary.tabs()[secondary_tab]
                        secondary_exp = self.experiment_notebook_secondary.tab(secondary_tab_id, 'text')
                        old_current = self.current_experiment
                        self.current_experiment = secondary_exp
                        self.update_conditional_sections()
                        self.current_experiment = old_current
                except:
                    pass

    def update_view_button(self):
        """Update view toggle button to reflect current state"""
        if self.view_mode == "single":
            # Currently in single view, show option to switch to side-by-side
            self.view_mode_button.config(
                text="⊞ Side-by-Side View",
                bg=self.colors['accent']
            )
        else:
            # Currently in side-by-side view, show option to switch to single
            self.view_mode_button.config(
                text="⊟ Single View",
                bg=self.colors['success']
            )

    def toggle_view_mode(self):
        """Toggle between single and side-by-side experiment view"""
        if self.view_mode == "single":
            # Switch to side-by-side view
            self.view_mode = "side-by-side"

            # Create secondary notebook if it doesn't exist
            if self.experiment_notebook_secondary is None:
                # Secondary notebook container with navigation arrows
                self.secondary_notebook_container = tk.Frame(self.paned_window, bg='white')
                self.paned_window.add(self.secondary_notebook_container, stretch='always', minsize=300)

                # Navigation frame for secondary notebook
                self.secondary_nav_frame = tk.Frame(self.secondary_notebook_container, bg=self.colors['light'], height=35)
                self.secondary_nav_frame.pack(fill='x', side='top')
                self.secondary_nav_frame.pack_propagate(False)

                tk.Button(self.secondary_nav_frame, text="◄", font=('Segoe UI', 10, 'bold'),
                         bg=self.colors['primary'], fg='white',
                         command=lambda: self.scroll_tabs('secondary', -1),
                         relief='flat', width=3).pack(side='left', padx=2, pady=2)

                tk.Label(self.secondary_nav_frame, text="Compare With",
                        font=('Segoe UI', 9, 'bold'), bg=self.colors['light'],
                        fg=self.colors['primary']).pack(side='left', padx=10)

                tk.Button(self.secondary_nav_frame, text="►", font=('Segoe UI', 10, 'bold'),
                         bg=self.colors['primary'], fg='white',
                         command=lambda: self.scroll_tabs('secondary', 1),
                         relief='flat', width=3).pack(side='left', padx=2, pady=2)

                # Copy controls in navigation bar
                tk.Label(self.secondary_nav_frame, text="|",
                        font=('Segoe UI', 10), bg=self.colors['light'],
                        fg='#cccccc').pack(side='left', padx=10)

                tk.Button(self.secondary_nav_frame, text="←", font=('Segoe UI', 14, 'bold'),
                         bg=self.colors['accent'], fg='white',
                         command=self.copy_right_to_left,
                         relief='flat', width=3).pack(side='left', padx=2, pady=2)

                tk.Button(self.secondary_nav_frame, text="→", font=('Segoe UI', 14, 'bold'),
                         bg=self.colors['accent'], fg='white',
                         command=self.copy_left_to_right,
                         relief='flat', width=3).pack(side='left', padx=2, pady=2)

                tk.Label(self.secondary_nav_frame, text="|",
                        font=('Segoe UI', 10), bg=self.colors['light'],
                        fg='#cccccc').pack(side='left', padx=10)

                tk.Button(self.secondary_nav_frame, text="☑ All", font=('Segoe UI', 8),
                         bg=self.colors['success'], fg='white',
                         command=self.select_all_fields,
                         relief='flat', padx=8, pady=3).pack(side='left', padx=2, pady=2)

                tk.Button(self.secondary_nav_frame, text="☐ None", font=('Segoe UI', 8),
                         bg=self.colors['warning'], fg='white',
                         command=self.deselect_all_fields,
                         relief='flat', padx=8, pady=3).pack(side='left', padx=2, pady=2)

                tk.Button(self.secondary_nav_frame, text="≡", font=('Segoe UI', 12, 'bold'),
                         bg=self.colors['accent'], fg='white',
                         command=lambda: self.show_tab_menu('secondary'),
                         relief='flat', width=3).pack(side='right', padx=2, pady=2)

                # Create custom scrollable tabs container for secondary
                self.secondary_tabs_container = tk.Frame(self.secondary_notebook_container, bg='white')
                self.secondary_tabs_container.pack(fill='x', side='top')

                # Scrollable tabs canvas
                self.secondary_tabs_canvas = tk.Canvas(self.secondary_tabs_container, bg=self.colors['light'], height=35, highlightthickness=0)
                self.secondary_tabs_canvas.pack(fill='x', expand=True)

                # Frame inside canvas to hold tab buttons
                self.secondary_tabs_frame = tk.Frame(self.secondary_tabs_canvas, bg=self.colors['light'])
                self.secondary_tabs_window = self.secondary_tabs_canvas.create_window((0, 0), window=self.secondary_tabs_frame, anchor='nw')

                # Bind canvas configuration
                self.secondary_tabs_frame.bind('<Configure>', lambda e: self.secondary_tabs_canvas.configure(scrollregion=self.secondary_tabs_canvas.bbox('all')))

                # Track tab buttons for secondary notebook
                self.secondary_tab_buttons = {}

                self.experiment_notebook_secondary = ttk.Notebook(self.secondary_notebook_container)
                self.experiment_notebook_secondary.pack(fill='both', expand=True)
                self.experiment_notebook_secondary.bind('<<NotebookTabChanged>>', self.on_secondary_tab_change)

                # Enable mouse wheel scrolling on secondary notebook tabs
                self.experiment_notebook_secondary.bind('<Button-4>', lambda e: self._scroll_secondary_notebook(e, -1))
                self.experiment_notebook_secondary.bind('<Button-5>', lambda e: self._scroll_secondary_notebook(e, 1))
                self.experiment_notebook_secondary.bind('<MouseWheel>', self._on_secondary_notebook_mousewheel)

                # Copy all experiments to secondary notebook
                for exp_name in self.experiments.keys():
                    self.create_experiment_tab_in_notebook(exp_name, self.experiment_notebook_secondary)
            else:
                # Show the secondary notebook
                self.experiment_notebook_secondary.pack(fill='both', expand=True, side='right')

            # Update button to reflect new state
            self.update_view_button()
        else:
            # Switch back to single view
            self.view_mode = "single"

            # Remove secondary notebook from PanedWindow
            if hasattr(self, 'secondary_notebook_container') and self.secondary_notebook_container:
                try:
                    self.paned_window.remove(self.secondary_notebook_container)
                    self.secondary_notebook_container.destroy()
                    self.secondary_notebook_container = None
                except:
                    pass
            if self.experiment_notebook_secondary:
                self.experiment_notebook_secondary = None

            # Update button to reflect new state
            self.update_view_button()

    def create_experiment_tab_in_notebook(self, exp_name, notebook):
        """Create experiment tab in a specific notebook (for side-by-side view)"""
        # Check if experiment exists
        if exp_name not in self.experiments:
            return

        # Create main frame for this experiment
        tab_frame = tk.Frame(notebook, bg='white')
        notebook.add(tab_frame, text=exp_name)

        # Store frame reference for secondary notebook
        if notebook == self.experiment_notebook_secondary:
            self.secondary_notebook_frames[exp_name] = tab_frame

        # Create scrollable canvas
        canvas = tk.Canvas(tab_frame, bg='white')
        scrollbar = tk.Scrollbar(tab_frame, orient='vertical', command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='white')

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        # Get widgets from main experiment - create live view sharing same variables
        widgets = self.experiments[exp_name]['widgets']

        # Group fields by section for better organization
        sections = {
            'Basic Information': ['Experiment', 'Test Name', 'Test Mode', 'Test Type'],
            'Advanced Configuration': ['Configuration (Mask)', 'Boot Breakpoint', 'Pseudo Config',
                                      'Disable 2 Cores', 'Disable 1 Core', 'Check Core', 'Stop on Fail',
                                      'Fuse File', 'Bios File'],
            'Test Configuration': ['TTL Folder', 'Scripts File', 'Post Process', 'Pass String',
                                  'Fail String', 'Test Number', 'Test Time', 'Loops', 'Reset',
                                  'Reset on PASS', 'FastBoot', 'Core License', '600W Unit'],
            'Voltage & Frequency': ['Voltage Type', 'Voltage IA', 'Voltage CFC', 'Frequency IA', 'Frequency CFC']
        }

        row = 0
        for section_name, field_list in sections.items():
            section = ttk.LabelFrame(scrollable_frame, text=f"🔹 {section_name}",
                                    style='Section.TLabelframe', padding=15)
            section.grid(row=row, column=0, sticky='ew', padx=10, pady=10)
            scrollable_frame.grid_columnconfigure(0, weight=1)

            # Add fields from list that exist in widgets
            field_row = 0
            for field_name in field_list:
                if field_name in widgets:
                    field_info = widgets[field_name]

                    # Checkbox for selection
                    checkbox_var = tk.BooleanVar(value=False)
                    checkbox = tk.Checkbutton(section, variable=checkbox_var, bg='white')
                    checkbox.grid(row=field_row, column=0, sticky='w', padx=2, pady=3)

                    # Store checkbox reference
                    if exp_name not in self.secondary_field_checkboxes:
                        self.secondary_field_checkboxes[exp_name] = {}
                    self.secondary_field_checkboxes[exp_name][field_name] = checkbox_var

                    lbl = tk.Label(section, text=f"{field_name}:",
                                  font=('Segoe UI', 9), bg='white', anchor='w', width=23)
                    lbl.grid(row=field_row, column=1, sticky='w', padx=5, pady=3)

                    # Create read-only entry bound to same variable for live updates
                    if field_info['type'] == 'bool':
                        value_widget = tk.Checkbutton(section, variable=field_info['var'],
                                                     bg='white', state='disabled')
                    else:
                        value_widget = tk.Entry(section, textvariable=field_info['var'],
                                              font=('Segoe UI', 9), state='readonly',
                                              width=50, bg='#f9f9f9')

                    value_widget.grid(row=field_row, column=2, sticky='ew', padx=5, pady=3)
                    section.grid_columnconfigure(2, weight=1)
                    field_row += 1

            row += 1

        # Create custom tab button if this is for secondary notebook
        if notebook == self.experiment_notebook_secondary and hasattr(self, 'secondary_tab_buttons'):
            self.create_custom_tab_button(exp_name, 'secondary')

    def select_all_fields(self):
        """Select all fields in the secondary (right) panel"""
        if not self.experiment_notebook_secondary:
            return

        try:
            current_tab = self.experiment_notebook_secondary.index('current')
            if current_tab >= 0:
                tab_id = self.experiment_notebook_secondary.tabs()[current_tab]
                exp_name = self.experiment_notebook_secondary.tab(tab_id, 'text')

                if exp_name in self.secondary_field_checkboxes:
                    for checkbox_var in self.secondary_field_checkboxes[exp_name].values():
                        checkbox_var.set(True)
        except:
            pass

    def deselect_all_fields(self):
        """Deselect all fields in the secondary (right) panel"""
        if not self.experiment_notebook_secondary:
            return

        try:
            current_tab = self.experiment_notebook_secondary.index('current')
            if current_tab >= 0:
                tab_id = self.experiment_notebook_secondary.tabs()[current_tab]
                exp_name = self.experiment_notebook_secondary.tab(tab_id, 'text')

                if exp_name in self.secondary_field_checkboxes:
                    for checkbox_var in self.secondary_field_checkboxes[exp_name].values():
                        checkbox_var.set(False)
        except:
            pass

    def copy_left_to_right(self):
        """Copy selected fields from left experiment to right experiment"""
        if not self.experiment_notebook_secondary:
            return

        try:
            # Get current experiments
            left_tab = self.experiment_notebook.index('current')
            right_tab = self.experiment_notebook_secondary.index('current')

            if left_tab >= 0 and right_tab >= 0:
                left_exp = self.experiment_notebook.tab(self.experiment_notebook.tabs()[left_tab], 'text')
                right_exp = self.experiment_notebook_secondary.tab(
                    self.experiment_notebook_secondary.tabs()[right_tab], 'text')

                if left_exp not in self.experiments or right_exp not in self.experiments:
                    return

                # Get selected fields
                selected_fields = []
                if right_exp in self.secondary_field_checkboxes:
                    for field_name, checkbox_var in self.secondary_field_checkboxes[right_exp].items():
                        if checkbox_var.get():
                            selected_fields.append(field_name)

                if not selected_fields:
                    show_warning("No Selection", "Please select fields to copy.")
                    return

                # Copy data
                left_widgets = self.experiments[left_exp]['widgets']
                right_widgets = self.experiments[right_exp]['widgets']

                copied_count = 0
                for field_name in selected_fields:
                    if field_name in left_widgets and field_name in right_widgets:
                        value = left_widgets[field_name]['var'].get()
                        right_widgets[field_name]['var'].set(value)
                        copied_count += 1

                messagebox.showinfo("Copy Complete",
                                   f"Copied {copied_count} field(s) from '{left_exp}' to '{right_exp}'.")
        except Exception as e:
            show_error("Copy Error", f"Error copying fields: {str(e)}")

    def copy_right_to_left(self):
        """Copy selected fields from right experiment to left experiment"""
        if not self.experiment_notebook_secondary:
            return

        try:
            # Get current experiments
            left_tab = self.experiment_notebook.index('current')
            right_tab = self.experiment_notebook_secondary.index('current')

            if left_tab >= 0 and right_tab >= 0:
                left_exp = self.experiment_notebook.tab(self.experiment_notebook.tabs()[left_tab], 'text')
                right_exp = self.experiment_notebook_secondary.tab(
                    self.experiment_notebook_secondary.tabs()[right_tab], 'text')

                if left_exp not in self.experiments or right_exp not in self.experiments:
                    return

                # Get selected fields
                selected_fields = []
                if right_exp in self.secondary_field_checkboxes:
                    for field_name, checkbox_var in self.secondary_field_checkboxes[right_exp].items():
                        if checkbox_var.get():
                            selected_fields.append(field_name)

                if not selected_fields:
                    show_warning("No Selection", "Please select fields to copy.")
                    return

                # Copy data
                left_widgets = self.experiments[left_exp]['widgets']
                right_widgets = self.experiments[right_exp]['widgets']

                copied_count = 0
                for field_name in selected_fields:
                    if field_name in left_widgets and field_name in right_widgets:
                        value = right_widgets[field_name]['var'].get()
                        left_widgets[field_name]['var'].set(value)
                        copied_count += 1

                messagebox.showinfo("Copy Complete",
                                   f"Copied {copied_count} field(s) from '{right_exp}' to '{left_exp}'.")
        except Exception as e:
            show_error("Copy Error", f"Error copying fields: {str(e)}")

    def _scroll_secondary_notebook(self, event, direction):
        """Scroll through secondary notebook tabs (Linux)"""
        try:
            if self.experiment_notebook_secondary:
                current = self.experiment_notebook_secondary.index('current')
                tabs_count = len(self.experiment_notebook_secondary.tabs())
                new_index = (current + direction) % tabs_count
                self.experiment_notebook_secondary.select(new_index)
        except:
            pass

    def _on_secondary_notebook_mousewheel(self, event):
        """Scroll through secondary notebook tabs (Windows)"""
        try:
            if self.experiment_notebook_secondary:
                current = self.experiment_notebook_secondary.index('current')
                tabs_count = len(self.experiment_notebook_secondary.tabs())
                direction = -1 if event.delta > 0 else 1
                new_index = (current + direction) % tabs_count
                self.experiment_notebook_secondary.select(new_index)
        except:
            pass

    def on_secondary_tab_change(self, event):
        """Handle secondary notebook tab change"""
        try:
            if self.experiment_notebook_secondary:
                current_tab = self.experiment_notebook_secondary.index('current')
                tab_id = self.experiment_notebook_secondary.tabs()[current_tab]
                exp_name = self.experiment_notebook_secondary.tab(tab_id, 'text')

                # Update custom tab button appearances
                if hasattr(self, 'secondary_tab_buttons'):
                    for name, btn in self.secondary_tab_buttons.items():
                        is_disabled = self.is_experiment_disabled(name)
                        if name == exp_name:
                            # Selected tab
                            if is_disabled:
                                btn.configure(bg=self.colors['disabled_tab'], fg=self.colors['disabled_tab_text'], relief='sunken')
                            else:
                                btn.configure(bg=self.colors['primary'], fg='white', relief='sunken')
                        else:
                            # Non-selected tab
                            if is_disabled:
                                btn.configure(bg='#cbd5e1', fg='#64748b', relief='raised')
                            else:
                                btn.configure(bg=self.colors['light'], fg='black', relief='raised')

                    # Scroll to show the selected tab button
                    if exp_name in self.secondary_tab_buttons:
                        btn = self.secondary_tab_buttons[exp_name]
                        self.secondary_tabs_canvas.update_idletasks()
                        x = btn.winfo_x()
                        canvas_width = self.secondary_tabs_canvas.winfo_width()
                        if self.secondary_tabs_frame.winfo_width() > 0:
                            self.secondary_tabs_canvas.xview_moveto(max(0, (x - canvas_width/2) / self.secondary_tabs_frame.winfo_width()))

                if exp_name in self.experiments:
                    # Refresh status indicator to match current state
                    self.refresh_status_indicator(exp_name)

                    # Get the widgets for this experiment and trigger conditional update
                    old_current = self.current_experiment
                    self.current_experiment = exp_name
                    self.update_conditional_sections()
                    self.current_experiment = old_current  # Restore primary tab's experiment
        except:
            pass

    def update_conditional_sections(self):
        """Enable/disable sections based on Test Type, Content, and Product"""
        if not self.current_experiment or self.current_experiment not in self.experiments:
            return

        # Prevent recursion
        if self._updating_sections:
            return

        self._updating_sections = True
        try:
            self._do_update_conditional_sections()
        finally:
            self._updating_sections = False

    def _do_update_conditional_sections(self):
        """Internal method that does the actual section updates"""
        widgets = self.experiments[self.current_experiment]['widgets']

        # Get current values
        test_type = widgets.get('Test Type', {}).get('var', tk.StringVar()).get()
        content = widgets.get('Content', {}).get('var', tk.StringVar()).get()
        current_product = self.product_var.get()

        # Test Type conditional sections: Loops, Sweep, Shmoo
        # Loops section - only enabled when Test Type is Loops
        if 'loops_section' in widgets:
            section = widgets['loops_section']
            if test_type == 'Loops':
                self.enable_section(section, widgets, ['Loops'])
                section.configure(style='Section.TLabelframe')
            else:
                self.disable_section(section, widgets, ['Loops'])
                section.configure(style='Disabled.TLabelframe')

        # Sweep section - only enabled when Test Type is Sweep
        if 'sweep_section' in widgets:
            section = widgets['sweep_section']
            if test_type == 'Sweep':
                self.enable_section(section, widgets, ['Type', 'Domain', 'Start', 'End', 'Steps'])
                section.configure(style='Section.TLabelframe')
            else:
                self.disable_section(section, widgets, ['Type', 'Domain', 'Start', 'End', 'Steps'])
                section.configure(style='Disabled.TLabelframe')

        # Shmoo section - only enabled when Test Type is Shmoo
        if 'shmoo_section' in widgets:
            section = widgets['shmoo_section']
            if test_type == 'Shmoo':
                self.enable_section(section, widgets, ['ShmooFile', 'ShmooLabel'])
                section.configure(style='Section.TLabelframe')
            else:
                self.disable_section(section, widgets, ['ShmooFile', 'ShmooLabel'])
                section.configure(style='Disabled.TLabelframe')

        # Content Selection: Gray out non-selected content sections
        # Linux section - only enabled when Content is Linux
        if 'linux_section' in widgets:
            section = widgets['linux_section']
            linux_fields = [f'Linux {x}' for x in ['Path', 'Pre Command', 'Post Command', 'Pass String', 'Fail String', 'Content Wait Time']]
            linux_fields += ['Startup Linux'] + [f'Linux Content Line {i}' for i in range(10)]

            if content == 'Linux':
                self.enable_section(section, widgets, linux_fields)
                section.configure(style='Section.TLabelframe')
            else:
                self.disable_section(section, widgets, linux_fields)
                section.configure(style='Disabled.TLabelframe')

        # Dragon section - only enabled when Content is Dragon
        if 'dragon_section' in widgets:
            section = widgets['dragon_section']
            dragon_fields = ['Dragon Content Path', 'Startup Dragon', 'Dragon Pre Command', 'Dragon Post Command',
                           'ULX Path', 'ULX CPU', 'Product Chop', 'VVAR0', 'VVAR1', 'VVAR2', 'VVAR3',
                           'VVAR_EXTRA', 'Dragon Content Line']

            if content == 'Dragon':
                self.enable_section(section, widgets, dragon_fields)
                section.configure(style='Section.TLabelframe')
            else:
                self.disable_section(section, widgets, dragon_fields)
                section.configure(style='Disabled.TLabelframe')

        # Merlin section - only enabled when Content is Dragon (tied to Dragon)
        if 'merlin_section' in widgets:
            section = widgets['merlin_section']
            merlin_fields = ['Merlin Name', 'Merlin Drive', 'Merlin Path']

            if content == 'Dragon':
                self.enable_section(section, widgets, merlin_fields)
                section.configure(style='Section.TLabelframe')
            else:
                self.disable_section(section, widgets, merlin_fields)
                section.configure(style='Disabled.TLabelframe')

        # Gray out product-specific fields based on config
        field_enable_config = self.config_template.get('field_enable_config', {})

        # Apply product-specific field enabling from config
        for field_name, field_info in widgets.items():
            # Skip section markers (any key ending with _section)
            if field_name.endswith('_section'):
                continue

            # Ensure field_info has required keys
            if not isinstance(field_info, dict) or 'widget' not in field_info:
                continue

            if field_name in field_enable_config:
                enabled_products = field_enable_config[field_name]
                widget = field_info['widget']
                if current_product in enabled_products:
                    # Enable based on widget type
                    if isinstance(widget, ttk.Combobox):
                        widget.configure(state='readonly')
                    else:
                        widget.configure(state='normal')
                    field_info['label'].configure(fg='black')
                else:
                    # Disable properly for all widget types
                    if isinstance(widget, ttk.Combobox):
                        widget.configure(state='disabled')
                    elif isinstance(widget, tk.Checkbutton):
                        widget.configure(state='disabled')
                    else:
                        widget.configure(state='disabled')
                    field_info['label'].configure(fg='#cccccc')

        # Legacy fallback for backwards compatibility
        if not field_enable_config:
            # Pseudo Config - GNR only
            if 'Pseudo Config' in widgets:
                widget = widgets['Pseudo Config']['widget']
                if current_product == 'GNR':
                    widget.configure(state='normal')
                    widgets['Pseudo Config']['label'].configure(fg='black')
                else:
                    widget.configure(state='disabled')
                    widgets['Pseudo Config']['label'].configure(fg='#cccccc')

            # Disable 2 Cores - CWF only
            if 'Disable 2 Cores' in widgets:
                widget = widgets['Disable 2 Cores']['widget']
                if current_product == 'CWF':
                    if isinstance(widget, ttk.Combobox):
                        widget.configure(state='readonly')
                    else:
                        widget.configure(state='normal')
                    widgets['Disable 2 Cores']['label'].configure(fg='black')
                else:
                    if isinstance(widget, ttk.Combobox):
                        widget.configure(state='disabled')
                    else:
                        widget.configure(state='disabled')
                    widgets['Disable 2 Cores']['label'].configure(fg='#cccccc')

            # Disable 1 Core - DMR only
            if 'Disable 1 Core' in widgets:
                if current_product == 'DMR':
                    widgets['Disable 1 Core']['widget'].configure(state='normal')
                    widgets['Disable 1 Core']['label'].configure(fg='black')
                else:
                    widgets['Disable 1 Core']['widget'].configure(state='disabled')
                    widgets['Disable 1 Core']['label'].configure(fg='#cccccc')

            # Core License - GNR/DMR only (has dropdown for these products)
            if 'Core License' in widgets:
                if current_product in ['GNR', 'DMR']:
                    widgets['Core License']['widget'].configure(state='readonly')
                    widgets['Core License']['label'].configure(fg='black')
                else:
                    widgets['Core License']['widget'].configure(state='disabled')
                    widgets['Core License']['label'].configure(fg='#cccccc')

    def enable_section(self, section, widgets, field_names):
        """Enable a section and its fields"""
        for field in field_names:
            if field in widgets:
                try:
                    widget = widgets[field]['widget']
                    # Handle different widget types
                    if isinstance(widget, ttk.Combobox):
                        widget.configure(state='readonly')  # Combobox uses 'readonly' not 'normal'
                    elif isinstance(widget, tk.Checkbutton):
                        widget.configure(state='normal')
                    else:
                        widget.configure(state='normal')
                    widgets[field]['label'].configure(fg='black')
                except Exception as e:
                    print(f"Could not enable {field}: {e}")

    def disable_section(self, section, widgets, field_names):
        """Disable a section and its fields"""
        for field in field_names:
            if field in widgets:
                try:
                    widget = widgets[field]['widget']
                    # Handle different widget types
                    if isinstance(widget, ttk.Combobox):
                        widget.configure(state='disabled')
                    elif isinstance(widget, tk.Checkbutton):
                        widget.configure(state='disabled')
                    else:
                        widget.configure(state='disabled')
                    widgets[field]['label'].configure(fg='#cccccc')
                except Exception as e:
                    print(f"Could not disable {field}: {e}")

    def gray_out_experiment(self, widgets):
        """Gray out entire experiment when Disabled"""
        for field_name, field_info in widgets.items():
            # Skip section markers (any key ending with _section)
            if field_name.endswith('_section'):
                continue

            if field_name == 'Experiment':
                continue  # Keep Experiment toggle enabled so user can re-enable

            # Ensure field_info has required keys
            if not isinstance(field_info, dict):
                continue

            try:
                if 'widget' in field_info:
                    widget = field_info['widget']
                    # Properly disable all widget types
                    if isinstance(widget, ttk.Combobox):
                        widget.configure(state='disabled')
                    elif isinstance(widget, tk.Checkbutton):
                        widget.configure(state='disabled')
                    else:
                        widget.configure(state='disabled')
                # Don't change label color for Experiment field (status label has custom styling)
                if 'label' in field_info and field_name != 'Experiment':
                    field_info['label'].configure(fg='#cccccc')
            except:
                pass

    def enable_experiment(self, widgets):
        """Enable entire experiment when Enabled - but conditional sections will override as needed"""
        for field_name, field_info in widgets.items():
            # Skip section markers (any key ending with _section)
            if field_name.endswith('_section'):
                continue

            # Ensure field_info has required keys
            if not isinstance(field_info, dict):
                continue

            try:
                if 'widget' in field_info:
                    widget = field_info['widget']
                    # Properly enable widget based on type
                    if isinstance(widget, ttk.Combobox):
                        widget.configure(state='readonly')
                    elif isinstance(widget, tk.Checkbutton):
                        widget.configure(state='normal')
                    else:
                        widget.configure(state='normal')
                # Don't change label color for Experiment field (status label has custom styling)
                if 'label' in field_info and field_name != 'Experiment':
                    field_info['label'].configure(fg='black')
            except:
                pass

        # Note: update_conditional_sections() will be called after this
        # to properly disable sections that shouldn't be active based on Test Type/Content

    def format_experiment_data_for_export(self, data):
        """Format experiment data for export with proper types and null handling"""
        formatted_data = {}

        # Get field configs for type information
        field_configs = self.config_template.get('field_configs', {})

        for field_name, value in data.items():
            # Get field type from config
            field_config = field_configs.get(field_name, {})
            field_type = field_config.get('type', 'str')

            # Check for conditional type based on another field's value
            conditional_options = field_config.get('conditional_options', {})
            if conditional_options:
                # Get the controlling field name (e.g., "Test Mode")
                control_field = conditional_options.get('field')
                if control_field and control_field in data:
                    # Get the current value of the controlling field from data
                    control_value = data.get(control_field)

                    # Get the conditional config for this value
                    conditional_config = conditional_options.get(control_value, {})
                    if isinstance(conditional_config, dict) and 'type' in conditional_config:
                        # Override field_type with the conditional type
                        field_type = conditional_config['type']

            # Handle Experiment field specially - convert bool to "Enabled"/"Disabled"
            if field_name == 'Experiment':
                if isinstance(value, bool):
                    formatted_data[field_name] = "Enabled" if value else "Disabled"
                else:
                    formatted_data[field_name] = value
                continue

            # Handle empty values - convert to null
            if value == '' or value == "" or (isinstance(value, str) and value.strip() == ''):
                formatted_data[field_name] = None
                continue

            # Convert based on type
            try:
                if field_type == 'int':
                    formatted_data[field_name] = int(value) if value not in [None, ''] else None
                elif field_type == 'float':
                    formatted_data[field_name] = float(value) if value not in [None, ''] else None
                elif field_type == 'bool':
                    formatted_data[field_name] = bool(value) if value is not None else False
                else:  # str
                    # Keep as string, but check if it looks like a number that should stay as string
                    # (like "0x3", "169", etc. that are defined as strings in config)
                    formatted_data[field_name] = str(value) if value not in [None, ''] else None
            except (ValueError, TypeError):
                # If conversion fails, keep as is or set to None
                formatted_data[field_name] = None if value == '' else value

        return formatted_data

    def normalize_experiment_data_for_import(self, data):
        """Normalize imported experiment data - convert null to empty strings and Enabled/Disabled to bool"""
        normalized_data = {}

        for field_name, value in data.items():
            # Handle Experiment field - convert "Enabled"/"Disabled" to boolean
            if field_name == 'Experiment':
                if isinstance(value, str):
                    normalized_data[field_name] = (value == "Enabled")
                else:
                    normalized_data[field_name] = bool(value) if value is not None else True
                continue

            # Convert null to empty string for internal storage
            if value is None:
                normalized_data[field_name] = ''
            else:
                normalized_data[field_name] = value

        return normalized_data

    def refresh_status_indicator(self, exp_name):
        """Refresh the status indicator label to match current experiment state"""
        if exp_name not in self.experiments:
            return

        widgets = self.experiments[exp_name].get('widgets', {})
        experiment_var = widgets.get('Experiment', {}).get('var', None)
        status_label = widgets.get('Experiment', {}).get('label', None)

        if experiment_var and status_label:
            enabled = experiment_var.get()
            if enabled:
                status_label.configure(text="✓ ENABLED", bg=self.colors['success'], fg='white')
            else:
                status_label.configure(text="✕ DISABLED", bg=self.colors['danger'], fg='white')

    def is_experiment_disabled(self, exp_name):
        """Check if an experiment is disabled"""
        if exp_name not in self.experiments:
            return False

        widgets = self.experiments[exp_name].get('widgets', {})
        experiment_var = widgets.get('Experiment', {}).get('var', None)

        if experiment_var:
            return not experiment_var.get()  # Returns True if disabled (False value)

        return False  # Default to enabled

    def update_tab_color(self, exp_name, enabled=True):
        """Update tab button color based on experiment status"""
        # Update primary notebook tab button
        if exp_name in self.primary_tab_buttons:
            btn = self.primary_tab_buttons[exp_name]
            # Check if this is the currently selected tab
            current_tab_idx = self.experiment_notebook.index('current')
            current_tab_id = self.experiment_notebook.tabs()[current_tab_idx]
            current_tab_name = self.experiment_notebook.tab(current_tab_id, 'text')
            is_selected = (current_tab_name == exp_name)

            if is_selected:
                if enabled:
                    btn.configure(bg=self.colors['primary'], fg='white', relief='sunken')
                else:
                    btn.configure(bg=self.colors['disabled_tab'], fg=self.colors['disabled_tab_text'], relief='sunken')
            else:
                if enabled:
                    btn.configure(bg=self.colors['light'], fg='black', relief='raised')
                else:
                    btn.configure(bg='#cbd5e1', fg='#64748b', relief='raised')

        # Update secondary notebook tab button if in side-by-side mode
        if exp_name in self.secondary_tab_buttons and self.experiment_notebook_secondary:
            btn = self.secondary_tab_buttons[exp_name]
            # Check if this is the currently selected tab in secondary notebook
            try:
                current_tab_idx = self.experiment_notebook_secondary.index('current')
                current_tab_id = self.experiment_notebook_secondary.tabs()[current_tab_idx]
                current_tab_name = self.experiment_notebook_secondary.tab(current_tab_id, 'text')
                is_selected = (current_tab_name == exp_name)

                if is_selected:
                    if enabled:
                        btn.configure(bg=self.colors['primary'], fg='white', relief='sunken')
                    else:
                        btn.configure(bg=self.colors['disabled_tab'], fg=self.colors['disabled_tab_text'], relief='sunken')
                else:
                    if enabled:
                        btn.configure(bg=self.colors['light'], fg='black', relief='raised')
                    else:
                        btn.configure(bg='#cbd5e1', fg='#64748b', relief='raised')
            except:
                pass

    def apply_unit_data_to_experiment_data(self, exp_name):
        """Apply unit data to a specific experiment's data dictionary"""
        if exp_name not in self.experiments:
            return

        data = self.experiments[exp_name]['data']

        # Get unit data from widgets and apply to data
        for field_name, field_info in self.unit_data_widgets.items():
            try:
                value = field_info['var'].get()
                field_type = field_info['type']

                if field_type == 'int':
                    data[field_name] = int(value) if value else 0
                elif field_type == 'float':
                    data[field_name] = float(value) if value else 0.0
                else:
                    data[field_name] = str(value)
            except:
                pass

    def auto_apply_unit_data(self):
        """Automatically apply unit data to all experiments"""
        # Get unit data from widgets
        unit_data = {}
        for field_name, field_info in self.unit_data_widgets.items():
            unit_data[field_name] = field_info['var'].get()

        # Apply to all experiments' widgets
        for exp_name, exp_info in self.experiments.items():
            widgets = exp_info['widgets']
            for field_name, value in unit_data.items():
                if field_name in widgets:
                    widgets[field_name]['var'].set(value)

            # Also apply to data dictionary
            self.apply_unit_data_to_experiment_data(exp_name)

    def create_new_template(self):
        """Create a new template from scratch, clearing all current experiments"""
        # Confirm action with user
        if self.experiments:
            response = messagebox.askyesno(
                "Create New Template",
                "This will clear all current experiments and create a fresh template.\n\n"
                "Any unsaved changes will be lost. Do you want to continue?",
                icon='warning'
            )
            if not response:
                return

        print("INFO - Creating new template from scratch")

        # Clear all existing experiments
        for exp_name in list(self.experiments.keys()):
            if exp_name in self.experiments:
                # Remove from notebook
                for i, tab_id in enumerate(self.experiment_notebook.tabs()):
                    if self.experiment_notebook.tab(tab_id, 'text') == exp_name:
                        self.experiment_notebook.forget(i)
                        break

                # Remove tab button
                if exp_name in self.primary_tab_buttons:
                    self.primary_tab_buttons[exp_name].destroy()
                    del self.primary_tab_buttons[exp_name]

                # Remove from data structure
                del self.experiments[exp_name]

        # Clear unit data fields
        for field_name, field_info in self.unit_data_widgets.items():
            try:
                field_type = field_info.get('type', 'str')
                if field_type == 'bool':
                    field_info['var'].set(False)
                else:
                    field_info['var'].set('')
            except:
                pass

        # Clear templates
        self.templates.clear()
        self.refresh_template_list()

        # Create a single default experiment
        self.add_new_experiment()

        # Switch to first tab
        if self.experiment_notebook.tabs():
            self.experiment_notebook.select(0)

        messagebox.showinfo(
            "New Template Created",
            "A new blank template has been created with one default experiment."
        )

    def save_configuration(self):
        """Save complete tool configuration (Unit Data + all experiments) to .tpl file"""
        if not self.experiments:
            show_warning("No Data", "No experiments to save.")
            return

        # Save current experiment first
        self.save_current_experiment()

        # Apply Unit Data to all experiments
        for exp_name in self.experiments.keys():
            self.apply_unit_data_to_experiment_data(exp_name)

        file_path = filedialog.asksaveasfilename(
            title="Save Configuration",
            defaultextension=".tpl",
            filetypes=[("Template files", "*.tpl"), ("All files", "*.*")]
        )

        if file_path:
            try:
                # Get Unit Data
                unit_data = {}
                for field_name, field_info in self.unit_data_widgets.items():
                    unit_data[field_name] = field_info['var'].get()

                # Collect all experiment data with proper formatting
                experiments_data = []
                for exp_name, exp_info in self.experiments.items():
                    formatted_data = self.format_experiment_data_for_export(exp_info['data'])
                    experiments_data.append(formatted_data)

                # Create output structure
                output = {
                    "unit_data": unit_data,
                    "experiments": experiments_data,
                    "templates": self.templates,
                    "saved_date": datetime.now().isoformat(),
                    "tool": "PPV Experiment Builder v2.5",
                    "file_type": "configuration"
                }

                with open(file_path, 'w') as f:
                    json.dump(output, f, indent=4)

                messagebox.showinfo("Success",
                                   "Configuration saved successfully!\n\n"
                                   f"Unit Data: {len(unit_data)} fields\n"
                                   f"Experiments: {len(experiments_data)}\n"
                                   f"Templates: {len(self.templates)}\n\n"
                                   f"File: {file_path}")
            except Exception as e:
                show_error("Save Error", f"Failed to save configuration:\n{str(e)}")

    def load_configuration(self):
        """Load complete tool configuration (Unit Data + all experiments) from .tpl file"""
        file_path = filedialog.askopenfilename(
            title="Load Configuration",
            filetypes=[("Template files", "*.tpl"), ("All files", "*.*")]
        )

        if file_path:
            try:
                with open(file_path, 'r') as f:
                    config = json.load(f)

                # Validate file type
                if config.get('file_type') != 'configuration':
                    show_warning("Invalid File",
                                         "This file does not appear to be a valid configuration file.")
                    return

                # Clear existing experiments
                for exp_name in list(self.experiments.keys()):
                    self.experiments.pop(exp_name)

                # Clear notebook tabs
                for tab_id in self.experiment_notebook.tabs():
                    self.experiment_notebook.forget(tab_id)

                # Clear custom tab buttons
                for btn in self.primary_tab_buttons.values():
                    btn.destroy()
                self.primary_tab_buttons.clear()

                # Clear secondary notebook if in side-by-side view
                if self.view_mode == "side-by-side" and self.experiment_notebook_secondary:
                    for tab_id in self.experiment_notebook_secondary.tabs():
                        self.experiment_notebook_secondary.forget(tab_id)
                    for btn in self.secondary_tab_buttons.values():
                        btn.destroy()
                    self.secondary_tab_buttons.clear()

                # Load Unit Data
                if 'unit_data' in config:
                    for field_name, value in config['unit_data'].items():
                        if field_name in self.unit_data_widgets:
                            self.unit_data_widgets[field_name]['var'].set(value)

                # Load Templates
                if 'templates' in config:
                    self.templates = config['templates']
                    self.refresh_template_list()

                # Load Experiments
                experiments = config.get('experiments', [])
                for exp_data in experiments:
                    # Normalize the data (handle null, Enabled/Disabled, etc.)
                    normalized_data = self.normalize_experiment_data_for_import(exp_data.copy())

                    exp_name = normalized_data.get('Test Name', 'Unnamed')

                    # Create experiment
                    self.experiments[exp_name] = {
                        'data': normalized_data,
                        'widgets': {},
                        'modified': False
                    }

                    self.create_experiment_tab(exp_name)
                    self.load_experiment_data(exp_name)

                # Select first experiment
                if self.experiments:
                    first_exp = list(self.experiments.keys())[0]
                    self.current_experiment = first_exp
                    self.experiment_notebook.select(0)
                    self.select_custom_tab(first_exp, 'primary')

                messagebox.showinfo("Success",
                                   "Configuration loaded successfully!\n\n"
                                   f"Unit Data: {len(config.get('unit_data', {}))} fields\n"
                                   f"Experiments: {len(experiments)}\n"
                                   f"Templates: {len(config.get('templates', {}))}")
            except Exception as e:
                show_error("Load Error", f"Failed to load configuration:\n{str(e)}")

    def save_as_template(self):
        """Save current experiment as a template"""
        if not self.current_experiment:
            show_warning("No Experiment", "Please select an experiment first.")
            return

        self.save_current_experiment()

        # Ask for template name
        template_name = tk.simpledialog.askstring("Save Template",
                                                  "Enter template name:",
                                                  initialvalue=self.current_experiment)
        if template_name:
            self.templates[template_name] = self.experiments[self.current_experiment]['data'].copy()
            self.refresh_template_list()
            messagebox.showinfo("Saved", f"Template '{template_name}' saved successfully.")

    def load_template(self):
        """Load selected template into current experiment"""
        selection = self.template_listbox.curselection()
        if not selection:
            show_warning("No Selection", "Please select a template to load.")
            return

        if not self.current_experiment:
            show_warning("No Experiment", "Please select an experiment first.")
            return

        template_name = self.template_listbox.get(selection[0])
        template_data = self.templates[template_name].copy()

        # Keep current experiment name
        template_data["Test Name"] = self.current_experiment

        # Update experiment data
        self.experiments[self.current_experiment]['data'] = template_data

        # Reload widgets
        self.load_experiment_data(self.current_experiment)

        messagebox.showinfo("Loaded", f"Template '{template_name}' loaded successfully.")

    def delete_template(self):
        """Delete selected template"""
        selection = self.template_listbox.curselection()
        if not selection:
            show_warning("No Selection", "Please select a template to delete.")
            return

        template_name = self.template_listbox.get(selection[0])

        if messagebox.askyesno("Confirm Delete", f"Delete template '{template_name}'?"):
            del self.templates[template_name]
            self.refresh_template_list()

    def refresh_template_list(self):
        """Refresh the template listbox"""
        self.template_listbox.delete(0, tk.END)
        for template_name in sorted(self.templates.keys()):
            self.template_listbox.insert(tk.END, template_name)

    def create_default_experiment_data(self):
        """Create default experiment data structure"""
        data = {}

        # Support both new field_configs and old data_types format
        if 'field_configs' in self.config_template:
            # New format - use default values from field_configs
            for field, field_config in self.config_template['field_configs'].items():
                default_value = field_config.get('default', '')
                # Use configured default, or apply special defaults
                if field == "Experiment":
                    data[field] = "Enabled"
                else:
                    data[field] = default_value
        else:
            # Old format - maintain backward compatibility
            data_types = self.config_template.get('data_types', {})

            for field, field_types in data_types.items():
                field_type = field_types[0]

                if field_type == "str":
                    if field == "Experiment":
                        data[field] = "Enabled"
                    elif field == "Test Mode":
                        data[field] = "Mesh"
                    elif field == "Test Type":
                        data[field] = "Loops"
                    elif field == "Content":
                        data[field] = "Dragon"
                    elif field == "Voltage Type":
                        data[field] = "vbump"
                    else:
                        data[field] = ""
                elif field_type == "int":
                    # Fields that should default to None/empty
                    if field in ["Frequency IA", "Frequency CFC", "Check Core"]:
                        data[field] = ""
                    else:
                        data[field] = 0
                elif field_type == "float":
                    # Fields that should default to None/empty
                    if field in ["Voltage IA", "Voltage CFC"]:
                        data[field] = ""
                    else:
                        data[field] = 0.0
                elif field_type == "bool":
                    data[field] = False
                elif field_type == "dict":
                    data[field] = {}

        return data

    def load_experiment_data(self, exp_name):
        """Load experiment data into the form"""
        if exp_name not in self.experiments:
            return

        # Temporarily disable change tracking while loading
        old_track_state = self._track_changes
        self._track_changes = False

        data = self.experiments[exp_name]['data']
        widgets = self.experiments[exp_name]['widgets']

        for field_name, field_info in widgets.items():
            # Skip section markers (any key ending with _section)
            if field_name.endswith('_section'):
                continue

            # Ensure field_info has required keys
            if not isinstance(field_info, dict) or 'var' not in field_info:
                continue

            value = data.get(field_name, "")
            var = field_info['var']

            try:
                if field_info.get('type') == 'bool':
                    var.set(bool(value))
                else:
                    var.set(str(value))
            except:
                pass

        # Update tab color based on experiment status
        experiment_enabled = widgets.get('Experiment', {}).get('var', tk.BooleanVar(value=True)).get()
        self.update_tab_color(exp_name, enabled=experiment_enabled)

        # Re-enable change tracking
        self._track_changes = old_track_state

    def import_templates(self):
        """Import templates from JSON folder or Excel file"""
        # Ask user what to import
        choice = messagebox.askquestion("Import Templates",
                                       "Import from folder of JSON files?\n\nYes = JSON folder\nNo = Excel file",
                                       icon='question')

        if choice == 'yes':
            # Import from JSON folder
            folder_path = filedialog.askdirectory(title="Select Folder with JSON Template Files")
            if folder_path:
                self.import_templates_from_folder(folder_path)
        else:
            # Import from Excel file
            file_path = filedialog.askopenfilename(
                title="Select Excel File with Templates",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            if file_path:
                self.import_templates_from_excel(file_path)

    def import_templates_from_folder(self, folder_path):
        """Import templates from JSON files in a folder"""
        import glob

        json_files = glob.glob(os.path.join(folder_path, "*.json"))

        if not json_files:
            show_warning("No Files", "No JSON files found in the selected folder.")
            return

        imported_count = 0
        for json_file in json_files:
            try:
                with open(json_file, 'r') as f:
                    template_data = json.load(f)

                # Get template name from filename (without extension)
                template_name = os.path.splitext(os.path.basename(json_file))[0]

                # Check if it's a valid template (has required fields)
                if isinstance(template_data, dict) and 'Test Name' in template_data:
                    # Store template
                    self.templates[template_name] = template_data
                    imported_count += 1
            except Exception as e:
                print(f"Error importing {json_file}: {str(e)}")

        # Refresh template listbox
        self.refresh_template_list()

        messagebox.showinfo("Import Complete",
                           f"Imported {imported_count} template(s) from folder.")

    def import_templates_from_excel(self, file_path):
        """Import templates from Excel file (each sheet = one template)"""
        try:
            wb = openpyxl.load_workbook(file_path)
            imported_count = 0

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                template_data = {}

                # Get field list from config (support both old and new format)
                valid_fields = self.config_template.get('field_configs', {}).keys() if 'field_configs' in self.config_template else self.config_template.get('data_types', {}).keys()

                # Read field-value pairs from columns A and B
                for row in range(1, sheet.max_row + 1):
                    field = sheet.cell(row, 1).value
                    value = sheet.cell(row, 2).value

                    if field and field in valid_fields:
                        template_data[field] = value if value is not None else ""

                if template_data and 'Test Name' in template_data:
                    # Use sheet name as template name
                    self.templates[sheet_name] = template_data
                    imported_count += 1

            # Refresh template listbox
            self.refresh_template_list()

            messagebox.showinfo("Import Complete",
                               f"Imported {imported_count} template(s) from Excel file.")
        except Exception as e:
            show_error("Import Error", f"Failed to import templates:\n{str(e)}")

    def save_current_experiment(self):
        """Save current experiment data from widgets"""
        if not self.current_experiment or self.current_experiment not in self.experiments:
            return

        widgets = self.experiments[self.current_experiment]['widgets']
        data = self.experiments[self.current_experiment]['data']

        for field_name, field_info in widgets.items():
            # Skip section markers (any key ending with _section)
            if field_name.endswith('_section'):
                continue

            # Ensure field_info has required keys
            if not isinstance(field_info, dict) or 'var' not in field_info:
                continue

            try:
                value = field_info['var'].get()
                field_type = field_info['type']

                # Check for conditional type based on another field's value
                field_config = self.get_field_config(field_name)
                conditional_options = field_config.get('conditional_options', {})

                if conditional_options:
                    # Get the controlling field name (e.g., "Test Mode")
                    control_field = conditional_options.get('field')
                    if control_field and control_field in widgets:
                        # Get the current value of the controlling field
                        control_value = widgets[control_field]['var'].get()

                        # Get the conditional config for this value
                        conditional_config = conditional_options.get(control_value, {})
                        if isinstance(conditional_config, dict) and 'type' in conditional_config:
                            # Override field_type with the conditional type
                            field_type = conditional_config['type']

                # Store raw values - formatting will happen on export
                if field_type == 'int':
                    try:
                        data[field_name] = int(value) if value != '' else ''
                    except (ValueError, TypeError):
                        data[field_name] = ''
                elif field_type == 'float':
                    try:
                        data[field_name] = float(value) if value != '' else ''
                    except (ValueError, TypeError):
                        data[field_name] = ''
                elif field_type == 'bool':
                    data[field_name] = bool(value)
                else:
                    data[field_name] = str(value) if value else ''
            except Exception as e:
                print(f"Error saving field {field_name}: {e}")
                pass

        # Apply Unit Data overrides
        self.apply_unit_data_to_experiment_data(self.current_experiment)

    def restore_current_experiment(self):
        """Restore current experiment to last saved state"""
        if not self.current_experiment or self.current_experiment not in self.experiments:
            show_warning("No Experiment", "Please select an experiment first.")
            return

        # Check if there are unsaved changes
        if not self.experiments[self.current_experiment].get('modified', False):
            messagebox.showinfo("No Changes", "This experiment has no unsaved changes.")
            return

        # Confirm restore
        result = messagebox.askyesno(
            "Restore Experiment",
            f"Are you sure you want to discard all unsaved changes for '{self.current_experiment}'?\n\n"
            "This will reload the last saved state.",
            icon='warning'
        )

        if result:
            # Disable change tracking temporarily
            self._track_changes = False

            # Reload the experiment data
            self.load_experiment_data(self.current_experiment)

            # Clear modified flag
            self.experiments[self.current_experiment]['modified'] = False
            self.update_tab_modified_indicator(self.current_experiment)

            # Re-enable change tracking
            self._track_changes = True

            messagebox.showinfo("Restored", f"Experiment '{self.current_experiment}' has been restored to its last saved state.")

    def save_and_update_tab(self):
        """Save experiment and update tab name based on Test Name"""
        if not self.current_experiment or self.current_experiment not in self.experiments:
            show_warning("No Experiment", "Please select an experiment first.")
            return

        # Save current data
        self.save_current_experiment()

        # Clear modified flag
        self.experiments[self.current_experiment]['modified'] = False
        self.update_tab_modified_indicator(self.current_experiment)

        # Get new test name
        widgets = self.experiments[self.current_experiment]['widgets']
        if 'Test Name' in widgets:
            new_name = widgets['Test Name']['var'].get().strip()

            if not new_name:
                show_warning("Invalid Name", "Test Name cannot be empty.")
                return

            # Check if name already exists (and it's not the current experiment)
            if new_name != self.current_experiment and new_name in self.experiments:
                show_error("Duplicate Name", f"An experiment with name '{new_name}' already exists.")
                return

            # Update experiment name
            if new_name != self.current_experiment:
                # Save old name
                old_name = self.current_experiment

                # Rename in experiments dict
                self.experiments[new_name] = self.experiments.pop(old_name)

                # Update tab text
                current_tab = self.experiment_notebook.index('current')
                tab_id = self.experiment_notebook.tabs()[current_tab]
                self.experiment_notebook.tab(tab_id, text=new_name)

                # Update custom tab button text in primary
                if old_name in self.primary_tab_buttons:
                    btn = self.primary_tab_buttons[old_name]
                    btn.configure(text=new_name, command=lambda n=new_name: self.select_custom_tab(n, 'primary'))
                    self.primary_tab_buttons[new_name] = self.primary_tab_buttons.pop(old_name)

                # Update current experiment reference
                self.current_experiment = new_name

                # Update secondary notebook if in side-by-side view
                if self.view_mode == "side-by-side" and self.experiment_notebook_secondary:
                    # Update secondary field checkboxes
                    if old_name in self.secondary_field_checkboxes:
                        self.secondary_field_checkboxes[new_name] = self.secondary_field_checkboxes.pop(old_name)
                    # Update secondary frames
                    if old_name in self.secondary_notebook_frames:
                        self.secondary_notebook_frames[new_name] = self.secondary_notebook_frames.pop(old_name)
                    # Update tab in secondary notebook
                    for i, tab_id in enumerate(self.experiment_notebook_secondary.tabs()):
                        if self.experiment_notebook_secondary.tab(tab_id, 'text') == old_name:
                            self.experiment_notebook_secondary.tab(tab_id, text=new_name)
                            break

                    # Update custom tab button text in secondary
                    if hasattr(self, 'secondary_tab_buttons') and old_name in self.secondary_tab_buttons:
                        btn = self.secondary_tab_buttons[old_name]
                        btn.configure(text=new_name, command=lambda n=new_name: self.select_custom_tab(n, 'secondary'))
                        self.secondary_tab_buttons[new_name] = self.secondary_tab_buttons.pop(old_name)

                messagebox.showinfo("Saved", f"Experiment saved and renamed to '{new_name}'.")
            else:
                messagebox.showinfo("Saved", f"Experiment '{new_name}' saved successfully.")
        else:
            messagebox.showinfo("Saved", "Experiment saved successfully.")

    def browse_path(self, field_name, var):
        """Browse for file or folder path"""
        if "Folder" in field_name or "Path" in field_name:
            path = filedialog.askdirectory()
        else:
            path = filedialog.askopenfilename()

        if path:
            var.set(path)

    def generate_unique_experiment_name(self, base_name):
        """Generate a unique experiment name by appending a counter if needed"""
        if base_name not in self.experiments:
            return base_name

        # Name already exists, append counter
        counter = 1
        while f"{base_name}_{counter}" in self.experiments:
            counter += 1

        return f"{base_name}_{counter}"

    def import_from_excel(self):
        """Import experiments from Excel file"""
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )

        if file_path:
            try:
                self.process_excel_file(file_path)
                messagebox.showinfo("Success", "Excel file imported successfully!")
            except Exception as e:
                show_error("Import Error", f"Failed to import Excel file:\n{str(e)}")

    def process_excel_file(self, file_path):
        """Process Excel file and create experiments"""
        wb = openpyxl.load_workbook(file_path)

        # Get field list from config (support both old and new format)
        valid_fields = self.config_template.get('field_configs', {}).keys() if 'field_configs' in self.config_template else self.config_template.get('data_types', {}).keys()

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            exp_data = {}

            # Read field-value pairs from columns A and B
            for row in range(1, sheet.max_row + 1):
                field = sheet.cell(row, 1).value
                value = sheet.cell(row, 2).value

                if field and field in valid_fields:
                    exp_data[field] = value if value is not None else ""

            if exp_data:
                # Normalize the data (handle null, Enabled/Disabled, etc.)
                normalized_data = self.normalize_experiment_data_for_import(exp_data)

                # Create experiment with unique name
                base_name = normalized_data.get("Test Name", sheet_name)
                exp_name = self.generate_unique_experiment_name(base_name)

                # Update the Test Name in the data to match the unique name
                normalized_data["Test Name"] = exp_name

                self.experiments[exp_name] = {
                    'data': normalized_data,
                    'widgets': {},
                    'tab_frame': None,
                    'modified': False
                }

                self.create_experiment_tab(exp_name)
                self.load_experiment_data(exp_name)

    def import_from_json(self):
        """Import experiments from JSON file"""
        file_path = filedialog.askopenfilename(
            title="Select JSON File",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )

        if file_path:
            try:
                with open(file_path, 'r') as f:
                    data = json.load(f)

                # Handle multiple formats:
                # 1. New format: {TestName1: {data}, TestName2: {data}, ...}
                # 2. Old format with array: {"experiments": [...]}
                # 3. Single experiment: {field: value, ...}
                # 4. Array of experiments: [{...}, {...}]
                experiments = []

                if isinstance(data, dict):
                    if "experiments" in data:
                        # Old format with "experiments" key
                        exp_list = data["experiments"]
                        if isinstance(exp_list, list):
                            experiments = exp_list
                        elif isinstance(exp_list, dict):
                            # It's a dictionary of experiments
                            experiments = list(exp_list.values())
                    elif "Test Name" in data:
                        # Single experiment
                        experiments = [data]
                    else:
                        # New format: dictionary with test names as keys
                        # Check if values are experiment dictionaries
                        first_value = next(iter(data.values()), None)
                        if isinstance(first_value, dict) and "Test Name" in first_value:
                            experiments = list(data.values())
                        else:
                            # Treat as single experiment
                            experiments = [data]
                elif isinstance(data, list):
                    # Array of experiments
                    experiments = data

                for exp_data in experiments:
                    # Normalize the data (handle null, Enabled/Disabled, etc.)
                    normalized_data = self.normalize_experiment_data_for_import(exp_data)

                    # Create experiment with unique name
                    base_name = normalized_data.get("Test Name", "Imported_Experiment")
                    exp_name = self.generate_unique_experiment_name(base_name)

                    # Update the Test Name in the data to match the unique name
                    normalized_data["Test Name"] = exp_name

                    self.experiments[exp_name] = {
                        'data': normalized_data,
                        'widgets': {},
                        'tab_frame': None,
                        'modified': False
                    }

                    self.create_experiment_tab(exp_name)
                    self.load_experiment_data(exp_name)

                messagebox.showinfo("Success", f"Imported {len(experiments)} experiment(s)!")
            except Exception as e:
                show_error("Import Error", f"Failed to import JSON:\n{str(e)}")

    def export_to_json(self):
        """Export all experiments to JSON file"""
        if not self.experiments:
            show_warning("No Data", "No experiments to export.")
            return

        # Check for unsaved changes before exporting
        unsaved_experiments = [name for name, exp in self.experiments.items() if exp.get('modified', False)]

        if unsaved_experiments:
            unsaved_list = "\n".join([f"• {name}" for name in unsaved_experiments])
            message = f"The following experiments have unsaved changes:\n\n{unsaved_list}\n\nDo you want to save all changes before exporting?"

            response = messagebox.askyesnocancel(
                "Unsaved Changes",
                message,
                icon='warning'
            )

            if response is None:  # Cancel
                return
            elif response:  # Yes - save changes
                # Save current experiment first
                self.save_current_experiment()
                # Then save all other modified experiments
                current_exp = self.current_experiment
                for exp_name in unsaved_experiments:
                    if exp_name != current_exp:
                        self.current_experiment = exp_name
                        self.save_current_experiment()
                # Restore current experiment
                self.current_experiment = current_exp
            # If response is False (No), continue without saving
        else:
            # Save current experiment if no unsaved changes detected
            self.save_current_experiment()

        # Apply Unit Data to all experiments before export
        for exp_name in self.experiments.keys():
            self.apply_unit_data_to_experiment_data(exp_name)

        file_path = filedialog.asksaveasfilename(
            title="Save JSON File",
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )

        if file_path:
            try:
                # Collect all experiment data with proper formatting
                # Format: {TestName1: {experiment_data}, TestName2: {experiment_data}, ...}
                experiments_data = {}
                for exp_name, exp_info in self.experiments.items():
                    formatted_data = self.format_experiment_data_for_export(exp_info['data'])
                    # Use Test Name as the key
                    test_name = formatted_data.get('Test Name', exp_name)
                    experiments_data[test_name] = formatted_data

                with open(file_path, 'w') as f:
                    json.dump(experiments_data, f, indent=4)

                messagebox.showinfo("Success", f"Exported {len(experiments_data)} experiments to:\n{file_path}")
            except Exception as e:
                show_error("Export Error", f"Failed to export:\n{str(e)}")

    def validate_all_experiments(self):
        """Validate all experiments"""
        self.save_current_experiment()

        errors = []
        warnings = []

        for exp_name, exp_info in self.experiments.items():
            exp_errors, exp_warnings = self.validate_experiment(exp_name, exp_info['data'])
            if exp_errors:
                errors.extend([f"{exp_name}: {e}" for e in exp_errors])
            if exp_warnings:
                warnings.extend([f"{exp_name}: {w}" for w in exp_warnings])

        message = "Validation Results:\n\n"
        message += f"Experiments: {len(self.experiments)}\n"
        message += f"Errors: {len(errors)}\n"
        message += f"Warnings: {len(warnings)}\n\n"

        if errors:
            message += "ERRORS:\n" + "\n".join(errors[:10])
            if len(errors) > 10:
                message += f"\n... and {len(errors) - 10} more"

        if warnings:
            message += "\n\nWARNINGS:\n" + "\n".join(warnings[:10])
            if len(warnings) > 10:
                message += f"\n... and {len(warnings) - 10} more"

        if not errors and not warnings:
            message += "✓ All experiments validated successfully!"
            messagebox.showinfo("Validation Complete", message)
        else:
            show_warning("Validation Issues", message)

    def validate_experiment(self, exp_name, exp_data):
        """Validate a single experiment"""
        errors = []
        warnings = []

        # Required fields
        if not exp_data.get("Test Name"):
            errors.append("Test Name is required")

        # IP Address format
        ip = exp_data.get("IP Address", "")
        if ip and not self.validate_ip(ip):
            errors.append(f"Invalid IP Address format: {ip}")

        # Test Type specific validation
        test_type = exp_data.get("Test Type", "")
        if test_type in ["Sweep", "Shmoo"]:
            if not exp_data.get("Start"):
                warnings.append("Start value should be set for Sweep/Shmoo")
            if not exp_data.get("End"):
                warnings.append("End value should be set for Sweep/Shmoo")

        return errors, warnings

    def validate_ip(self, ip):
        """Validate IP address format"""
        parts = ip.split('.')
        if len(parts) != 4:
            return False
        try:
            return all(0 <= int(part) <= 255 for part in parts)
        except:
            return False

    def launch_developer_environment(self):
        """Launch the standalone Developer Environment tool"""
        try:
            import subprocess
            # Developer Environment is now in utils folder, not gui folder
            dev_env_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'utils', 'DeveloperEnvironment.py')

            if os.path.exists(dev_env_path):
                # Launch in separate process
                subprocess.Popen([sys.executable, dev_env_path])
                messagebox.showinfo("Developer Environment",
                                   "Developer Environment launched successfully!\n\n"
                                   "The tool will open in a separate window.")
            else:
                show_error("Not Found",
                                   f"Developer Environment not found at:\n{dev_env_path}")
        except Exception as e:
            show_error("Launch Error", f"Failed to launch Developer Environment:\n{str(e)}")

    def show_about(self):
        """Show about dialog"""
        about_text = (
            "PPV Experiment Builder v2.5\n\n"
            "Excel-like interface for Control Panel Config Generator\n\n"
            "Features:\n"
            "• Multiple experiment management with tabs\n"
            "• Shared Unit Data configuration\n"
            "• Side-by-side experiment comparison\n"
            "• Template library\n"
            "• JSON/Excel import/export\n"
            "• Field copy functionality\n"
            "• Configuration save/load (.tpl files)\n\n"
            "Part of THR Debug Tools Hub\n"
            "© 2025"
        )
        messagebox.showinfo("About", about_text)

    def on_closing(self):
        """Handle window close"""
        self.save_current_experiment()
        self.root.destroy()

    def run(self):
        """Run the application"""
        self.root.mainloop()


def main():
    """Main entry point"""
    app = ExperimentBuilderGUI()
    app.run()


if __name__ == "__main__":
    main()

