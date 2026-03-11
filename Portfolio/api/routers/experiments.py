"""
Experiment Builder REST endpoints.
  GET  /api/experiments/config/{product}       — ControlPanelConfig.json
  GET  /api/experiments/products               — available products
  POST /api/experiments/build                  — build/export experiment list as .tpl
  POST /api/experiments/save_to_folder         — save .tpl file to a server folder path
  POST /api/experiments/templates/import-excel — import templates from .xlsx
"""
from __future__ import annotations
import io
import json
import os
import sys
import traceback
from typing import List

from fastapi import APIRouter, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

router = APIRouter()

_CONFIGS_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "THRTools", "configs"
)


@router.get("/products")
async def get_products():
    """Return list of products that have a ControlPanelConfig.json."""
    products = []
    for fname in sorted(os.listdir(_CONFIGS_DIR)):
        if fname.endswith("ControlPanelConfig.json"):
            products.append(fname.replace("ControlPanelConfig.json", ""))
    return {"products": products}


@router.get("/config/{product}")
async def get_config(product: str):
    """Return the ControlPanelConfig.json for a product (drives the dynamic form).

    Returns the full config with field_configs for the frontend to render sections,
    field types, defaults, options, required flags, and descriptions.
    """
    fname = os.path.join(_CONFIGS_DIR, f"{product.upper()}ControlPanelConfig.json")
    if not os.path.isfile(fname):
        # Fall back to generic
        fname = os.path.join(_CONFIGS_DIR, "ControlPanelConfig.json")
    if not os.path.isfile(fname):
        raise HTTPException(status_code=404, detail=f"Config not found for product '{product}'")
    with open(fname, encoding="utf-8") as f:
        raw = json.load(f)

    # Support both formats: {"field_configs": {...}} and flat {"key": value}
    field_configs = raw.get("field_configs", raw)
    field_enable_config = raw.get("field_enable_config", {})
    return {
        "product": product.upper(),
        "config": raw,
        "field_configs": field_configs,
        "field_enable_config": field_enable_config,
    }


class ExperimentBuildRequest(BaseModel):
    experiments: List[dict] = Field(..., description="List of experiment parameter dicts")
    filename: str = "experiments.tpl"


@router.post("/build")
async def build_experiments(req: ExperimentBuildRequest):
    """Export experiment list as a .tpl (JSON) file."""
    content = json.dumps(req.experiments, indent=2).encode("utf-8")
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{req.filename}"'},
    )


class SaveTplToFolderRequest(BaseModel):
    folder_path: str
    filename: str = "experiments"
    experiments: List[dict] = Field(default_factory=list)
    unit_data: dict = Field(default_factory=dict)
    templates: dict = Field(default_factory=dict)


@router.post("/save_to_folder")
async def save_tpl_to_folder(req: SaveTplToFolderRequest):
    """Save .tpl and .json files directly to a server-side folder path."""
    folder = req.folder_path
    if not os.path.isdir(folder):
        try:
            os.makedirs(folder, exist_ok=True)
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Cannot create folder: {exc}")

    from datetime import datetime
    tpl_output = {
        "unit_data":   req.unit_data,
        "experiments": req.experiments,
        "templates":   req.templates,
        "saved_date":  datetime.now().isoformat(),
        "tool":        "THR Experiment Builder",
        "file_type":   "configuration",
    }
    # Production JSON: { "Test Name": flatFields, ... }
    json_output: dict = {
        str(exp.get("Test Name") or f"Experiment_{i + 1}"): exp
        for i, exp in enumerate(req.experiments)
    }
    tpl_fname = f"{req.filename}.tpl"
    json_fname = f"{req.filename}.json"
    try:
        with open(os.path.join(folder, tpl_fname), "w", encoding="utf-8") as fh:
            json.dump(tpl_output, fh, indent=2)
        with open(os.path.join(folder, json_fname), "w", encoding="utf-8") as fh:
            json.dump(json_output, fh, indent=4)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))
    return {"success": True, "saved_files": [tpl_fname, json_fname], "folder": folder}


@router.post("/templates/import-excel")
async def import_templates_from_excel(file: UploadFile = File(...)):
    """Import templates from an Excel file (.xlsx).

    Each worksheet becomes one template (sheet name → template name).
    Columns A and B of each sheet are read as field→value pairs.
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl is not installed")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Cannot read Excel file: {exc}")

    templates: dict = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        template_data: dict = {}
        for row in range(1, sheet.max_row + 1):
            field = sheet.cell(row, 1).value
            value = sheet.cell(row, 2).value
            if field and isinstance(field, str):
                template_data[field] = str(value) if value is not None else ""
        if template_data:
            templates[sheet_name] = template_data

    return {"templates": templates}
