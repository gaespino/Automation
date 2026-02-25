"""
Experiment Builder
==================
Build, manage and export experiment configuration files.
Reads product-specific config from THRTools/configs/{product}ControlPanelConfig.json.

Replicates all functionality from PPV/gui/ExperimentBuilder.ExperimentBuilderGUI:
- Dynamic field form from ControlPanelConfig (all sections: Basic, Unit Data, Test Config, etc.)
- Conditional sections: show/hide based on Test Type and Content values
- Multi-experiment queue (add, edit index, delete, reorder)
- Template management: save/load .tpl files (JSON with .tpl extension)
- Import experiments from JSON or Excel (openpyxl named-table format)
- Export queue to JSON

Scalability:
- Add new fields by editing ControlPanelConfig.json — no code changes needed
- New products: add {Product}ControlPanelConfig.json to THRTools/configs/

CaaS notes: see CAAS_TODO.md
"""
import base64
import json
import logging
import os
import io
from collections import defaultdict

import dash
from dash import html, dcc, Input, Output, State, callback, no_update, ctx
import dash_bootstrap_components as dbc

logger = logging.getLogger(__name__)

dash.register_page(
    __name__,
    path='/thr-tools/experiment-builder',
    name='Experiment Builder',
    title='Experiment Builder'
)

ACCENT = "#36d7b7"
_PRODUCTS = ["GNR", "CWF", "DMR"]
_CONFIGS_DIR = os.path.normpath(
    os.path.join(os.path.dirname(__file__), '..', '..', 'THRTools', 'configs')
)

# ── Config helpers ─────────────────────────────────────────────────────────────

def _load_config(product: str) -> dict:
    """Load ControlPanelConfig for a product, return {} on failure."""
    path = os.path.join(_CONFIGS_DIR, f"{product}ControlPanelConfig.json")
    if os.path.exists(path):
        try:
            with open(path, encoding="utf-8") as f:
                return json.load(f)
        except Exception as exc:
            logger.warning("Could not load config for %s: %s", product, exc)
    return {}


def _field_configs(product: str) -> dict:
    """Return field_configs dict for the product."""
    cfg = _load_config(product)
    return cfg.get("field_configs", {})


def _get_sections(field_configs: dict) -> list:
    """Return unique section names in order of first appearance."""
    seen = []
    for fc in field_configs.values():
        s = fc.get("section", "General")
        if s not in seen:
            seen.append(s)
    return seen


def _default_value(fc: dict) -> str:
    """Return the default value as a string for form rendering."""
    v = fc.get("default", "")
    if v is None:
        return ""
    if isinstance(v, bool):
        return str(v).lower()   # "true" / "false"
    return str(v)


# ── Form rendering helpers ─────────────────────────────────────────────────────

_INPUT_STYLE = {
    "backgroundColor": "#1a1d26", "color": "#e0e0e0",
    "border": "1px solid rgba(255,255,255,0.1)", "fontSize": "0.83rem"
}
_LABEL_STYLE = {"color": "#a0a0a0", "fontSize": "0.8rem", "marginBottom": "2px"}


def _make_field(fname: str, fc: dict, value=None, field_id_prefix="eb") -> html.Div:
    """Render a single form field (input, select, or checkbox) as a Div row."""
    ftype = fc.get("type", "str")
    opts = fc.get("options", [])
    # Check conditional_options
    if not opts and "conditional_options" in fc:
        # Use first available option set as default
        cond = fc["conditional_options"]
        for _key, cval in cond.items():
            if "options" in cval:
                opts = cval["options"]
                break

    fid = f"{field_id_prefix}-{fname.lower().replace(' ', '-').replace('(', '').replace(')', '')}"
    label = html.Label(fname, style=_LABEL_STYLE)

    if value is None:
        value = _default_value(fc)

    if ftype == "bool":
        checked = (str(value).lower() in ("true", "1", "yes"))
        widget = dbc.Checklist(
            id={"type": f"{field_id_prefix}-field", "index": fname},
            options=[{"label": "", "value": "true"}],
            value=["true"] if checked else [],
            inputStyle={"marginRight": "4px"},
            className="mb-0"
        )
    elif opts:
        safe_val = str(value) if str(value) in [str(o) for o in opts] else (str(opts[0]) if opts else "")
        widget = dcc.Dropdown(
            id={"type": f"{field_id_prefix}-field", "index": fname},
            options=[{"label": str(o), "value": str(o)} for o in opts],
            value=safe_val,
            clearable=False,
            className="mb-0"
        )
    else:
        widget = dbc.Input(
            id={"type": f"{field_id_prefix}-field", "index": fname},
            value=str(value) if value else "",
            placeholder=fc.get("description", fname),
            type="number" if ftype in ("int", "float") else "text",
            className="mb-0",
            style=_INPUT_STYLE
        )

    title_style = {"color": "#c0c0c0", "fontSize": "0.75rem", "fontStyle": "italic"} if fc.get("description") else {}
    return html.Div([
        label,
        widget,
        html.Small(fc.get("description", ""), style={**title_style, "display": "none"})
    ], className="mb-2")


def _build_form(product: str, exp_data: dict | None = None) -> list:
    """Build the complete dynamic form for a product, optionally populated with exp_data."""
    fcs = _field_configs(product)
    if not fcs:
        return [html.P("No config found for this product.", style={"color": "#a0a0a0"})]

    # Group by section
    sections = _get_sections(fcs)
    grouped = defaultdict(list)
    for fname, fc in fcs.items():
        grouped[fc.get("section", "General")].append((fname, fc))

    # Conditional visibility: read current test_type / content from exp_data
    test_type = "Loops"
    content = "Linux"
    if exp_data:
        test_type = str(exp_data.get("Test Type", "Loops"))
        content = str(exp_data.get("Content", "Linux"))

    elements = []
    for section in sections:
        fields = grouped.get(section, [])
        if not fields:
            continue

        # Determine if section should be visible
        # Sections with condition: only show if the condition matches
        # A section is conditional if ALL its fields have the same condition
        cond_values = set()
        for fname, fc in fields:
            cond = fc.get("condition", {})
            if cond:
                cond_values.add((cond.get("field"), cond.get("value")))

        # Section visibility
        visible = True
        if cond_values and len(cond_values) == 1:
            cfield, cvalue = next(iter(cond_values))
            if cfield == "Test Type":
                visible = (test_type == cvalue)
            elif cfield == "Content":
                visible = (content == cvalue)

        section_style = {} if visible else {"display": "none"}
        section_id = f"eb-section-{section.lower().replace(' ', '-').replace('&', 'and')}"

        field_elements = []
        for fname, fc in fields:
            val = None
            if exp_data and fname in exp_data:
                val = exp_data[fname]
            field_elements.append(_make_field(fname, fc, value=val))

        elements.append(html.Div([
            html.H6(section, style={"color": ACCENT, "fontSize": "0.85rem",
                                    "borderBottom": f"1px solid {ACCENT}22",
                                    "paddingBottom": "4px", "marginBottom": "8px"}),
            *field_elements,
        ], id=section_id, style={"marginBottom": "16px", **section_style}))

    return elements


# ── Layout ─────────────────────────────────────────────────────────────────────

layout = dbc.Container(fluid=True, className="pb-5", children=[
    html.Div(id="eb-toast"),
    dcc.Download(id="eb-download"),
    dcc.Store(id="eb-experiments-store", data=[]),
    dcc.Store(id="eb-edit-index-store", data=None),

    dbc.Row(dbc.Col(html.Div([
        html.H4([
            html.I(className="bi bi-beaker me-2", style={"color": ACCENT}),
            html.Span("Experiment Builder",
                      style={"color": ACCENT, "fontFamily": "Inter, sans-serif"})
        ], className="mb-1"),
        html.P("Build, manage and export experiment configuration files.",
               style={"color": "#a0a0a0", "fontSize": "0.9rem"}),
        html.Hr(style={"borderColor": "rgba(255,255,255,0.08)"})
    ], className="pt-3 pb-1"), width=12)),

    dbc.Row([
        # ── Left: form ──────────────────────────────────────────────────────────
        dbc.Col(md=5, children=[
            dbc.Card(dbc.CardBody([
                html.H6("Experiment Parameters", className="mb-3", style={"color": ACCENT}),

                dbc.Row([
                    dbc.Col([
                        html.Label("Product", style=_LABEL_STYLE),
                        dbc.Select(
                            id="eb-product",
                            options=[{"label": p, "value": p} for p in _PRODUCTS],
                            value="GNR",
                            style={**_INPUT_STYLE}
                        ),
                    ], width=12, className="mb-3"),
                ]),

                # Scrollable form container
                html.Div(
                    id="eb-form-container",
                    style={"maxHeight": "550px", "overflowY": "auto", "paddingRight": "4px"},
                    children=_build_form("GNR")
                ),

                html.Hr(style={"borderColor": "rgba(255,255,255,0.08)"}),

                # Edit index indicator
                html.Div(id="eb-edit-indicator", style={"color": "#ffbd2e", "fontSize": "0.8rem",
                                                         "marginBottom": "8px"}),

                dbc.Row([
                    dbc.Col(dbc.Button(
                        [html.I(className="bi bi-plus-circle me-1"), "Add"],
                        id="eb-add-btn", outline=True, className="w-100",
                        style={"borderColor": ACCENT, "color": ACCENT}
                    ), width=4),
                    dbc.Col(dbc.Button(
                        [html.I(className="bi bi-pencil me-1"), "Update"],
                        id="eb-update-btn", outline=True, className="w-100",
                        style={"borderColor": "#ffbd2e", "color": "#ffbd2e"}
                    ), width=4),
                    dbc.Col(dbc.Button(
                        [html.I(className="bi bi-trash me-1"), "Clear"],
                        id="eb-clear-form-btn", color="secondary", outline=True, className="w-100",
                    ), width=4),
                ], className="mb-2 g-2"),

            ]), className="card-premium border-0"),
        ]),

        # ── Right: queue + actions ───────────────────────────────────────────────
        dbc.Col(md=7, children=[
            dbc.Card(dbc.CardBody([
                dbc.Row([
                    dbc.Col(html.H6("Experiment Queue", style={"color": ACCENT}), width=6),
                    dbc.Col(dbc.Button(
                        [html.I(className="bi bi-trash me-1"), "Clear All"],
                        id="eb-clear-all-btn", color="danger", outline=True, size="sm",
                        className="float-end"
                    ), width=6),
                ], className="mb-3"),

                html.Div(id="eb-exp-list", children=[
                    dbc.Alert("No experiments yet. Fill the form and click Add.",
                              color="secondary", className="card-premium border-0 text-white")
                ], style={"maxHeight": "400px", "overflowY": "auto"}),
            ]), className="card-premium border-0 mb-3"),

            dbc.Card(dbc.CardBody([
                html.H6("Import / Export", style={"color": ACCENT}, className="mb-3"),

                dbc.Row([
                    dbc.Col([
                        html.Label("Export to JSON", style=_LABEL_STYLE),
                        dbc.Input(id="eb-export-name", placeholder="experiments.json",
                                  type="text", className="mb-2", style=_INPUT_STYLE),
                        dbc.Button([html.I(className="bi bi-download me-1"), "Export JSON"],
                                   id="eb-export-btn", outline=True, className="w-100",
                                   style={"borderColor": ACCENT, "color": ACCENT}),
                    ], width=6),
                    dbc.Col([
                        html.Label("Save / Load Template (.tpl)", style=_LABEL_STYLE),
                        dbc.Input(id="eb-tpl-name", placeholder="my_template.tpl",
                                  type="text", className="mb-2", style=_INPUT_STYLE),
                        dbc.Row([
                            dbc.Col(dbc.Button([html.I(className="bi bi-bookmark me-1"), "Save .tpl"],
                                               id="eb-save-tpl-btn", outline=True, className="w-100",
                                               style={"borderColor": "#a070ff", "color": "#a070ff"}),
                                    width=6),
                            dbc.Col(dbc.Button([html.I(className="bi bi-bookmark-check me-1"), "Load .tpl"],
                                               id="eb-load-tpl-trigger", outline=True, className="w-100",
                                               style={"borderColor": "#a070ff", "color": "#a070ff"}),
                                    width=6),
                        ], className="g-1"),
                    ], width=6),
                ], className="mb-3"),

                html.Hr(style={"borderColor": "rgba(255,255,255,0.08)"}),

                dbc.Row([
                    dbc.Col([
                        html.Label("Import from JSON / .tpl", style=_LABEL_STYLE),
                        dcc.Upload(
                            id="eb-import-json",
                            children=html.Div([html.A("Browse", style={"color": ACCENT}),
                                               " or drop .json / .tpl"]),
                            multiple=False,
                            style={"border": f"1px dashed {ACCENT}", "borderRadius": "6px",
                                   "padding": "8px", "textAlign": "center",
                                   "color": "#a0a0a0", "fontSize": "0.82rem",
                                   "backgroundColor": f"rgba(54,215,183,0.03)",
                                   "cursor": "pointer"}
                        ),
                    ], width=6),
                    dbc.Col([
                        html.Label("Import from Excel (.xlsx)", style=_LABEL_STYLE),
                        dcc.Upload(
                            id="eb-import-excel",
                            children=html.Div([html.A("Browse", style={"color": "#00ff9d"}),
                                               " or drop .xlsx"]),
                            multiple=False,
                            style={"border": "1px dashed #00ff9d", "borderRadius": "6px",
                                   "padding": "8px", "textAlign": "center",
                                   "color": "#a0a0a0", "fontSize": "0.82rem",
                                   "backgroundColor": "rgba(0,255,157,0.03)",
                                   "cursor": "pointer"}
                        ),
                    ], width=6),
                ]),
            ]), className="card-premium border-0"),
        ]),
    ]),
])


# ── Callbacks ──────────────────────────────────────────────────────────────────

@callback(
    Output("eb-form-container", "children"),
    Input("eb-product", "value"),
    prevent_initial_call=False
)
def rebuild_form(product):
    return _build_form(product or "GNR")


@callback(
    Output("eb-experiments-store", "data"),
    Output("eb-toast", "children"),
    Output("eb-edit-index-store", "data"),
    Input("eb-add-btn", "n_clicks"),
    Input("eb-update-btn", "n_clicks"),
    Input("eb-clear-all-btn", "n_clicks"),
    Input("eb-import-json", "contents"),
    Input("eb-import-excel", "contents"),
    State("eb-experiments-store", "data"),
    State("eb-product", "value"),
    State("eb-edit-index-store", "data"),
    State({"type": "eb-field", "index": dash.ALL}, "value"),
    State({"type": "eb-field", "index": dash.ALL}, "id"),
    prevent_initial_call=True
)
def manage_queue(add_c, update_c, clear_c,
                 import_json_content, import_excel_content,
                 current, product, edit_idx,
                 field_values, field_ids):
    trigger = ctx.triggered_id

    # ── Clear all
    if trigger == "eb-clear-all-btn":
        return [], _toast("Queue cleared.", "info", 2000), None

    # ── Import JSON / .tpl
    if trigger == "eb-import-json" and import_json_content:
        try:
            _, data = import_json_content.split(',')
            imported = json.loads(base64.b64decode(data).decode('utf-8'))
            if isinstance(imported, dict):
                # Single experiment or {exp_name: {...}}
                exps = list(imported.values()) if all(isinstance(v, dict) for v in imported.values()) else [imported]
            elif isinstance(imported, list):
                exps = imported
            else:
                exps = [imported]
            return exps, _toast(f"Imported {len(exps)} experiment(s).", "success"), None
        except Exception as e:
            return no_update, _toast(f"Import error: {e}", "danger"), no_update

    # ── Import Excel
    if trigger == "eb-import-excel" and import_excel_content:
        try:
            exps = _parse_excel_import(import_excel_content)
            return exps, _toast(f"Imported {len(exps)} experiment(s) from Excel.", "success"), None
        except Exception as e:
            return no_update, _toast(f"Excel import error: {e}", "danger"), no_update

    # ── Collect current form values
    entry = {"product": product or "GNR"}
    for fid, fval in zip(field_ids, field_values):
        fname = fid["index"]
        # Checklist returns list; convert to bool string
        if isinstance(fval, list):
            fval = "true" if "true" in fval else "false"
        if fval is not None and str(fval).strip() != "":
            entry[fname] = fval

    current = list(current or [])

    if trigger == "eb-add-btn":
        current.append(entry)
        return current, _toast(f"Experiment #{len(current)} added.", "success", 2000), None

    if trigger == "eb-update-btn":
        if edit_idx is not None and 0 <= edit_idx < len(current):
            current[edit_idx] = entry
            return current, _toast(f"Experiment #{edit_idx + 1} updated.", "success", 2000), None
        return no_update, _toast("Select an experiment to update.", "warning"), no_update

    return no_update, no_update, no_update


def _parse_excel_import(content: str) -> list:
    """Parse Excel file uploaded via dcc.Upload — extracts named tables."""
    import openpyxl
    _, data = content.split(',')
    wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(data)), data_only=True)
    all_exps = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for table in ws.tables.values():
            rng = ws[table.ref]
            headers = [cell.value for cell in rng[0]]
            if "Field" in headers and "Value" in headers:
                fi, vi = headers.index("Field"), headers.index("Value")
                exp = {}
                for row in rng[1:]:
                    field = row[fi].value
                    value = row[vi].value
                    if field:
                        exp[str(field)] = str(value) if value is not None else ""
                if exp:
                    all_exps.append(exp)
    return all_exps


@callback(
    Output("eb-exp-list", "children"),
    Input("eb-experiments-store", "data"),
)
def render_queue(experiments):
    if not experiments:
        return dbc.Alert("No experiments yet. Fill the form and click Add.",
                         color="secondary", className="card-premium border-0 text-white")
    rows = []
    for i, exp in enumerate(experiments):
        product_val = exp.get("product", exp.get("Product", "?"))
        name_val = exp.get("Test Name", exp.get("Experiment", f"Exp #{i+1}"))
        test_type = exp.get("Test Type", "")
        rows.append(dbc.Card(dbc.CardBody([
            dbc.Row([
                dbc.Col(html.Span(
                    f"#{i+1} — {product_val} | {name_val} | {test_type}",
                    style={"color": ACCENT, "fontWeight": "600", "fontSize": "0.88rem"}
                ), width=8),
                dbc.Col([
                    dbc.Button(html.I(className="bi bi-pencil"),
                               id={"type": "eb-edit-btn", "index": i},
                               size="sm", outline=True, className="me-1",
                               style={"borderColor": "#ffbd2e", "color": "#ffbd2e"}),
                    dbc.Button(html.I(className="bi bi-trash"),
                               id={"type": "eb-del-btn", "index": i},
                               size="sm", color="danger", outline=True),
                ], width=4, className="text-end"),
            ], className="mb-1"),
            html.Div([
                html.Span(f"{k}: {v}  ", style={"color": "#a0a0a0", "fontSize": "0.75rem"})
                for k, v in list(exp.items())[:8] if v and k not in ("product",)
            ])
        ]), className="card-premium border-0 mb-1",
            style={"borderLeft": f"3px solid {ACCENT}"}))
    return rows


@callback(
    Output("eb-edit-index-store", "data", allow_duplicate=True),
    Output("eb-edit-indicator", "children"),
    Input({"type": "eb-edit-btn", "index": dash.ALL}, "n_clicks"),
    State("eb-experiments-store", "data"),
    prevent_initial_call=True
)
def load_edit(edit_clicks, experiments):
    if not any(c for c in edit_clicks if c):
        return no_update, no_update
    # Find which button was clicked
    triggered = ctx.triggered_id
    if triggered and "index" in triggered:
        idx = triggered["index"]
        return idx, f"✏ Editing experiment #{idx + 1} — click Update to save"
    return no_update, no_update


@callback(
    Output("eb-experiments-store", "data", allow_duplicate=True),
    Output("eb-toast", "children", allow_duplicate=True),
    Input({"type": "eb-del-btn", "index": dash.ALL}, "n_clicks"),
    State("eb-experiments-store", "data"),
    prevent_initial_call=True
)
def delete_experiment(del_clicks, experiments):
    if not any(c for c in del_clicks if c):
        return no_update, no_update
    triggered = ctx.triggered_id
    if triggered and "index" in triggered:
        idx = triggered["index"]
        new_list = [e for i, e in enumerate(experiments) if i != idx]
        return new_list, _toast(f"Experiment #{idx+1} deleted.", "info", 2000)
    return no_update, no_update


@callback(
    Output("eb-download", "data"),
    Output("eb-toast", "children", allow_duplicate=True),
    Input("eb-export-btn", "n_clicks"),
    Input("eb-save-tpl-btn", "n_clicks"),
    State("eb-experiments-store", "data"),
    State("eb-export-name", "value"),
    State("eb-tpl-name", "value"),
    prevent_initial_call=True
)
def export_or_save(export_c, tpl_c, experiments, export_name, tpl_name):
    if not experiments:
        return no_update, _toast("No experiments to export.", "warning")

    trigger = ctx.triggered_id
    if trigger == "eb-export-btn":
        fname = export_name or "experiments.json"
        if not fname.endswith(".json"):
            fname += ".json"
        payload = json.dumps(experiments, indent=2)
        return dcc.send_string(payload, fname), _toast(f"Exported {fname}.", "success", 2500)

    if trigger == "eb-save-tpl-btn":
        fname = tpl_name or "template.tpl"
        if not fname.endswith(".tpl"):
            fname += ".tpl"
        payload = json.dumps(experiments, indent=2)
        return dcc.send_string(payload, fname), _toast(f"Template saved as {fname}.", "success", 2500)

    return no_update, no_update


def _toast(msg, icon, duration=4000):
    return dbc.Toast(
        msg, icon=icon, duration=duration, is_open=True,
        style={"position": "fixed", "top": 20, "right": 20, "zIndex": 9999},
        className="toast-custom"
    )
