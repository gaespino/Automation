"""
Author: jfnavarr/gaespino
Last update: 28/1/25
Works with Logger UI to create reports
"""

import openpyxl
import zipfile
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter


from namednodes import sv
from ipccli.stdiolog import log
from ipccli.stdiolog import nolog
import ipccli
import users.gaespino.dev.S2T.ConfigsLoader as pe
import users.gaespino.dev.DebugFramework.FileHandler as fh

SELECTED_PRODUCT = pe.SELECTED_PRODUCT
SELECTED_VARIANT = pe.PRODUCT_VARIANT
MAXPHYSICAL = pe.MAXPHYSICAL


try: 
	if SELECTED_PRODUCT == 'GNR':
		import mc.cwfDimmInfo as Dimm ## Not available for CWF at the moment
	elif  SELECTED_PRODUCT == 'CWF':
		import mc.gnrDimmInfo as Dimm ## Not available for CWF at the moment
except: 
	print( f'Error importing DimmInfo scripts for {SELECTED_PRODUCT}')

try: 
	from core import debug as cd
except: 
	print( 'Error importing CORE DEBUG scripts')


try: 
	from core import debug as cd
except: 
	print( 'Error importing CORE DEBUG scripts')

from pysvtools.server_ip_debug.cha import cha
from pysvtools.server_ip_debug.pm import pm
from toolext.server_ip_debug.punit import punit
from pysvtools.server_ip_debug.ubox import ubox

try: 
	from toolext.server_ip_debug.llc import llc ## Not available for CWF at the moment
except: 
	print( 'Error importing server_ip_debug LLC decoder')

try: 
	from pysvtools import debug_mca ## Not available for CWF at the moment
except: 
	print( 'Error importing DEBUG MCA scripts')

import pysvtools.server_wave_4.tileview as t
import os
import sys
import contextlib
import random
import users.gaespino.dev.S2T.dpmChecks as dpm
import users.gaespino.dev.S2T.CoreManipulation as gcm

default_folder = "C:\\temp\\"

## Dragon DR Collection -- CWF WIP
DRGDUMP_ENABLED = False
if SELECTED_PRODUCT == 'GNR': DRGDUMP_ENABLED = True
elif  SELECTED_PRODUCT == 'CWF': DRGDUMP_ENABLED = False

# GNR uses bigcore IP -- labeled as Core with 2 threads
# CWF uses atomcore IP -- which is labeled as a module with 4 Atom cores inside (4 threads)
#core_product = dpm.product_str()
core_ip = {'GNRAP':'CORE', 'CWFAP':'MOD','GNRSP':'CORE', 'CWFSP':'MOD'}

# VARIABLES TO USE IN FRAMEWORK REPORT -- TO BE CALLED BY UI

LICENSE_DICT = pe.LICENSE_S2T_MENU

FRAMEWORK_CORELIC = [f"{k}:{v}" for k,v in LICENSE_DICT.items()]
FRAMEWORK_VTYPES = ["VBUMP", "FIXED", "PPVC"]
FRAMEWORK_RUNSTATUS = ["PASS", "FAIL"]
FRAMEWORK_CONTENT = ["Dragon", "Linux", "PYSVConsole"]

FRAMEWORK_VARS = {	'qdf':'', 
			   		'tnum':'', 
					'mask':'', 
					'corelic':'', 
					'bumps':'', 
					'htdis':'', 
					'dis2CPM':'', 
					'freqIA':'', 
					'voltIA':'', 
					'freqCFC':'', 
					'voltCFC':'', 
					'content':'', 
					'passstring':'', 
					'failstring':'', 
					'runName':'', 
					'runStatus':'', 
					'scratchpad':'', 
					'seed':'',
					'ttlog':  None
					}

def dragon_dump(logging, logger=None, mcadata=None):
	log(logging,'a')
	
	gcm._core_dr_registers(logger, mcadata)
	
	nolog()

def load_workbook():
	"""
	Load the workbook template to a variable
	"""
	scriptHome = os.path.dirname(os.path.realpath(__file__))
	parentdir = os.path.dirname(scriptHome)
	templatePath = os.path.join(parentdir, 'product_specific',f'{SELECTED_PRODUCT.lower()}', 'LoggerFiles', 'TileView_Template.xlsx')
	#print(scriptHome)
	# template for tileview
	workbook_template = openpyxl.load_workbook(templatePath)
	workbook = workbook_template
	workbook_template.close()
	return workbook

def save_workbook(workbook, filepath):
	"""
	Save the workbook into a file in the filepath
	"""
	workbook.save(filepath)
	return

def find_value(sheet, value):
	"""
	Find the cell where is located the value
	"""
	if value == 'ubox':
		value = 'ubox_0'

	for row in sheet.iter_rows():
		for cell in row:
			if cell.value == value:
				return cell

def add_error(sheet, cell, error):
	"""
	Mark and write an error into the sheet
	"""
	# mark error
	if cell != None:
		error_color = "FFA3A3" # light red
		fill = PatternFill(start_color=error_color, end_color=error_color, fill_type="solid")
		cell.fill = fill

	# write error
	current_error_content = sheet["X8"].value
	if not current_error_content:
		sheet["X8"] = error

	else:
		current_error_content += ("\n" + error)
		sheet["X8"] = current_error_content

	return

def mark_disable(cell, able = 0):
	"""
	Mark the disable IP's into the sheet
	"""
	if able:
		color = "00FFC000" # yellow
	else:
		color = "AEAAAA" # gray
	fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
	cell.fill = fill
	return

def get_mask(mask1, mask2, mask3):
	"""
	Combine the 2 parts of core/cha mask
	"""
	_mask1 = mask1 if mask1 != None else 0xFFFFFFFFFFFFFFF
	_mask2 = mask2 if mask2 != None else 0xFFFFFFFFFFFFFFF
	_mask3 = mask3 if mask3 != None else 0xFFFFFFFFFFFFFFF

	mask1_str = hex(_mask1)[2:] 
	mask2_str = hex(_mask2)[2:] 
	mask3_str = hex(_mask3)[2:]

	# complete the lenght of the mask in case it has zeros at the beginning
	mask1_str = mask1_str.zfill(15)
	mask2_str = mask2_str.zfill(15)
	mask3_str = mask3_str.zfill(15)

	# combine both parts
	mask = mask3_str + mask2_str + mask1_str
	hex_mask = '0x' + mask

	# hex2bin mask
	bin_mask = bin(int(mask, 16))[2:]

	# complete the lenght of the mask in case it has zeros at the beginning
	bin_mask = bin_mask.zfill(MAXPHYSICAL)

	return bin_mask, hex_mask

def review_mask(mask, IP, sheet, product):
	"""
	Review the mask to find disable IP's
	"""
	productip = core_ip[product]
	if IP == 'm2iosf':
		ip_num = 0
		cell_names = ['M2IOSF8', 'M2IOSF1', 'M2IOSF2', 'M2IOSF3', 'M2IOSF9', 'M2IOSF10', 'M2IOSF5', 'M2IOSF0', 'M2IOSF4', 'M2IOSF11']

		for ip_val in reversed(mask):

			if ip_val == '1':
				cell_name = cell_names[ip_num]
				cell = find_value(sheet, cell_name)
				mark_disable(cell)

			ip_num = ip_num + 1

			if ip_num > 9:
				break

	else: # for core, cha and core/cha
		ip_num = 0

		for ip_val in reversed(mask):

			if ip_val == '1':
				if IP == 'core_cha':
					full_cell_name = productip + str(ip_num)
					cell = find_value(sheet, full_cell_name)
					cell = cell.offset(row=0, column =-1)
				else:
					full_cell_name = IP + str(ip_num)
					cell = find_value(sheet, full_cell_name)
				#print(full_cell_name)    
				mark_disable(cell)
			#print(ip_num)
			ip_num = ip_num + 1
			
def common_mask(mask1, mask2):
	"""
	Makes an AND operation between core and cha mask
	"""
	mask1_int = int(mask1, 2)
	mask2_int = int(mask2, 2)

	common_mask_int = mask1_int & mask2_int
	common_mask = bin(common_mask_int)[2:]
	common_mask = common_mask.zfill(MAXPHYSICAL)
	return common_mask

def get_m2iosf_mask(m2iosf_mask):
	"""
	Converts the m2iosf mask (int) to a binary number
	"""
	bin_m2iosf_mask = bin(int(m2iosf_mask))[2:]
	bin_m2iosf_mask = bin_m2iosf_mask.zfill(9)

	return bin_m2iosf_mask

def check_MCA(sheet, mca_path, mca_debug_path, debug_mca_log, mca_report_path, dr_dump, visual = '', loop = '', WW = 'WW', Bucket = 'UNCORE', product = 'GNR', QDF = '', bindesc= '', logger = None):
	"""
	Reads and save the MCA registers, and mark the in the workbook
	"""
	productip = core_ip[product]
	mcadata = {'TestName':[],	'TestValue':[]}

	log(mca_path)
	try:
		print("Reading MCA...")
		
		if 'GNR' in product: mcadata, pysvdecode = mca_dump_gnr()
		elif 'CWF' in product: mcadata, pysvdecode = mca_dump_cwf()

		if pysvdecode['cha']:
			print("Decoding CHA MCA...")
			cha.Cha().show_mca_status()
		if pysvdecode['ubox']:
			print("Decoding Ubox MCA...")
			ubox.Ubox().show_mca_status()
		if pysvdecode['punit']:
			print("Decoding PM MCA...")
			pm.PM().show_mca_status(source='reg')
		if pysvdecode['pm']:
			print("Decoding Punit MCA...")
			punit.Punit().show_mca_status()
		#if pysvdecode['llc']: # Not available for CWF
		#	print("Decoding LLC MCA...")
		#	llc.Llc().show_mca_status()

		
	# zip the logs		
		print("MCA READ SUCCESFULLY")

	except:
		print("MCA COULD NOT BE READ")
		return
	nolog()

	if DRGDUMP_ENABLED and dr_dump:
		try:
			print("DUMPING DRs DATA INTO MCA FILE")
			dragon_dump(mca_path, logger, mcadata)
		except Exception as e:
			print(f"ERROR !! COULD NOT DUMP DRs DATA INTO MCA FILE: {e}")

	mcreport = mca_report(mcadata=mcadata, visual = visual, lot = f'RVP{loop}', WW = WW, Bucket = Bucket, product = product, QDF = QDF, bindesc = bindesc, lotseq=loop)
		
	mca_file(mca_report_path, mcreport)
	
	log(mca_debug_path)

	if debug_mca_log:

		#try:
		#	print("Reading MCA debug...")
		if 'CWF' in product:
			print('Debug MCA script not available for CWF')
		else:
			debug_mca.analyze(scope=r'socket\d\.',variant='AP')
		
		print("MCA DEBUG READ SUCCESFULLY")

		#except:
		#	print("MCA DEBUG COULD NOT BE READ")
	else: 
		log(mca_debug_path)
		print("Debug MCA Script disabled by default, use key debug_mca = True, to enable it if needed")
		#nolog()
	nolog()
	
	try:
		with open(mca_path, 'r') as file:
			MCAs = file.readlines()

		if 'FOUND VALID MCA\n' in MCAs:
			# mark FAILED in the workbook
			sheet["X2"] = 'FAILED'
			cell_color = "FF0000"

			# look for the MCA's
			start_search = MCAs.index('FOUND VALID MCA\n')
			for num_line in range(start_search, len(MCAs)):
				line = MCAs[num_line]
				
				if line[0] == 's':
					parts = line.split(".")
					IP = parts[2]
					if IP[0] == 'c':
						IP = productip + IP[4:]
					if parts[3] == 'cha':
						IP = parts[4]
						IP = 'CHA' + IP[3:]
					cell = find_value(sheet, IP)
					add_error(sheet, cell, line)

		else:
			# mark PASS in the workbook
			sheet["X2"] = 'PASS'
			cell_color = "00B050"
			print('DID NOT FIND VALID MCA')

		fill = PatternFill(start_color=cell_color, end_color=cell_color, fill_type="solid")
		sheet["X2"].fill = fill
		sheet["B66"].fill = fill
		print("MCA WRITTEN SUCCESFULLY")
		return
	
	except:
		print("COULD NOT WRITE MCA")
		return

def check_mem(sheet, config_path, product):
	"""
	Check memory configuration, mark it in the workbook and save in a log
	"""
	log(config_path)
	ipc.halt()
	if 'GNR' in product: Dimm.gnr_dimm_info()
	ipc.go()
	nolog()

	with open(config_path, 'r') as file:
		mem = file.readlines()

	slt_position = []
	for position, value in enumerate(mem):
		if value[:6] == '  0  |':
			slt_position.append(position)

	# review each MC
	count = 0
	for MC in slt_position:
		cell_name = 'MC' + str(count)

		# cell of the MC
		cell = find_value(sheet, cell_name)

		# move the cell to the column of the MC's channels
		next_column_index = cell.column + 1
		next_column = get_column_letter(next_column_index)
		ch0_cell = sheet[next_column + str(cell.row)]
		ch1_cell = sheet[next_column + str(cell.row+1)]

		# strings to review if the channel is populated or not
		ch0 = mem[MC][5:63]
		ch1 = mem[MC][62:120]
		
		# review CH0
		if ch0 == '|                     Not installed                      |':
			mark_disable(ch0_cell)
		else:
			mark_disable(ch0_cell, 1) # mark the cell able
		# review CH1
		if ch1 == '|                     Not installed                      |':
			mark_disable(ch1_cell)
		else:
			mark_disable(ch1_cell, 1) # mark the cell able

		# check if a full MC is disable
		ch0_color = Color(ch0_cell.fill.start_color.rgb)
		ch1_color = Color(ch1_cell.fill.start_color.rgb)
		if ch0_color.rgb == "00AEAAAA" and ch1_color.rgb == "00AEAAAA":
			mark_disable(cell)
		if ch0_color.rgb == "00FFC000" and ch1_color.rgb == "00FFC000":
			mark_disable(cell, 1)
		count += 1

def define_paths(folder, test):
	# Define paths
	if folder[-1] != "\\":
		folder = folder + "\\"

	if test == '':
		file_path = folder + "TileView.xlsx"
		mca_path = folder + "MCAs.txt"
		mca_debug_path = folder + "MCA_debug.txt"
		mca_report_path = folder + "MCA_Report.xlsx"
		config_path = folder + "config.txt"
		zip_path = folder + "log.zip"
	else:
		file_path = folder + test + "_TileView"  + ".xlsx"
		mca_path = folder + test + "_MCAs"  + ".txt"
		mca_debug_path = folder + test + "_MCA_debug"  + ".txt"
		mca_report_path = folder + test + "_MCA_Report" + ".xlsx"
		config_path = folder + test + "_config"  + ".txt"
		zip_path = folder + test + "_log"  + ".zip"

	return file_path, mca_path, mca_debug_path, config_path, zip_path, mca_report_path

def generate_temporal(handler, files, serial_log):
	# Generate temporal folder
	handler.generate_temporal()

	# Copy Files to Temporal
	handler.copy_to_temporal(files)

	# if Serial log File copy to temporal
	handler.copy_to_log_to_temporal(serial_log)

def default_framework_vars():

	frame_vars = FRAMEWORK_VARS
	
	return frame_vars

def generate_frameworklog(handler, frame_vars):
	qdf = frame_vars.get('qdf', '')
	tnum = frame_vars.get('tnum', '')
	mask = frame_vars.get('mask', '')
	corelic = frame_vars.get('corelic', '')
	bumps = frame_vars.get('bumps', 'vbump')
	htdis = frame_vars.get('htdis', False)
	dis2CPM = frame_vars.get('dis2CPM', False)
	freqIA = frame_vars.get('freqIA', '')
	voltIA = frame_vars.get('voltIA', '')
	freqCFC = frame_vars.get('freqCFC', '')
	voltCFC = frame_vars.get('voltCFC', '')
	content = frame_vars.get('content', 'Dragon')
	passstring = frame_vars.get('passstring', 'Test Complete')
	failstring = frame_vars.get('failstring', 'Test Failed')
	runName = frame_vars.get('runName', 'NotDefined')
	runStatus = frame_vars.get('runStatus', 'NotDefined')
	scratchpad = frame_vars.get('scratchpad', 'NotDefined')
	seed = frame_vars.get('seed', 'NotDefined')
	

	# Starts recording console to save a FrameworkLogging format file
	handler.initlog()
	handler.TestBanner(qdf, tnum, mask, corelic, bumps, htdis, dis2CPM, freqIA, voltIA, freqCFC, voltCFC, content, passstring, failstring)
	handler.Iteration_end(tnum, runName, runStatus, scratchpad, seed)

	# Generate Summary
	handler.generate_summary()

def data_upload(handler):

	# Upload Data
	handler.upload_data()

	# Remove Temporal Folder
	handler.remove_temporal()

def run(visual ='', Testnumber='', TestName='', chkmem = 0, debug_mca = 0, dr_dump = False, folder=None, WW = 'WW', Bucket = 'UNCORE', product = 'GNR', QDF = '', logger =None, upload_to_disk = False, upload_to_danta = False, framework_data = None):
	"""
	Main function of the script
	"""
	test = f'{Testnumber}_{visual}_{TestName}'

	global ipc 
	ipc = ipccli.baseaccess()
	productip = core_ip[product]
	print(f'{"+"*20} Collecting System data for {product}:{productip} {"+"*20} ')
	## Initializing the default folder
	if folder == None:
		folder = default_folder
	# Define paths
	file_path, mca_path, mca_debug_path, config_path, zip_path, mca_report_path = define_paths(folder, test)
	
	# Generate Upload Data object

	datahandler = fh.TestUpload(folder=folder, vid=visual, name=TestName, bucket=Bucket, WW=WW, product=product, logger = logger, from_Framework = False, upload_to_disk=upload_to_disk, upload_to_danta=upload_to_danta)
	files = [file_path, mca_report_path, zip_path]
	framework_vars = default_framework_vars() if framework_data == None else framework_data

	# Load workbook
	try:
		workbook = load_workbook()
		sheet = workbook['Template']
		print("WORKBOOK LOADED SUCCESFULLY")
	except:
		print("WORKBOOK COULD NOT BE LOADED")
		return
	
	# check Mem
	try:
		if chkmem and 'GNR' in product: # Not available for CWF
			check_mem(sheet, config_path)
			print("MEMORY READ SUCCESFULLY")
	
	except:
		print("MEMORY COULD NOT BE READ")

	# Reading Registers
	try:
		#sv.sockets.computes.fuses.load_fuse_ram()
		#sv.sockets.ios.fuses.load_fuse_ram()

		masks = dpm.fuses()

		# Reference for the above dictionary
		#masks = {	'ia_compute_0' : iamask_c0,
		#		'ia_compute_1' : iamask_c1,
		#		'ia_compute_2' : iamask_c2,
		#		'llc_compute_0' : llcmask_c0,
		#		'llc_compute_1' : llcmask_c1,
		#		'llc_compute_2' : llcmask_c2,
		#		}

		# read core mask
		core_mask_c0 = masks['ia_compute_0']
		core_mask_c1 = masks['ia_compute_1']
		core_mask_c2 = masks['ia_compute_2']

		# read cha mask
		cha_mask_c0 = masks['llc_compute_0']
		cha_mask_c1 = masks['llc_compute_1']
		cha_mask_c2 = masks['llc_compute_2']

		# read m2iosf mask
		#iocoh_mask = sv.socket0.tile0.fuses.punit.pcode_sa_disable_scf_io #iocoh
		#pi5_mask = sv.socket0.tile0.fuses.punit.pcode_sa_disable_pi5 #pi5
		#fuse1 = sv.socket0.tile0.fuses.punit.capid_ptpcioregs_capid10_capid10_1_fuse
		#fuse2 = sv.socket0.tile0.fuses.punit.capid_ptpcioregs_capid10_capid10_2_fuse
		#fuse3 = sv.socket0.tile0.fuses.punit.capid_ptpcioregs_capid10_capid10_3_fuse
		#fuse4 = sv.socket0.tile0.fuses.punit.capid_ptpcioregs_capid10_capid10_4_fuse
		#fuse5 = sv.socket0.tile0.fuses.punit.capid_ptpcioregs_capid10_capid10_5_fuse
		#fuse8 = sv.socket0.tile0.fuses.punit.capid_ptpcioregs_capid10_capid10_8_fuse
		#fuse9 = sv.socket0.tile0.fuses.punit.capid_ptpcioregs_capid10_capid10_9_fuse
		#fuse10 = sv.socket0.tile0.fuses.punit.capid_ptpcioregs_capid10_capid10_10_fuse
		#fuse11 = sv.socket0.tile0.fuses.punit.capid_ptpcioregs_capid10_capid10_11_fuse

		# Read PC
		PC = readscratchpad() #str(sv.socket0.io0.uncore.ubox.ncdecs.biosnonstickyscratchpad7_cfg)

		print("REGISTERS READ SUCCESFULLY")

	except:
		print("REGISTERS COULD NOT BE READ")
		return

	# Saving TileView
	log(config_path, "a")
	try:
		
		print("\nTile View\n")
		t.tileview(sv.socket0).show_soc()
		print("TILEVIEW SAVED SUCCESFULLY")

	except:
		print("TILEVIEW COULD NOT BE SAVED")
	try:
		print("\nCFC Voltage/Ratios for current System \n")
		CFC = dpm.cfc(fuse_ram = False)
		CFC.ratios()
		CFC.voltage()

		print("\nCORE(IA) Voltage/Ratios for current System \n")
		IA = dpm.ia(fuse_ram = False)
		IA.ratios()
		IA.voltage()
	except:
		print("CFC/iA Ratios/Voltages for current system cannot be read....")
	nolog()

	# Saving Registers
	try:
		with open(config_path, "a") as file:

			# save PC
			file.write("PC\n")
			file.write("sv.socket0.uncore.ubox.ncdecs.biosnonstickyscratchpad7_cfg = " + PC + "\n\n")

			file.write("\nConfig Registers\n")

			# save core mask
			file.write("CORE Mask\n")
			file.write("sv.socket0.compute0.fuses.hwrs_top_rom.ip_disable_fuses_dword6_core_disable.read() = " + str(core_mask_c0) + "\n")
			file.write("sv.socket0.compute1.fuses.hwrs_top_rom.ip_disable_fuses_dword6_core_disable.read() = " + str(core_mask_c1) + "\n")
			file.write("sv.socket0.compute2.fuses.hwrs_top_rom.ip_disable_fuses_dword6_core_disable.read() = " + str(core_mask_c2) + "\n\n")

			# save cha mask
			file.write("CHA Mask\n")
			file.write("sv.socket0.compute0.fuses.hwrs_top_rom.ip_disable_fuses_dword2_llc_disable.read() = " + str(cha_mask_c0) + "\n")
			file.write("sv.socket0.compute1.fuses.hwrs_top_rom.ip_disable_fuses_dword2_llc_disable.read() = " + str(cha_mask_c1) + "\n")
			file.write("sv.socket0.compute2.fuses.hwrs_top_rom.ip_disable_fuses_dword2_llc_disable.read() = " + str(cha_mask_c2) + "\n\n")

			# read m2iosf mask
			#file.write("M2IOSF Mask\n")
			#file.write("sv.socket0.tile0.fuses.punit.pcode_sa_disable_scf_io = " + str(iocoh_mask) + "\n") #iocoh
			#file.write("sv.socket0.tile0.fuses.punit.pcode_sa_disable_pi5 = " + str(pi5_mask) + "\n") #pi5
			#file.write("sv.socket0.tile0.fuses.punit.capid_ptpcioregs_capid10_capid10_1_fuse = " +str(fuse1) + "\n")
			#file.write("sv.socket0.tile0.fuses.punit.capid_ptpcioregs_capid10_capid10_2_fuse = " +str(fuse2) + "\n")
			#file.write("sv.socket0.tile0.fuses.punit.capid_ptpcioregs_capid10_capid10_3_fuse = " +str(fuse3) + "\n")
			#file.write("sv.socket0.tile0.fuses.punit.capid_ptpcioregs_capid10_capid10_4_fuse = " +str(fuse4) + "\n") 
			#file.write("sv.socket0.tile0.fuses.punit.capid_ptpcioregs_capid10_capid10_5_fuse = " +str(fuse5) + "\n")
			#file.write("sv.socket0.tile0.fuses.punit.capid_ptpcioregs_capid10_capid10_8_fuse = " +str(fuse8) + "\n")
			#file.write("sv.socket0.tile0.fuses.punit.capid_ptpcioregs_capid10_capid10_9_fuse = " +str(fuse9) + "\n")
			#file.write("sv.socket0.tile0.fuses.punit.capid_ptpcioregs_capid10_capid10_10_fuse = " +str(fuse10) + "\n")
			#file.write("sv.socket0.tile0.fuses.punit.capid_ptpcioregs_capid10_capid10_11_fuse = " +str(fuse11) + "\n\n")

		print("REGISTERS SAVED SUCCESFULLY")

	except:
		print("REGISTERS COULD NOT BE SAVED")


	# get cha mask
	bin_cha_mask, hex_cha_mask = get_mask(cha_mask_c0, cha_mask_c1, cha_mask_c2)

	# get core mask
	bin_core_mask, hex_core_mask = get_mask(core_mask_c0, core_mask_c1, core_mask_c2)

	# get core/cha common mask
	common_core_cha_mask = common_mask(bin_cha_mask, bin_core_mask)

	# get bin m2iosf mask
	#bin_m2iosf_mask = get_m2iosf_mask(iocoh_mask)
	
	# review slices disable
	try:
	
		review_mask(bin_cha_mask, 'CHA', sheet, product)
	
		review_mask(bin_core_mask, productip, sheet, product) ## In the case of CWF CORE Refers to MODULE --
	
		review_mask(common_core_cha_mask, 'core_cha', sheet, product) ## In the case of CWF CORE Refers to MODULE -- 
		#review_mask(bin_m2iosf_mask, 'm2iosf', sheet)
		print("DISABLE SLICES WRITTEN SUCCESFULLY")
	except:
		print("COULD NOT WRITE DISABLE SLICES")
		return

	# write PC
	sheet["B66"] = 'PC - ' + PC

	# Check MCA
	#check_MCA(sheet, mca_path, mca_debug_path, debug_mca, mca_report_path, visual = visual, WW=WW, loop=Testnumber)
	check_MCA(sheet, mca_path, mca_debug_path, debug_mca, mca_report_path, dr_dump=dr_dump, visual = visual, loop = Testnumber, WW = WW, Bucket = Bucket, product = product, QDF = QDF, bindesc=test, logger=logger)
	# write bootscript
	#bootscript = 'b.go(llc_slice_disable=' + hex_cha_mask + ', ia_core_disable=' + hex_core_mask + ", fuse_str=['punit.pcode_mapout_0_core_id=" + str(hex(malgudi_0)) + "','punit.pcode_mapout_1_core_id=" + str(hex(malgudi_1)) + "'])"
	#sheet["C4"] = bootscript

	# write name
	test_name = sheet["B2"].value + ' ' + test
	sheet["B2"] = test_name
	

	try:
		with zipfile.ZipFile(zip_path, "w") as zipf:
			zipf.write(mca_path, arcname = "MCAs.txt")
			zipf.write(mca_debug_path, arcname = "MCA_debug.txt")
			zipf.write(config_path, arcname = "Config.txt")
			#zipf.write(mca_report_path, arcname = "MCA_report.xlsx")
		
		os.remove(mca_path)
		os.remove(mca_debug_path)
		os.remove(config_path)
		#os.remove(mca_report_path)
		print("ZIP FILE CREATED SUCCESFULLY")
	except:
		print("COULD NOT CREATE ZIP FILE")
	
	try:
		save_workbook(workbook, file_path)
		print("WORKBOOK SAVED SUCCESFULLY")
		
		if upload_to_disk or upload_to_danta:
			print("DATA UPLOAD IN PROGRESS...")
			
			# Define data
			framework_vars['qdf'] = QDF
			framework_vars['tnum'] = Testnumber
			framework_vars['runName'] = TestName
			framework_vars['scratchpad'] = readscratchpad()
			
			serial_log = framework_vars['ttlog'] if framework_vars['ttlog'] != None else None
			generate_temporal(handler=datahandler, files=files, serial_log=serial_log)
			generate_frameworklog(handler=datahandler, frame_vars=framework_vars)
			data_upload(handler=datahandler)

		print(
"""
  ____                      _
 |  _ \  ___  _ __   ___   | |
 | | | |/ _ \| '_ \ / _ \  |_|
 | |_| | (_) | | | |  __/   _ 
 |____/ \___/|_| |_|\___|  |_|

"""
		)
		return
	
	except:
		print("WORKBOOK COULD NOT BE SAVED")

def mca_report(mcadata, visual, lot='RVP', operation='DPM', WW = 'WW', Bucket = 'UNCORE', product = 'GNR', QDF = '', lotseq = None, seqkey= None, bindesc=''):
	mcreport = {'VisualId':[], 'Lot':[], 'LatoStartWW':[], 'LotsSeqKey':[],	'UnitTestingSeqKey':[],	'TestName':[], 'TestValue':[], 'Operation':[], 'TestNameNumber':[],	'TestNameWithoutNumeric':[],}

	# This is to avoid duplicates during testing
	if lotseq == None: lotseq = random.getrandbits(12)
	if seqkey == None: seqkey = random.getrandbits(4)
	
	if mcadata['TestName'] != []:
		for i, mca in enumerate(mcadata['TestName']):
			#print(f'Index:{i} -- data{mca}' )
			mcreport['VisualId'].append(visual)
			mcreport['Lot'].append(lot)			
			mcreport['LotsSeqKey'].append(lotseq)  
			mcreport['UnitTestingSeqKey'].append(seqkey)	  			
			mcreport['LatoStartWW'].append(WW)		
			mcreport['TestName'].append(mca.replace(".","__").upper()) # Adapt format for the Other report
			mcreport['TestValue'].append(mcadata['TestValue'][i])
			mcreport['Operation'].append(operation)
			mcreport['TestNameNumber'].append('0')
			mcreport['TestNameWithoutNumeric'].append(mca)

	
	results = {
						'VisualId'                  :[visual],
						'Lot'                       :[lot],
						'LatoStartWW'               :[WW],
						'LotsSeqKey'                :[lotseq],
						'UnitTestingSeqKey'         :[''],
						'DpmBucket'                 :[Bucket],
						'DecimaSite'                :['CR'],
						'DecimaWW'                  :[WW],
						'DecimaBucket'              :[Bucket],
						'Accuracy'                  :[''],
						'Operation'                 :[operation],
						'ProductConfigurationName'  :[product],
						'Program'                   :[''],
						'SSPEC'                     :[QDF],
						'DevRevStep'                :[''],
						'TestName'                  :[mcadata['TestName'][0] if mcadata['TestName'] != [] else ''],                        
						'TestValue'                 :[mcadata['TestValue'][0] if mcadata['TestName'] != [] else ''],                     
						'DB'                        :[''],   
						'BinDesc'                   :[bindesc],   
					}
		
	final_bin = {   
						'VisualId'                  :[visual],
						'DpmBucket'                 :[Bucket],
						'DecimaSite'                :['CR'],
						'DecimaWW'                  :[WW],
						'DecimaBucket'              :[Bucket],
						'DpmBucketAccuracy'         :[''],
						'ProductConfigurationName'  :[product]}


	data = {'raw_data':mcreport, 'final_bucket':final_bin,'results':results}
	return data		

def mca_file(pathfile, data):
	dfraw = pd.DataFrame(data['raw_data'])
	dffinal = pd.DataFrame(data['final_bucket'])
	dfresults = pd.DataFrame(data['results'])
	#df = pd.DataFrame(data)

	# Save the DataFrame to an Excel file
	with pd.ExcelWriter(pathfile) as writer:   
		dfraw.to_excel(writer, sheet_name='raw_data', index=False)
		dffinal.to_excel(writer, sheet_name='final_bucket', index=False)
		dfresults.to_excel(writer, sheet_name='results', index=False)

# Portion of code extracted from core debug
def mca_dump_gnr(verbose=True):

	def print_valid(i,a,mc,e,b=63,save=False):
		try:
			if verbose: print("%s = 0x%x" %(i.path, i.read()))
			if i.bits(b,1): 
				a.append(i)
				mc.append(i)
				
				return True
			else:
				if save: mc.append(i)
				return False
		except:
			#print("Access failed to: {}".format(i.path))
			print("%s = <error reading>" %(i.path))
			e.append(i)
			return False

	import ipccli
	itp = ipccli.baseaccess()
	unlock()
	if not sv.sockets: sv.refresh()
	a=[] #mcas
	e=[] #errors
	mc=[]
	#drg = []

	mcadata = {'TestName':[],	'TestValue':[]}

	sv.sockets.cpu.cores.setaccess('crb') #no halt required
	## Changed Core data request to use the compute register of directly from cpu.core, this in order to keep format
	## there shouldnt be any difference in both registers, will keep looking for a similar approach in the future, for now this works.

	with itp.device_locker():
		for i in sv.sockets.computes.cpu.cores.ml2_cr_mc3_status:
			if print_valid(i,a,mc,e):
				if i.bits(58,1): 
					a.append(i.parent.ml2_cr_mc3_addr)
					mc.append(i.parent.ml2_cr_mc3_addr)
				if i.bits(59,1): 
					mc.append(i.parent.ml2_cr_mc3_misc)
		for i in sv.sockets.computes.cpu.cores.dtlb_cr_mc2_status:
			if print_valid(i,a,mc,e):
				if i.bits(58,1): 
					a.append(i.parent.dtlb_cr_mc2_addr) #CRIF issue named as dcu instead of dtlb
					mc.append(i.parent.dtlb_cr_mc2_addr)
				if i.bits(59,1):
					mc.append(i.parent.dtlb_cr_mc2_misc)
		for i in sv.sockets.computes.cpu.cores.threads.dcu_cr_mc1_status:
			if print_valid(i,a,mc,e):
				if i.bits(58,1): 
					a.append(i.parent.dcu_cr_mc1_addr)
					mc.append(i.parent.dcu_cr_mc1_addr)
				if i.bits(59,1):
					mc.append(i.parent.dcu_cr_mc1_misc)
		for i in sv.sockets.computes.cpu.cores.threads.ifu_cr_mc0_status:
			if print_valid(i,a,mc,e):
				if i.bits(58,1): 
					a.append(i.parent.ifu_cr_mc0_addr)
					mc.append(i.parent.ifu_cr_mc0_addr)
				if i.bits(59,1):
					mc.append(i.parent.ifu_cr_mc0_misc)

	for i in sv.sockets.computes.uncore.core_pmsb.core_pmsbs.core_pmsb_instance.pmsb_top.pma_core.error_report:
			if print_valid(i,a,mc,e,b=0):
				mc.append(i.parent.pma_debug)
				mc.append(i.parent.pma_debug2)
				mc.append(i.parent.pma_debug3)
	#for i in sv.sockets.soc.chalogical.chas.util.mc_status:
	for i in sv.sockets.computes.uncore.cha.chas.util.mc_status:
		if print_valid(i,a,mc,e):
			if i.bits(58,1): 
				a.append(i.parent.mc_addr)
			mc.append(i.parent.mc_addr)
			if i.bits(59,1): 
				mc.append(i.parent.mc_misc)
				mc.append(i.parent.mc_misc2)
				mc.append(i.parent.mc_misc3)
	#sometimes the iMC paths not showing in computes..use soc backup method
	try:
		for i in sv.sockets.computes.uncore.memss.mcs.chs.mcchan.imc0_mc_status:
			if print_valid(i,a,mc,e):
				if i.bits(58,1): 
					a.append(i.parent.imc0_mc8_addr)
					mc.append(i.parent.imc0_mc8_addr)
				if i.bits(59,1): 
					mc.append(i.parent.imc0_mc_misc)
	except:
		for i in sv.sockets.soc.memss.mcs.chs.mcchan.imc0_mc_status:
			if print_valid(i,a,mc,e):
				if i.bits(58,1): 
					a.append(i.parent.imc0_mc8_addr)
					mc.append(i.parent.imc0_mc8_addr)
				if i.bits(59,1):
					mc.append(i.parent.imc0_mc_misc)
	try:
		for i in sv.sockets.computes.uncore.memss.b2cmis.mci_status:
			if print_valid(i,a,mc,e):
				if i.bits(58,1): 
					a.append(i.parent.mci_addr)
					mc.append(i.parent.mci_addr)
				if i.bits(59,1):
					mc.append(i.parent.mci_misc)
	except:
		for i in sv.sockets.soc.memss.b2cmis.mci_status:
			if print_valid(i,a,mc,e):
				if i.bits(58,1): 
					a.append(i.parent.mci_addr)
					mc.append(i.parent.mci_addr)
				if i.bits(59,1):
					mc.append(i.parent.mci_misc)
	try:
		for i in sv.sockets.computes.uncore.memss.mcs.chs.mse.mse_mci_status:
			if print_valid(i,a,mc,e):
				if i.bits(58,1): 
					a.append(i.parent.mse_mci_addr)
					mc.append(i.parent.mse_mci_addr)
				if i.bits(59,1):
					mc.append(i.parent.mci_misc)
	except:
		for i in sv.sockets.soc.memss.mcs.chs.mse.mse_mci_status:
			if print_valid(i,a,mc,e):
				if i.bits(58,1): 
					a.append(i.parent.mse_mci_addr)
					mc.append(i.parent.mse_mci_addr)
				if i.bits(59,1):
					mc.append(i.parent.mci_misc)
	for i in sv.sockets.computes.uncore.scf.scf_llc.scf_llcs.mci_status:
		if print_valid(i,a,mc,e):
			if i.bits(58,1): 
				a.append(i.parent.mci_addr)
				mc.append(i.parent.mci_addr)
			if i.bits(59,1):
				mc.append(i.parent.mci_misc)
	try:
		for i in sv.sockets.ios.uncore.upi.upis.upi_regs.kti_mc_st:
			if print_valid(i,a,mc,e):
				if i.bits(58,1): 
					a.append(i.parent.kti_mc_ad)
					mc.append(i.parent.kti_mc_ad)
	except:
		print('WARNING: skipping kti, not working - sv.sockets.ios.uncore.upi.upis.upi_regs.kti_mc_st')
	##cant find MDF in GNR
	#try:
	#    for i in sv.sockets.soc.mdf.mdfvs.mca_status_type:
	#        print_valid(i,a,e)
	#    for i in sv.sockets.soc.mdf.mdfhs.mca_status_type:
	#        print_valid(i,a,e)
	#    for i in sv.sockets.soc.mdf.mdf_multi.mca_status_type:
	#        print_valid(i,a,e)
	#except:
	#    print('WARNING: skipping mdf, not working - sv.sockets.soc.mdf.mdfvs.mca_status_type')
	try:
		for i in sv.sockets.io0.uncore.ubox.ncevents.ncevents_cr_ubox_mci_status:
			print_valid(i,a,mc,e)
			#mc.append(i)
		for i in sv.sockets.io0.uncore.ubox.ncevents.ierrloggingreg:
			print_valid(i,a,mc,e,b=16) #firstierrsrcvalid #GNR changed from 10->16
			#mc.append(i)
		for i in sv.sockets.io0.uncore.ubox.ncevents.ncuevdbgsts:
			mc.append(i)
		for i in sv.sockets.io0.uncore.ubox.ncevents.ncevents_cr_ubox_mci_ctl:
			mc.append(i)
		for i in sv.sockets.io0.uncore.ubox.ncevents.mcerrloggingreg:
			mc.append(i)
		for i in sv.sockets.io0.uncore.ubox.ncdecs.biosnonstickyscratchpad7_cfg:
			mc.append(i)
	except:
		print('WARNING: skipping ubox, not working - sv.sockets.io0.uncore.ubox.ncevents.ncevents_cr_ubox_mci_status')
	try:
		#for i in sv.sockets.soc.punit_multi.ptpcfsms.ptpcfsms.mc_status:
		for i in sv.sockets.computes.uncore.punit.ptpcfsms.ptpcfsms.mc_status:
			if print_valid(i,a,mc,e):
				if i.bits(59,1): 
					a.append(i.parent.mc_misc)
					mc.append(i.parent.mc_misc)
				a.append(i.parent.parent.parent.parent.pcodeio_map.io_firmware_mca_command)
				mc.append(i.parent.parent.parent.parent.pcodeio_map.io_firmware_mca_command)
		for i in sv.sockets.ios.uncore.punit.ptpcfsms.ptpcfsms.mc_status:
			if print_valid(i,a,mc,e):
				if i.bits(59,1): 
					a.append(i.parent.mc_misc)
					mc.append(i.parent.mc_misc)
				a.append(i.parent.parent.parent.parent.pcodeio_map.io_firmware_mca_command)
				mc.append(i.parent.parent.parent.parent.pcodeio_map.io_firmware_mca_command)

	except:
		print('WARNING: skipping punit, not working - sv.sockets.computes.uncore.punit.ptpcfsms.ptpcfsms.mc_status')
	### update for SPR
	## for the DMI HDD silent data corruption
	#if sv.sockets.soc.dmi.conf.miscctrlsts_0.disable_ob_parity_check.read()==0:
	#for i in sv.sockets.soc.dmi.conf.xpuncerrsts:
	#    print("%s= 0x%x" %(i.path, i.read()))
	#    if i.outbound_poisoned_data or i.outbound_switch_fifo_data_parity_error_detected: a.append(i)
	pysvdecode = {'cha':False, 'llc':False, 'ubox':False, 'punit':False, 'pm':False, 'core':False  }
	if mc != []:
		for i in mc:
			mcadata['TestName'].append("%s" % i.path)
			mcadata['TestValue'].append("0x%x" % i.get_value())
			
	if a != []:
		print('\nFOUND VALID MCA')
		for i in a:
			print("%s = 0x%x" %(i.path, i.get_value()))
			if 'cha' in i.path: pysvdecode['cha'] = True
			if 'scf_llc' in i.path: pysvdecode['llc'] = True
			if 'ubox' in i.path: pysvdecode['ubox'] = True
			if 'punit' in i.path: pysvdecode['punit'] = True
			if 'punit' in i.path: pysvdecode['pm'] = True ## Need to check the proper for punit and pm
			if 'cpu.core' in i.path: pysvdecode['core'] = True
			#mcadata['TestName'].append("%s" % i.path)
			#mcadata['TestValue'].append("0x%x" % i.get_value())
			if verbose: print("%s \n"%i.show()) #i.nodes
	else:
		print('did not find valid MCA')
	
	if e != []:
		print('errors found during mca_dump. see above')
	
	sv.sockets.cpu.cores.setaccess('default') #restore
	return mcadata, pysvdecode
	#return a

# Portion of code extracted from core debug
def mca_dump_cwf(verbose=True):

	def print_valid(i,a,mc,e,b=63,save=False):
		try:
			if verbose: print("%s = 0x%x" %(i.path, i.read()))
			if i.bits(b,1): 
				a.append(i)
				mc.append(i)
				
				return True
			else:
				if save: mc.append(i)
				return False
		except:
			#print("Access failed to: {}".format(i.path))
			print("%s = <error reading>" %(i.path))
			e.append(i)
			return False


	import ipccli
	itp = ipccli.baseaccess()
	unlock()
	if not sv.sockets: sv.refresh()
	a=[] #mcas
	e=[] #errors
	mc=[]
	mcadata = {'TestName':[],	'TestValue':[]}

	sv.sockets.computes.cpu.modules.setaccess('crb') #no halt required
	## Changed Core data request to use the compute register of directly from cpu.core, this in order to keep format
	## there shouldnt be any difference in both registers, will keep looking for a similar approach in the future, for now this works.

	with itp.device_locker():
		for i in sv.sockets.computes.cpu.modules.cores.mec_cr_mci_status: ## MEC Bank MCE Status dump
			if print_valid(i,a,mc,e):
				for th in i.parent.threads:
					mc.append(th.mec_cr_mcg_contain)
					#mc.append(th.msrs.mec_cr_mci_misc)
					if i.bits(59,1): 
						mc.append(th.mec_cr_mci_misc)
				if i.bits(58,1): 
					a.append(i.parent.mec_cr_mci_addr)
					mc.append(i.parent.mec_cr_mci_addr)
				
		for i in sv.sockets.computes.cpu.modules.bus_cr_mci_status: ## BUS Bank MCE Status dump
			if print_valid(i,a,mc,e):
				if i.bits(58,1): 
					a.append(i.parent.bus_cr_mci_addr) 
					mc.append(i.parent.bus_cr_mci_addr)
				if i.bits(59,1):
					mc.append(i.parent.bus_cr_mci_misc)
		for i in sv.sockets.computes.cpu.modules.bbl_cr_mci_status: ## BBL Bank MCE Status dump there is no Addr/misc
			if print_valid(i,a,mc,e):
				pass
				#if i.bits(58,1): 
				#	a.append(i.parent.dcu_cr_mc1_addr)
				#	mc.append(i.parent.dcu_cr_mc1_addr)
				#if i.bits(59,1):
				#	mc.append(i.parent.dcu_cr_mc1_misc)
		for i in sv.sockets.computes.cpu.modules.cores.ic_cr_mci_status: ## IC Bank MCE Status dump
			if print_valid(i,a,mc,e):
				if i.bits(58,1): 
					a.append(i.parent.ic_cr_mci_addr)
					mc.append(i.parent.ic_cr_mci_addr)
				if i.bits(59,1):
					mc.append(i.parent.ic_cr_mci_misc)

		for i in sv.sockets.computes.cpu.modules.l2_cr_mci_status: ## L2 Bank MCE Status dump
			if print_valid(i,a,mc,e):
				if i.bits(58,1): 
					a.append(i.parent.l2_cr_mci_addr)
					mc.append(i.parent.l2_cr_mci_addr)
				if i.bits(59,1):
					mc.append(i.parent.l2_cr_mci_misc)
	## There is no PMSB for AtomCores
	#for i in sv.sockets.computes.uncore.core_pmsb.core_pmsbs.core_pmsb_instance.pmsb_top.pma_core.error_report:
	#		print_valid(i,a,mc,e,b=0)
	#		mc.append(i.parent.pma_debug)
	#		mc.append(i.parent.pma_debug2)
	#		mc.append(i.parent.pma_debug3)
	#for i in sv.sockets.soc.chalogical.chas.util.mc_status:
	for i in sv.sockets.computes.uncore.cha.chas.util.mc_status:
		if print_valid(i,a,mc,e):
			if i.bits(58,1): 
				a.append(i.parent.mc_addr)
			mc.append(i.parent.mc_addr)
			if i.bits(59,1): 
				mc.append(i.parent.mc_misc)
				mc.append(i.parent.mc_misc2)
				mc.append(i.parent.mc_misc3)
	#sometimes the iMC paths not showing in computes..use soc backup method
	try:
		for i in sv.sockets.computes.uncore.memss.mcs.chs.mcchan.imc0_mc_status:
			if print_valid(i,a,mc,e):
				if i.bits(58,1): 
					a.append(i.parent.imc0_mc8_addr)
					mc.append(i.parent.imc0_mc8_addr)
				if i.bits(59,1): 
					mc.append(i.parent.imc0_mc_misc)
	except:
		for i in sv.sockets.soc.memss.mcs.chs.mcchan.imc0_mc_status:
			if print_valid(i,a,mc,e):
				if i.bits(58,1): 
					a.append(i.parent.imc0_mc8_addr)
					mc.append(i.parent.imc0_mc8_addr)
				if i.bits(59,1):
					mc.append(i.parent.imc0_mc_misc)
	try:
		for i in sv.sockets.computes.uncore.memss.b2cmis.mci_status:
			if print_valid(i,a,mc,e):
				if i.bits(58,1): 
					a.append(i.parent.mci_addr)
					mc.append(i.parent.mci_addr)
				if i.bits(59,1):
					mc.append(i.parent.mci_misc)
	except:
		for i in sv.sockets.soc.memss.b2cmis.mci_status:
			if print_valid(i,a,mc,e):
				if i.bits(58,1): 
					a.append(i.parent.mci_addr)
					mc.append(i.parent.mci_addr)
				if i.bits(59,1):
					mc.append(i.parent.mci_misc)
	try:
		for i in sv.sockets.computes.uncore.memss.mcs.chs.mse.mse_mci_status:
			if print_valid(i,a,mc,e):
				if i.bits(58,1): 
					a.append(i.parent.mse_mci_addr)
					mc.append(i.parent.mse_mci_addr)
				if i.bits(59,1):
					mc.append(i.parent.mci_misc)
	except:
		for i in sv.sockets.soc.memss.mcs.chs.mse.mse_mci_status:
			if print_valid(i,a,mc,e):
				if i.bits(58,1): 
					a.append(i.parent.mse_mci_addr)
					mc.append(i.parent.mse_mci_addr)
				if i.bits(59,1):
					mc.append(i.parent.mci_misc)
	for i in sv.sockets.computes.uncore.scf.scf_llc.scf_llcs.mci_status:
		if print_valid(i,a,mc,e):
			if i.bits(58,1): 
				a.append(i.parent.mci_addr)
				mc.append(i.parent.mci_addr)
			if i.bits(59,1):
				mc.append(i.parent.mci_misc)
	try:
		for i in sv.sockets.ios.uncore.upi.upis.upi_regs.kti_mc_st:
			if print_valid(i,a,mc,e):
				if i.bits(58,1): 
					a.append(i.parent.kti_mc_ad)
					mc.append(i.parent.kti_mc_ad)
	except:
		print('WARNING: skipping kti, not working - sv.sockets.ios.uncore.upi.upis.upi_regs.kti_mc_st')
	##cant find MDF in GNR
	#try:
	#    for i in sv.sockets.soc.mdf.mdfvs.mca_status_type:
	#        print_valid(i,a,e)
	#    for i in sv.sockets.soc.mdf.mdfhs.mca_status_type:
	#        print_valid(i,a,e)
	#    for i in sv.sockets.soc.mdf.mdf_multi.mca_status_type:
	#        print_valid(i,a,e)
	#except:
	#    print('WARNING: skipping mdf, not working - sv.sockets.soc.mdf.mdfvs.mca_status_type')
	try:
		for i in sv.sockets.io0.uncore.ubox.ncevents.ncevents_cr_ubox_mci_status:
			print_valid(i,a,mc,e)
			#mc.append(i)
		for i in sv.sockets.io0.uncore.ubox.ncevents.ierrloggingreg:
			print_valid(i,a,mc,e,b=16) #firstierrsrcvalid #GNR changed from 10->16
			#mc.append(i)
		for i in sv.sockets.io0.uncore.ubox.ncevents.ncuevdbgsts:
			mc.append(i)
		for i in sv.sockets.io0.uncore.ubox.ncevents.ncevents_cr_ubox_mci_ctl:
			mc.append(i)
		for i in sv.sockets.io0.uncore.ubox.ncevents.mcerrloggingreg:
			mc.append(i)
		for i in sv.sockets.io0.uncore.ubox.ncdecs.biosnonstickyscratchpad7_cfg:
			mc.append(i)
	except:
		print('WARNING: skipping ubox, not working - sv.sockets.io0.uncore.ubox.ncevents.ncevents_cr_ubox_mci_status')
	try:
		#for i in sv.sockets.soc.punit_multi.ptpcfsms.ptpcfsms.mc_status:
		for i in sv.sockets.computes.uncore.punit.ptpcfsms.ptpcfsms.mc_status:
			if print_valid(i,a,mc,e):
				if i.bits(59,1): 
					a.append(i.parent.mc_misc)
					mc.append(i.parent.mc_misc)
				a.append(i.parent.parent.parent.parent.pcodeio_map.io_firmware_mca_command)
				mc.append(i.parent.parent.parent.parent.pcodeio_map.io_firmware_mca_command)
		for i in sv.sockets.ios.uncore.punit.ptpcfsms.ptpcfsms.mc_status:
			if print_valid(i,a,mc,e):
				if i.bits(59,1): 
					a.append(i.parent.mc_misc)
					mc.append(i.parent.mc_misc)
				a.append(i.parent.parent.parent.parent.pcodeio_map.io_firmware_mca_command)
				mc.append(i.parent.parent.parent.parent.pcodeio_map.io_firmware_mca_command)

	except:
		print('WARNING: skipping punit, not working - sv.sockets.computes.uncore.punit.ptpcfsms.ptpcfsms.mc_status')
	### update for SPR
	## for the DMI HDD silent data corruption
	#if sv.sockets.soc.dmi.conf.miscctrlsts_0.disable_ob_parity_check.read()==0:
	#for i in sv.sockets.soc.dmi.conf.xpuncerrsts:
	#    print("%s= 0x%x" %(i.path, i.read()))
	#    if i.outbound_poisoned_data or i.outbound_switch_fifo_data_parity_error_detected: a.append(i)
	pysvdecode = {'cha':False, 'llc':False, 'ubox':False, 'punit':False, 'pm':False, 'core':False  }
	if mc != []:
		for i in mc:
			mcadata['TestName'].append("%s" % i.path)
			mcadata['TestValue'].append("0x%x" % i.get_value())
	if a != []:
		print('\nFOUND VALID MCA')
		for i in a:
			print("%s = 0x%x" %(i.path, i.get_value()))
			if 'cha' in i.path: pysvdecode['cha'] = True
			if 'scf_llc' in i.path: pysvdecode['llc'] = True
			if 'ubox' in i.path: pysvdecode['ubox'] = True
			if 'punit' in i.path: pysvdecode['punit'] = True
			if 'punit' in i.path: pysvdecode['pm'] = True ## Need to check the proper for punit and pm
			if 'cpu.module' in i.path: pysvdecode['core'] = True
			#mcadata['TestName'].append("%s" % i.path)
			#mcadata['TestValue'].append("0x%x" % i.get_value())
			if verbose: print("%s \n"%i.show()) #i.nodes
	else:
		print('did not find valid MCA')
	
	if e != []:
		print('errors found during mca_dump. see above')
	
	sv.sockets.computes.cpu.modules.setaccess('default') #restore
	return mcadata, pysvdecode
	#return a

def readscratchpad():
	scratchpad = str(sv.socket0.io0.uncore.ubox.ncdecs.biosnonstickyscratchpad7_cfg)
	return scratchpad

def unlock():
	was_locked=False
	import ipccli
	base_ipc = ipccli.baseaccess()
	for uncore in base_ipc.chipleveltaps:
		if "MTP" in uncore.name: continue #filter out PCH for GNR-WS (MTP0_CLTAP0)
		if base_ipc.islocked(uncore.name): 
			was_locked=True
			base_ipc.unlock(uncore.name)
			if base_ipc.islocked(uncore.name): #ensure unlocked
				print('Can not unlock %s..' % uncore.name)
				raise 
	return was_locked