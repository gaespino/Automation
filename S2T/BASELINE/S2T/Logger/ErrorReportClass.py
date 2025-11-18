"""
ErrorReport Class - Refactored for Multi-Product Support
Author: jfnavarr/gaespino
Last update: 17/11/25

Refactored class-based implementation that supports multiple products (GNR, CWF, DMR)
Uses product-specific configurations and mca_banks for:
- MCA decoding and reporting
- TileView generation
- Register reading and masking
- Excel report generation
"""

import openpyxl
import zipfile
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
import os
import sys
import contextlib
import random
from importlib import import_module

from namednodes import sv
from ipccli.stdiolog import log
from ipccli.stdiolog import nolog
import ipccli


# ANSI Color codes for console output (Python 3.8+ compatible)
class Colors:
	"""ANSI color codes for terminal output."""
	GREEN = '\033[92m'
	RED = '\033[91m'
	BLUE = '\033[94m'
	CYAN = '\033[96m'
	YELLOW = '\033[93m'
	MAGENTA = '\033[95m'
	BOLD = '\033[1m'
	RESET = '\033[0m'
	
	@staticmethod
	def success(text):
		return f"{Colors.GREEN}{text}{Colors.RESET}"
	
	@staticmethod
	def error(text):
		return f"{Colors.RED}{text}{Colors.RESET}"
	
	@staticmethod
	def info(text):
		return f"{Colors.CYAN}{text}{Colors.RESET}"
	
	@staticmethod
	def warning(text):
		return f"{Colors.YELLOW}{text}{Colors.RESET}"


class ErrorReportGenerator:
	"""
	Multi-product error report generator for hardware debugging.
	Supports GNR, CWF, and DMR products with product-specific configurations.
	"""
		
	def __init__(self, product='GNR', variant='AP', config_loader=None, logger=None, 
				core_manipulation=None, file_handler=None, dpm_checks=None):
		"""
		Initialize ErrorReportGenerator with product-specific configurations.
		
		Args:
			product (str): Product name ('GNR', 'CWF', 'DMR')
			variant (str): Product variant ('AP', 'SP', etc.)
			config_loader: Configuration loader object (ConfigsLoader)
			logger: Logger instance for debug output
			core_manipulation: CoreManipulation module for dragon dumps
			file_handler: FileHandler module for data uploads
			dpm_checks: dpmChecks module for fuse/voltage operations
		"""
		self.product = product.upper()
		self.variant = variant.upper()
		self.logger = logger
		self.config_loader = config_loader
		
		# Store external module dependencies
		self.core_manipulation = core_manipulation
		self.file_handler = file_handler
		self.dpm_checks = dpm_checks
		
		# Initialize IPC connection
		self.ipc = ipccli.baseaccess()
		
		# Load product-specific configurations and modules
		self._load_product_configs()
		self._load_product_modules()
		
		# Framework variables
		self.framework_vars = self._init_framework_vars()
		
		print(f'{"+"*20} ErrorReport initialized for {self.product}:{self.variant} {"+"*20}')

	def _load_product_configs(self):
		"""Load product-specific configuration parameters."""
		try:
			# Determine base path based on dev mode flag
			dev_mode = getattr(self.config_loader, 'DEV_MODE', True) if self.config_loader else True
			
			if dev_mode:
				self.base_path = 'users.gaespino.dev'
			else:
				# Get base path from config_loader, fallback to THR path
				self.base_path = getattr(self.config_loader, 'BASE_PATH', 'users.THR.PythonScript.thr')
			
			# Import product-specific configs module
			configs_module = import_module(f'{self.base_path}.S2T.product_specific.{self.product.lower()}.configs')
			self.product_configs = configs_module.configurations(self.product)
			
			# Get product-specific settings
			self.config = self.product_configs.init_product_specific()
			self.fuse_config = self.product_configs.init_product_fuses()
			
			# Extract commonly used config values
			self.core_string = self.config.get('CORESTRING', 'CORE')
			self.cha_string = self.config.get('CHASTRING', 'CHA')
			self.max_physical = self.config.get('MAXPHYSICAL', 60)
			self.max_logical = self.config.get('MAXLOGICAL', 60)
			self.log2phy = self.config.get('LOG2PHY', {})
			self.phy2log = self.config.get('PHY2LOG', {})
			
			print(f"  {Colors.success('[+]')} Loaded product configuration for {Colors.BOLD}{self.product}{Colors.RESET} (base: {self.base_path})")
			
		except Exception as e:
			print(f"  {Colors.error('[X]')} Error loading product configs: {e}")
			# Set defaults
			self.base_path = 'users.gaespino.dev.S2T'
			self.core_string = 'CORE'
			self.cha_string = 'CHA'
			self.max_physical = 60
			self.max_logical = 60
			self.log2phy = {}
			self.phy2log = {}
	
	def _load_product_modules(self):
		"""Dynamically load product-specific modules (imports, decoders, etc.)."""
		try:
			# Load MCA banks module using base_path
			mca_module = import_module(f'{self.base_path}.S2T.product_specific.{self.product.lower()}.mca_banks')
			self.mca_banks = mca_module
			self.mca_dump_func = getattr(mca_module, 'mca_dump', None)
			
			# Instantiate the MCADecoders class
			mca_decoders_class = getattr(mca_module, 'MCADecoders', None)
			if mca_decoders_class:
				self.mca_decoders = mca_decoders_class()
				print(f"  {Colors.success('[+]')} Loaded MCA decoders for {Colors.BOLD}{self.product}{Colors.RESET}")
			else:
				print(f"  {Colors.warning('[!]')} MCADecoders class not found in {self.product} mca_banks")
				self.mca_decoders = None
			
			print(f"  {Colors.success('[+]')} Loaded MCA banks for {Colors.BOLD}{self.product}{Colors.RESET}")
			
		except Exception as e:
			print(f"  {Colors.error('[X]')} Error loading MCA banks: {e}")
			self.mca_banks = None
			self.mca_dump_func = None
			self.mca_decoders = None
	
	def _init_framework_vars(self):
		"""Initialize framework variables dictionary."""
		return {
			'qdf': '', 
			'tnum': '', 
			'mask': '', 
			'corelic': '', 
			'bumps': '', 
			'htdis': '', 
			'dis2CPM': '', 
			'dis1CPM': '',
			'freqIA': '', 
			'voltIA': '', 
			'freqCFC': '', 
			'voltCFC': '', 
			'content': '', 
			'passstring': '', 
			'failstring': '', 
			'runName': '', 
			'runStatus': '', 
			'scratchpad': '', 
			'seed': '',
			'ttlog': None
		}
	
	# =================================================================
	# WORKBOOK OPERATIONS
	# =================================================================
		
	def load_workbook(self):
		"""Load the workbook template for the current product."""
		try:
			script_home = os.path.dirname(os.path.realpath(__file__))
			parent_dir = os.path.dirname(script_home)
			template_path = os.path.join(
				parent_dir, 
				'product_specific', 
				f'{self.product.lower()}', 
				'LoggerFiles', 
				'TileView_Template.xlsx'
			)
			
			workbook = openpyxl.load_workbook(template_path)
			print(f"  {Colors.success('[+]')} Workbook loaded from: {Colors.BOLD}{template_path}{Colors.RESET}")
			return workbook
			
		except Exception as e:
			print(f"  {Colors.error('[X]')} Error loading workbook: {e}")
			return None

	def save_workbook(self, workbook, filepath):
		"""Save the workbook to specified filepath."""
		try:
			workbook.save(filepath)
			print(f"  {Colors.success('[+]')} Workbook saved to: {Colors.BOLD}{filepath}{Colors.RESET}")
			return True
		except Exception as e:
			print(f"  {Colors.error('[X]')} Error saving workbook: {e}")
			return False
	
	def find_value(self, sheet, value):
		"""Find the cell where a value is located in the sheet."""
		if value == 'ubox':
			value = 'ubox_0'
		
		for row in sheet.iter_rows():
			for cell in row:
				if cell.value == value:
					return cell
		return None
	
	def add_error(self, sheet, cell, error):
		"""Mark and write an error into the sheet."""
		# Mark error cell
		if cell is not None:
			error_color = "FFA3A3"  # light red
			fill = PatternFill(start_color=error_color, end_color=error_color, fill_type="solid")
			cell.fill = fill
		
		# Write error to error log cell
		current_error_content = sheet["X8"].value
		if not current_error_content:
			sheet["X8"] = error
		else:
			current_error_content += ("\n" + error)
			sheet["X8"] = current_error_content
	
	def mark_disable(self, cell, able=0):
		"""Mark disabled IP's in the sheet."""
		if able:
			color = "00FFC000"  # yellow
		else:
			color = "AEAAAA"  # gray
		fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
		cell.fill = fill
	
	# =================================================================
	# MASK OPERATIONS
	# =================================================================
	
	def get_mask(self, mask1, mask2, mask3=None):
		"""Combine mask parts into binary and hex format."""
		_mask1 = mask1 if mask1 is not None else 0xFFFFFFFFFFFFFFF
		_mask2 = mask2 if mask2 is not None else 0xFFFFFFFFFFFFFFF
		_mask3 = mask3 if mask3 is not None else 0xFFFFFFFFFFFFFFF
		
		mask1_str = hex(_mask1)[2:].zfill(15)
		mask2_str = hex(_mask2)[2:].zfill(15)
		mask3_str = hex(_mask3)[2:].zfill(15)
		
		# Combine all parts
		mask = mask3_str + mask2_str + mask1_str
		hex_mask = '0x' + mask
		
		# Convert to binary
		bin_mask = bin(int(mask, 16))[2:]
		bin_mask = bin_mask.zfill(self.max_physical)
		
		return bin_mask, hex_mask
	
	def common_mask(self, mask1, mask2):
		"""Perform AND operation between two masks."""
		mask1_int = int(mask1, 2)
		mask2_int = int(mask2, 2)
		
		common_mask_int = mask1_int & mask2_int
		common_mask = bin(common_mask_int)[2:]
		common_mask = common_mask.zfill(self.max_physical)
		
		return common_mask
	
	def get_m2iosf_mask(self, m2iosf_mask):
		"""Convert m2iosf mask to binary number."""
		bin_m2iosf_mask = bin(int(m2iosf_mask))[2:]
		bin_m2iosf_mask = bin_m2iosf_mask.zfill(9)
		return bin_m2iosf_mask
	
	def review_mask(self, mask, IP, sheet):
		"""Review mask to find disabled IPs and mark them in sheet."""
		# Get product-specific IP string
		core_ip_map = {
			'GNRAP': 'CORE', 
			'CWFAP': 'MOD',
			'GNRSP': 'CORE', 
			'CWFSP': 'MOD',
			'DMRAP': 'MOD',
			'DMRSP': 'MOD'
		}
		product_key = f"{self.product}{self.variant}"
		productip = core_ip_map.get(product_key, self.core_string)
		
		if IP == 'm2iosf':
			ip_num = 0
			cell_names = ['M2IOSF8', 'M2IOSF1', 'M2IOSF2', 'M2IOSF3', 'M2IOSF9', 
						  'M2IOSF10', 'M2IOSF5', 'M2IOSF0', 'M2IOSF4', 'M2IOSF11']
			
			for ip_val in reversed(mask):
				if ip_val == '1':
					cell_name = cell_names[ip_num]
					cell = self.find_value(sheet, cell_name)
					if cell:
						self.mark_disable(cell)
				
				ip_num += 1
				if ip_num > 9:
					break
		
		else:  # for core, cha and core/cha
			ip_num = 0
			
			for ip_val in reversed(mask):
				if ip_val == '1':
					if IP == 'core_cha':
						full_cell_name = productip + str(ip_num)
						cell = self.find_value(sheet, full_cell_name)
						if cell:
							cell = cell.offset(row=0, column=-1)
					else:
						full_cell_name = IP + str(ip_num)
						cell = self.find_value(sheet, full_cell_name)
					
					if cell:
						self.mark_disable(cell)
				
				ip_num += 1
	
	# =================================================================
	# MCA OPERATIONS
	# =================================================================
	
	def mca_init(self):
		"""Initialize MCA operations - unlock and refresh."""
		self.ipc.unlock()
		if not sv.sockets:
			sv.refresh()
		return self.ipc
	
	def mca_dump(self, verbose=True):
		"""
		Perform product-specific MCA dump.
		Delegates to product-specific mca_dump function from mca_banks module.
		"""
		if self.mca_dump_func is None:
			print(f"[X] MCA dump not available for {self.product}")
			return {}, {}
		
		try:
			itp = self.mca_init()
			mcadata, pysvdecode = self.mca_dump_func(sv, itp=itp, verbose=verbose)
			return mcadata, pysvdecode
		except Exception as e:
			print(f"[X] Error during MCA dump: {e}")
			return {}, {}
	
	def decode_mca_bank(self, register_path, status_value=None):
		"""
		Decode MCA bank information from register path and status value.
		Uses product-specific MCA bank decoder if available.
		"""
		if self.mca_banks is None:
			return ""
		
		try:
			return self.mca_banks.decode_mca_bank(register_path, status_value)
		except:
			return ""
	
	def check_MCA(self, sheet, mca_path, mca_debug_path, debug_mca_log, 
				  mca_report_path, dr_dump, visual='', loop='', WW='WW', 
				  Bucket='UNCORE', QDF='', bindesc=''):
		"""
		Read and save MCA registers, mark them in workbook.
		"""
		mcadata = {'TestName': [], 'TestValue': []}
		pysvdecode = {'cha': False, 'ubox': False, 'punit': False, 
					  'pm': False, 'llc': False, 'core': False}
			
		log(mca_path)
		
		try:
			print(f"{Colors.info('[*]')} Reading MCA...")
			mcadata, pysvdecode = self.mca_dump()
			print(f"  {Colors.success('[+]')} {Colors.BOLD}MCA READ SUCCESSFULLY{Colors.RESET}")
		except Exception as e:
			print(f"  {Colors.error('[X]')} {Colors.BOLD}MCA COULD NOT BE READ:{Colors.RESET} {e}")
			nolog()
			return
		
		# Decode MCAs using pysvtools
		try:
			if self.mca_decoders:
				if pysvdecode['cha'] and hasattr(self.mca_decoders, 'cha'):
					print(f"{Colors.info('[*]')} Decoding CHA MCA...")
					self.mca_decoders.cha.Cha().show_mca_status()
				if pysvdecode['ubox'] and hasattr(self.mca_decoders, 'ubox'):
					print(f"{Colors.info('[*]')} Decoding Ubox MCA...")
					self.mca_decoders.ubox.Ubox().show_mca_status()
				if pysvdecode['punit'] and hasattr(self.mca_decoders, 'punit'):
					print(f"{Colors.info('[*]')} Decoding Punit MCA...")
					self.mca_decoders.punit.Punit().show_mca_status()
				if pysvdecode['pm'] and hasattr(self.mca_decoders, 'pm'):
					print(f"{Colors.info('[*]')} Decoding PM MCA...")
					self.mca_decoders.pm.PM().show_mca_status(source='reg')
			if pysvdecode['llc'] and hasattr(self.mca_decoders, 'llc'):
				print(f"{Colors.info('[*]')} Decoding LLC MCA...")
				self.mca_decoders.llc.Llc().show_mca_status()
			
			print(f"  {Colors.success('[+]')} {Colors.BOLD}MCA DECODE SUCCESSFUL{Colors.RESET}")
		except Exception as e:
			print(f"  {Colors.error('[X]')} {Colors.BOLD}MCA DECODE FAILED:{Colors.RESET} {e}")
		
		nolog()
		
		# Dragon dump if enabled
		if self._is_dragon_dump_enabled() and dr_dump:
			try:
				print(f"{Colors.info('[*]')} DUMPING DRs DATA INTO MCA FILE")
				self.dragon_dump(mca_path, mcadata)
			except Exception as e:
				print(f"  {Colors.error('[X]')} ERROR DUMPING DRs DATA: {e}")
		
		# Generate MCA report
		mcreport = self.mca_report(
			mcadata=mcadata, visual=visual, lot=f'RVP{loop}', 
			WW=WW, Bucket=Bucket, QDF=QDF, bindesc=bindesc, lotseq=loop
		)
		self.mca_file(mca_report_path, mcreport)
		
		# Debug MCA logging
		log(mca_debug_path)
			
		if debug_mca_log and self.mca_decoders and self.mca_decoders.debug_mca:
			if self.product == 'CWF':
				print(f"{Colors.warning('[!]')} Debug MCA script not available for CWF")
			else:
				try:
					self.mca_decoders.debug_mca.analyze(scope=r'socket\d\.', variant=self.variant)
					print(f"  {Colors.success('[+]')} {Colors.BOLD}MCA DEBUG READ SUCCESSFULLY{Colors.RESET}")
				except Exception as e:
					print(f"  {Colors.error('[X]')} {Colors.BOLD}MCA DEBUG FAILED:{Colors.RESET} {e}")
		else:
			print(f"{Colors.warning('[!]')} Debug MCA Script disabled by default")
		
		nolog()
		
		# Parse and mark MCAs in workbook
		self._parse_mca_log_and_mark_sheet(mca_path, sheet)

	def _is_dragon_dump_enabled(self):
		"""Check if dragon dump is enabled for current product."""
		# Currently only enabled for GNR
		return self.product in ['GNR', 'CWF']
	
	def _parse_mca_log_and_mark_sheet(self, mca_path, sheet):
		"""Parse MCA log file and mark errors in sheet."""
		try:
			with open(mca_path, 'r') as file:
				MCAs = file.readlines()
			
			# Get product-specific IP string
			core_ip_map = {
				'GNRAP': 'CORE', 
				'CWFAP': 'MOD',
				'GNRSP': 'CORE', 
				'CWFSP': 'MOD',
				'DMRAP': 'MOD',
				'DMRSP': 'MOD'
			}
			product_key = f"{self.product}{self.variant}"
			productip = core_ip_map.get(product_key, self.core_string)
			
			if 'FOUND VALID MCA\n' in MCAs:
				# Mark FAILED
				sheet["X2"] = 'FAILED'
				cell_color = "FF0000"
				
				# Look for MCAs
				start_search = MCAs.index('FOUND VALID MCA\n')
				for num_line in range(start_search, len(MCAs)):
					line = MCAs[num_line]
					
					if line and line[0] == 's':
						parts = line.split(".")
						if len(parts) > 2:
							IP = parts[2]
							if IP and IP[0] == 'c':
								IP = productip + IP[4:]
							if len(parts) > 3 and parts[3] == 'cha':
								IP = parts[4]
								IP = 'CHA' + IP[3:]
							cell = self.find_value(sheet, IP)
							self.add_error(sheet, cell, line)
			else:
				# Mark PASS
				sheet["X2"] = 'PASS'
				cell_color = "00B050"
				print('Did not find valid MCA')
		
			# Color status cells
			fill = PatternFill(start_color=cell_color, end_color=cell_color, fill_type="solid")
			sheet["X2"].fill = fill
			sheet["B66"].fill = fill
			
			print(f"  {Colors.success('[+]')} {Colors.BOLD}MCA WRITTEN SUCCESSFULLY{Colors.RESET}")
				
		except Exception as e:
			print(f"  {Colors.error('[X]')} {Colors.BOLD}COULD NOT WRITE MCA:{Colors.RESET} {e}")

	def dragon_dump(self, logging, mcadata=None):
		"""Perform dragon DR collection (GNR specific)."""
		if self.core_manipulation is None:
			print(f"  {Colors.warning('[!]')} CoreManipulation module not provided")
			return
		
		log(logging, 'a')
		try:
			self.core_manipulation._core_dr_registers(self.logger, mcadata)
		except Exception as e:
			print(f"  {Colors.error('[X]')} Dragon dump error: {e}")
		nolog()

	def mca_report(self, mcadata, visual, lot='RVP', operation='DPM', 
				   WW='WW', Bucket='UNCORE', QDF='', lotseq=None, 
				   seqkey=None, bindesc=''):
		"""Generate MCA report data structure."""
		mcreport = {
			'VisualId': [], 'Lot': [], 'LatoStartWW': [], 
			'LotsSeqKey': [], 'UnitTestingSeqKey': [], 
			'TestName': [], 'TestValue': [], 'Operation': [], 
			'TestNameNumber': [], 'TestNameWithoutNumeric': []
		}
		
		# Generate unique keys
		if lotseq is None:
			lotseq = random.getrandbits(12)
		if seqkey is None:
			seqkey = random.getrandbits(4)
		
		# Fill report data
		if mcadata['TestName'] != []:
			for i, mca in enumerate(mcadata['TestName']):
				mcreport['VisualId'].append(visual)
				mcreport['Lot'].append(lot)
				mcreport['LotsSeqKey'].append(lotseq)
				mcreport['UnitTestingSeqKey'].append(seqkey)
				mcreport['LatoStartWW'].append(WW)
				mcreport['TestName'].append(mca.replace(".", "__").upper())
				mcreport['TestValue'].append(mcadata['TestValue'][i])
				mcreport['Operation'].append(operation)
				mcreport['TestNameNumber'].append('0')
				mcreport['TestNameWithoutNumeric'].append(mca)
		
		results = {
			'VisualId': [visual],
			'Lot': [lot],
			'LatoStartWW': [WW],
			'LotsSeqKey': [lotseq],
			'UnitTestingSeqKey': [''],
			'DpmBucket': [Bucket],
			'DecimaSite': ['CR'],
			'DecimaWW': [WW],
			'DecimaBucket': [Bucket],
			'Accuracy': [''],
			'Operation': [operation],
			'ProductConfigurationName': [self.product],
			'Program': [''],
			'SSPEC': [QDF],
			'DevRevStep': [''],
			'TestName': [mcadata['TestName'][0] if mcadata['TestName'] != [] else ''],
			'TestValue': [mcadata['TestValue'][0] if mcadata['TestName'] != [] else ''],
			'DB': [''],
			'BinDesc': [bindesc],
		}
		
		final_bin = {
			'VisualId': [visual],
			'DpmBucket': [Bucket],
			'DecimaSite': ['CR'],
			'DecimaWW': [WW],
			'DecimaBucket': [Bucket],
			'DpmBucketAccuracy': [''],
			'ProductConfigurationName': [self.product]
		}
		
		return {'raw_data': mcreport, 'final_bucket': final_bin, 'results': results}
	
	def mca_file(self, pathfile, data):
		"""Save MCA report to Excel file."""
		try:
			dfraw = pd.DataFrame(data['raw_data'])
			dffinal = pd.DataFrame(data['final_bucket'])
			dfresults = pd.DataFrame(data['results'])
			
			with pd.ExcelWriter(pathfile) as writer:
				dfraw.to_excel(writer, sheet_name='raw_data', index=False)
				dffinal.to_excel(writer, sheet_name='final_bucket', index=False)
				dfresults.to_excel(writer, sheet_name='results', index=False)
			
			print(f"[+] MCA report saved to: {pathfile}")
		except Exception as e:
			print(f"[X] Error saving MCA file: {e}")
	
	# =================================================================
	# REGISTER AND SYSTEM OPERATIONS
	# =================================================================
	
	def check_mem(self, sheet, config_path):
		"""Check memory configuration and mark in workbook."""
		if self.product != 'GNR' or not self.mca_decoders or self.mca_decoders.dimm_info is None:
			print(f"Memory check not available for {self.product}")
			return
		
		log(config_path)
		self.ipc.halt()
		
		try:
			self.mca_decoders.dimm_info.gnr_dimm_info()
		except Exception as e:
			print(f"[X] Error reading DIMM info: {e}")
		
		self.ipc.go()
		nolog()
		
		# Parse and mark memory configuration
		try:
			with open(config_path, 'r') as file:
				mem = file.readlines()
			
			slt_position = []
			for position, value in enumerate(mem):
				if value[:6] == '  0  |':
					slt_position.append(position)
			
			# Review each MC
			count = 0
			for MC in slt_position:
				cell_name = 'MC' + str(count)
				cell = self.find_value(sheet, cell_name)
				
				if cell:
					# Move to channel columns
					next_column_index = cell.column + 1
					next_column = get_column_letter(next_column_index)
					ch0_cell = sheet[next_column + str(cell.row)]
					ch1_cell = sheet[next_column + str(cell.row + 1)]
					
					# Check channel status
					ch0 = mem[MC][5:63]
					ch1 = mem[MC][62:120]
					
					# Mark CH0
					if ch0 == '|                     Not installed                      |':
						self.mark_disable(ch0_cell)
					else:
						self.mark_disable(ch0_cell, 1)
					
					# Mark CH1
					if ch1 == '|                     Not installed                      |':
						self.mark_disable(ch1_cell)
					else:
						self.mark_disable(ch1_cell, 1)
					
					# Check if full MC is disabled
					ch0_color = Color(ch0_cell.fill.start_color.rgb)
					ch1_color = Color(ch1_cell.fill.start_color.rgb)
					if ch0_color.rgb == "00AEAAAA" and ch1_color.rgb == "00AEAAAA":
						self.mark_disable(cell)
					if ch0_color.rgb == "00FFC000" and ch1_color.rgb == "00FFC000":
						self.mark_disable(cell, 1)
				
				count += 1
			
			print("[+] Memory configuration marked successfully")
			
		except Exception as e:
			print(f"[X] Error parsing memory config: {e}")
	
	def read_scratchpad(self):
		"""Read scratchpad register value."""
		try:
			scratchpad = str(sv.socket0.io0.uncore.ubox.ncdecs.biosnonstickyscratchpad7_cfg)
			return scratchpad
		except Exception as e:
			print(f"[X] Error reading scratchpad: {e}")
			return "ERROR"
	
	def unlock(self):
		"""Unlock all chip level taps."""
		was_locked = False
		base_ipc = self.ipc
		
		for uncore in base_ipc.chipleveltaps:
			if "MTP" in uncore.name:
				continue  # Filter out PCH for GNR-WS
			if base_ipc.islocked(uncore.name):
				was_locked = True
				base_ipc.unlock(uncore.name)
				if base_ipc.islocked(uncore.name):
					print(f'Cannot unlock {uncore.name}')
					raise Exception(f"Failed to unlock {uncore.name}")
		
		return was_locked
	
	# =================================================================
	# PATH AND FILE OPERATIONS
	# =================================================================
	
	def define_paths(self, folder, test):
		"""Define all file paths for the test."""
		if folder[-1] != "\\":
			folder = folder + "\\"
		
		if test == '':
			file_path = folder + "TileView.xlsx"
			mca_path = folder + "MCAs.txt"
			mca_debug_path = folder + "MCA_debug.txt"
			mca_report_path = folder + "MCA_Report.xlsx"
			config_path = folder + "config.txt"
			zip_path = folder + "log.zip"
		else:
			file_path = folder + test + "_TileView" + ".xlsx"
			mca_path = folder + test + "_MCAs" + ".txt"
			mca_debug_path = folder + test + "_MCA_debug" + ".txt"
			mca_report_path = folder + test + "_MCA_Report" + ".xlsx"
			config_path = folder + test + "_config" + ".txt"
			zip_path = folder + test + "_log" + ".zip"
		
		return file_path, mca_path, mca_debug_path, config_path, zip_path, mca_report_path
	
	# =================================================================
	# MAIN EXECUTION HELPER METHODS
	# =================================================================
	
	def _initialize_data_handler(self, folder, visual, TestName, Bucket, WW, logger, 
								  upload_to_disk, upload_to_danta):
		"""Initialize and return data handler for uploads."""
		if self.file_handler is None:
			print(f"  {Colors.warning('[!]')} FileHandler module not provided - upload features disabled")
			return None
		
		return self.file_handler.TestUpload(
			folder=folder, vid=visual, name=TestName, bucket=Bucket, 
			WW=WW, product=self.product, logger=logger, from_Framework=False,
			upload_to_disk=upload_to_disk, upload_to_danta=upload_to_danta
		)
	
	def _read_fuse_registers(self):
		"""Read fuse registers and return masks and PC value."""
		if self.dpm_checks is None:
			print(f"  {Colors.warning('[!]')} dpmChecks module not provided - skipping fuse operations")
			return None
		
		try:
			masks = self.dpm_checks.fuses()
			
			core_mask_c0 = masks['ia_compute_0']
			core_mask_c1 = masks['ia_compute_1']
			core_mask_c2 = masks['ia_compute_2']
			
			cha_mask_c0 = masks['llc_compute_0']
			cha_mask_c1 = masks['llc_compute_1']
			cha_mask_c2 = masks['llc_compute_2']
			
			PC = self.read_scratchpad()
			
			print("[+] REGISTERS READ SUCCESSFULLY")
			
			return {
				'core_masks': (core_mask_c0, core_mask_c1, core_mask_c2),
				'cha_masks': (cha_mask_c0, cha_mask_c1, cha_mask_c2),
				'PC': PC
			}
			
		except Exception as e:
			print(f"[X] REGISTERS COULD NOT BE READ: {e}")
			return None
	
	def _save_tileview(self, config_path):
		"""Save TileView to config file."""
		log(config_path, "a")
		
		try:
			print("\nTile View\n")
			if self.mca_decoders and self.mca_decoders.tileview:
				self.mca_decoders.tileview.tileview(sv.socket0).show_soc()
			print("[+] TILEVIEW SAVED SUCCESSFULLY")
		except Exception as e:
			print(f"[X] TILEVIEW COULD NOT BE SAVED: {e}")
	
	def _save_voltage_ratios(self):
		"""Save CFC and IA voltage/ratio information."""
		try:
			if self.dpm_checks:
				print("\nCFC Voltage/Ratios for current System\n")
				CFC = self.dpm_checks.cfc(fuse_ram=False)
				CFC.ratios()
				CFC.voltage()
				
				print("\nCORE(IA) Voltage/Ratios for current System\n")
				IA = self.dpm_checks.ia(fuse_ram=False)
				IA.ratios()
				IA.voltage()
			else:
				print(f"  {Colors.warning('[!]')} dpmChecks module not available - skipping voltage/ratio info")
		except Exception as e:
			print(f"[X] CFC/IA Ratios/Voltages could not be read: {e}")
	
	def _save_registers_to_file(self, config_path, PC, core_masks, cha_masks):
		"""Save register values to config file."""
		try:
			core_mask_c0, core_mask_c1, core_mask_c2 = core_masks
			cha_mask_c0, cha_mask_c1, cha_mask_c2 = cha_masks
			
			with open(config_path, "a") as file:
				file.write("PC\n")
				file.write(f"sv.socket0.uncore.ubox.ncdecs.biosnonstickyscratchpad7_cfg = {PC}\n\n")
				
				file.write("\nConfig Registers\n")
				
				file.write("CORE Mask\n")
				file.write(f"sv.socket0.compute0.fuses.hwrs_top_rom.ip_disable_fuses_dword6_core_disable.read() = {core_mask_c0}\n")
				file.write(f"sv.socket0.compute1.fuses.hwrs_top_rom.ip_disable_fuses_dword6_core_disable.read() = {core_mask_c1}\n")
				file.write(f"sv.socket0.compute2.fuses.hwrs_top_rom.ip_disable_fuses_dword6_core_disable.read() = {core_mask_c2}\n\n")
				
				file.write("CHA Mask\n")
				file.write(f"sv.socket0.compute0.fuses.hwrs_top_rom.ip_disable_fuses_dword2_llc_disable.read() = {cha_mask_c0}\n")
				file.write(f"sv.socket0.compute1.fuses.hwrs_top_rom.ip_disable_fuses_dword2_llc_disable.read() = {cha_mask_c1}\n")
				file.write(f"sv.socket0.compute2.fuses.hwrs_top_rom.ip_disable_fuses_dword2_llc_disable.read() = {cha_mask_c2}\n\n")
			
			print("[+] REGISTERS SAVED SUCCESSFULLY")
			return True
			
		except Exception as e:
			print(f"[X] REGISTERS COULD NOT BE SAVED: {e}")
			return False
	
	def _process_masks(self, sheet, core_masks, cha_masks):
		"""Process and review masks, marking them in the sheet."""
		try:
			core_mask_c0, core_mask_c1, core_mask_c2 = core_masks
			cha_mask_c0, cha_mask_c1, cha_mask_c2 = cha_masks
			
			bin_cha_mask, hex_cha_mask = self.get_mask(cha_mask_c0, cha_mask_c1, cha_mask_c2)
			bin_core_mask, hex_core_mask = self.get_mask(core_mask_c0, core_mask_c1, core_mask_c2)
			common_core_cha_mask = self.common_mask(bin_cha_mask, bin_core_mask)
			
			self.review_mask(bin_cha_mask, self.cha_string, sheet)
			self.review_mask(bin_core_mask, self.core_string, sheet)
			self.review_mask(common_core_cha_mask, 'core_cha', sheet)
			
			print("[+] DISABLE SLICES WRITTEN SUCCESSFULLY")
			return True
		except Exception as e:
			print(f"[X] COULD NOT WRITE DISABLE SLICES: {e}")
			return False
	
	def _create_log_archive(self, zip_path, mca_path, mca_debug_path, config_path):
		"""Create zip archive from log files and clean up."""
		try:
			with zipfile.ZipFile(zip_path, "w") as zipf:
				zipf.write(mca_path, arcname="MCAs.txt")
				zipf.write(mca_debug_path, arcname="MCA_debug.txt")
				zipf.write(config_path, arcname="Config.txt")
			
			os.remove(mca_path)
			os.remove(mca_debug_path)
			os.remove(config_path)
			
			print("[+] ZIP FILE CREATED SUCCESSFULLY")
			return True
		except Exception as e:
			print(f"[X] COULD NOT CREATE ZIP FILE: {e}")
			return False
	
	def _upload_data(self, datahandler, files, framework_vars, QDF, Testnumber, TestName):
		"""Handle data upload operations."""
		if datahandler is None:
			return
		
		print("DATA UPLOAD IN PROGRESS...")
		
		framework_vars['qdf'] = QDF
		framework_vars['tnum'] = Testnumber
		framework_vars['runName'] = TestName
		framework_vars['scratchpad'] = self.read_scratchpad()
		
		serial_log = framework_vars['ttlog'] if framework_vars['ttlog'] is not None else None
		
		# Generate temporal and upload
		datahandler.generate_temporal()
		datahandler.copy_to_temporal(files)
		if serial_log:
			datahandler.copy_to_log_to_temporal(serial_log)
		
		datahandler.initlog()
		datahandler.TestBanner(
			framework_vars.get('qdf', ''),
			framework_vars.get('tnum', ''),
			framework_vars.get('mask', ''),
			framework_vars.get('corelic', ''),
			framework_vars.get('bumps', 'vbump'),
			framework_vars.get('htdis', False),
			framework_vars.get('dis1CPM', False),
			framework_vars.get('dis2CPM', False),
			framework_vars.get('freqIA', ''),
			framework_vars.get('voltIA', ''),
			framework_vars.get('freqCFC', ''),
			framework_vars.get('voltCFC', ''),
			framework_vars.get('content', 'Dragon'),
			framework_vars.get('passstring', 'Test Complete'),
			framework_vars.get('failstring', 'Test Failed')
		)
		datahandler.Iteration_end(
			framework_vars.get('tnum', ''),
			framework_vars.get('runName', 'NotDefined'),
			framework_vars.get('runStatus', 'NotDefined'),
			framework_vars.get('scratchpad', 'NotDefined'),
			framework_vars.get('seed', 'NotDefined')
		)
		datahandler.generate_summary()
		datahandler.upload_data()
		datahandler.remove_temporal()
	
	# =================================================================
	# MAIN EXECUTION
	# =================================================================
	
	def run(self, visual='', Testnumber='', TestName='', chkmem=0, 
			debug_mca=0, dr_dump=False, folder=None, WW='WW', 
			Bucket='UNCORE', QDF='', logger=None, upload_to_disk=False, 
			upload_to_danta=False, framework_data=None):
		"""
		Main function to generate error report.
		
		Args:
			visual: Visual ID
			Testnumber: Test number
			TestName: Test name
			chkmem: Check memory configuration (0/1)
			debug_mca: Enable debug MCA logging (0/1)
			dr_dump: Enable dragon dump (True/False)
			folder: Output folder path
			WW: Work week
			Bucket: DPM bucket
			QDF: QDF identifier
			logger: Logger instance
			upload_to_disk: Upload results to disk
			upload_to_danta: Upload results to Danta
			framework_data: Framework configuration data
		"""
		test = f'{Testnumber}_{visual}_{TestName}'
		
		print(f'{"+"*20} Collecting System data for {self.product}:{self.variant} {"+"*20}')
		
		# Set default folder
		if folder is None:
			folder = "C:\\temp\\"
		
		# Define paths
		file_path, mca_path, mca_debug_path, config_path, zip_path, mca_report_path = \
			self.define_paths(folder, test)
		
		# Initialize data handler
		datahandler = self._initialize_data_handler(
			folder, visual, TestName, Bucket, WW, logger, 
			upload_to_disk, upload_to_danta
		)
		
		files = [file_path, mca_report_path, zip_path]
		framework_vars = self._init_framework_vars() if framework_data is None else framework_data
		
		# Load workbook
		try:
			workbook = self.load_workbook()
			sheet = workbook['Template']
			print("[+] WORKBOOK LOADED SUCCESSFULLY")
		except Exception as e:
			print(f"[X] WORKBOOK COULD NOT BE LOADED: {e}")
			return
		
		# Check memory configuration
		if chkmem and self.product == 'GNR':
			try:
				self.check_mem(sheet, config_path)
				print("[+] MEMORY READ SUCCESSFULLY")
			except Exception as e:
				print(f"[X] MEMORY COULD NOT BE READ: {e}")
		
		# Read fuse registers
		register_data = self._read_fuse_registers()
		if register_data is None:
			return
		
		core_masks = register_data['core_masks']
		cha_masks = register_data['cha_masks']
		PC = register_data['PC']
		
		# Save TileView
		self._save_tileview(config_path)
		
		# Save voltage/ratio info
		self._save_voltage_ratios()
		
		nolog()
		
		# Save registers to config file
		self._save_registers_to_file(config_path, PC, core_masks, cha_masks)
		
		# Process and review masks
		if not self._process_masks(sheet, core_masks, cha_masks):
			return
		
		# Write PC to sheet
		sheet["B66"] = 'PC - ' + PC
		
		# Check MCA
		self.check_MCA(
			sheet, mca_path, mca_debug_path, debug_mca, mca_report_path,
			dr_dump=dr_dump, visual=visual, loop=Testnumber, WW=WW,
			Bucket=Bucket, QDF=QDF, bindesc=test
		)
		
		# Write test name
		test_name = sheet["B2"].value + ' ' + test
		sheet["B2"] = test_name
		
		# Create zip archive
		self._create_log_archive(zip_path, mca_path, mca_debug_path, config_path)
		
		# Save workbook and upload data
		try:
			self.save_workbook(workbook, file_path)
			print("[+] WORKBOOK SAVED SUCCESSFULLY")
			
			# Upload data if requested
			if upload_to_disk or upload_to_danta:
				self._upload_data(datahandler, files, framework_vars, QDF, Testnumber, TestName)
			
			print(r"""
  ____                      _
 |  _ \  ___  _ __   ___   | |
 | | | |/ _ \| '_ \ / _ \  |_|
 | |_| | (_) | | | |  __/   _ 
 |____/ \___/|_| |_|\___|  |_|

			""")
			return True
			
		except Exception as e:
			print(f"[X] WORKBOOK COULD NOT BE SAVED: {e}")
			return False

# =================================================================
# BACKWARD COMPATIBILITY WRAPPER
# =================================================================

def run(visual='', Testnumber='', TestName='', chkmem=0, debug_mca=0, 
		dr_dump=False, folder=None, WW='WW', Bucket='UNCORE', 
		product='GNR', QDF='', logger=None, upload_to_disk=False, 
		upload_to_danta=False, framework_data=None):
	"""
	Backward compatibility wrapper for the run() function.
	Creates an ErrorReportGenerator instance and calls its run method.
	
	Note: This wrapper imports required modules and passes them to the class.
	For better performance, consider importing modules once and passing them directly.
	"""
	# Import configuration loader - try both paths
	pe = None
	dev_mode = True
	base_path = None
	
	try:
		import users.gaespino.dev.S2T.ConfigsLoader as pe
		selected_product = pe.SELECTED_PRODUCT
		selected_variant = pe.PRODUCT_VARIANT
		dev_mode = getattr(pe, 'DEV_MODE', True)
		base_path = getattr(pe, 'BASE_PATH', None)
	except:
		try:
			import users.THR.PythonScript.thr.S2T.ConfigsLoader as pe
			selected_product = pe.SELECTED_PRODUCT
			selected_variant = pe.PRODUCT_VARIANT
			dev_mode = getattr(pe, 'DEV_MODE', False)
			base_path = getattr(pe, 'BASE_PATH', None)
		except:
			selected_product = product
			selected_variant = 'AP'
	
	# Determine base path
	if dev_mode:
		base_path = 'users.gaespino.dev'
		debug_path = 'users.gaespino.dev.DebugFramework'
	else:
		# Use BASE_PATH from config or fallback to default
		if base_path is None:
			base_path = 'users.THR.PythonScript.thr'
		# Derive debug path by replacing S2T with DebugFramework
		debug_path = 'users.gaespino.DebugFramework'
	
	# Import required modules
	core_manipulation = None
	file_handler = None
	dpm_checks = None
	
	try:
		core_manipulation = import_module(f'{base_path}.CoreManipulation')
	except ImportError:
		print(f"  {Colors.warning('[!]')} CoreManipulation module not available")
	
	try:
		file_handler = import_module(f'{debug_path}.FileHandler')
	except ImportError:
		print(f"  {Colors.warning('[!]')} FileHandler module not available")
	
	try:
		dpm_checks = import_module(f'{base_path}.dpmChecks')
	except ImportError:
		print(f"  {Colors.warning('[!]')} dpmChecks module not available")
	
	# Create generator instance
	generator = ErrorReportGenerator(
		product=selected_product,
		variant=selected_variant,
		config_loader=pe,
		logger=logger,
		core_manipulation=core_manipulation,
		file_handler=file_handler,
		dpm_checks=dpm_checks
	)
	
	# Run report generation
	return generator.run(
		visual=visual,
		Testnumber=Testnumber,
		TestName=TestName,
		chkmem=chkmem,
		debug_mca=debug_mca,
		dr_dump=dr_dump,
		folder=folder,
		WW=WW,
		Bucket=Bucket,
		QDF=QDF,
		logger=logger,
		upload_to_disk=upload_to_disk,
		upload_to_danta=upload_to_danta,
		framework_data=framework_data
	)
