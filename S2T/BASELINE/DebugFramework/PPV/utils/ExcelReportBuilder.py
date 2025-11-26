"""
ExcelReportBuilder - Flexible Excel report generation class
Refactored from save_to_excel function for better maintainability and configurability
"""

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Font, Alignment


from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

class SheetConfig:
	"""Configuration for a single Excel sheet"""
	
	def __init__(self, name, dataframe_key, table_name, table_style='TableStyleLight1', 
	             color_rules=None, custom_formatter=None):
		"""
		Initialize sheet configuration.
		
		Args:
			name: Sheet display name
			dataframe_key: Key to retrieve DataFrame from data dict
			table_name: Excel table name (must be unique)
			table_style: Excel table style name
			color_rules: List of dicts with color formatting rules
			custom_formatter: Custom function(worksheet, dataframe, row_idx, col_idx, value, cell)
		"""
		self.name = name
		self.dataframe_key = dataframe_key
		self.table_name = table_name
		self.table_style = table_style
		self.color_rules = color_rules or []
		self.custom_formatter = custom_formatter


class ExcelReportBuilder:
	"""
	Flexible Excel report builder with configurable sheets and styling.
	
	Features:
	- Configurable sheet order
	- Customizable table styles per sheet
	- Color formatting rules
	- Column width auto-adjustment
	- Reusable formatting logic
	"""
	
	# Predefined table styles
	STYLES = {
		'data': TableStyleInfo(name="TableStyleLight8", showFirstColumn=False,
		                       showLastColumn=False, showRowStripes=False, showColumnStripes=False),
		'report': TableStyleInfo(name="TableStyleDark8", showFirstColumn=False,
		                         showLastColumn=False, showRowStripes=False, showColumnStripes=False),
		'file': TableStyleInfo(name="TableStyleLight1", showFirstColumn=False,
		                       showLastColumn=False, showRowStripes=False, showColumnStripes=False),
		'fails': TableStyleInfo(name="TableStyleLight9", showFirstColumn=False,
		                        showLastColumn=False, showRowStripes=False, showColumnStripes=False),
		'mcas': TableStyleInfo(name="TableStyleLight10", showFirstColumn=False,
		                       showLastColumn=False, showRowStripes=False, showColumnStripes=False),
		'overview': TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
		                           showLastColumn=False, showRowStripes=True, showColumnStripes=False),
	}
	
	# Predefined color schemes
	COLORS = {
		'pass': {'fill': 'C6EFCE', 'font': '006100'},
		'fail': {'fill': 'FFC7CE', 'font': '9C0006'},
		'mixed': {'fill': 'FFE4B5', 'font': 'CC6600'},
		'yes': {'fill': 'C6EFCE', 'font': '006100'},
		'no': {'fill': 'FFC7CE', 'font': '9C0006'},
	}
	
	def __init__(self, sheet_configs=None):
		"""
		Initialize ExcelReportBuilder.
		
		Args:
			sheet_configs: List of SheetConfig objects (defines sheet order)
		"""
		self.sheet_configs = sheet_configs or []
		self.workbook = None
	
	def build(self, data_dict, filename):
		"""
		Build Excel file from data dictionary.
		
		Args:
			data_dict: Dictionary mapping dataframe_keys to DataFrames
			filename: Output Excel file path
		"""
		self.workbook = Workbook()
		
		# Remove default sheet if we have configurations
		if self.sheet_configs and 'Sheet' in self.workbook.sheetnames:
			self.workbook.remove(self.workbook['Sheet'])
		
		# Create sheets in configured order
		for config in self.sheet_configs:
			df = data_dict.get(config.dataframe_key)
			
			# Skip if DataFrame is None or empty
			if df is None or df.empty:
				continue
			
			self._create_sheet(config, df)
		
		# Save workbook
		self.workbook.save(filename)
	
	def _create_sheet(self, config, df):
		"""Create a single worksheet with data and formatting"""
		# Create worksheet
		ws = self.workbook.create_sheet(title=config.name)
		
		# Special handling for Overview sheet
		if config.name == 'Overview':
			self._create_overview_sheet(ws, df, config)
			return
		
		# Write data
		self._write_data(ws, df, config)
		
		# Add headers
		self._write_headers(ws, df)
		
		# Adjust column widths
		self._adjust_column_widths(ws)
		
		# Create table
		self._create_table(ws, df, config)
	
	def _create_overview_sheet(self, ws, df, config):
		"""Create Overview sheet with professional presentation formatting (6-column layout)"""
		from openpyxl.styles import Border, Side
		
		# Track cells to merge after writing all data
		merge_ranges = []
		
		# Write data to cells (skip first row for headers later)
		for r_idx, row in enumerate(df.itertuples(), start=1):
			for c_idx, value in enumerate(row[1:], start=1):
				cell = ws.cell(row=r_idx+1, column=c_idx, value=value)
				
				# Get column name and section
				col_name = df.columns[c_idx-1]
				section = row.Section if hasattr(row, 'Section') else ''
				category = row.Category if hasattr(row, 'Category') else ''
				metric = row.Metric if hasattr(row, 'Metric') else ''
				
				# HEADER SECTION - Large title row
				if section == 'HEADER' and r_idx == 1:
					if col_name == 'Metric':  # Title in Metric column
						cell.font = Font(size=18, bold=True, color='FFFFFF')
						cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
						cell.alignment = Alignment(horizontal='center', vertical='center')
						# Mark for merging later
						merge_ranges.append((r_idx+1, 1, r_idx+1, 6))
					continue
				
				# SECTION HEADERS (Category column) - Blue background
				if category and not metric and col_name == 'Category':
					cell.font = Font(size=13, bold=True, color='FFFFFF')
					cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
					cell.alignment = Alignment(horizontal='left', vertical='center')
					# Mark for merging later
					merge_ranges.append((r_idx+1, 1, r_idx+1, 6))
					continue
				
				# METRIC LABELS - Light blue background
				if metric and col_name == 'Metric':
					cell.font = Font(size=10, bold=True)
					cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
					cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
				
				# VALUE column - Apply status colors
				elif col_name == 'Value':
					cell.font = Font(size=10, bold=False)
					cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
					# Apply color rules for values
					self._apply_color_rules(cell, df, c_idx, value, config.color_rules)
				
				# STATUS column - Color indicators
				elif col_name == 'Status':
					if value:
						status_colors = {
							'YES': {'fill': '00B050', 'font': 'FFFFFF'},  # Green
							'PASS': {'fill': '00B050', 'font': 'FFFFFF'},  # Green
							'NO': {'fill': 'C00000', 'font': 'FFFFFF'},  # Red
							'FAIL': {'fill': 'C00000', 'font': 'FFFFFF'},  # Red
							'FLAKY': {'fill': 'FFC000', 'font': '000000'},  # Orange
							'NOT_RUN': {'fill': 'D9D9D9', 'font': '000000'}  # Gray
						}
						
						status_str = str(value).strip().upper()
						if status_str in status_colors:
							colors = status_colors[status_str]
							cell.fill = PatternFill(start_color=colors['fill'], end_color=colors['fill'], fill_type='solid')
							cell.font = Font(size=10, bold=True, color=colors['font'])
							cell.alignment = Alignment(horizontal='center', vertical='center')
					else:
						cell.alignment = Alignment(horizontal='center', vertical='center')
				
				# DETAILS column
				elif col_name == 'Details':
					cell.font = Font(size=9, italic=True, color='666666')
					cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
				
				# DEFAULT styling for other columns
				else:
					cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
				
				# Add borders to all non-blank rows
				if section != '':  # Not a blank row
					thin_border = Border(
						left=Side(style='thin', color='CCCCCC'),
						right=Side(style='thin', color='CCCCCC'),
						top=Side(style='thin', color='CCCCCC'),
						bottom=Side(style='thin', color='CCCCCC')
					)
					cell.border = thin_border
		
		# NOW perform all merge operations
		for start_row, start_col, end_row, end_col in merge_ranges:
			ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
		
		# Add headers in first row
		headers = ['Section', 'Category', 'Metric', 'Value', 'Status', 'Details']
		for c_idx, header in enumerate(headers, start=1):
			cell = ws.cell(row=1, column=c_idx, value=header)
			cell.font = Font(size=11, bold=True, color='FFFFFF')
			cell.fill = PatternFill(start_color='203864', end_color='203864', fill_type='solid')
			cell.alignment = Alignment(horizontal='center', vertical='center')
			cell.border = Border(
				left=Side(style='medium'),
				right=Side(style='medium'),
				top=Side(style='medium'),
				bottom=Side(style='medium')
			)
		
		# Adjust column widths for professional presentation
		ws.column_dimensions['A'].width = 15  # Section
		ws.column_dimensions['B'].width = 25  # Category
		ws.column_dimensions['C'].width = 30  # Metric
		ws.column_dimensions['D'].width = 40  # Value
		ws.column_dimensions['E'].width = 12  # Status
		ws.column_dimensions['F'].width = 50  # Details
		
		# Freeze top row and first column
		ws.freeze_panes = 'A2'
	
	def _write_data(self, ws, df, config):
		"""Write DataFrame data to worksheet with formatting"""
		for r_idx, row in enumerate(df.itertuples(), start=1):
			for c_idx, value in enumerate(row[1:], start=1):
				cell = ws.cell(row=r_idx+1, column=c_idx, value=value)
				
				# Apply custom formatter if provided
				if config.custom_formatter:
					config.custom_formatter(ws, df, r_idx, c_idx, value, cell)
				
				# Apply color rules
				self._apply_color_rules(cell, df, c_idx, value, config.color_rules)
	
	def _apply_color_rules(self, cell, df, c_idx, value, color_rules):
		"""Apply color formatting rules to a cell"""
		if not color_rules:
			return
		
		col_name = df.columns[c_idx-1]
		
		for rule in color_rules:
			# Check if rule applies to this column
			if rule.get('column') != col_name:
				continue
			
			# Get rule type and apply formatting
			rule_type = rule.get('type', 'value_match')
			
			if rule_type == 'value_match':
				self._apply_value_match_rule(cell, value, rule)
			elif rule_type == 'contains':
				self._apply_contains_rule(cell, value, rule)
			elif rule_type == 'mixed_status':
				self._apply_mixed_status_rule(cell, value, rule)
			elif rule_type == 'partial_text_color':
				self._apply_partial_text_color_rule(cell, value, rule)
	
	def _apply_value_match_rule(self, cell, value, rule):
		"""Apply formatting when value matches specific values"""
		value_str = str(value).upper() if value else ''
		
		for match_value, color_key in rule.get('mappings', {}).items():
			if value_str == match_value.upper():
				color = self.COLORS.get(color_key, {})
				if 'fill' in color:
					cell.fill = PatternFill(start_color=color['fill'], 
					                        end_color=color['fill'], fill_type='solid')
				if 'font' in color:
					cell.font = Font(color=color['font'])
				break
	
	def _apply_contains_rule(self, cell, value, rule):
		"""Apply formatting when value contains specific text"""
		value_str = str(value).upper() if value else ''
		
		for search_text, color_key in rule.get('mappings', {}).items():
			if search_text.upper() in value_str:
				color = self.COLORS.get(color_key, {})
				# Apply both fill (background) and font color for whole cell formatting
				if 'fill' in color:
					cell.fill = PatternFill(start_color=color['fill'], 
					                        end_color=color['fill'], fill_type='solid')
				if 'font' in color:
					cell.font = Font(color=color['font'])
				break
	
	def _apply_mixed_status_rule(self, cell, value, rule):
		"""Apply formatting for mixed status (e.g., P15 | F3)"""
		value_str = str(value) if value else ''
		
		# Check if it matches mixed pattern (contains both P and F with |)
		if 'P' in value_str and 'F' in value_str and '|' in value_str:
			color = self.COLORS.get('mixed', {})
			if 'fill' in color:
				cell.fill = PatternFill(start_color=color['fill'], 
				                        end_color=color['fill'], fill_type='solid')
			if 'font' in color:
				cell.font = Font(color=color['font'])
	
	def _apply_partial_text_color_rule(self, cell, value, rule):
		"""
		Apply different colors to specific words within the same cell.
		Example: "MCAs: YES | VVARs: NO" - color YES in red, NO in green
		"""
		value_str = str(value) if value else ''
		if not value_str:
			return
		
		# Get mappings: {'YES': 'fail', 'NO': 'pass'}
		mappings = rule.get('mappings', {})
		if not mappings:
			return
		
		# Build rich text parts
		rich_parts = []
		current_pos = 0
		matches_found = []
		
		# Find all matches and their positions
		for search_text, color_key in mappings.items():
			pos = 0
			while True:
				pos = value_str.find(search_text, pos)
				if pos == -1:
					break
				matches_found.append((pos, pos + len(search_text), search_text, color_key))
				pos += len(search_text)
		
		# Sort matches by position
		matches_found.sort()
		
		# Build rich text
		if not matches_found:
			# No matches, just use plain text
			return
		
		for start, end, text, color_key in matches_found:
			# Add text before match (if any)
			if current_pos < start:
				rich_parts.append(TextBlock(InlineFont(), value_str[current_pos:start]))
			
			# Add colored match
			color = self.COLORS.get(color_key, {})
			font_color = color.get('font', '000000')
			from openpyxl.styles.colors import Color
			rich_parts.append(TextBlock(InlineFont(color=Color(rgb=font_color)), text))
			current_pos = end
		
		# Add remaining text (if any)
		if current_pos < len(value_str):
			rich_parts.append(TextBlock(InlineFont(), value_str[current_pos:]))
		
		# Apply rich text to cell
		if rich_parts:
			cell.value = CellRichText(*rich_parts)
	
	def _write_headers(self, ws, df):
		"""Write column headers"""
		for c_idx, header in enumerate(df.columns, start=1):
			ws.cell(row=1, column=c_idx, value=header)
	
	def _adjust_column_widths(self, ws, max_width=50):
		"""Auto-adjust column widths based on content"""
		for column in ws.columns:
			max_length = 0
			for cell in column:
				if cell.value:
					max_length = max(max_length, len(str(cell.value)))
			
			adjusted_width = min(max_length + 2, max_width)
			ws.column_dimensions[column[0].column_letter].width = adjusted_width
	
	def _create_table(self, ws, df, config):
		"""Create Excel table for the data range"""
		# Get table style
		if isinstance(config.table_style, str):
			# If string, look up in predefined styles or create new
			table_style = self.STYLES.get(config.table_style)
			if not table_style:
				table_style = TableStyleInfo(name=config.table_style, 
				                             showFirstColumn=False,
				                             showLastColumn=False, 
				                             showRowStripes=False, 
				                             showColumnStripes=False)
		else:
			table_style = config.table_style
		
		# Create table
		last_cell = ws.cell(row=len(df)+1, column=len(df.columns))
		table = Table(displayName=config.table_name, ref=f"A1:{last_cell.coordinate}")
		table.tableStyleInfo = table_style
		ws.add_table(table)
	
	@staticmethod
	def create_default_config():
		"""Create default sheet configuration (backward compatible with save_to_excel)"""
		return [
			# Sheet order: Overview FIRST, then ExperimentSummary, ExperimentReport, etc.
			SheetConfig(
				name='Overview',
				dataframe_key='overview_df',
				table_name='Overview',
				table_style='overview',
				color_rules=[
					{
						'column': 'Status',
						'type': 'value_match',
						'mappings': {
							'YES': 'pass',
							'PASS': 'pass',
							'NO': 'fail',
							'FAIL': 'fail',
							'FLAKY': 'mixed',
							'NOT_RUN': 'no'
						}
					}
				],
				custom_formatter=None  # Special formatting handled in _create_overview_sheet
			),
			SheetConfig(
				name='ExperimentSummary',
				dataframe_key='experiment_summary_df',
				table_name='ExperimentSummary',
				table_style='report',
				color_rules=[
					{
						'column': 'Status',
						'type': 'contains',  # Changed from value_match to handle "PASS 3", "FAIL 2", etc.
						'mappings': {'PASS': 'pass', 'FAIL': 'fail'}
					},
					{
						'column': 'Status',
						'type': 'mixed_status',
						'mappings': {}
					},
					{
						'column': 'Results',
						'type': 'partial_text_color',
						'mappings': {'YES': 'fail', 'NO': 'pass'}  # YES=red, NO=green
					},
					{
						'column': 'Characterization',
						'type': 'contains',
						'mappings': {'PASS': 'pass', 'FAIL': 'fail'}
					}
				]
			),
			SheetConfig(
				name='ExperimentReport',
				dataframe_key='summary_df',
				table_name='FrameworkSummary',
				table_style='report',
				color_rules=[
					{
						'column': 'Status',
						'type': 'value_match',
						'mappings': {'PASS': 'pass', 'FAIL': 'fail'}
					},
					{
						'column': 'Content Results',
						'type': 'value_match',
						'mappings': {'PASS': 'pass', 'FAIL': 'fail'}
					},
					{
						'column': 'MCAs',
						'type': 'value_match',
						'mappings': {'YES': 'yes', 'NO': 'no'}
					}
				]
			),
			SheetConfig(
				name='CoreData',
				dataframe_key='core_data_df',
				table_name='CoreData',
				table_style='report'
			),
			SheetConfig(
				name='DragonData',
				dataframe_key='vvar_df',
				table_name='DragonData',
				table_style='data'
			),
			SheetConfig(
				name='UniqueFails',
				dataframe_key='unique_fails_df',
				table_name='UniqueFails',
				table_style='fails'
			),
			SheetConfig(
				name='FrameworkData',
				dataframe_key='test_df',
				table_name='FrameworkData',
				table_style='data',
				color_rules=[
					{
						'column': 'Result',
						'type': 'value_match',
						'mappings': {'PASS': 'pass', 'FAIL': 'fail'}
					}
				]
			),
			SheetConfig(
				name='FrameworkFails',
				dataframe_key='fail_info_df',
				table_name='FrameworkFails',
				table_style='fails'
			),
			SheetConfig(
				name='FrameworkFiles',
				dataframe_key='initial_df',
				table_name='FrameworkFiles',
				table_style='file'
			),
		]


# Backward compatible function
def save_to_excel(initial_df, test_df, summary_df, fail_info_df=None, unique_fails_df=None, 
                  unique_mcas_df=None, vvar_df=None, core_data_df=None, 
                  experiment_summary_df=None, overview_df=None, filename='output.xlsx'):
	"""
	Backward compatible wrapper for ExcelReportBuilder.
	
	This function maintains the original signature but uses the new class internally.
	"""
	# Prepare data dictionary
	data_dict = {
		'overview_df': overview_df,
		'initial_df': initial_df,
		'test_df': test_df,
		'summary_df': summary_df,
		'fail_info_df': fail_info_df,
		'unique_fails_df': unique_fails_df,
		'unique_mcas_df': unique_mcas_df,
		'vvar_df': vvar_df,
		'core_data_df': core_data_df,
		'experiment_summary_df': experiment_summary_df,
	}
	
	# Create builder with default configuration
	builder = ExcelReportBuilder(sheet_configs=ExcelReportBuilder.create_default_config())
	
	# Build Excel file
	builder.build(data_dict, filename)
