import os
import shutil
import zipfile
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook

def replace_word_in_file(file_path, old_word, new_word):
    with open(file_path, 'r', encoding='utf-8') as file:
        content = file.read()
    content = content.replace(old_word, new_word)
    with open(file_path, 'w', encoding='utf-8') as file:
        file.write(content)

def replace_word_in_excel(file_path, old_word, new_word):
    wb = load_workbook(file_path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    cell.value = cell.value.replace(old_word, new_word)
    wb.save(file_path)
    wb.close()

def process_zip_file(zip_path, old_word, new_word, temp_dir):
    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        zip_temp_dir = os.path.join(temp_dir, 'zip_temp')
        os.makedirs(zip_temp_dir, exist_ok=True)
        zip_ref.extractall(zip_temp_dir)

        for root, dirs, files in os.walk(zip_temp_dir):
            for file in files:
                file_path = os.path.join(root, file)
                new_file_name = file.replace(old_word, new_word)
                new_file_path = os.path.join(root, new_file_name)

                if file.endswith('.txt') or file.endswith('.log'):
                    replace_word_in_file(file_path, old_word, new_word)
                elif file.endswith('.xlsx'):
                    replace_word_in_excel(file_path, old_word, new_word)

                os.rename(file_path, new_file_path)

        new_zip_path = os.path.join(temp_dir, os.path.basename(zip_path).replace(old_word, new_word))
        with zipfile.ZipFile(new_zip_path, 'w') as zip_ref:
            for root, dirs, files in os.walk(zip_temp_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    zip_ref.write(file_path, os.path.relpath(file_path, zip_temp_dir))

        # Clean up temporary directory
        shutil.rmtree(zip_temp_dir)

def replace_word_in_folder(folder_path, old_word, new_word):
    temp_dir = 'C:\\Temp\\FrameworkFileFix'
    os.makedirs(temp_dir, exist_ok=True)

    for root, dirs, files in os.walk(folder_path):
        for file in files:
            file_path = os.path.join(root, file)
            new_file_name = file.replace(old_word, new_word)
            new_file_path = os.path.join(temp_dir, new_file_name)

            if file.endswith('.log'):
                shutil.copy(file_path, new_file_path)
                replace_word_in_file(new_file_path, old_word, new_word)
            elif file.endswith('.zip'):
                process_zip_file(file_path, old_word, new_word, temp_dir)
            elif file.endswith('.xlsx'):
                shutil.copy(file_path, new_file_path)
                replace_word_in_excel(new_file_path, old_word, new_word)

    # Move files from temp directory to original folder
    for file in os.listdir(temp_dir):
        new_file_path = os.path.join(temp_dir, file)
        shutil.move(new_file_path, os.path.join(folder_path, file))
        
        # Only delete the original Excel files
        if file.endswith('.xlsx'):
            original_file_path = os.path.join(folder_path, file.replace(new_word, old_word))
            if os.path.exists(original_file_path):
                os.remove(original_file_path)

    # Clean up temporary directory
    shutil.rmtree(temp_dir)

class WordReplacementApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Framework File Visual ID Fix")
        
        self.create_widgets()

    def create_widgets(self):
        header_label = tk.Label(self.root, text="Framework Files Visual ID Fix", font=("Arial", 16))
        header_label.pack(pady=10)

        folder_frame = tk.Frame(self.root)
        folder_frame.pack(pady=5)

        self.folder_path_entry = tk.Entry(folder_frame, width=100)
        self.folder_path_entry.pack(side=tk.LEFT, padx=10)

        browse_button = tk.Button(folder_frame, text="Browse", command=self.browse_folder)
        browse_button.pack(side=tk.LEFT, padx=10)

        old_word_frame = tk.Frame(self.root)
        old_word_frame.pack(pady=5)

        old_word_label = tk.Label(old_word_frame, text="Old Word:")
        old_word_label.pack(side=tk.LEFT, padx=10)

        self.old_word_entry = tk.Entry(old_word_frame, width=30)
        self.old_word_entry.pack(side=tk.LEFT)

        new_word_frame = tk.Frame(self.root)
        new_word_frame.pack(pady=5)

        new_word_label = tk.Label(new_word_frame, text="New Word:")
        new_word_label.pack(side=tk.LEFT, padx=10)

        self.new_word_entry = tk.Entry(new_word_frame, width=30)
        self.new_word_entry.pack(side=tk.LEFT)

        button_frame = tk.Frame(self.root)
        button_frame.pack(pady=20)

        submit_button = tk.Button(button_frame, text="Submit", command=self.submit_action)
        submit_button.pack(side=tk.LEFT, padx=10)

        exit_button = tk.Button(button_frame, text="Exit", command=self.exit_action)
        exit_button.pack(side=tk.LEFT, padx=10)

    def browse_folder(self):
        folder_selected = filedialog.askdirectory()
        self.folder_path_entry.delete(0, tk.END)
        self.folder_path_entry.insert(0, folder_selected)

    def submit_action(self):
        folder_path = self.folder_path_entry.get()
        old_word = self.old_word_entry.get()
        new_word = self.new_word_entry.get()

        if not folder_path or not old_word or not new_word:
            messagebox.showwarning("Input Error", "Please fill all fields.")
            return

        try:
            replace_word_in_folder(folder_path, old_word, new_word)
            messagebox.showinfo("Success", "Word replacement completed successfully.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")

    def exit_action(self):
        self.root.quit()

if __name__ == "__main__":
    root = tk.Tk()
    app = WordReplacementApp(root)
    root.mainloop()