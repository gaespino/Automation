from openpyxl import load_workbook
import socket
from datetime import datetime
import os
from zipfile import BadZipFile


def create_reduced_object_excel_representation(file_path):


    # Load the Excel file
    workbook = load_workbook(filename=file_path, data_only=True)


    # Dictionary to store data from each sheet
    result = {}
	
    
    # Iterate over each sheet in the Excel file
    for sheet_name in workbook.sheetnames:
        # Get the current sheet
        sheet = workbook[sheet_name]

        # Get the column names from the first row
        columns = [cell.value for cell in sheet[1]]

        # Get the data from the rows starting from the second row
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(list(row))
            print(f"row is {row}")

        # Create the object with the specified format
        obj = {
            "columns": columns,
            "data": data
        }

        # Add the object to the dictionary with the sheet name as the key
        result[sheet_name] = obj

    return result 

def find_excel_files(directory):
    """
    Recursively search for valid .xlsx files in the given directory and its subdirectories.
    A valid file is one that can be successfully opened with openpyxl.

    Args:
        directory (str): The root directory to start the search from.

    Returns:
        List[str]: A list containing the full paths of all valid .xlsx files found.
    """
    excel_files = []

    for root, _, files in os.walk(directory):
        for file in files:
            if file.lower().endswith(".xlsx"):
                full_path = os.path.join(root, file)
                try:
                    # Try to open the file to ensure it's a valid Excel file
                    load_workbook(full_path, read_only=True)
                    excel_files.append(full_path)
                except (BadZipFile, OSError, IOError):
                    # Skip invalid or unreadable files
                    continue

    return excel_files  

def generate_ID(visual_id,key):

    hostname = socket.gethostname()


    # Generating Timestamp
    now = datetime.now()
    timestamp = now.strftime("%Y%m%d")


    #Generating unique ID
    id=f"{visual_id}_{hostname}_{timestamp}_{key}"
    return id

def get_mongo_database_date_format():
    now=datetime.now()
    date=now.strftime("%d_%m_%Y_%H_%M_%S")
    return date