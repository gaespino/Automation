import tkinter as tk
from tkinter import ttk, messagebox
import os
import shutil
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy
import socket

# Constants for dropdown selections
TEST_MODES = ['Mesh', 'Slice']
TEST_TYPES = ['Loops', 'Sweep', 'Shmoo']
MASK_OPTIONS = ['RowPass1', 'RowPass2', 'RowPass3', 'FirstPass', 'SecondPass', 'ThirdPass', '']
TYPES = ['Voltage', 'Frequency']
DOMAINS = ['CFC', 'IA']
CONTENT_OPTIONS = ['Linux', 'Dragon']
CORE_LICENSE_OPTIONS = ['', '1: SSE/128', '2: AVX2/256 Light', '3: AVX2/256 Heavy', '4: AVX3/512 Light', '5: AVX3/512 Heavy', '6: TMUL Light', '7: TMUL Heavy']

# Descriptions for fields
field_descriptions = {
    "Experiment": "Enable or Disable Experiment",
    "Test Name": "Name of Test to be executed",
    "Test Mode": "Test Mode: Mesh or Slice",
    "Test Type": "Type of test: Loops, Sweep or Shmoo",
    "Visual ID": "Unit Visual ID",
    "Bucket": "Unit Assigned Bucket",
    "COM Port": "Serial Port to communicate to MB",
    "IP Address": "IP Address to communicate to MB",
    "Content": "Select your content: Linux or Dragon",
    "Test Number": "First Number to be assigned to your test iteration",
    "Test Time": "Wait time for watchdog monitoring",
    "Reset": "Reset Unit before Starting first experiment",
    "Reset on PASS": "Reset Unit after each PASS",
    "FastBoot": "If Mesh Masking is selected, set this to False",
    "Core License": "License type or leave Blank if not used",
    "600W Unit": "True or False",
    "Pseudo Config": "Set if using Pseudo Mesh Content",
    "Configuration (Mask)": "Mesh Options: RowPass1, RowPass2, ... Blank for FullDie",
    "Check Core": "Used in Mesh Mode to check core data",
    "Voltage Type": "vbump or fixed",
    "Voltage IA": "Not Configured",
    "Voltage CFC": "Not Configured",
    "Frequency IA": "Not Configured",
    "Frequency CFC": "Not Configured",
    "Start": "Start point of Sweep",
    "End": "End point of Sweep",
    "Steps": "Size of each step",
    "ShmooFile": "Shmoo configuration file",
    "ShmooLabel": "Label for Shmoo experiment",
    "TTL Folder": "TTL Directory path",
    "Scripts File": "Scripts File path",
    "Pass String": "Pass indicator",
    "Fail String": "Fail indicator"
}

fields_to_hide = {
    'Loops': ['Type', 'Domain', 'Start', 'End', 'Steps', 'ShmooFile', 'ShmooLabel'],
    'Sweep': ['Loops', 'ShmooFile', 'ShmooLabel'],
    'Shmoo': ['Loops', 'Type', 'Domain', 'Start', 'End', 'Steps']
}

mandatory_fields = ["Visual ID", "Test Time", "Check Core", "Scripts File", "Fail String", 
                    "Test Name", "Test Number", "TTL Folder", "Pass String"]

test_type_specific_mandatory = {
    'Shmoo': ["ShmooFile", "ShmooLabel"],
    'Sweep': ["Type", "Domain", "Start", "End", "Steps"]
}

class ExperimentApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Experiment Interface")

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(expand=True, fill='both', padx=10, pady=10)

        self.experiments_list = []  # List to hold experiment instances

        self.last_values = {
            "Visual ID": '',
            "COM Port": '8',
            "Test Time": '45',
            "IP Address": '192.168.0.2',
            "Check Core": '0',
            "Scripts File": r'C:\SystemDebug\ipc_powerdowns.txt',
            "TTL Folder": r'C:\SystemDebug\TTL',
            "Pass String": 'Test Complete',
            "Fail String": 'Test Failed'
        }

        self.create_new_experiment_tab()
        self.add_plus_tab()

        self.submit_button = ttk.Button(self, text="Collect Data", command=self.submit_experiments)
        self.submit_button.pack(pady=10)

    def create_new_experiment_tab(self):
        tab_frame = ExperimentTab(self.notebook, self.rename_tab, self.last_values)
        self.notebook.add(tab_frame, text=tab_frame.get_test_name())
        self.notebook.select(tab_frame)

    def add_plus_tab(self):
        plus_tab = ttk.Frame(self.notebook)
        self.notebook.add(plus_tab, text="+")
        self.notebook.bind("<<NotebookTabChanged>>", self.check_add_tab)

    def check_add_tab(self, event):
        selected_tab = self.notebook.index(self.notebook.select())
        if self.notebook.tab(selected_tab, "text") == "+":
            if len(self.notebook.tabs()) > 1:
                last_tab_frame_id = self.notebook.tabs()[-2]  # Get the last tab before "+"
                last_tab_frame = self.notebook.nametowidget(last_tab_frame_id)
                if isinstance(last_tab_frame, ExperimentTab):
                    self.last_values.update(last_tab_frame.get_values())
            self.notebook.forget(selected_tab)
            self.create_new_experiment_tab()
            self.add_plus_tab()

    def submit_experiments(self):
        self.experiments_list.clear()

        error_messages = []

        for tab_frame_id in self.notebook.tabs():
            tab_frame = self.notebook.nametowidget(tab_frame_id)
            if isinstance(tab_frame, ExperimentTab):
                missing_fields = tab_frame.highlight_missing_mandatory_fields()
                validation_errors = tab_frame.validate_fields()

                if missing_fields:
                    error_messages.append(f"Tab '{tab_frame.get_test_name()}': Missing fields - {', '.join(missing_fields)}")

                if validation_errors:
                    error_messages.append(f"Tab '{tab_frame.get_test_name()}': Invalid entries in - {', '.join(validation_errors)}")

                experiment = tab_frame.create_experiment_instance()
                self.experiments_list.append(experiment)

        if error_messages:
            messagebox.showwarning("Data Input Errors", "\n".join(error_messages))
        else:
            print("Submitted Experiments:", self.experiments_list)
            create_excels_from_dicts(self.experiments_list, "DebugFramework/RecipesTemplate.xlsx", "DebugFramework/RecipesTemplateOut.xlsx")

    def rename_tab(self, tab_frame):
        self.notebook.tab(tab_frame, text=tab_frame.get_test_name())

class ExperimentTab(ttk.Frame):
    def __init__(self, parent, rename_callback, last_values):
        super().__init__(parent)

        self.rename_callback = rename_callback

        self.input_vars = {
            "Visual ID": tk.StringVar(value=last_values["Visual ID"]),
            "COM Port": tk.StringVar(value=last_values["COM Port"]),
            "Test Time": tk.StringVar(value=last_values["Test Time"]),
            "IP Address": tk.StringVar(value=last_values["IP Address"]),
            "Check Core": tk.StringVar(value=last_values["Check Core"]),
            "Scripts File": tk.StringVar(value=last_values["Scripts File"]),
            "TTL Folder": tk.StringVar(value=last_values["TTL Folder"]),
            "Pass String": tk.StringVar(value=last_values["Pass String"]),
            "Fail String": tk.StringVar(value=last_values["Fail String"]),
            "Bucket": tk.StringVar(value='DebugFramework'),
            "Test Number": tk.StringVar(value='1'),
            "Voltage Type": tk.StringVar(value='vbump')
        }

        for field in field_descriptions:
            if field not in self.input_vars:
                self.input_vars[field] = tk.StringVar(value='')

        self.test_type_var = tk.StringVar(value=TEST_TYPES[0])
        self.input_vars['Test Type'] = self.test_type_var

        # Add boolean variables for check buttons
        self.bool_vars = {field: tk.StringVar(value='True' if field == "Experiment" else 'False') for field in ['Reset', 'Reset on PASS', 'FastBoot', '600W Unit', 'Pseudo Config', 'Experiment']}
        self.input_vars.update(self.bool_vars)

        self.test_name_var = self.input_vars['Test Name']
        self.test_name_var.trace_add("write", lambda *args: rename_callback(self))

        self.create_form()
        self.update_display_fields()

    def create_form(self):
        self.widgets = {}
        row = 0

        for idx, field in enumerate(field_descriptions):
            column = (idx % 2) * 2
            label = ttk.Label(self, text=field)
            label.grid(row=row, column=column, sticky='W', padx=5, pady=5)

            # Dropdown and check button setup
            if field == 'Voltage Type':
                entry = ttk.Combobox(self, textvariable=self.input_vars[field], values=['vbump', 'fixed'])
            elif field == 'Experiment':
                entry = tk.Checkbutton(self, variable=self.input_vars[field], onvalue='True', offvalue='False')
            elif field in ['Content', 'Core License']:
                options = CONTENT_OPTIONS if field == 'Content' else CORE_LICENSE_OPTIONS
                entry = ttk.Combobox(self, textvariable=self.input_vars[field], values=options)
                entry.set(self.input_vars[field].get() or options[0])
            elif field in ['Test Mode', 'Test Type', 'Configuration (Mask)', 'Type', 'Domain']:
                options = (
                    TEST_MODES if field == 'Test Mode' else TEST_TYPES if field == 'Test Type' else MASK_OPTIONS if field == 'Configuration (Mask)' else TYPES if field == 'Type' else DOMAINS
                )
                entry = ttk.Combobox(self, textvariable=self.input_vars[field], values=options)
                entry.set(self.input_vars[field].get() or options[0])
                if field == 'Test Type':
                    entry.bind("<<ComboboxSelected>>", self.update_display_fields)
            elif field in ['Reset', 'Reset on PASS', 'FastBoot', '600W Unit', 'Pseudo Config']:
                entry = tk.Checkbutton(self, variable=self.input_vars[field], onvalue='True', offvalue='False')
            else:
                entry = ttk.Entry(self, textvariable=self.input_vars[field])

            entry.grid(row=row, column=column + 1, sticky='EW', padx=5, pady=5)
            
            self.create_tooltip(label, field_descriptions[field])
            self.widgets[field] = (label, entry)

            # Add trace to update label color
            self.input_vars[field].trace_add("write", lambda *args, fld=field: self.update_label_color(fld))

            if column == 0:
                column = 1  # Switch to next column
            else:
                column = 0  # Reset to first column for next row
                row += 1

    def get_test_name(self):
        return self.test_name_var.get() or "Experiment"

    def get_values(self):
        return {field: var.get() for field, var in self.input_vars.items()}

    def highlight_missing_mandatory_fields(self):
        missing_fields = []
        
        for field in mandatory_fields:
            if not self.input_vars[field].get().strip():
                self.widgets[field][0].config(foreground="red")
                missing_fields.append(field)

        selected_type = self.test_type_var.get()
        if selected_type in test_type_specific_mandatory:
            for field in test_type_specific_mandatory[selected_type]:
                if not self.input_vars[field].get().strip():
                    self.widgets[field][0].config(foreground="red")
                    missing_fields.append(field)
        
        return missing_fields

    def update_label_color(self, field):
        # Update the label color to black when value is entered
        if self.input_vars[field].get().strip():
            self.widgets[field][0].config(foreground="black")

    def update_display_fields(self, event=None):
        selected_type = self.test_type_var.get()
        hide_fields = fields_to_hide.get(selected_type, [])

        for field, (label, entry) in self.widgets.items():
            if field in hide_fields:
                label.grid_remove()
                entry.grid_remove()
            else:
                label.grid()
                entry.grid()

    def create_tooltip(self, widget, text):
        def on_enter(event):
            tooltip = tk.Label(self, text=text, background="lightyellow", relief="solid", borderwidth=1)
            tooltip.place(x=widget.winfo_x() + widget.winfo_reqwidth(), y=widget.winfo_y())
            widget.tooltip = tooltip

        def on_leave(event):
            if hasattr(widget, 'tooltip'):
                widget.tooltip.destroy()
                del widget.tooltip

        widget.bind("<Enter>", on_enter)
        widget.bind("<Leave>", on_leave)

    def create_experiment_instance(self):
        return {
            field: self.input_vars[field].get() for field in field_descriptions
        }

    def validate_fields(self):
        error_fields = []

        def is_valid_ip(ip):
            try:
                socket.inet_aton(ip)
                return True
            except socket.error:
                return False

        def validate_com_port(value):
            try:
                port_num = int(value)
                return 0 <= port_num <= 256
            except ValueError:
                return False

        def validate_positive_integer(value):
            return value.isdigit() and int(value) > 0

        def validate_non_negative_integer(value):
            return value.isdigit() and int(value) >= 0

        def validate_existing_file(path):
            return os.path.isfile(path)

        def validate_existing_directory(path):
            return os.path.isdir(path)

        # Add the checks here and highlight fields only if visible
        checks = {
            "COM Port": validate_com_port,
            "IP Address": is_valid_ip,
            "Test Number": validate_positive_integer,
            "Test Time": validate_positive_integer,
            "Check Core": validate_non_negative_integer,
            "ShmooFile": validate_existing_file,
            "TTL Folder": validate_existing_directory,
            "Scripts File": validate_existing_file
        }

        hide_fields = fields_to_hide.get(self.test_type_var.get(), [])

        for field, check in checks.items():
            if field in hide_fields:
                continue

            value = self.input_vars[field].get()
            if not check(value):
                self.widgets[field][0].config(foreground="red")
                error_fields.append(field)
            else:
                self.widgets[field][0].config(foreground="black")

        return error_fields

#Need to move function to filehandler
def create_excels_from_dicts(data_dicts, reference_excel_path, output_excel_path):
    ref_ext = os.path.splitext(reference_excel_path)[1].lower()
    out_ext = os.path.splitext(output_excel_path)[1].lower()
    if ref_ext != out_ext:
        raise ValueError("Reference and output file extensions must match (.xlsx or .xlsm)")

    shutil.copy(reference_excel_path, output_excel_path)

    wb = load_workbook(output_excel_path, keep_vba=(ref_ext == ".xlsm"))
    hidden_sheet_name = "MiscData"

    template_sheet = None
    for sheet in wb.worksheets:
        if sheet.title != hidden_sheet_name and sheet.sheet_state == 'visible':
            template_sheet = sheet
            break

    if template_sheet is None:
        raise ValueError("No visible template sheet found (excluding 'MiscData').")

    if not template_sheet.tables:
        raise ValueError("Template sheet does not contain a table structure.")

    template_table = list(template_sheet.tables.values())[0]
    table_range = template_table.ref
    base_table_name = template_table.name
    template_validations = list(template_sheet.data_validations.dataValidation)

    for sheet in wb.worksheets[:]:
        if sheet.title != hidden_sheet_name:
            wb.remove(sheet)

    for i, param_dict in enumerate(data_dicts):
        new_sheet = wb.copy_worksheet(template_sheet)
        new_sheet.title = f"Experiment_{i}"
        new_table_name = f"{base_table_name}_{i}"  # Shorter unique name

        # Fix broken formulas by updating table names in formula strings
        for row in new_sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    if f"{base_table_name}[" in cell.value:
                        cell.value = cell.value.replace(f"{base_table_name}[", f"{new_table_name}[")

        # Re-apply data validation
        for dv in template_validations:
            new_dv = DataValidation(
                type=dv.type,
                formula1=dv.formula1,
                formula2=dv.formula2,
                allow_blank=dv.allow_blank,
                showDropDown=dv.showDropDown,
                showInputMessage=dv.showInputMessage,
                promptTitle=dv.promptTitle,
                prompt=dv.prompt,
                showErrorMessage=dv.showErrorMessage,
                errorTitle=dv.errorTitle,
                error=dv.error
            )
            for cell_range in dv.cells.ranges:
                new_dv.add(cell_range.coord)
            new_sheet.add_data_validation(new_dv)

        # Create new table with updated name
        new_table = Table(displayName=new_table_name, ref=table_range)
        new_sheet.add_table(new_table)

        # Fill column B based on keys in column A
        for row in range(3, new_sheet.max_row + 1):
            key = new_sheet[f"A{row}"].value
            if key in param_dict:
                new_sheet[f"B{row}"].value = param_dict[key]

    wb.save(output_excel_path)


if __name__ == "__main__":
    app = ExperimentApp()
    app.mainloop()