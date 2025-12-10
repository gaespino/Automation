import threading
import random
import time
import struct
import ipccli
import namednodes
import csv
from datetime import datetime
import os
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not available. Install with: pip install openpyxl")

# PythonSV imports
from namednodes import sv
sv.initialize()

# IPC object for thread access
ipc = ipccli.baseaccess()

# Supported data sizes (in bytes)
DATA_SIZES = {
    1: 'BYTE (8-bit)',
    2: 'WORD (16-bit)',
    4: 'DWORD (32-bit)',
    8: 'QWORD (64-bit)',
    16: '128-bit (XMM)',
    32: '256-bit (YMM)',
    64: '512-bit (ZMM) - 2 BL pumps'
}

# Default patterns - can be modified or extended
DEFAULT_PATTERNS = [
    0xA5A5A5A5A5A5A5A5,  # Walking ones pattern
    0x5A5A5A5A5A5A5A5A,  # Walking zeros pattern
    0xFFFFFFFFFFFFFFFF,  # All ones
    0x0000000000000000,  # All zeros
    0xAAAAAAAAAAAAAAAA,  # Alternating pattern
    0x5555555555555555,  # Alternating pattern (inverse)
    0x0F0F0F0F0F0F0F0F,  # Nibble pattern
    0xF0F0F0F0F0F0F0F0,  # Nibble pattern (inverse)
]

# CHA Address Parity Stress Modes
CHA_STRESS_MODES = {
    'sequential': 'Sequential address access (baseline)',
    'walking_addr': 'Walking 1s through address bits',
    'addr_inversion': 'Alternating address bit patterns',
    'cacheline_stride': 'Jump between cache lines',
    'same_set': 'Target same cache set, different ways',
    'bank_interleave': 'Stress address interleaving across banks',
    'random_addr': 'Pseudo-random address patterns',
}

# Operation types for memory manipulation
OPERATIONS = {
    'write': 'Simple write pattern to memory',
    'add': 'Add value to existing memory content',
    'xor': 'XOR pattern with existing memory content',
    'multiply': 'Multiply existing memory content (treating as unsigned int)',
    'rotate': 'Rotate bits in existing memory content',
    'atomic_add': 'Atomic add operation (lock prefix)',
    'atomic_xchg': 'Atomic exchange operation',
    'atomic_cmpxchg': 'Atomic compare-and-exchange',
}


def mask_value(value, data_size):
    """Ensure value fits in specified data size."""
    if data_size == 1:
        return value & 0xFF
    elif data_size == 2:
        return value & 0xFFFF
    elif data_size == 4:
        return value & 0xFFFFFFFF
    elif data_size == 8:
        return value & 0xFFFFFFFFFFFFFFFF
    else:
        # For larger sizes, mask to appropriate bit width
        mask = (1 << (data_size * 8)) - 1
        return value & mask


def mask_64bit(value):
    """Ensure value fits in 64-bit unsigned integer (backward compatibility)."""
    return value & 0xFFFFFFFFFFFFFFFF


def generate_cha_stress_addresses(addr_start, addr_end, addr_step, mode='sequential', count=None):
    """
    Generate address patterns specifically for CHA address parity stress testing.
    
    Args:
        addr_start: Starting address
        addr_end: Ending address
        addr_step: Base step size
        mode: CHA stress mode (from CHA_STRESS_MODES)
        count: Number of addresses to generate (None = use full range)
    
    Returns:
        list: Generated addresses
    """
    if mode == 'sequential':
        # Standard sequential access
        return list(range(addr_start, addr_end, addr_step))
    
    elif mode == 'walking_addr':
        # Walking 1s through address bits - stresses address parity checking
        addrs = []
        base_addr = addr_start & ~0xFFF  # Align to 4KB
        for bit in range(12, 48):  # Address bits 12-47 (skip page offset)
            addr = base_addr | (1 << bit)
            if addr_start <= addr < addr_end:
                addrs.append(addr)
        return addrs
    
    elif mode == 'addr_inversion':
        # Alternating address patterns - stresses address bit flips
        addrs = []
        for addr in range(addr_start, addr_end, addr_step):
            addrs.append(addr)
            # Flip address bits in different patterns
            flipped = addr ^ 0x0000000001C0  # Flip bits in address field
            if addr_start <= flipped < addr_end:
                addrs.append(flipped)
        return addrs[:count] if count else addrs
    
    elif mode == 'cacheline_stride':
        # Large strides to stress different cache lines
        # Use prime number stride to avoid set aliasing
        stride = 4096 + 64  # Page + cacheline offset
        return list(range(addr_start, addr_end, stride))
    
    elif mode == 'same_set':
        # Target same cache set with different tags - stresses set indexing
        # Assuming 11-bit set index (2048 sets), 6-bit offset (64B line)
        addrs = []
        base = addr_start & ~0x3F  # Align to cache line
        set_bits = base & 0x1FFC0  # Extract set index bits
        for way in range(64):  # Create conflicts in same set
            addr = (base & ~0x1FFC0) | set_bits | (way << 17)
            if addr_start <= addr < addr_end:
                addrs.append(addr)
        return addrs
    
    elif mode == 'bank_interleave':
        # Access patterns that stress memory bank interleaving
        addrs = []
        bank_stride = 256  # Typical bank interleave granularity
        for offset in range(0, min(addr_end - addr_start, 64 * 1024), bank_stride):
            for bank in range(8):  # Cycle through banks
                addr = addr_start + offset + (bank * 64)
                if addr < addr_end:
                    addrs.append(addr)
        return addrs[:count] if count else addrs
    
    elif mode == 'random_addr':
        # Pseudo-random addresses within range
        import hashlib
        addrs = []
        addr_range = addr_end - addr_start
        num_addrs = count if count else min(100000, addr_range // addr_step)
        for i in range(num_addrs):
            # Use hash for deterministic pseudo-random
            hash_val = int(hashlib.md5(str(i).encode()).hexdigest()[:16], 16)
            addr = addr_start + ((hash_val % (addr_range // 64)) * 64)
            addrs.append(addr)
        return addrs
    
    else:
        # Default to sequential
        return list(range(addr_start, addr_end, addr_step))


def perform_operation(thread, addr, pattern, operation='write', original_value=None, data_size=8):
    """
    Perform specified operation on memory address.
    
    Args:
        thread: Thread object to perform operation
        addr: Memory address
        pattern: Pattern/value to use in operation
        operation: Type of operation to perform
        original_value: Original memory value (for operations that need it)
        data_size: Size of data transfer in bytes (1, 2, 4, 8, 16, 32, 64)
    
    Returns:
        tuple: (expected_value, actual_value_written)
    """
    # Convert address to hex string for IPC API
    addr_str = hex(addr) if isinstance(addr, int) else addr
    
    # Mask pattern to data size
    pattern = mask_value(pattern, data_size)
    
    if operation == 'write':
        # For sizes > 8 bytes, write multiple 8-byte chunks
        if data_size <= 8:
            thread.mem(addr_str, data_size, pattern)
            expected = pattern
            actual = pattern
        else:
            # Write in 8-byte chunks for larger sizes
            for offset in range(0, data_size, 8):
                chunk_addr = hex(addr + offset) if isinstance(addr, int) else hex(int(addr, 16) + offset)
                # For each chunk, use pattern (could be extended to vary per chunk)
                thread.mem(chunk_addr, 8, pattern)
            expected = pattern
            actual = pattern
        
    elif operation == 'add':
        # Read current value, add pattern, write back
        read_size = min(data_size, 8)  # Limit to 8 bytes for arithmetic ops
        current = original_value if original_value is not None else thread.mem(addr_str, read_size)
        result = mask_value(current + pattern, data_size)
        thread.mem(addr_str, read_size, result)
        expected = result
        actual = result
        
    elif operation == 'xor':
        # XOR pattern with current value
        read_size = min(data_size, 8)
        current = original_value if original_value is not None else thread.mem(addr_str, read_size)
        result = mask_value(current ^ pattern, data_size)
        thread.mem(addr_str, read_size, result)
        expected = result
        actual = result
        
    elif operation == 'multiply':
        # Multiply (use lower 32-bits to avoid overflow issues)
        read_size = min(data_size, 8)
        current = original_value if original_value is not None else thread.mem(addr_str, read_size)
        # Use lower 32-bits for multiplication to keep it reasonable
        low_current = current & 0xFFFFFFFF
        low_pattern = pattern & 0xFFFFFFFF
        result = mask_value(low_current * low_pattern, data_size)
        thread.mem(addr_str, read_size, result)
        expected = result
        actual = result
        
    elif operation == 'rotate':
        # Rotate left by pattern & appropriate bit count for data size
        read_size = min(data_size, 8)
        current = original_value if original_value is not None else thread.mem(addr_str, read_size)
        bit_width = data_size * 8 if data_size <= 8 else 64
        shift = pattern & (bit_width - 1)
        result = mask_value((current << shift) | (current >> (bit_width - shift)), data_size)
        thread.mem(addr_str, read_size, result)
        expected = result
        actual = result
        
    elif operation == 'atomic_add':
        # Atomic add using lock xadd instruction
        # Note: This is approximated - actual implementation may vary
        read_size = min(data_size, 8)
        size_suffix = {1: 'byte', 2: 'word', 4: 'dword', 8: 'qword'}.get(read_size, 'qword')
        try:
            current = thread.mem(addr_str, read_size)
            # Use inline assembly for true atomic operation
            thread.asm("$", f"lock xadd {size_suffix} ptr [{addr_str}], {pattern}")
            expected = mask_value(current + pattern, data_size)
            actual = thread.mem(addr_str, read_size)
        except:
            # Fallback to regular add if assembly fails
            current = thread.mem(addr_str, read_size)
            result = mask_value(current + pattern, data_size)
            thread.mem(addr_str, read_size, result)
            expected = result
            actual = result
            
    elif operation == 'atomic_xchg':
        # Atomic exchange
        read_size = min(data_size, 8)
        size_suffix = {1: 'byte', 2: 'word', 4: 'dword', 8: 'qword'}.get(read_size, 'qword')
        try:
            thread.asm("$", f"xchg {size_suffix} ptr [{addr_str}], {pattern}")
            expected = pattern
            actual = thread.mem(addr_str, read_size)
        except:
            thread.mem(addr_str, read_size, pattern)
            expected = pattern
            actual = pattern
            
    elif operation == 'atomic_cmpxchg':
        # Atomic compare and exchange (compare with original_value, exchange with pattern)
        read_size = min(data_size, 8)
        size_suffix = {1: 'byte', 2: 'word', 4: 'dword', 8: 'qword'}.get(read_size, 'qword')
        try:
            compare_val = original_value if original_value is not None else 0
            thread.asm("$", f"lock cmpxchg {size_suffix} ptr [{addr_str}], {pattern}")
            actual = thread.mem(addr_str, read_size)
            # If original value matched compare_val, it will be replaced with pattern
            expected = pattern if compare_val == original_value else original_value
        except:
            thread.mem(addr_str, read_size, pattern)
            expected = pattern
            actual = pattern
    else:
        raise ValueError(f"Unknown operation: {operation}")
    
    return expected, actual

def stress_worker(thread_idx, threads_list, addr_list, patterns, iterations, 
                  operation='write', verify=False, use_wbinvd=False, 
                  randomize=False, verbose=True, original_values=None, restore_memory=False,
                  log_data=None, race_tolerant=False, valid_patterns=None, 
                  last_writer_tracker=None, deterministic_seed=None, data_size=8):
    """
    Worker function for each thread to stress memory.
    
    Args:
        thread_idx: Index of the thread
        threads_list: List of all available threads
        addr_list: List of addresses this thread should stress
        patterns: List of 64-bit patterns to write
        iterations: Number of iterations to perform
        operation: Type of operation to perform (write, add, xor, multiply, rotate, atomic_*)
        verify: If True, read back and verify each operation (slower)
        use_wbinvd: If True, flush cache after each write (forces memory traffic, very slow)
        randomize: If True, randomize address order each iteration (reduces performance)
        verbose: If True, print progress and errors
        original_values: Dict mapping addresses to their original values
        restore_memory: If True, restore original values at the end
        log_data: Shared list for logging operation results
        last_writer_tracker: Shared dict to track last write per address (for predictable verification)
        deterministic_seed: Seed for deterministic pattern selection
    """
    thread = threads_list[thread_idx]
    failures = 0
    operation_count = 0
    thread_log = []  # Local log for this thread
    
    # Use deterministic random if seed provided
    if deterministic_seed is not None:
        local_rng = random.Random(deterministic_seed + thread_idx)
    
    for iteration in range(iterations):
        if verbose and iteration % 10 == 0 and len(addr_list) > 0:
            print(f"Thread {thread_idx}: Iteration {iteration+1}/{iterations}")
        
        # Optionally randomize address order
        if randomize:
            random.shuffle(addr_list)
        
        for addr in addr_list:
            # Use deterministic pattern selection if seed provided
            if deterministic_seed is not None:
                pattern = local_rng.choice(patterns)
            else:
                pattern = random.choice(patterns)
            
            original_val = original_values.get(addr, None) if original_values else None
            
            # Perform the specified operation
            expected, actual = perform_operation(
                thread, addr, pattern, operation, original_val, data_size
            )
            operation_count += 1
            
            # Track last writer for this address (thread-safe dict update)
            if last_writer_tracker is not None:
                last_writer_tracker[addr] = {
                    'thread': thread_idx,
                    'pattern': pattern,
                    'expected': expected,
                    'iteration': iteration,
                    'operation': operation
                }
            
            # Optional: flush cache to force memory traffic
            if use_wbinvd:
                thread.wbinvd()
            
            # Verify operation result
            addr_hex = hex(addr) if isinstance(addr, int) else addr
            status = 'PASS'
            read_val = None
            
            if verify:
                read_size = min(data_size, 8)  # Read up to 8 bytes for verification
                read_val = thread.mem(addr_hex, read_size)
                
                # Determine expected value based on last writer if tracking enabled
                if last_writer_tracker is not None and addr in last_writer_tracker:
                    last_write = last_writer_tracker[addr]
                    actual_expected = last_write['expected']
                    # Check if we're the last writer or someone else
                    if last_write['thread'] != thread_idx:
                        # Another thread wrote after us - this is expected race
                        if read_val == actual_expected:
                            status = 'RACE_EXPECTED'  # Read shows last writer's value
                        elif race_tolerant and valid_patterns and read_val in valid_patterns:
                            status = 'RACE'  # Read shows some other valid pattern
                        else:
                            failures += 1
                            status = 'FAIL'  # Truly unexpected value
                    else:
                        # We're the last writer - should read our value
                        if read_val == expected:
                            status = 'PASS'
                        elif race_tolerant and valid_patterns and read_val in valid_patterns:
                            status = 'RACE'  # Another thread wrote after our tracking update
                        else:
                            failures += 1
                            status = 'FAIL'
                else:
                    # No tracking - use original logic
                    if read_val != expected:
                        # Check if this is a race condition (read value is a valid pattern from another thread)
                        if race_tolerant and valid_patterns and read_val in valid_patterns:
                            status = 'RACE'  # Race condition - another thread wrote first
                        else:
                            failures += 1
                            status = 'FAIL'  # True failure - unexpected value
                    
                if verbose and len(addr_list) <= 10 and status == 'FAIL':  # Only print real failures
                    print(f"  Thread {thread_idx} Iter {iteration}: {addr_hex} "
                          f"wrote {hex(pattern)}, expected {hex(expected)}, read {hex(read_val)} - {status}")
            
            # Log operation details
            if log_data is not None:
                thread_log.append({
                    'thread': thread_idx,
                    'iteration': iteration,
                    'address': addr_hex,
                    'original_value': hex(original_val) if original_val is not None else 'N/A',
                    'pattern': hex(pattern),
                    'operation': operation,
                    'expected': hex(expected),
                    'read_value': hex(read_val) if read_val is not None else 'N/A',
                    'status': status
                })
    
    # Restore original memory values if requested
    if restore_memory and original_values:
        if verbose and len(addr_list) > 0:
            print(f"Thread {thread_idx}: Restoring {len(addr_list)} original values...")
        for addr in addr_list:
            if addr in original_values:
                restore_size = min(data_size, 8)
                thread.mem(hex(addr), restore_size, original_values[addr])
    
    # Add thread log to shared log
    if log_data is not None:
        log_data.extend(thread_log)
    
    if verbose and len(addr_list) > 0:
        status = f"Thread {thread_idx} complete: {operation_count} operations"
        if verify:
            status += f", {failures} failures"
        if restore_memory:
            status += ", memory restored"
        print(status)
    
    return failures


def validate_memory_integrity(addr_list, threads, test_pattern=0xDEADBEEFDEADBEEF, verbose=False):
    """
    Validate memory integrity by writing a known pattern and verifying all threads read consistently.
    Use before/after stress testing to detect real hardware failures.
    
    Args:
        addr_list: List of addresses to test
        threads: List of thread objects
        test_pattern: Pattern to write (default: 0xDEADBEEFDEADBEEF)
        verbose: Print detailed results
    
    Returns:
        list: Failures detected (empty list = PASS)
    """
    failures = []
    
    if verbose:
        print(f"\nValidating memory integrity on {len(addr_list):,} addresses...")
    
    for addr in addr_list:
        addr_hex = hex(addr) if isinstance(addr, int) else addr
        
        # Write pattern with thread 0
        threads[0].mem(addr_hex, 8, test_pattern)
        
        # Verify all threads read the same value
        for idx, thread in enumerate(threads):
            read_val = thread.mem(addr_hex, 8)
            if read_val != test_pattern:
                failures.append({
                    'address': addr_hex,
                    'thread': idx,
                    'expected': hex(test_pattern),
                    'read': hex(read_val),
                    'type': 'COHERENCY_FAILURE'
                })
                if verbose:
                    print(f"  FAIL: Thread {idx} @ {addr_hex}: expected {hex(test_pattern)}, read {hex(read_val)}")
    
    if verbose:
        if failures:
            print(f"  Result: FAILED - {len(failures)} coherency errors detected")
        else:
            print(f"  Result: PASSED - All threads read consistently")
    
    return failures


def run_memory_stress(addr_start, addr_end, addr_step=0x40, 
                     patterns=None, iterations=100,
                     operation='write', socket=0, num_threads=None,
                     verify=False, use_wbinvd=False, randomize=False, 
                     restore_memory=True, verbose=True,
                     cha_stress_mode=None, interleave_threads=True,
                     log_to_csv=True, csv_filename=None,
                     race_tolerant=True, pre_post_validation=False,
                     deterministic=False, track_last_writer=False,
                     use_excel=True, use_txt=False, data_size=8):
    """
    Run memory stress test across multiple threads.
    
    Args:
        addr_start: Starting address (hex or int)
        addr_end: Ending address (hex or int)
        addr_step: Address stride/step size (default 0x40 = 64 bytes, cache line size)
        patterns: List of 64-bit patterns to use (default: DEFAULT_PATTERNS)
        iterations: Number of iterations per thread (default: 100)
        operation: Type of operation - 'write', 'add', 'xor', 'multiply', 'rotate',
                  'atomic_add', 'atomic_xchg', 'atomic_cmpxchg' (default: 'write')
        socket: Socket number to use (default: 0)
        num_threads: Number of threads to use (default: all available on socket)
        verify: Read back and verify each operation (default: False, reduces performance)
        use_wbinvd: Flush cache after each write (default: False, greatly reduces performance)
        randomize: Randomize address order each iteration (default: False, reduces performance)
        restore_memory: Restore original memory values at end (default: True)
        verbose: Print progress messages (default: True)
        cha_stress_mode: CHA address parity stress mode - 'walking_addr', 'addr_inversion',
                        'cacheline_stride', 'same_set', 'bank_interleave', 'random_addr'
                        (default: None = sequential)
        interleave_threads: Interleave addresses across threads vs block assignment (default: True)
        log_to_csv: Save detailed operation log to file (default: True)
        csv_filename: Custom filename (default: auto-generated with timestamp, .xlsx or .csv)
        race_tolerant: When True, don't count race conditions as failures (default: True)
                      Race = read value is valid pattern from another thread
        pre_post_validation: Run memory integrity checks before/after stress (default: False)
                            Detects real hardware failures vs race conditions
        deterministic: Use deterministic pattern selection (reproducible results) (default: False)
        track_last_writer: Track last writer per address for predictable verification (default: False)
                          When enabled, can predict expected value based on write ordering
        use_excel: Save as Excel (.xlsx) with multiple sheets and formatting (default: True)
                  Falls back to CSV if openpyxl not available. TXT always generated.
        use_txt: Deprecated - TXT report now always generated for quick viewing
        data_size: Size of data transfer in bytes: 1, 2, 4, 8, 16, 32, 64 (default: 8)
                  64 bytes = 512-bit (2 BL ring pumps), 32 bytes = 256-bit, etc.
    
    Returns:
        dict: Statistics about the stress test including report_file path, pre/post validation results
    """
    if patterns is None:
        patterns = DEFAULT_PATTERNS
    
    # Validate parameters
    if operation not in OPERATIONS:
        raise ValueError(f"Unknown operation '{operation}'. Valid operations: {list(OPERATIONS.keys())}")
    
    if data_size not in DATA_SIZES:
        raise ValueError(f"Invalid data_size '{data_size}'. Valid sizes: {list(DATA_SIZES.keys())} bytes")
    
    # Get threads for the specified socket
    if socket == 0:
        ipc_obj = ipc
    elif socket == 1:
        ipc_obj = ipc
    else:
        raise ValueError(f"Invalid socket number: {socket}")
    
    # Get available threads
    if num_threads is None:
        # Use all available threads
        threads_list = ipc_obj.threads
    else:
        # Limit to specified number of threads
        threads_list = ipc_obj.threads[:num_threads]
    
    num_threads_actual = len(threads_list)
    
    # Build address list using CHA stress mode if specified
    if cha_stress_mode and cha_stress_mode in CHA_STRESS_MODES:
        if verbose:
            print(f"Using CHA stress mode: {cha_stress_mode} - {CHA_STRESS_MODES[cha_stress_mode]}")
        addr_list = generate_cha_stress_addresses(addr_start, addr_end, addr_step, cha_stress_mode)
    else:
        addr_list = list(range(addr_start, addr_end, addr_step))
    
    total_addresses = len(addr_list)
    
    # Pre-stress validation
    pre_validation_failures = []
    if pre_post_validation:
        if verbose:
            print(f"\n{'='*60}")
            print(f"PRE-STRESS VALIDATION")
            print(f"{'='*60}")
        pre_validation_failures = validate_memory_integrity(addr_list, threads_list, verbose=verbose)
    
    # Save original memory values if restore is requested
    original_values = {}
    if restore_memory:
        if verbose:
            print(f"Reading original memory values from {total_addresses:,} addresses...")
        # Use first thread to read original values
        read_size = min(data_size, 8)
        for addr in addr_list:
            original_values[addr] = threads_list[0].mem(hex(addr), read_size)
        if verbose:
            print(f"Original values saved for {len(original_values):,} addresses\n")
    
    if verbose:
        print(f"\n{'='*60}")
        print(f"Memory Stress Configuration:")
        print(f"  Address Range: {hex(addr_start)} to {hex(addr_end)}")
        print(f"  Address Step: {hex(addr_step)} ({addr_step} bytes)")
        print(f"  Total Addresses: {total_addresses:,}")
        print(f"  Memory Range: {(addr_end - addr_start) / (1024**3):.2f} GB")
        print(f"  Operation: {operation} - {OPERATIONS[operation]}")
        print(f"  Data Size: {data_size} bytes - {DATA_SIZES[data_size]}")
        print(f"  Patterns: {len(patterns)} patterns")
        print(f"  Iterations: {iterations}")
        print(f"  Threads: {num_threads_actual}")
        print(f"  Verify: {verify}")
        print(f"  Cache Flush: {use_wbinvd}")
        print(f"  Randomize: {randomize}")
        print(f"  Restore Memory: {restore_memory}")
        if total_addresses < num_threads_actual and interleave_threads:
            print(f"  Mode: All threads on same addresses (max contention)")
        print(f"  Total Operations: {total_addresses * iterations * num_threads_actual:,}")
        print(f"{'='*60}\n")
    
    # Initialize logging
    log_data = [] if log_to_csv or verify else None
    
    # Prepare set of valid patterns for race detection
    valid_patterns_set = set(patterns) if race_tolerant and verify else None
    
    # Initialize last writer tracking (shared across threads)
    last_writer_tracker = {} if track_last_writer else None
    deterministic_seed = hash((addr_start, addr_end, iterations)) if deterministic else None
    
    # Generate filename if needed
    timestamp = None
    if log_to_csv and csv_filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        # Ensure C:\Temp directory exists
        temp_dir = r"C:\Temp"
        if not os.path.exists(temp_dir):
            os.makedirs(temp_dir)
        
        # Use Excel if available and requested, otherwise CSV
        if use_excel and EXCEL_AVAILABLE:
            csv_filename = os.path.join(temp_dir, f"{timestamp}_memstress.xlsx")
        else:
            csv_filename = os.path.join(temp_dir, f"{timestamp}_memstress.csv")
            if use_excel and not EXCEL_AVAILABLE:
                print("Note: Excel format requested but openpyxl not available. Using CSV format.")
    
    # Split address list among threads
    thread_objects = []
    
    # Adjust thread count if we have fewer addresses than threads
    if total_addresses < num_threads_actual:
        if verbose:
            print(f"Note: Only {total_addresses} addresses, using {total_addresses} threads instead of {num_threads_actual}")
        # For single/few addresses, have all threads work on same addresses for maximum stress
        if interleave_threads:
            # All threads get all addresses for maximum contention
            thread_addr_lists = [addr_list[:] for _ in range(num_threads_actual)]
        else:
            # Assign one address per thread, remaining threads get nothing
            thread_addr_lists = [[addr_list[i]] if i < total_addresses else [] for i in range(num_threads_actual)]
    else:
        # Normal case: more addresses than threads
        if interleave_threads:
            # Round-robin distribution - thread 0 gets addr[0,n,2n...], thread 1 gets addr[1,n+1,2n+1...]
            thread_addr_lists = [addr_list[i::num_threads_actual] for i in range(num_threads_actual)]
        else:
            # Block distribution - divide address space into contiguous chunks
            chunk_size = total_addresses // num_threads_actual
            thread_addr_lists = []
            for i in range(num_threads_actual):
                if i < num_threads_actual - 1:
                    thread_addr_lists.append(addr_list[i*chunk_size:(i+1)*chunk_size])
                else:
                    thread_addr_lists.append(addr_list[i*chunk_size:])
    
    start_time = time.time()
    
    for i in range(num_threads_actual):
        chunk = thread_addr_lists[i]
        
        t = threading.Thread(
            target=stress_worker, 
            args=(i, threads_list, chunk, patterns, iterations, 
                  operation, verify, use_wbinvd, randomize, verbose, 
                  original_values, restore_memory, log_data, 
                  race_tolerant, valid_patterns_set,
                  last_writer_tracker, deterministic_seed, data_size)
        )
        thread_objects.append(t)
        t.start()
    
    # Wait for all threads to complete
    for t in thread_objects:
        t.join()
    
    end_time = time.time()
    elapsed = end_time - start_time
    
    total_ops = total_addresses * iterations * num_threads_actual
    ops_per_sec = total_ops / elapsed if elapsed > 0 else 0
    
    # Post-stress validation
    post_validation_failures = []
    if pre_post_validation:
        if verbose:
            print(f"\n{'='*60}")
            print(f"POST-STRESS VALIDATION")
            print(f"{'='*60}")
        post_validation_failures = validate_memory_integrity(addr_list, threads_list, verbose=verbose)
    
    # Calculate pass/fail/race statistics
    total_failures = 0
    total_passes = 0
    total_races = 0
    total_races_expected = 0
    failure_analysis = None
    if log_data:
        total_failures = sum(1 for entry in log_data if entry['status'] == 'FAIL')
        total_passes = sum(1 for entry in log_data if entry['status'] == 'PASS')
        total_races = sum(1 for entry in log_data if entry['status'] == 'RACE')
        total_races_expected = sum(1 for entry in log_data if entry['status'] == 'RACE_EXPECTED')
        
        # Analyze failure patterns if we have real failures
        if total_failures > 0:
            from collections import Counter
            failures_list = [e for e in log_data if e['status'] == 'FAIL']
            
            addr_failures = Counter(e['address'] for e in failures_list)
            thread_failures = Counter(e['thread'] for e in failures_list)
            
            # Detect hot spots (>10% of failures at same location)
            hot_addresses = [addr for addr, count in addr_failures.items() 
                           if count > total_failures * 0.1]
            hot_threads = [tid for tid, count in thread_failures.items()
                         if count > total_failures * 0.1]
            
            failure_analysis = {
                'hot_addresses': hot_addresses,
                'hot_threads': hot_threads,
                'addr_distribution': dict(addr_failures.most_common(5)),
                'thread_distribution': dict(thread_failures.most_common(5))
            }
    
    # Determine final verdict (must be calculated before report_data)
    if pre_post_validation:
        # Pre/post validation is the ground truth
        if pre_validation_failures:
            test_verdict = "FAIL - PRE-VALIDATION"
        elif post_validation_failures:
            test_verdict = "FAIL - POST-VALIDATION (Hardware failure detected)"
        else:
            test_verdict = "PASS"
    elif verify:
        test_verdict = "PASS" if total_failures == 0 else "FAIL"
    else:
        test_verdict = "NOT VERIFIED"
    
    # Prepare comprehensive report data
    report_data = {
        'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        'config': {
            'addr_start': hex(addr_start),
            'addr_end': hex(addr_end),
            'addr_step': hex(addr_step),
            'total_addresses': total_addresses,
            'memory_range_gb': (addr_end - addr_start) / (1024**3),
            'operation': operation,
            'data_size': f"{data_size} bytes - {DATA_SIZES[data_size]}",
            'patterns_count': len(patterns),
            'iterations': iterations,
            'threads': num_threads_actual,
            'verify': verify,
            'cache_flush': use_wbinvd,
            'randomize': randomize,
            'restore_memory': restore_memory,
            'cha_stress_mode': cha_stress_mode if cha_stress_mode else 'sequential',
            'interleave_threads': interleave_threads,
            'race_tolerant': race_tolerant,
            'pre_post_validation': pre_post_validation,
            'deterministic': deterministic,
            'track_last_writer': track_last_writer,
        },
        'results': {
            'elapsed_time': elapsed,
            'total_operations': total_ops,
            'ops_per_second': ops_per_sec,
            'test_verdict': test_verdict,
            'passes': total_passes,
            'failures': total_failures,
            'races': total_races,
            'races_expected': total_races_expected,
            'pre_validation_failures': len(pre_validation_failures) if pre_post_validation else 0,
            'post_validation_failures': len(post_validation_failures) if pre_post_validation else 0,
        },
        'failure_analysis': failure_analysis,
        'log_data': log_data,
        'pre_validation_failures': pre_validation_failures,
        'post_validation_failures': post_validation_failures,
        'last_writer_tracker': last_writer_tracker,
        'patterns': patterns,
    }
    
    # Write to Excel/CSV and always write TXT
    report_path = None
    txt_report_path = None
    if log_to_csv:
        report_path = os.path.abspath(csv_filename)
        
        # Write primary format (Excel or CSV)
        if use_excel and EXCEL_AVAILABLE and report_path.endswith('.xlsx'):
            _write_excel_report(report_path, report_data, verbose)
        else:
            _write_csv_report(report_path, log_data, verbose)
        
        # Always write TXT report for quick viewing (especially on stations without Excel)
        if timestamp is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        temp_dir = r"C:\Temp"
        if not os.path.exists(temp_dir):
            os.makedirs(temp_dir)
        txt_report_path = os.path.join(temp_dir, f"{timestamp}_memstress.txt")
        _write_txt_report(txt_report_path, report_data, verbose)
    
    if verbose:
        print(f"\n{'='*60}")
        print(f"Memory Stress Complete:")
        print(f"  Elapsed Time: {elapsed:.2f} seconds")
        print(f"  Operations: {total_ops:,}")
        print(f"  Ops/Second: {ops_per_sec:,.0f}")
        if verify:
            print(f"  Passes: {total_passes:,}")
            if track_last_writer and total_races_expected > 0:
                print(f"  Expected Races: {total_races_expected:,} (read showed last writer's value)")
            if race_tolerant and total_races > 0:
                print(f"  Races: {total_races:,} (normal for multi-thread same address)")
            print(f"  Failures: {total_failures:,}")
            total_checked = total_passes + total_failures + total_races + total_races_expected
            if total_checked > 0:
                success_count = total_passes + total_races + total_races_expected
                print(f"  Success Rate: {(success_count/total_checked*100):.2f}%")
        if restore_memory:
            print(f"  Memory Restored: Yes ({len(original_values):,} addresses)")
        print(f"")
        
        # Show validation results
        if pre_post_validation:
            print(f"  Pre-Validation: {'PASS' if not pre_validation_failures else f'FAIL ({len(pre_validation_failures)} errors)'}")
            print(f"  Post-Validation: {'PASS' if not post_validation_failures else f'FAIL ({len(post_validation_failures)} errors)'}")
        
        print(f"  FINAL VERDICT: {test_verdict}")
        
        if test_verdict.startswith("FAIL") and verbose:
            if pre_validation_failures:
                print(f"\n  WARNING: Memory was already corrupted before stress test!")
            elif post_validation_failures:
                print(f"\n  HARDWARE FAILURE: Memory coherency broken after stress!")
                print(f"  Addresses with issues: {len(post_validation_failures)}")
            elif total_failures > 0:
                print(f"  Failure Details: {total_failures} operations failed verification")
                if failure_analysis:
                    print(f"\n  Failure Analysis:")
                    if failure_analysis['hot_addresses']:
                        print(f"    Hot Addresses (>10% failures): {failure_analysis['hot_addresses'][:3]}")
                    if failure_analysis['hot_threads']:
                        print(f"    Hot Threads (>10% failures): {failure_analysis['hot_threads'][:5]}")
                if race_tolerant and total_races > 0:
                    print(f"  Note: {total_races} race conditions detected (expected for simultaneous access)")
            if log_to_csv:
                print(f"  Check detailed report: {report_path}")
                if txt_report_path:
                    print(f"  Quick view TXT report: {txt_report_path}")
        print(f"{'='*60}\n")
    
    return {
        'elapsed_time': elapsed,
        'total_operations': total_ops,
        'ops_per_second': ops_per_sec,
        'num_threads': num_threads_actual,
        'iterations': iterations,
        'addresses': total_addresses,
        'operation': operation,
        'memory_restored': restore_memory,
        'test_verdict': test_verdict,
        'passes': total_passes if verify else None,
        'failures': total_failures if verify else None,
        'races': total_races if verify else None,
        'races_expected': total_races_expected if (verify and track_last_writer) else None,
        'pre_validation_failures': len(pre_validation_failures) if pre_post_validation else None,
        'post_validation_failures': len(post_validation_failures) if pre_post_validation else None,
        'failure_analysis': failure_analysis,
        'last_writer_tracker': last_writer_tracker if track_last_writer else None,
        'report_file': report_path,
        'txt_report_file': txt_report_path,
        'csv_file': report_path  # Backward compatibility
    }


def main():
    """
    Example usage of the memory stress function.
    Modify these parameters as needed.
    """
    # Example 1: CHA Address Parity Stress - Walking address bits
    run_memory_stress(
        addr_start=0x1F0EAE6980,
        addr_end=0x1FB4A7D600,
        addr_step=0x40,
        iterations=100,
        operation='write',
        cha_stress_mode='walking_addr',
        interleave_threads=True,
        restore_memory=True
    )
    
    # Example 1b: Basic write stress test (memory will be restored)
    # run_memory_stress(
    #     addr_start=0x1F0EAE6980,
    #     addr_end=0x1FB4A7D600,
    #     addr_step=0x40,
    #     iterations=100,
    #     operation='write',
    #     restore_memory=True
    # )
    
    # Example 2: XOR operation with verification
    # run_memory_stress(
    #     addr_start=0x1F0EAE6980,
    #     addr_end=0x1FB4A7D600,
    #     addr_step=0x40,
    #     operation='xor',
    #     iterations=50,
    #     verify=True,
    #     restore_memory=True
    # )
    
    # Example 3: Atomic add operations for cache coherency stress
    # run_memory_stress(
    #     addr_start=0x1F0EAE6980,
    #     addr_end=0x1FB4A7D600,
    #     addr_step=0x40,
    #     operation='atomic_add',
    #     iterations=100,
    #     verify=True,
    #     restore_memory=True
    # )
    
    # Example 4: Multiply operations with custom patterns
    # custom_patterns = [
    #     0x0000000000000002,  # Multiply by 2
    #     0x0000000000000003,  # Multiply by 3
    #     0x0000000000000005,  # Multiply by 5
    # ]
    # run_memory_stress(
    #     addr_start=0x1F0EAE6980,
    #     addr_end=0x1FB4A7D600,
    #     operation='multiply',
    #     patterns=custom_patterns,
    #     iterations=50,
    #     verify=True,
    #     restore_memory=True
    # )
    
    # Example 5: Rotate operations (bit manipulation)
    # run_memory_stress(
    #     addr_start=0x1F0EAE6980,
    #     addr_end=0x1FB4A7D600,
    #     operation='rotate',
    #     patterns=[1, 8, 16, 32],  # Rotate by these bit positions
    #     iterations=50,
    #     verify=True,
    #     restore_memory=True
    # )
    
    # Example 6: Maximum throughput (no verification, no restore, no cache flush)
    # run_memory_stress(
    #     addr_start=0x1F0EAE6980,
    #     addr_end=0x1FB4A7D600,
    #     addr_step=0x40,
    #     iterations=1000,
    #     operation='write',
    #     verify=False,
    #     use_wbinvd=False,
    #     randomize=False,
    #     restore_memory=False
    # )
    
    # Example 7: Force memory traffic with cache flushes
    # run_memory_stress(
    #     addr_start=0x1F0EAE6980,
    #     addr_end=0x1FB4A7D600,
    #     addr_step=0x40,
    #     operation='write',
    #     iterations=50,
    #     use_wbinvd=True,
    #     verify=True,
    #     restore_memory=True
    # )
    
    # Example 8: CHA Address Parity - Address bit inversion
    # run_memory_stress(
    #     addr_start=0x1F0EAE6980,
    #     addr_end=0x1FB4A7D600,
    #     cha_stress_mode='addr_inversion',
    #     iterations=200,
    #     verify=True,
    #     interleave_threads=True
    # )
    
    # Example 9: CHA stress - Same cache set conflicts
    # run_memory_stress(
    #     addr_start=0x1F0EAE6980,
    #     addr_end=0x1FB4A7D600,
    #     cha_stress_mode='same_set',
    #     operation='atomic_add',
    #     iterations=500,
    #     verify=True
    # )
    
    # Example 10: CHA stress - Bank interleaving patterns
    # run_memory_stress(
    #     addr_start=0x1F0EAE6980,
    #     addr_end=0x1FB4A7D600,
    #     cha_stress_mode='bank_interleave',
    #     iterations=300,
    #     use_wbinvd=True,
    #     verify=True
    # )


def _write_csv_report(filepath, log_data, verbose=False):
    """Write simple CSV report with operation log."""
    if not log_data:
        return
    
    with open(filepath, 'w', newline='') as csvfile:
        fieldnames = ['thread', 'iteration', 'address', 'original_value', 'pattern', 
                     'operation', 'expected', 'read_value', 'status']
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(log_data)
    
    if verbose:
        print(f"\nDetailed log saved to: {filepath}")


def _write_txt_report(filepath, report_data, verbose=False):
    """Write comprehensive TXT report formatted for Consolas font in Notepad."""
    with open(filepath, 'w', encoding='utf-8') as f:
        # Header
        f.write("="* 100 + "\n")
        f.write("  MEMORY STRESS TEST REPORT\n")
        f.write("=" * 100 + "\n")
        f.write(f"Timestamp: {report_data['timestamp']}\n\n")
        
        # Configuration Section
        f.write("-" * 100 + "\n")
        f.write("  TEST CONFIGURATION\n")
        f.write("-" * 100 + "\n")
        config = report_data['config']
        for key, value in config.items():
            label = key.replace('_', ' ').title()
            f.write(f"  {label:.<35} {str(value):<60}\n")
        f.write("\n")
        
        # Results Section
        f.write("-" * 100 + "\n")
        f.write("  TEST RESULTS\n")
        f.write("-" * 100 + "\n")
        results = report_data['results']
        
        verdict = results['test_verdict']
        status_mark = "[PASS]" if verdict.startswith('PASS') else "[FAIL]"
        f.write(f"  Test Verdict:............................ {verdict} {status_mark}\n")
        f.write(f"  Elapsed Time:............................ {results['elapsed_time']:.2f} seconds\n")
        f.write(f"  Total Operations:........................ {results['total_operations']:,}\n")
        f.write(f"  Operations/Second:....................... {results['ops_per_second']:,.0f}\n")
        
        if config['verify']:
            f.write(f"\n  VERIFICATION STATISTICS:\n")
            f.write(f"    Passes:................................ {results['passes']:,}\n")
            
            if report_data['config']['track_last_writer'] and results['races_expected'] > 0:
                f.write(f"    Expected Races:........................ {results['races_expected']:,} (read correct last writer)\n")
            
            if results['races'] > 0:
                f.write(f"    Races (Unexpected Pattern):............ {results['races']:,}\n")
            
            f.write(f"    Failures:.............................. {results['failures']:,}")
            if results['failures'] == 0:
                f.write(" [OK]\n")
            else:
                f.write(" [ERROR]\n")
            
            total_checked = results['passes'] + results['failures'] + results['races'] + results['races_expected']
            if total_checked > 0:
                success_count = results['passes'] + results['races'] + results['races_expected']
                f.write(f"    Success Rate:.......................... {(success_count/total_checked*100):.2f}%\n")
        
        if config['restore_memory']:
            f.write(f"\n  Memory Restored:......................... Yes\n")
        
        # Validation Results
        if config['pre_post_validation']:
            f.write("\n" + "-" * 100 + "\n")
            f.write("  VALIDATION RESULTS\n")
            f.write("-" * 100 + "\n")
            
            pre_status = "[PASS]" if results['pre_validation_failures'] == 0 else f"[FAIL] ({results['pre_validation_failures']} errors)"
            post_status = "[PASS]" if results['post_validation_failures'] == 0 else f"[FAIL] ({results['post_validation_failures']} errors)"
            
            f.write(f"  Pre-Stress Validation:................... {pre_status}\n")
            f.write(f"  Post-Stress Validation:.................. {post_status}\n")
        
        # Failure Analysis
        if report_data['failure_analysis']:
            f.write("\n" + "-" * 100 + "\n")
            f.write("  FAILURE ANALYSIS\n")
            f.write("-" * 100 + "\n")
            fa = report_data['failure_analysis']
            
            if fa['hot_addresses']:
                f.write(f"  Hot Addresses (>10% failures):\n")
                for addr in fa['hot_addresses'][:5]:
                    f.write(f"    > {addr}\n")
            
            if fa['hot_threads']:
                f.write(f"  Hot Threads (>10% failures):\n")
                f.write(f"    Threads: {', '.join(str(t) for t in fa['hot_threads'][:10])}\n")
            
            f.write(f"\n  Top 5 Addresses by Failure Count:\n")
            for addr, count in fa['addr_distribution'].items():
                f.write(f"    {addr:.<50} {count:>5} failures\n")
        
        # Operation Log (sample)
        if report_data['log_data']:
            f.write("\n" + "-" * 100 + "\n")
            f.write("  OPERATION LOG (Sample - First 100 operations)\n")
            f.write("-" * 100 + "\n")
            f.write(f"{'Thread':<8}{'Iter':<6}{'Address':<20}{'Pattern':<20}{'Operation':<10}{'Expected':<20}{'Read':<20}{'Status':<15}\n")
            f.write("-" * 100 + "\n")
            
            for entry in report_data['log_data'][:100]:
                status_mark = {'PASS': '[OK]', 'FAIL': '[ERR]', 'RACE': '[~]', 'RACE_EXPECTED': '[=]'}.get(entry['status'], '[?]')
                f.write(f"{entry['thread']:<8}{entry['iteration']:<6}{entry['address']:<20}{entry['pattern'][:18]:<20}"
                       f"{entry['operation']:<10}{entry['expected'][:18]:<20}{entry['read_value'][:18]:<20}"
                       f"{status_mark} {entry['status']:<13}\n")
            
            if len(report_data['log_data']) > 100:
                f.write(f"\n  ... ({len(report_data['log_data']) - 100:,} more operations not shown)\n")
        
        # Pre-Validation Failures
        if report_data['pre_validation_failures']:
            f.write("\n" + "-" * 100 + "\n")
            f.write("  PRE-VALIDATION FAILURES\n")
            f.write("-" * 100 + "\n")
            f.write(f"{'Address':<20}{'Thread':<10}{'Expected':<20}{'Read':<20}{'Type':<30}\n")
            f.write("-" * 100 + "\n")
            
            for failure in report_data['pre_validation_failures'][:50]:
                f.write(f"{failure['address']:<20}{failure['thread']:<10}{failure['expected']:<20}"
                       f"{failure['read']:<20}{failure['type']:<30}\n")
            
            if len(report_data['pre_validation_failures']) > 50:
                f.write(f"\n  ... ({len(report_data['pre_validation_failures']) - 50} more failures not shown)\n")
        
        # Post-Validation Failures
        if report_data['post_validation_failures']:
            f.write("\n" + "-" * 100 + "\n")
            f.write("  POST-VALIDATION FAILURES (HARDWARE ISSUES!)\n")
            f.write("-" * 100 + "\n")
            f.write(f"{'Address':<20}{'Thread':<10}{'Expected':<20}{'Read':<20}{'Type':<30}\n")
            f.write("-" * 100 + "\n")
            
            for failure in report_data['post_validation_failures'][:50]:
                f.write(f"{failure['address']:<20}{failure['thread']:<10}{failure['expected']:<20}"
                       f"{failure['read']:<20}{failure['type']:<30}\n")
            
            if len(report_data['post_validation_failures']) > 50:
                f.write(f"\n  ... ({len(report_data['post_validation_failures']) - 50} more failures not shown)\n")
        
        # Last Writer Tracker (sample)
        if report_data['last_writer_tracker']:
            f.write("\n" + "-" * 100 + "\n")
            f.write("  LAST WRITER TRACKER (Sample - First 50 addresses)\n")
            f.write("-" * 100 + "\n")
            f.write(f"{'Address':<20}{'Thread':<10}{'Pattern':<20}{'Expected':<20}{'Iteration':<12}{'Operation':<10}\n")
            f.write("-" * 100 + "\n")
            
            for idx, (addr, info) in enumerate(report_data['last_writer_tracker'].items()):
                if idx >= 50:
                    break
                addr_str = hex(addr) if isinstance(addr, int) else addr
                f.write(f"{addr_str:<20}{info['thread']:<10}{hex(info['pattern'])[:18]:<20}"
                       f"{hex(info['expected'])[:18]:<20}{info['iteration']:<12}{info['operation']:<10}\n")
            
            if len(report_data['last_writer_tracker']) > 50:
                f.write(f"\n  ... ({len(report_data['last_writer_tracker']) - 50:,} more addresses not shown)\n")
        
        # Patterns Used
        f.write("\n" + "-" * 100 + "\n")
        f.write("  PATTERNS USED\n")
        f.write("-" * 100 + "\n")
        f.write(f"{'Index':<10}{'Hex Value':<20}{'Decimal Value':<25}\n")
        f.write("-" * 100 + "\n")
        
        for idx, pattern in enumerate(report_data['patterns']):
            f.write(f"{idx:<10}{hex(pattern):<20}{pattern:<25}\n")
        
        # Footer
        f.write("\n" + "=" * 100 + "\n")
        f.write("  END OF REPORT\n")
        f.write("=" * 100 + "\n\n")
        f.write("Legend:\n")
        f.write("  [OK]  PASS           - Operation succeeded, value matches expected\n")
        f.write("  [ERR] FAIL           - Operation failed, unexpected value (HARDWARE ISSUE!)\n")
        f.write("  [~]   RACE           - Read valid pattern but not from last writer (timing variance)\n")
        f.write("  [=]   RACE_EXPECTED  - Read correct last writer's value (expected race condition)\n")
        f.write("\nNote: View this file in Notepad with Consolas font for best formatting.\n")
    
    if verbose:
        print(f"\nComprehensive TXT report saved to: {filepath}")


def _write_txt_report(filepath, report_data, verbose=False):
    """Write comprehensive TXT report formatted for Consolas font in Notepad."""
    with open(filepath, 'w', encoding='utf-8') as f:
        # Header
        f.write("═" * 100 + "\n")
        f.write("  MEMORY STRESS TEST REPORT\n")
        f.write("═" * 100 + "\n")
        f.write(f"Timestamp: {report_data['timestamp']}\n")
        f.write("\n")
        
        # Configuration Section
        f.write("─" * 100 + "\n")
        f.write("  TEST CONFIGURATION\n")
        f.write("─" * 100 + "\n")
        config = report_data['config']
        for key, value in config.items():
            label = key.replace('_', ' ').title()
            f.write(f"  {label:.<35} {str(value):<60}\n")
        f.write("\n")
        
        # Results Section
        f.write("─" * 100 + "\n")
        f.write("  TEST RESULTS\n")
        f.write("─" * 100 + "\n")
        results = report_data['results']
        
        verdict = results['test_verdict']
        status_mark = "✓ PASS" if verdict.startswith('PASS') else "✗ FAIL"
        f.write(f"  Test Verdict:............................ {verdict} [{status_mark}]\n")
        f.write(f"  Elapsed Time:............................ {results['elapsed_time']:.2f} seconds\n")
        f.write(f"  Total Operations:........................ {results['total_operations']:,}\n")
        f.write(f"  Operations/Second:....................... {results['ops_per_second']:,.0f}\n")
        
        if config['verify']:
            f.write(f"\n  VERIFICATION STATISTICS:\n")
            f.write(f"    Passes:................................ {results['passes']:,}\n")
            
            if report_data['config']['track_last_writer'] and results['races_expected'] > 0:
                f.write(f"    Expected Races:........................ {results['races_expected']:,} (read correct last writer)\n")
            
            if results['races'] > 0:
                f.write(f"    Races (Unexpected Pattern):............ {results['races']:,}\n")
            
            f.write(f"    Failures:.............................. {results['failures']:,}")
            if results['failures'] == 0:
                f.write(" ✓\n")
            else:
                f.write(" ✗\n")
            
            total_checked = results['passes'] + results['failures'] + results['races'] + results['races_expected']
            if total_checked > 0:
                success_count = results['passes'] + results['races'] + results['races_expected']
                f.write(f"    Success Rate:.......................... {(success_count/total_checked*100):.2f}%\n")
        
        if config['restore_memory']:
            f.write(f"\n  Memory Restored:......................... Yes\n")
        
        # Validation Results
        if config['pre_post_validation']:
            f.write("\n")
            f.write("─" * 100 + "\n")
            f.write("  VALIDATION RESULTS\n")
            f.write("─" * 100 + "\n")
            
            pre_status = "✓ PASS" if results['pre_validation_failures'] == 0 else f"✗ FAIL ({results['pre_validation_failures']} errors)"
            post_status = "✓ PASS" if results['post_validation_failures'] == 0 else f"✗ FAIL ({results['post_validation_failures']} errors)"
            
            f.write(f"  Pre-Stress Validation:................... {pre_status}\n")
            f.write(f"  Post-Stress Validation:.................. {post_status}\n")
        
        # Failure Analysis
        if report_data['failure_analysis']:
            f.write("\n")
            f.write("─" * 100 + "\n")
            f.write("  FAILURE ANALYSIS\n")
            f.write("─" * 100 + "\n")
            fa = report_data['failure_analysis']
            
            if fa['hot_addresses']:
                f.write(f"  Hot Addresses (>10% failures):\n")
                for addr in fa['hot_addresses'][:5]:
                    f.write(f"    • {addr}\n")
            
            if fa['hot_threads']:
                f.write(f"  Hot Threads (>10% failures):\n")
                f.write(f"    Threads: {', '.join(str(t) for t in fa['hot_threads'][:10])}\n")
            
            f.write(f"\n  Top 5 Addresses by Failure Count:\n")
            for addr, count in fa['addr_distribution'].items():
                f.write(f"    {addr:.<50} {count:>5} failures\n")
        
        # Operation Log (sample)
        if report_data['log_data']:
            f.write("\n")
            f.write("─" * 100 + "\n")
            f.write("  OPERATION LOG (Sample - First 100 operations)\n")
            f.write("─" * 100 + "\n")
            f.write(f"{'Thread':<8}{'Iter':<6}{'Address':<20}{'Pattern':<20}{'Operation':<10}{'Expected':<20}{'Read':<20}{'Status':<15}\n")
            f.write("─" * 100 + "\n")
            
            for entry in report_data['log_data'][:100]:
                status_mark = {'PASS': '✓', 'FAIL': '✗', 'RACE': '~', 'RACE_EXPECTED': '≈'}.get(entry['status'], '?')
                f.write(f"{entry['thread']:<8}{entry['iteration']:<6}{entry['address']:<20}{entry['pattern'][:18]:<20}"
                       f"{entry['operation']:<10}{entry['expected'][:18]:<20}{entry['read_value'][:18]:<20}"
                       f"{status_mark} {entry['status']:<13}\n")
            
            if len(report_data['log_data']) > 100:
                f.write(f"\n  ... ({len(report_data['log_data']) - 100:,} more operations not shown)\n")
        
        # Pre-Validation Failures
        if report_data['pre_validation_failures']:
            f.write("\n")
            f.write("─" * 100 + "\n")
            f.write("  PRE-VALIDATION FAILURES\n")
            f.write("─" * 100 + "\n")
            f.write(f"{'Address':<20}{'Thread':<10}{'Expected':<20}{'Read':<20}{'Type':<30}\n")
            f.write("─" * 100 + "\n")
            
            for failure in report_data['pre_validation_failures'][:50]:
                f.write(f"{failure['address']:<20}{failure['thread']:<10}{failure['expected']:<20}"
                       f"{failure['read']:<20}{failure['type']:<30}\n")
            
            if len(report_data['pre_validation_failures']) > 50:
                f.write(f"\n  ... ({len(report_data['pre_validation_failures']) - 50} more failures not shown)\n")
        
        # Post-Validation Failures
        if report_data['post_validation_failures']:
            f.write("\n")
            f.write("─" * 100 + "\n")
            f.write("  POST-VALIDATION FAILURES (HARDWARE ISSUES!)\n")
            f.write("─" * 100 + "\n")
            f.write(f"{'Address':<20}{'Thread':<10}{'Expected':<20}{'Read':<20}{'Type':<30}\n")
            f.write("─" * 100 + "\n")
            
            for failure in report_data['post_validation_failures'][:50]:
                f.write(f"{failure['address']:<20}{failure['thread']:<10}{failure['expected']:<20}"
                       f"{failure['read']:<20}{failure['type']:<30}\n")
            
            if len(report_data['post_validation_failures']) > 50:
                f.write(f"\n  ... ({len(report_data['post_validation_failures']) - 50} more failures not shown)\n")
        
        # Last Writer Tracker (sample)
        if report_data['last_writer_tracker']:
            f.write("\n")
            f.write("─" * 100 + "\n")
            f.write("  LAST WRITER TRACKER (Sample - First 50 addresses)\n")
            f.write("─" * 100 + "\n")
            f.write(f"{'Address':<20}{'Thread':<10}{'Pattern':<20}{'Expected':<20}{'Iteration':<12}{'Operation':<10}\n")
            f.write("─" * 100 + "\n")
            
            for idx, (addr, info) in enumerate(report_data['last_writer_tracker'].items()):
                if idx >= 50:
                    break
                addr_str = hex(addr) if isinstance(addr, int) else addr
                f.write(f"{addr_str:<20}{info['thread']:<10}{hex(info['pattern'])[:18]:<20}"
                       f"{hex(info['expected'])[:18]:<20}{info['iteration']:<12}{info['operation']:<10}\n")
            
            if len(report_data['last_writer_tracker']) > 50:
                f.write(f"\n  ... ({len(report_data['last_writer_tracker']) - 50:,} more addresses not shown)\n")
        
        # Patterns Used
        f.write("\n")
        f.write("─" * 100 + "\n")
        f.write("  PATTERNS USED\n")
        f.write("─" * 100 + "\n")
        f.write(f"{'Index':<10}{'Hex Value':<20}{'Decimal Value':<25}\n")
        f.write("─" * 100 + "\n")
        
        for idx, pattern in enumerate(report_data['patterns']):
            f.write(f"{idx:<10}{hex(pattern):<20}{pattern:<25}\n")
        
        # Footer
        f.write("\n")
        f.write("═" * 100 + "\n")
        f.write("  END OF REPORT\n")
        f.write("═" * 100 + "\n")
        f.write("\n")
        f.write("Legend:\n")
        f.write("  ✓ PASS           - Operation succeeded, value matches expected\n")
        f.write("  ✗ FAIL           - Operation failed, unexpected value (HARDWARE ISSUE!)\n")
        f.write("  ~ RACE           - Read valid pattern but not from last writer (timing variance)\n")
        f.write("  ≈ RACE_EXPECTED  - Read correct last writer's value (expected race condition)\n")
        f.write("\n")
        f.write("Note: View this file in Notepad with Consolas font for best formatting.\n")
    
    if verbose:
        print(f"\nComprehensive TXT report saved to: {filepath}")


def _write_excel_report(filepath, report_data, verbose=False):
    """Write comprehensive Excel report with multiple sheets."""
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    subheader_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    subheader_font = Font(bold=True)
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    race_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: Summary
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 40
    
    row = 1
    ws_summary[f'A{row}'] = "MEMORY STRESS TEST REPORT"
    ws_summary[f'A{row}'].font = Font(bold=True, size=14)
    row += 1
    
    ws_summary[f'A{row}'] = "Timestamp:"
    ws_summary[f'B{row}'] = report_data['timestamp']
    row += 2
    
    # Configuration Section
    ws_summary[f'A{row}'] = "TEST CONFIGURATION"
    ws_summary[f'A{row}'].font = subheader_font
    ws_summary[f'A{row}'].fill = subheader_fill
    row += 1
    
    config = report_data['config']
    for key, value in config.items():
        ws_summary[f'A{row}'] = key.replace('_', ' ').title()
        ws_summary[f'B{row}'] = str(value)
        row += 1
    
    row += 1
    
    # Results Section
    ws_summary[f'A{row}'] = "TEST RESULTS"
    ws_summary[f'A{row}'].font = subheader_font
    ws_summary[f'A{row}'].fill = subheader_fill
    row += 1
    
    results = report_data['results']
    
    # Format results with colors
    ws_summary[f'A{row}'] = "Test Verdict"
    ws_summary[f'B{row}'] = results['test_verdict']
    if results['test_verdict'].startswith('PASS'):
        ws_summary[f'B{row}'].fill = pass_fill
    elif results['test_verdict'].startswith('FAIL'):
        ws_summary[f'B{row}'].fill = fail_fill
    row += 1
    
    ws_summary[f'A{row}'] = "Elapsed Time (seconds)"
    ws_summary[f'B{row}'] = f"{results['elapsed_time']:.2f}"
    row += 1
    
    ws_summary[f'A{row}'] = "Total Operations"
    ws_summary[f'B{row}'] = f"{results['total_operations']:,}"
    row += 1
    
    ws_summary[f'A{row}'] = "Operations/Second"
    ws_summary[f'B{row}'] = f"{results['ops_per_second']:,.0f}"
    row += 1
    
    if config['verify']:
        ws_summary[f'A{row}'] = "Passes"
        ws_summary[f'B{row}'] = f"{results['passes']:,}"
        ws_summary[f'B{row}'].fill = pass_fill if results['passes'] > 0 else None
        row += 1
        
        if report_data['config']['track_last_writer'] and results['races_expected'] > 0:
            ws_summary[f'A{row}'] = "Expected Races"
            ws_summary[f'B{row}'] = f"{results['races_expected']:,}"
            row += 1
        
        if results['races'] > 0:
            ws_summary[f'A{row}'] = "Races (Unexpected Pattern)"
            ws_summary[f'B{row}'] = f"{results['races']:,}"
            ws_summary[f'B{row}'].fill = race_fill
            row += 1
        
        ws_summary[f'A{row}'] = "Failures"
        ws_summary[f'B{row}'] = f"{results['failures']:,}"
        ws_summary[f'B{row}'].fill = fail_fill if results['failures'] > 0 else pass_fill
        row += 1
        
        total_checked = results['passes'] + results['failures'] + results['races'] + results['races_expected']
        if total_checked > 0:
            success_count = results['passes'] + results['races'] + results['races_expected']
            ws_summary[f'A{row}'] = "Success Rate"
            ws_summary[f'B{row}'] = f"{(success_count/total_checked*100):.2f}%"
            row += 1
    
    if config['pre_post_validation']:
        row += 1
        ws_summary[f'A{row}'] = "VALIDATION RESULTS"
        ws_summary[f'A{row}'].font = subheader_font
        ws_summary[f'A{row}'].fill = subheader_fill
        row += 1
        
        ws_summary[f'A{row}'] = "Pre-Validation"
        ws_summary[f'B{row}'] = f"PASS" if results['pre_validation_failures'] == 0 else f"FAIL ({results['pre_validation_failures']} errors)"
        ws_summary[f'B{row}'].fill = pass_fill if results['pre_validation_failures'] == 0 else fail_fill
        row += 1
        
        ws_summary[f'A{row}'] = "Post-Validation"
        ws_summary[f'B{row}'] = f"PASS" if results['post_validation_failures'] == 0 else f"FAIL ({results['post_validation_failures']} errors)"
        ws_summary[f'B{row}'].fill = pass_fill if results['post_validation_failures'] == 0 else fail_fill
        row += 1
    
    # Failure Analysis
    if report_data['failure_analysis']:
        row += 1
        ws_summary[f'A{row}'] = "FAILURE ANALYSIS"
        ws_summary[f'A{row}'].font = subheader_font
        ws_summary[f'A{row}'].fill = subheader_fill
        row += 1
        
        fa = report_data['failure_analysis']
        
        if fa['hot_addresses']:
            ws_summary[f'A{row}'] = "Hot Addresses (>10% failures)"
            ws_summary[f'B{row}'] = ', '.join(fa['hot_addresses'][:5])
            row += 1
        
        if fa['hot_threads']:
            ws_summary[f'A{row}'] = "Hot Threads (>10% failures)"
            ws_summary[f'B{row}'] = ', '.join(str(t) for t in fa['hot_threads'][:10])
            row += 1
        
        ws_summary[f'A{row}'] = "Top 5 Addresses by Failures"
        row += 1
        for addr, count in fa['addr_distribution'].items():
            ws_summary[f'A{row}'] = f"  {addr}"
            ws_summary[f'B{row}'] = count
            row += 1
    
    # Sheet 2: Operation Log
    if report_data['log_data']:
        ws_log = wb.create_sheet("Operation Log", 1)
        headers = ['Thread', 'Iteration', 'Address', 'Original Value', 'Pattern', 
                  'Operation', 'Expected', 'Read Value', 'Status']
        
        for col, header in enumerate(headers, 1):
            cell = ws_log.cell(1, col, header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, entry in enumerate(report_data['log_data'], 2):
            ws_log.cell(row_idx, 1, entry['thread'])
            ws_log.cell(row_idx, 2, entry['iteration'])
            ws_log.cell(row_idx, 3, entry['address'])
            ws_log.cell(row_idx, 4, entry['original_value'])
            ws_log.cell(row_idx, 5, entry['pattern'])
            ws_log.cell(row_idx, 6, entry['operation'])
            ws_log.cell(row_idx, 7, entry['expected'])
            ws_log.cell(row_idx, 8, entry['read_value'])
            
            status_cell = ws_log.cell(row_idx, 9, entry['status'])
            if entry['status'] == 'PASS':
                status_cell.fill = pass_fill
            elif entry['status'] == 'FAIL':
                status_cell.fill = fail_fill
            elif 'RACE' in entry['status']:
                status_cell.fill = race_fill
        
        # Auto-adjust column widths
        for col in range(1, 10):
            ws_log.column_dimensions[get_column_letter(col)].width = 18
    
    # Sheet 3: Pre-Validation Failures
    if report_data['pre_validation_failures']:
        ws_pre = wb.create_sheet("Pre-Validation Failures", 2)
        headers = ['Address', 'Thread', 'Expected', 'Read', 'Type']
        
        for col, header in enumerate(headers, 1):
            cell = ws_pre.cell(1, col, header)
            cell.font = header_font
            cell.fill = header_fill
        
        for row_idx, failure in enumerate(report_data['pre_validation_failures'], 2):
            ws_pre.cell(row_idx, 1, failure['address'])
            ws_pre.cell(row_idx, 2, failure['thread'])
            ws_pre.cell(row_idx, 3, failure['expected'])
            ws_pre.cell(row_idx, 4, failure['read'])
            ws_pre.cell(row_idx, 5, failure['type'])
    
    # Sheet 4: Post-Validation Failures
    if report_data['post_validation_failures']:
        ws_post = wb.create_sheet("Post-Validation Failures", 3)
        headers = ['Address', 'Thread', 'Expected', 'Read', 'Type']
        
        for col, header in enumerate(headers, 1):
            cell = ws_post.cell(1, col, header)
            cell.font = header_font
            cell.fill = header_fill
        
        for row_idx, failure in enumerate(report_data['post_validation_failures'], 2):
            ws_post.cell(row_idx, 1, failure['address'])
            ws_post.cell(row_idx, 2, failure['thread'])
            ws_post.cell(row_idx, 3, failure['expected'])
            ws_post.cell(row_idx, 4, failure['read'])
            ws_post.cell(row_idx, 5, failure['type'])
    
    # Sheet 5: Last Writer Tracker
    if report_data['last_writer_tracker']:
        ws_lw = wb.create_sheet("Last Writer Tracker", 4)
        headers = ['Address', 'Thread', 'Pattern', 'Expected', 'Iteration', 'Operation']
        
        for col, header in enumerate(headers, 1):
            cell = ws_lw.cell(1, col, header)
            cell.font = header_font
            cell.fill = header_fill
        
        for row_idx, (addr, info) in enumerate(report_data['last_writer_tracker'].items(), 2):
            ws_lw.cell(row_idx, 1, hex(addr) if isinstance(addr, int) else addr)
            ws_lw.cell(row_idx, 2, info['thread'])
            ws_lw.cell(row_idx, 3, hex(info['pattern']))
            ws_lw.cell(row_idx, 4, hex(info['expected']))
            ws_lw.cell(row_idx, 5, info['iteration'])
            ws_lw.cell(row_idx, 6, info['operation'])
        
        for col in range(1, 7):
            ws_lw.column_dimensions[get_column_letter(col)].width = 20
    
    # Sheet 6: Patterns Used
    ws_patterns = wb.create_sheet("Patterns", 5)
    ws_patterns['A1'] = "Pattern Index"
    ws_patterns['B1'] = "Hex Value"
    ws_patterns['C1'] = "Decimal Value"
    
    for i, cell in enumerate([ws_patterns['A1'], ws_patterns['B1'], ws_patterns['C1']]):
        cell.font = header_font
        cell.fill = header_fill
    
    for idx, pattern in enumerate(report_data['patterns'], 2):
        ws_patterns[f'A{idx}'] = idx - 2
        ws_patterns[f'B{idx}'] = hex(pattern)
        ws_patterns[f'C{idx}'] = pattern
    
    wb.save(filepath)
    
    if verbose:
        print(f"\nComprehensive Excel report saved to: {filepath}")
        print(f"  Sheets: Summary, Operation Log, ", end="")
        sheet_count = 2
        if report_data['pre_validation_failures']:
            print("Pre-Validation, ", end="")
            sheet_count += 1
        if report_data['post_validation_failures']:
            print("Post-Validation, ", end="")
            sheet_count += 1
        if report_data['last_writer_tracker']:
            print("Last Writer Tracker, ", end="")
            sheet_count += 1
        print("Patterns")


def print_operations_help():
    """Print available operations and their descriptions."""
    print("\nAvailable Memory Operations:")
    print("=" * 70)
    for op, desc in OPERATIONS.items():
        print(f"  {op:20s} - {desc}")
    print("=" * 70)
    print("\nUsage examples:")
    print("  write        : Direct pattern write (fastest)")
    print("  add          : Adds pattern to existing value (tests ALU)")
    print("  xor          : XORs pattern with existing value (tests logic)")
    print("  multiply     : Multiplies existing value (tests multiply units)")
    print("  rotate       : Rotates bits (tests shifter)")
    print("  atomic_add   : Atomic addition (tests cache coherency)")
    print("  atomic_xchg  : Atomic exchange (tests synchronization)")
    print("  atomic_cmpxchg: Compare-and-swap (tests lock-free algorithms)")
    print()


if __name__ == "__main__":
    # Uncomment to see available operations
    # print_operations_help()
    
    main()
    # Example 3: Force memory traffic with cache flushes
    # run_memory_stress(
    #     addr_start=0x1F0EAE6980,
    #     addr_end=0x1FB4A7D600,
    #     addr_step=0x40,
    #     iterations=50,
    #     use_wbinvd=True,
    #     verify=True
    # )
    
    # Example 4: Custom patterns
    # custom_patterns = [
    #     0xDEADBEEFDEADBEEF,
    #     0xCAFEBABECAFEBABE,
    #     0x1234567812345678,
    # ]
    # run_memory_stress(
    #     addr_start=0x1F0EAE6980,
    #     addr_end=0x1FB4A7D600,
    #     patterns=custom_patterns,
    #     iterations=100
    # )

